import os, re, io, sqlite3, json
from datetime import datetime, date, timedelta
from flask import Flask, abort, g, jsonify, render_template, request, send_file, send_from_directory, session, redirect
from werkzeug.security import generate_password_hash, check_password_hash
from assignment_helpers import (
    assign_employees_to_task, assign_employees_to_issue,
    get_task_assignees, get_issue_assignees,
    get_employee_tasks, get_employee_issues
)

# Crew Control System API
try:
    from crew_api import crew_bp
    CREW_API_AVAILABLE = True
except ImportError:
    CREW_API_AVAILABLE = False
    print("[INFO] Crew API module not available")

# Database path configuration
# Priority: 1. DB_PATH env var, 2. Persistent disk detection, 3. Default
if os.environ.get("DB_PATH"):
    # User explicitly set DB_PATH - use it
    DB_PATH = os.environ.get("DB_PATH")
elif os.environ.get("RENDER") or os.environ.get("RENDER_EXTERNAL_HOSTNAME"):
    # Render platform detected - try persistent disk paths
    if os.path.exists("/persistent"):
        DB_PATH = "/persistent/app.db"
    elif os.path.exists("/data"):
        DB_PATH = "/data/app.db"
    else:
        # Fallback to /tmp which is persistent on Render
        DB_PATH = "/tmp/app.db"
else:
    # Local development
    DB_PATH = "app.db"
SECRET_KEY = os.environ.get("SECRET_KEY", "dev-" + os.urandom(16).hex())
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "uploads")

app = Flask(__name__, static_folder=".", static_url_path="")
# Disable aggressive caching in development so UI settings (language/theme) apply immediately
app.config["SEND_FILE_MAX_AGE_DEFAULT"] = 0

# Register Crew Control System API blueprint
if CREW_API_AVAILABLE:
    app.register_blueprint(crew_bp)
    print("[INFO] Crew Control System API registered")

# Register Task API blueprints
try:
    import sys
    import os
    import importlib.util
    
    # Import blueprints directly without going through app module
    _tasks_path = os.path.join(os.path.dirname(__file__), 'app', 'routes', 'api', 'tasks.py')
    if os.path.exists(_tasks_path):
        spec = importlib.util.spec_from_file_location("tasks_api", _tasks_path)
        tasks_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(tasks_module)
        tasks_bp = tasks_module.tasks_bp
        app.register_blueprint(tasks_bp)
    
    _evidence_path = os.path.join(os.path.dirname(__file__), 'app', 'routes', 'api', 'task_evidence.py')
    if os.path.exists(_evidence_path):
        spec = importlib.util.spec_from_file_location("evidence_api", _evidence_path)
        evidence_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(evidence_module)
        evidence_bp = evidence_module.evidence_bp
        app.register_blueprint(evidence_bp)
    
    _sync_path = os.path.join(os.path.dirname(__file__), 'app', 'routes', 'api', 'task_sync.py')
    if os.path.exists(_sync_path):
        spec = importlib.util.spec_from_file_location("sync_api", _sync_path)
        sync_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(sync_module)
        sync_bp = sync_module.sync_bp
        app.register_blueprint(sync_bp)
    
    _audit_path = os.path.join(os.path.dirname(__file__), 'app', 'middleware', 'audit.py')
    if os.path.exists(_audit_path):
        spec = importlib.util.spec_from_file_location("audit", _audit_path)
        audit_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(audit_module)
        audit_module.setup_audit_middleware(app)
    
    print("[INFO] Task API routes registered")
except Exception as e:
    print(f"[WARNING] Task API routes not available: {e}")
    import traceback
    traceback.print_exc()

# Register AI Insights API blueprint
try:
    import sys
    import os
    import importlib.util
    
    _ai_insights_path = os.path.join(os.path.dirname(__file__), 'app', 'routes', 'api', 'ai_insights.py')
    if os.path.exists(_ai_insights_path):
        spec = importlib.util.spec_from_file_location("ai_insights", _ai_insights_path)
        ai_insights_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(ai_insights_module)
        ai_bp = ai_insights_module.ai_bp
        app.register_blueprint(ai_bp)
        print("[INFO] AI Insights API routes registered")
except Exception as e:
    print(f"[WARNING] AI Insights API routes not available: {e}")
    import traceback
    traceback.print_exc()

# Register Inventory API blueprint
try:
    import sys
    import os
    import importlib.util
    
    _inventory_path = os.path.join(os.path.dirname(__file__), 'app', 'routes', 'api', 'inventory.py')
    if os.path.exists(_inventory_path):
        spec = importlib.util.spec_from_file_location("inventory_api", _inventory_path)
        inventory_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(inventory_module)
        inventory_bp = inventory_module.inventory_bp
        app.register_blueprint(inventory_bp)
        print("[INFO] Inventory API routes registered")
except Exception as e:
    print(f"[WARNING] Inventory API routes not available: {e}")
    import traceback
    traceback.print_exc()

# Register QR API blueprint
try:
    import sys
    import os
    import importlib.util
    
    _qr_path = os.path.join(os.path.dirname(__file__), 'app', 'routes', 'api', 'qr.py')
    if os.path.exists(_qr_path):
        spec = importlib.util.spec_from_file_location("qr_api", _qr_path)
        qr_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(qr_module)
        qr_bp = qr_module.qr_bp
        app.register_blueprint(qr_bp)
        print("[INFO] QR API routes registered")
except Exception as e:
    print(f"[WARNING] QR API routes not available: {e}")
    import traceback
    traceback.print_exc()

# Register Permissions context processor
try:
    from app.utils.permissions import inject_permissions
    app.context_processor(inject_permissions)
    print("[INFO] Permissions context processor registered")
except Exception as e:
    print(f"[WARNING] Permissions context processor not available: {e}")
    import traceback
    traceback.print_exc()

# Register Asset Insights API blueprint
try:
    import sys
    import os
    import importlib.util
    
    _asset_insights_path = os.path.join(os.path.dirname(__file__), 'app', 'routes', 'api', 'asset_insights.py')
    if os.path.exists(_asset_insights_path):
        spec = importlib.util.spec_from_file_location("asset_insights_api", _asset_insights_path)
        asset_insights_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(asset_insights_module)
        asset_insights_bp = asset_insights_module.insights_bp
        app.register_blueprint(asset_insights_bp)
        print("[INFO] Asset Insights API routes registered")
except Exception as e:
    print(f"[WARNING] Asset Insights API routes not available: {e}")
    import traceback
    traceback.print_exc()

@app.after_request
def _disable_cache_for_static(resp):
    try:
        path = request.path or ""
        if path.startswith("/static/") or path.endswith(".js") or path.endswith(".css"):
            resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            resp.headers["Pragma"] = "no-cache"
            resp.headers["Expires"] = "0"
    except Exception:
        pass
    return resp

app.secret_key = SECRET_KEY

# ----------------- Role definitions -----------------
# Pozn.: role 'team_lead' byla v předchozích verzích; pro kompatibilitu ji mapujeme na 'lander'.
ROLES = ("owner", "admin", "manager", "lander", "worker")
WRITE_ROLES = ("owner", "admin", "manager", "lander")

# Role pro zaměstnance v UI (enterprise: jednotné a jednoduché)
EMPLOYEE_ROLES = ("owner", "manager", "lander", "worker")

def normalize_role(role):
    """Normalizace role pro zpětnou kompatibilitu a konzistentní autorizaci."""
    if not role:
        return "owner"
    role = str(role).strip().lower()
    if role == "team_lead":
        return "lander"
    return role


def normalize_employee_role(role: str) -> str:
    """Normalizace role zaměstnance.

    - zachová kompatibilitu se staršími hodnotami (admin -> owner, team_lead -> lander)
    - pokud je hodnota neznámá, vrací 'worker'
    """
    if not role:
        return "worker"
    r = str(role).strip().lower()
    if r in ("admin",):
        r = "owner"
    if r == "team_lead":
        r = "lander"
    return r if r in EMPLOYEE_ROLES else "worker"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# ----------------- Database utilities -----------------
def get_db():
    if "db" not in g:
        # Ensure directory exists for DB_PATH
        db_dir = os.path.dirname(DB_PATH)
        if db_dir and not os.path.exists(db_dir):
            try:
                os.makedirs(db_dir, exist_ok=True)
                print(f"[DB] Created directory: {db_dir}")
            except Exception as e:
                print(f"[DB] Warning: Could not create directory {db_dir}: {e}")
        
        # Log database path (only once at startup)
        if not hasattr(get_db, '_logged'):
            print(f"[DB] Using database: {DB_PATH}")
            get_db._logged = True
        
        # Connect with WAL mode for better concurrency
        # Use a small timeout to reduce 'database is locked' errors under concurrent requests
        g.db = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=5)
        g.db.row_factory = sqlite3.Row
        # Enable WAL mode for better performance and durability
        try:
            g.db.execute("PRAGMA journal_mode=WAL")
        except Exception:
            pass
    return g.db

def _table_has_column(db, table: str, column: str) -> bool:
    try:
        rows = db.execute(f"PRAGMA table_info({table})").fetchall()
        return any(r[1] == column for r in rows)
    except Exception:
        return False


def _table_exists(db, table: str) -> bool:
    """Return True if a table exists in the current SQLite database."""
    try:
        r = db.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1",
            (table,),
        ).fetchone()
        return bool(r)
    except Exception:
        return False


@app.teardown_appcontext
def close_db(exception=None):
    """Close the database connection at the end of the request."""
    db = g.pop('db', None)
    if db is not None:
        try:
            db.close()
        except Exception:
            pass

def apply_migrations():
    """Lightweight, non-breaking migration runner.

    Tracks applied versions in schema_migrations and applies idempotent ALTERs when needed.
    """
    db = get_db()
    db.execute("CREATE TABLE IF NOT EXISTS schema_migrations (version INTEGER PRIMARY KEY, applied_at TEXT NOT NULL DEFAULT (datetime('now')))")
    applied = {r[0] for r in db.execute("SELECT version FROM schema_migrations").fetchall()}

    migrations = [
        # v1: baseline (existing ensure_schema creates core tables)
        (1, []),

        # v2: employees extra contact fields (safe idempotent)
        (2, [
            ("employees", "phone",   "ALTER TABLE employees ADD COLUMN phone TEXT DEFAULT ''"),
            ("employees", "email",   "ALTER TABLE employees ADD COLUMN email TEXT DEFAULT ''"),
            ("employees", "address", "ALTER TABLE employees ADD COLUMN address TEXT DEFAULT ''"),
        ]),

        # v3: core search stability (jobs/tasks/issues timestamps + assignment tables)
        (3, [
            ("jobs", "created_at", "ALTER TABLE jobs ADD COLUMN created_at TEXT NOT NULL DEFAULT (datetime('now'))"),
            ("tasks", "created_at", "ALTER TABLE tasks ADD COLUMN created_at TEXT NOT NULL DEFAULT (datetime('now'))"),
            """
            CREATE TABLE IF NOT EXISTS task_assignments (
                task_id INTEGER NOT NULL,
                employee_id INTEGER NOT NULL,
                is_primary INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (task_id, employee_id)
            );
            CREATE TABLE IF NOT EXISTS issues (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id INTEGER,
                title TEXT NOT NULL,
                description TEXT DEFAULT '',
                type TEXT DEFAULT 'issue',
                status TEXT NOT NULL DEFAULT 'open',
                severity TEXT DEFAULT '',
                assigned_to INTEGER,
                created_by INTEGER,
                created_at TEXT NOT NULL DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS issue_assignments (
                issue_id INTEGER NOT NULL,
                employee_id INTEGER NOT NULL,
                is_primary INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (issue_id, employee_id)
            );
            """,
        ]),

        # v4: assignment tables primary flag (backward compatibility)
        (4, [
            ("task_assignments", "is_primary", "ALTER TABLE task_assignments ADD COLUMN is_primary INTEGER NOT NULL DEFAULT 0"),
            ("issue_assignments", "is_primary", "ALTER TABLE issue_assignments ADD COLUMN is_primary INTEGER NOT NULL DEFAULT 0"),
        ]),

        # v5: notifications (safe create)
        (5, [
            """
            CREATE TABLE IF NOT EXISTS notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                employee_id INTEGER,
                kind TEXT NOT NULL DEFAULT 'info',
                title TEXT NOT NULL DEFAULT '',
                body TEXT NOT NULL DEFAULT '',
                entity_type TEXT,
                entity_id INTEGER,
                is_read INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT (datetime('now'))
            );
            CREATE INDEX IF NOT EXISTS idx_notifications_user_read ON notifications(user_id, is_read, created_at);
            CREATE INDEX IF NOT EXISTS idx_notifications_emp_read ON notifications(employee_id, is_read, created_at);
            """,
        ]),

        # v6: tasks extra columns (deadline, depends_on, due_date)
        (6, [
            ("tasks", "deadline", "ALTER TABLE tasks ADD COLUMN deadline TEXT"),
            ("tasks", "depends_on", "ALTER TABLE tasks ADD COLUMN depends_on TEXT"),
            ("tasks", "due_date", "ALTER TABLE tasks ADD COLUMN due_date TEXT"),
        ]),

        # v7: jobs extra columns (invoiced, address, budget)
        (7, [
            ("jobs", "invoiced", "ALTER TABLE jobs ADD COLUMN invoiced INTEGER DEFAULT 0"),
            ("jobs", "address", "ALTER TABLE jobs ADD COLUMN address TEXT DEFAULT ''"),
            ("jobs", "budget", "ALTER TABLE jobs ADD COLUMN budget REAL DEFAULT 0"),
        ]),

        # v8: planning_assignments table
        (8, [
            """
            CREATE TABLE IF NOT EXISTS planning_assignments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                job_id INTEGER,
                task_id INTEGER,
                date TEXT NOT NULL,
                start_time TEXT,
                end_time TEXT,
                hours REAL DEFAULT 8,
                note TEXT DEFAULT '',
                status TEXT DEFAULT 'planned',
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (employee_id) REFERENCES employees(id),
                FOREIGN KEY (job_id) REFERENCES jobs(id),
                FOREIGN KEY (task_id) REFERENCES tasks(id)
            );
            CREATE INDEX IF NOT EXISTS idx_planning_date ON planning_assignments(date);
            CREATE INDEX IF NOT EXISTS idx_planning_employee ON planning_assignments(employee_id);
            """,
        ]),

        # v9: inventory/warehouse table
        (9, [
            """
            CREATE TABLE IF NOT EXISTS inventory (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                sku TEXT,
                category TEXT DEFAULT 'general',
                quantity REAL DEFAULT 0,
                unit TEXT DEFAULT 'ks',
                unit_price REAL DEFAULT 0,
                min_quantity REAL DEFAULT 0,
                location TEXT DEFAULT '',
                supplier TEXT DEFAULT '',
                note TEXT DEFAULT '',
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                updated_at TEXT NOT NULL DEFAULT (datetime('now'))
            );
            CREATE INDEX IF NOT EXISTS idx_inventory_category ON inventory(category);
            CREATE INDEX IF NOT EXISTS idx_inventory_sku ON inventory(sku);
            """,
        ]),

        # v10: attachments table
        (10, [
            """
            CREATE TABLE IF NOT EXISTS attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entity_type TEXT NOT NULL,
                entity_id INTEGER NOT NULL,
                filename TEXT NOT NULL,
                filepath TEXT NOT NULL,
                filesize INTEGER DEFAULT 0,
                mimetype TEXT DEFAULT '',
                uploaded_by INTEGER,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (uploaded_by) REFERENCES users(id)
            );
            CREATE INDEX IF NOT EXISTS idx_attachments_entity ON attachments(entity_type, entity_id);
            """,
        ]),

        # v11: warehouse_items table (alternative naming)
        (11, [
            """
            CREATE TABLE IF NOT EXISTS warehouse_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                sku TEXT,
                category TEXT DEFAULT 'general',
                quantity REAL DEFAULT 0,
                unit TEXT DEFAULT 'ks',
                unit_price REAL DEFAULT 0,
                min_quantity REAL DEFAULT 0,
                location TEXT DEFAULT '',
                supplier TEXT DEFAULT '',
                note TEXT DEFAULT '',
                job_id INTEGER,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                updated_at TEXT NOT NULL DEFAULT (datetime('now'))
            );
            """,
        ]),

        # v12: job_plan_proposals table (Ghost plán)
        (12, [
            """
            CREATE TABLE IF NOT EXISTS job_plan_proposals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id INTEGER NOT NULL,
                proposal_type TEXT NOT NULL DEFAULT 'ghost',
                title TEXT NOT NULL,
                description TEXT DEFAULT '',
                proposed_timeline JSON,
                proposed_resources JSON,
                proposed_budget REAL DEFAULT 0,
                risk_score INTEGER DEFAULT 0,
                health_score INTEGER DEFAULT 0,
                status TEXT NOT NULL DEFAULT 'pending',
                created_by INTEGER,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                reviewed_at TEXT,
                reviewed_by INTEGER,
                FOREIGN KEY (job_id) REFERENCES jobs(id),
                FOREIGN KEY (created_by) REFERENCES users(id),
                FOREIGN KEY (reviewed_by) REFERENCES users(id)
            );
            CREATE INDEX IF NOT EXISTS idx_job_plan_proposals_job ON job_plan_proposals(job_id, status);
            """,
        ]),

        # v13: timesheets extended (work logs as data package)
        (13, [
            ("timesheets", "user_id", "ALTER TABLE timesheets ADD COLUMN user_id INTEGER NULL"),
            ("timesheets", "duration_minutes", "ALTER TABLE timesheets ADD COLUMN duration_minutes INTEGER NULL"),
            ("timesheets", "work_type", "ALTER TABLE timesheets ADD COLUMN work_type TEXT DEFAULT 'manual'"),
            ("timesheets", "start_time", "ALTER TABLE timesheets ADD COLUMN start_time TEXT NULL"),
            ("timesheets", "end_time", "ALTER TABLE timesheets ADD COLUMN end_time TEXT NULL"),
            ("timesheets", "location", "ALTER TABLE timesheets ADD COLUMN location TEXT NULL"),
            ("timesheets", "task_id", "ALTER TABLE timesheets ADD COLUMN task_id INTEGER NULL"),
            ("timesheets", "material_used", "ALTER TABLE timesheets ADD COLUMN material_used TEXT NULL"),
            ("timesheets", "weather_snapshot", "ALTER TABLE timesheets ADD COLUMN weather_snapshot TEXT NULL"),
            ("timesheets", "performance_signal", "ALTER TABLE timesheets ADD COLUMN performance_signal TEXT DEFAULT 'normal'"),
            ("timesheets", "delay_reason", "ALTER TABLE timesheets ADD COLUMN delay_reason TEXT NULL"),
            ("timesheets", "photo_url", "ALTER TABLE timesheets ADD COLUMN photo_url TEXT NULL"),
            ("timesheets", "note", "ALTER TABLE timesheets ADD COLUMN note TEXT NULL"),
            ("timesheets", "ai_flags", "ALTER TABLE timesheets ADD COLUMN ai_flags TEXT NULL"),
            ("timesheets", "created_at", "ALTER TABLE timesheets ADD COLUMN created_at TEXT NOT NULL DEFAULT (datetime('now'))"),
            ("timesheets", "labor_cost", "ALTER TABLE timesheets ADD COLUMN labor_cost REAL NULL"),
            """
            -- Migrate existing hours to duration_minutes
            UPDATE timesheets SET duration_minutes = CAST(hours * 60 AS INTEGER) WHERE duration_minutes IS NULL AND hours IS NOT NULL;
            
            -- Migrate place to location
            UPDATE timesheets SET location = place WHERE location IS NULL AND place IS NOT NULL AND place != '';
            
            -- Migrate activity to note
            UPDATE timesheets SET note = activity WHERE note IS NULL AND activity IS NOT NULL AND activity != '';
            
            -- Create indexes for performance
            CREATE INDEX IF NOT EXISTS idx_timesheets_user_date ON timesheets(user_id, date);
            CREATE INDEX IF NOT EXISTS idx_timesheets_job_date ON timesheets(job_id, date);
            CREATE INDEX IF NOT EXISTS idx_timesheets_task ON timesheets(task_id);
            CREATE INDEX IF NOT EXISTS idx_timesheets_employee_date ON timesheets(employee_id, date);
            """,
        ]),

        # v14: timesheets delay_note (doplněk k delay_reason)
        (14, [
            ("timesheets", "delay_note", "ALTER TABLE timesheets ADD COLUMN delay_note TEXT NULL"),
        ]),

        # v15: trainings module (školení a vzdělávání)
        (15, [
            """
            CREATE TABLE IF NOT EXISTS trainings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                description TEXT,
                training_type TEXT DEFAULT 'external',
                category TEXT,
                provider TEXT,
                provider_type TEXT,
                date_start TEXT NOT NULL,
                date_end TEXT,
                duration_hours REAL,
                is_paid INTEGER DEFAULT 1,
                cost_training REAL DEFAULT 0,
                cost_travel REAL DEFAULT 0,
                cost_accommodation REAL DEFAULT 0,
                cost_meals REAL DEFAULT 0,
                cost_other REAL DEFAULT 0,
                cost_total REAL DEFAULT 0,
                cost_opportunity REAL DEFAULT 0,
                location TEXT,
                is_remote INTEGER DEFAULT 0,
                has_certificate INTEGER DEFAULT 0,
                certificate_name TEXT,
                certificate_valid_until TEXT,
                rating INTEGER,
                notes TEXT,
                skills_gained TEXT,
                skill_level_increase INTEGER DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                created_by INTEGER
            );
            CREATE TABLE IF NOT EXISTS training_attendees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                training_id INTEGER NOT NULL,
                employee_id INTEGER NOT NULL,
                status TEXT DEFAULT 'registered',
                attendance_confirmed INTEGER DEFAULT 0,
                test_score REAL,
                certificate_issued INTEGER DEFAULT 0,
                certificate_url TEXT,
                personal_rating INTEGER,
                personal_notes TEXT,
                FOREIGN KEY (training_id) REFERENCES trainings(id) ON DELETE CASCADE,
                FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_trainings_date ON trainings(date_start);
            CREATE INDEX IF NOT EXISTS idx_trainings_category ON trainings(category);
            CREATE INDEX IF NOT EXISTS idx_training_attendees_employee ON training_attendees(employee_id);
            CREATE INDEX IF NOT EXISTS idx_training_attendees_training ON training_attendees(training_id);
            """,
        ]),

        # v16: employees skills tracking
        (16, [
            ("employees", "skills", "ALTER TABLE employees ADD COLUMN skills TEXT NULL"),
            ("employees", "skill_score", "ALTER TABLE employees ADD COLUMN skill_score REAL DEFAULT 50"),
            ("employees", "training_hours_total", "ALTER TABLE employees ADD COLUMN training_hours_total REAL DEFAULT 0"),
            ("employees", "last_training_date", "ALTER TABLE employees ADD COLUMN last_training_date TEXT NULL"),
        ]),

        # v17: training compensation type (typ proplácení školení)
        (17, [
            ("trainings", "compensation_type", "ALTER TABLE trainings ADD COLUMN compensation_type TEXT DEFAULT 'paid_workday'"),
            ("trainings", "wage_cost", "ALTER TABLE trainings ADD COLUMN wage_cost REAL DEFAULT 0"),
            ("trainings", "wage_cost_per_person", "ALTER TABLE trainings ADD COLUMN wage_cost_per_person REAL NULL"),
        ]),

        # v18: worklogs training_id (propojení výkazů se školením)
        (18, [
            ("timesheets", "training_id", "ALTER TABLE timesheets ADD COLUMN training_id INTEGER NULL"),
            ("timesheets", "fk_timesheets_training", """
                CREATE INDEX IF NOT EXISTS idx_timesheets_training ON timesheets(training_id)
            """),
        ]),
        
        # v19: trainings participants field (účastníci jako JSON)
        (19, [
            ("trainings", "participants", "ALTER TABLE trainings ADD COLUMN participants TEXT DEFAULT '[]'"),
            ("trainings", "title", "ALTER TABLE trainings ADD COLUMN title TEXT NULL"),
            ("trainings", "date", "ALTER TABLE trainings ADD COLUMN date TEXT NULL"),
            ("trainings", "skills_improved", "ALTER TABLE trainings ADD COLUMN skills_improved TEXT NULL"),
            ("trainings", "skill_increase", "ALTER TABLE trainings ADD COLUMN skill_increase INTEGER DEFAULT 5"),
        ]),
        
        # v20: Crew Control System - team member profiles and capacity tracking
        (20, [
            """
            CREATE TABLE IF NOT EXISTS team_member_profile (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL UNIQUE,
                skills TEXT DEFAULT '[]',
                certifications TEXT DEFAULT '[]',
                weekly_capacity_hours REAL DEFAULT 40.0,
                preferred_work_types TEXT DEFAULT '[]',
                performance_stability_score REAL DEFAULT 0.5,
                ai_balance_score REAL DEFAULT 0.5,
                burnout_risk_level TEXT DEFAULT 'normal',
                total_jobs_completed INTEGER DEFAULT 0,
                current_active_jobs INTEGER DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                updated_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_team_profile_employee ON team_member_profile(employee_id);
            CREATE INDEX IF NOT EXISTS idx_team_profile_burnout ON team_member_profile(burnout_risk_level);
            
            CREATE TABLE IF NOT EXISTS team_capacity_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                planned_hours REAL DEFAULT 0,
                actual_hours REAL DEFAULT 0,
                capacity_status TEXT DEFAULT 'normal',
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
            );
            CREATE INDEX IF NOT EXISTS idx_capacity_log_employee_date ON team_capacity_log(employee_id, date);
            CREATE INDEX IF NOT EXISTS idx_capacity_log_date ON team_capacity_log(date);
            CREATE UNIQUE INDEX IF NOT EXISTS idx_capacity_log_unique ON team_capacity_log(employee_id, date);
            """,
        ]),
        
        # v21: Task Data Model - comprehensive Task model with TaskMaterial and TaskEvidence
        (21, [
            """
            -- Backup old tasks table structure (if needed for reference)
            CREATE TABLE IF NOT EXISTS tasks_legacy_backup AS 
            SELECT * FROM tasks WHERE 1=0;
            
            -- Create new comprehensive tasks table
            CREATE TABLE IF NOT EXISTS tasks_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                uuid TEXT UNIQUE NOT NULL DEFAULT (lower(hex(randomblob(4)) || '-' || hex(randomblob(2)) || '-4' || substr(hex(randomblob(2)), 2) || '-' || substr('89ab', abs(random()) % 4 + 1, 1) || substr(hex(randomblob(2)), 2) || '-' || hex(randomblob(6)))),
                
                -- IDENTIFIKACE
                title TEXT NOT NULL,
                description TEXT,
                task_type TEXT NOT NULL DEFAULT 'work',
                
                -- VAZBY
                job_id INTEGER NOT NULL,
                assigned_employee_id INTEGER NOT NULL,
                created_by_id INTEGER NOT NULL,
                
                -- ČASOVÉ OKNO
                planned_start TEXT NOT NULL,
                planned_end TEXT NOT NULL,
                planned_duration_minutes INTEGER NOT NULL,
                
                -- LOKACE
                location_type TEXT NOT NULL DEFAULT 'job_site',
                location_id INTEGER,
                location_name TEXT,
                gps_lat REAL,
                gps_lng REAL,
                
                -- OČEKÁVANÝ VÝSTUP
                expected_outcome TEXT NOT NULL,
                expected_outcome_type TEXT,
                expected_quantity REAL,
                expected_unit TEXT,
                
                -- SKUTEČNOST
                actual_start TEXT,
                actual_end TEXT,
                actual_duration_minutes INTEGER,
                
                -- STAV
                status TEXT NOT NULL DEFAULT 'planned',
                completion_state TEXT,
                completion_percentage INTEGER DEFAULT 0,
                
                -- ODCHYLKY
                time_deviation_minutes INTEGER DEFAULT 0,
                has_material_deviation INTEGER DEFAULT 0,
                has_workaround INTEGER DEFAULT 0,
                deviation_notes TEXT,
                
                -- INTEGRITA
                integrity_score REAL DEFAULT 100.0,
                integrity_flags TEXT DEFAULT '[]',
                
                -- PRIORITY & RIZIKO
                priority TEXT DEFAULT 'normal',
                risk_level TEXT DEFAULT 'low',
                risk_factors TEXT DEFAULT '[]',
                
                -- OFFLINE SUPPORT
                created_offline INTEGER DEFAULT 0,
                last_synced_at TEXT,
                offline_changes TEXT DEFAULT '{}',
                sync_conflict INTEGER DEFAULT 0,
                
                -- AUDIT
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                updated_at TEXT,
                version INTEGER DEFAULT 1,
                
                FOREIGN KEY (job_id) REFERENCES jobs(id),
                FOREIGN KEY (assigned_employee_id) REFERENCES employees(id),
                FOREIGN KEY (created_by_id) REFERENCES employees(id)
            );
            
            CREATE INDEX IF NOT EXISTS idx_task_job_status ON tasks_new(job_id, status);
            CREATE INDEX IF NOT EXISTS idx_task_employee_date ON tasks_new(assigned_employee_id, planned_start);
            CREATE INDEX IF NOT EXISTS idx_task_status_date ON tasks_new(status, planned_start);
            CREATE INDEX IF NOT EXISTS idx_task_uuid ON tasks_new(uuid);
            
            -- Create task_materials table
            CREATE TABLE IF NOT EXISTS task_materials (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER NOT NULL,
                material_id INTEGER,
                material_name TEXT NOT NULL,
                planned_quantity REAL NOT NULL,
                unit TEXT NOT NULL,
                actual_quantity REAL,
                was_available INTEGER,
                substitute_used INTEGER DEFAULT 0,
                substitute_material_id INTEGER,
                substitute_notes TEXT,
                reservation_id INTEGER,
                reservation_status TEXT,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (task_id) REFERENCES tasks_new(id) ON DELETE CASCADE,
                FOREIGN KEY (material_id) REFERENCES warehouse_items(id),
                FOREIGN KEY (substitute_material_id) REFERENCES warehouse_items(id)
            );
            
            CREATE INDEX IF NOT EXISTS idx_task_materials_task ON task_materials(task_id);
            CREATE INDEX IF NOT EXISTS idx_task_materials_material ON task_materials(material_id);
            
            -- Create task_evidence table
            CREATE TABLE IF NOT EXISTS task_evidence (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER NOT NULL,
                evidence_type TEXT NOT NULL,
                file_path TEXT,
                file_name TEXT,
                note_text TEXT,
                measurement_value REAL,
                measurement_unit TEXT,
                gps_lat REAL,
                gps_lng REAL,
                captured_at TEXT NOT NULL,
                captured_by_id INTEGER,
                captured_offline INTEGER DEFAULT 0,
                is_validated INTEGER DEFAULT 0,
                validated_by_id INTEGER,
                validated_at TEXT,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (task_id) REFERENCES tasks_new(id) ON DELETE CASCADE,
                FOREIGN KEY (captured_by_id) REFERENCES employees(id),
                FOREIGN KEY (validated_by_id) REFERENCES employees(id)
            );
            
            CREATE INDEX IF NOT EXISTS idx_task_evidence_task ON task_evidence(task_id);
            CREATE INDEX IF NOT EXISTS idx_task_evidence_type ON task_evidence(evidence_type);
            CREATE INDEX IF NOT EXISTS idx_task_evidence_captured ON task_evidence(captured_by_id);
            
            -- Create material_reservations table if it doesn't exist
            CREATE TABLE IF NOT EXISTS material_reservations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_id INTEGER,
                item_id INTEGER NOT NULL,
                quantity REAL NOT NULL,
                reserved_at TEXT NOT NULL DEFAULT (datetime('now')),
                status TEXT DEFAULT 'pending',
                FOREIGN KEY (item_id) REFERENCES warehouse_items(id)
            );
            
            CREATE INDEX IF NOT EXISTS idx_material_reservations_task ON material_reservations(task_id);
            CREATE INDEX IF NOT EXISTS idx_material_reservations_item ON material_reservations(item_id);
            """,
        ]),

        # v22: task_events table (event-driven system for Task entities)
        (22, [
            """
            CREATE TABLE IF NOT EXISTS task_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                uuid TEXT UNIQUE NOT NULL,
                
                task_id INTEGER NOT NULL,
                event_type TEXT NOT NULL,
                
                job_id INTEGER,
                employee_id INTEGER,
                
                payload TEXT NOT NULL DEFAULT '{}',
                
                occurred_at TEXT NOT NULL,
                recorded_at TEXT NOT NULL DEFAULT (datetime('now')),
                
                occurred_offline INTEGER NOT NULL DEFAULT 0,
                synced_at TEXT,
                
                source TEXT NOT NULL DEFAULT 'web_app',
                source_device_id TEXT,
                
                ai_processed INTEGER NOT NULL DEFAULT 0,
                ai_processed_at TEXT,
                ai_insights TEXT,
                
                FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE,
                FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE SET NULL,
                FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL
            );
            
            CREATE INDEX IF NOT EXISTS idx_event_task_type ON task_events(task_id, event_type);
            CREATE INDEX IF NOT EXISTS idx_event_job_time ON task_events(job_id, occurred_at);
            CREATE INDEX IF NOT EXISTS idx_event_type_time ON task_events(event_type, occurred_at);
            CREATE INDEX IF NOT EXISTS idx_event_ai_unprocessed ON task_events(ai_processed);
            CREATE INDEX IF NOT EXISTS idx_event_task_time ON task_events(task_id, occurred_at);
            CREATE INDEX IF NOT EXISTS idx_event_uuid ON task_events(uuid);
            CREATE INDEX IF NOT EXISTS idx_event_employee_time ON task_events(employee_id, occurred_at);
            """,
        ]),

        # v23: task_dependencies table (dependency system for tasks)
        (23, [
            """
            CREATE TABLE IF NOT EXISTS task_dependencies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                
                predecessor_task_id INTEGER NOT NULL,
                successor_task_id INTEGER NOT NULL,
                
                dependency_type TEXT NOT NULL,
                
                lag_minutes INTEGER DEFAULT 0,
                is_critical INTEGER DEFAULT 0,
                is_hard INTEGER DEFAULT 1,
                
                status TEXT DEFAULT 'pending',
                satisfied_at TEXT,
                violated_at TEXT,
                violation_reason TEXT,
                
                risk_weight REAL DEFAULT 1.0,
                current_risk_level TEXT DEFAULT 'low',
                
                material_id INTEGER,
                material_quantity REAL,
                
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                created_by_id INTEGER,
                
                auto_generated INTEGER DEFAULT 0,
                confidence REAL,
                validated INTEGER DEFAULT 0,
                
                FOREIGN KEY (predecessor_task_id) REFERENCES tasks(id) ON DELETE CASCADE,
                FOREIGN KEY (successor_task_id) REFERENCES tasks(id) ON DELETE CASCADE,
                FOREIGN KEY (material_id) REFERENCES warehouse_items(id) ON DELETE SET NULL,
                FOREIGN KEY (created_by_id) REFERENCES employees(id) ON DELETE SET NULL,
                
                CHECK (predecessor_task_id != successor_task_id)
            );
            
            CREATE UNIQUE INDEX IF NOT EXISTS idx_dep_unique ON task_dependencies(predecessor_task_id, successor_task_id, dependency_type);
            CREATE INDEX IF NOT EXISTS idx_dep_predecessor ON task_dependencies(predecessor_task_id);
            CREATE INDEX IF NOT EXISTS idx_dep_successor ON task_dependencies(successor_task_id);
            CREATE INDEX IF NOT EXISTS idx_dep_status ON task_dependencies(status);
            CREATE INDEX IF NOT EXISTS idx_dep_type ON task_dependencies(dependency_type);
            CREATE INDEX IF NOT EXISTS idx_dep_risk ON task_dependencies(current_risk_level);
            """,
        ]),

        # v24: audit_logs table (audit trail for API requests)
        (24, [
            """
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id TEXT NOT NULL,
                
                endpoint TEXT,
                method TEXT,
                
                user_id INTEGER,
                
                request_body TEXT,
                response_status INTEGER,
                
                duration_ms REAL,
                
                ip_address TEXT,
                user_agent TEXT,
                
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                
                FOREIGN KEY (user_id) REFERENCES employees(id) ON DELETE SET NULL
            );
            
            CREATE INDEX IF NOT EXISTS idx_audit_request_id ON audit_logs(request_id);
            CREATE INDEX IF NOT EXISTS idx_audit_endpoint ON audit_logs(endpoint);
            CREATE INDEX IF NOT EXISTS idx_audit_user_time ON audit_logs(user_id, created_at);
            CREATE INDEX IF NOT EXISTS idx_audit_created ON audit_logs(created_at);
            """,
        ]),

        # v25: ai_alerts and daily_briefings tables (AI Operator analytics)
        (25, [
            """
            CREATE TABLE IF NOT EXISTS ai_alerts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                
                alert_type TEXT NOT NULL,
                severity TEXT NOT NULL,
                
                title TEXT NOT NULL,
                summary TEXT,
                detail_data TEXT,
                
                job_id INTEGER,
                task_id INTEGER,
                employee_id INTEGER,
                
                status TEXT DEFAULT 'active',
                
                recommended_actions TEXT,
                action_taken TEXT,
                
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                acknowledged_at TEXT,
                acknowledged_by_id INTEGER,
                resolved_at TEXT,
                resolved_by_id INTEGER,
                
                FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE,
                FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE,
                FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL,
                FOREIGN KEY (acknowledged_by_id) REFERENCES employees(id) ON DELETE SET NULL,
                FOREIGN KEY (resolved_by_id) REFERENCES employees(id) ON DELETE SET NULL
            );
            
            CREATE INDEX IF NOT EXISTS idx_alert_status ON ai_alerts(status);
            CREATE INDEX IF NOT EXISTS idx_alert_severity ON ai_alerts(severity, status);
            CREATE INDEX IF NOT EXISTS idx_alert_job ON ai_alerts(job_id, status);
            CREATE INDEX IF NOT EXISTS idx_alert_type ON ai_alerts(alert_type);
            CREATE INDEX IF NOT EXISTS idx_alert_created ON ai_alerts(created_at);
            
            CREATE TABLE IF NOT EXISTS daily_briefings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                
                briefing_date TEXT NOT NULL UNIQUE,
                generated_at TEXT NOT NULL DEFAULT (datetime('now')),
                
                executive_summary TEXT,
                full_data TEXT,
                
                total_tasks INTEGER,
                critical_items INTEGER,
                integrity_average REAL,
                
                notified INTEGER DEFAULT 0,
                notified_at TEXT,
                
                CHECK (briefing_date IS NOT NULL)
            );
            
            CREATE INDEX IF NOT EXISTS idx_briefing_date ON daily_briefings(briefing_date);
            CREATE INDEX IF NOT EXISTS idx_briefing_generated ON daily_briefings(generated_at);
            """,
        ]),

        # v26: Asset Core tables (AssetType, Location, AssetLot, AssetReservation, ValuationSnapshot)
        (26, [
            """
            CREATE TABLE IF NOT EXISTS asset_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                uuid TEXT UNIQUE NOT NULL DEFAULT (lower(hex(randomblob(16)))),
                
                code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                latin_name TEXT,
                
                category TEXT NOT NULL,
                subcategory TEXT,
                
                growth_zone TEXT,
                sun_requirement TEXT,
                water_requirement TEXT,
                mature_height_cm INTEGER,
                mature_width_cm INTEGER,
                growth_rate TEXT,
                
                default_unit TEXT DEFAULT 'ks',
                default_purchase_price REAL,
                default_sell_price REAL,
                
                propagation_method TEXT,
                production_time_months INTEGER,
                
                planting_season_start INTEGER,
                planting_season_end INTEGER,
                sale_season_start INTEGER,
                sale_season_end INTEGER,
                
                primary_image_url TEXT,
                
                is_active INTEGER DEFAULT 1,
                
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT,
                
                CHECK (code IS NOT NULL AND code != ''),
                CHECK (name IS NOT NULL AND name != '')
            );
            
            CREATE INDEX IF NOT EXISTS idx_asset_type_category ON asset_types(category, is_active);
            CREATE INDEX IF NOT EXISTS idx_asset_type_code ON asset_types(code);
            CREATE INDEX IF NOT EXISTS idx_asset_type_active ON asset_types(is_active);
            
            CREATE TABLE IF NOT EXISTS locations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                uuid TEXT UNIQUE NOT NULL DEFAULT (lower(hex(randomblob(16)))),
                
                parent_id INTEGER,
                path TEXT,
                level INTEGER DEFAULT 0,
                
                code TEXT NOT NULL,
                name TEXT NOT NULL,
                location_type TEXT,
                
                capacity_units INTEGER,
                current_occupancy INTEGER DEFAULT 0,
                
                sun_exposure TEXT,
                irrigation INTEGER DEFAULT 0,
                heated INTEGER DEFAULT 0,
                
                gps_lat REAL,
                gps_lng REAL,
                
                qr_code TEXT UNIQUE,
                
                is_active INTEGER DEFAULT 1,
                
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT,
                
                FOREIGN KEY (parent_id) REFERENCES locations(id) ON DELETE SET NULL,
                CHECK (code IS NOT NULL AND code != ''),
                CHECK (name IS NOT NULL AND name != '')
            );
            
            CREATE INDEX IF NOT EXISTS idx_location_parent ON locations(parent_id);
            CREATE INDEX IF NOT EXISTS idx_location_code ON locations(code);
            CREATE INDEX IF NOT EXISTS idx_location_active ON locations(is_active);
            CREATE INDEX IF NOT EXISTS idx_location_type ON locations(location_type);
            
            CREATE TABLE IF NOT EXISTS asset_lots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                uuid TEXT UNIQUE NOT NULL DEFAULT (lower(hex(randomblob(16)))),
                
                asset_type_id INTEGER NOT NULL,
                location_id INTEGER NOT NULL,
                
                lot_code TEXT UNIQUE NOT NULL,
                batch_date TEXT,
                source TEXT,
                
                quantity_total INTEGER NOT NULL DEFAULT 0,
                quantity_available INTEGER NOT NULL DEFAULT 0,
                quantity_reserved INTEGER DEFAULT 0,
                quantity_damaged INTEGER DEFAULT 0,
                unit TEXT DEFAULT 'ks',
                
                lifecycle_stage TEXT NOT NULL DEFAULT 'growing',
                availability_status TEXT NOT NULL DEFAULT 'available',
                
                quality_grade TEXT DEFAULT 'A',
                quality_notes TEXT,
                last_quality_check TEXT,
                
                current_height_cm INTEGER,
                container_size TEXT,
                root_status TEXT,
                
                purchase_price_per_unit REAL,
                current_value_per_unit REAL,
                sell_price_per_unit REAL,
                
                qr_code TEXT UNIQUE,
                
                created_at TEXT DEFAULT (datetime('now')),
                updated_at TEXT,
                created_by_id INTEGER,
                
                last_synced_at TEXT,
                
                FOREIGN KEY (asset_type_id) REFERENCES asset_types(id) ON DELETE RESTRICT,
                FOREIGN KEY (location_id) REFERENCES locations(id) ON DELETE RESTRICT,
                FOREIGN KEY (created_by_id) REFERENCES employees(id) ON DELETE SET NULL,
                
                CHECK (quantity_available >= 0),
                CHECK (quantity_reserved >= 0),
                CHECK (quantity_total >= quantity_available + quantity_reserved),
                CHECK (lot_code IS NOT NULL AND lot_code != '')
            );
            
            CREATE INDEX IF NOT EXISTS idx_lot_type_location ON asset_lots(asset_type_id, location_id);
            CREATE INDEX IF NOT EXISTS idx_lot_availability ON asset_lots(availability_status, lifecycle_stage);
            CREATE INDEX IF NOT EXISTS idx_lot_qr ON asset_lots(qr_code);
            CREATE INDEX IF NOT EXISTS idx_lot_code ON asset_lots(lot_code);
            
            CREATE TABLE IF NOT EXISTS asset_reservations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                uuid TEXT UNIQUE NOT NULL DEFAULT (lower(hex(randomblob(16)))),
                
                lot_id INTEGER NOT NULL,
                job_id INTEGER NOT NULL,
                task_id INTEGER,
                
                quantity INTEGER NOT NULL,
                unit TEXT DEFAULT 'ks',
                
                status TEXT NOT NULL DEFAULT 'active',
                
                reserved_at TEXT DEFAULT (datetime('now')),
                needed_by TEXT,
                expires_at TEXT,
                
                fulfilled_quantity INTEGER DEFAULT 0,
                fulfilled_at TEXT,
                
                created_by_id INTEGER,
                notes TEXT,
                
                FOREIGN KEY (lot_id) REFERENCES asset_lots(id) ON DELETE CASCADE,
                FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE,
                FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE SET NULL,
                FOREIGN KEY (created_by_id) REFERENCES employees(id) ON DELETE SET NULL,
                
                CHECK (quantity > 0),
                CHECK (fulfilled_quantity >= 0)
            );
            
            CREATE INDEX IF NOT EXISTS idx_reservation_lot_status ON asset_reservations(lot_id, status);
            CREATE INDEX IF NOT EXISTS idx_reservation_job ON asset_reservations(job_id, status);
            CREATE INDEX IF NOT EXISTS idx_reservation_status ON asset_reservations(status);
            
            CREATE TABLE IF NOT EXISTS valuation_snapshots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                
                snapshot_type TEXT NOT NULL,
                snapshot_date TEXT NOT NULL,
                
                lot_id INTEGER,
                asset_type_id INTEGER,
                location_id INTEGER,
                
                total_quantity INTEGER,
                total_purchase_value REAL,
                total_current_value REAL,
                total_sell_value REAL,
                
                avg_quality_score REAL,
                dead_stock_quantity INTEGER,
                slow_moving_quantity INTEGER,
                
                created_at TEXT DEFAULT (datetime('now')),
                
                FOREIGN KEY (lot_id) REFERENCES asset_lots(id) ON DELETE CASCADE,
                FOREIGN KEY (asset_type_id) REFERENCES asset_types(id) ON DELETE CASCADE,
                FOREIGN KEY (location_id) REFERENCES locations(id) ON DELETE CASCADE,
                
                CHECK (snapshot_type IS NOT NULL),
                CHECK (snapshot_date IS NOT NULL)
            );
            
            CREATE INDEX IF NOT EXISTS idx_valuation_date ON valuation_snapshots(snapshot_date, snapshot_type);
            CREATE INDEX IF NOT EXISTS idx_valuation_lot ON valuation_snapshots(lot_id);
            CREATE INDEX IF NOT EXISTS idx_valuation_type ON valuation_snapshots(asset_type_id);
            CREATE INDEX IF NOT EXISTS idx_valuation_location ON valuation_snapshots(location_id);
            """,
        ]),

        # v27: asset_events table (Event-driven system for Asset Core)
        (27, [
            """
            CREATE TABLE IF NOT EXISTS asset_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                uuid TEXT UNIQUE NOT NULL DEFAULT (lower(hex(randomblob(16)))),
                
                event_type TEXT NOT NULL,
                
                lot_id INTEGER,
                asset_type_id INTEGER,
                location_id INTEGER,
                
                job_id INTEGER,
                task_id INTEGER,
                reservation_id INTEGER,
                
                payload TEXT NOT NULL DEFAULT '{}',
                
                quantity_before INTEGER,
                quantity_after INTEGER,
                quantity_delta INTEGER,
                
                from_location_id INTEGER,
                to_location_id INTEGER,
                
                occurred_at TEXT NOT NULL DEFAULT (datetime('now')),
                recorded_at TEXT DEFAULT (datetime('now')),
                
                triggered_by TEXT DEFAULT 'user',
                employee_id INTEGER,
                device_id TEXT,
                
                created_offline INTEGER DEFAULT 0,
                offline_uuid TEXT,
                synced_at TEXT,
                
                FOREIGN KEY (lot_id) REFERENCES asset_lots(id) ON DELETE CASCADE,
                FOREIGN KEY (asset_type_id) REFERENCES asset_types(id) ON DELETE CASCADE,
                FOREIGN KEY (location_id) REFERENCES locations(id) ON DELETE CASCADE,
                FOREIGN KEY (job_id) REFERENCES jobs(id) ON DELETE CASCADE,
                FOREIGN KEY (task_id) REFERENCES tasks(id) ON DELETE CASCADE,
                FOREIGN KEY (reservation_id) REFERENCES asset_reservations(id) ON DELETE CASCADE,
                FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE SET NULL,
                FOREIGN KEY (from_location_id) REFERENCES locations(id) ON DELETE SET NULL,
                FOREIGN KEY (to_location_id) REFERENCES locations(id) ON DELETE SET NULL,
                
                CHECK (event_type IS NOT NULL AND event_type != '')
            );
            
            CREATE INDEX IF NOT EXISTS idx_event_lot ON asset_events(lot_id, occurred_at);
            CREATE INDEX IF NOT EXISTS idx_event_type ON asset_events(event_type, occurred_at);
            CREATE INDEX IF NOT EXISTS idx_event_location ON asset_events(location_id, occurred_at);
            CREATE INDEX IF NOT EXISTS idx_event_job ON asset_events(job_id);
            CREATE INDEX IF NOT EXISTS idx_event_date ON asset_events(occurred_at);
            CREATE INDEX IF NOT EXISTS idx_event_from_location ON asset_events(from_location_id);
            CREATE INDEX IF NOT EXISTS idx_event_to_location ON asset_events(to_location_id);
            """,
        ]),

        # v28: Rozšíření rolí - přidání 'director' a 'lander', mapování starých rolí
        (28, [
            """
            -- Mapování existujících rolí na nové (před vytvořením nové tabulky)
            UPDATE users SET role = 'director' WHERE role = 'owner';
            UPDATE users SET role = 'director' WHERE role = 'admin';
            UPDATE users SET role = 'lander' WHERE role = 'team_lead';
            
            -- Vypnutí foreign key constraints pro bezpečné přejmenování tabulek
            PRAGMA foreign_keys = OFF;
            
            -- Vytvoření nové tabulky s rozšířeným CHECK constraintem
            CREATE TABLE IF NOT EXISTS users_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'worker' 
                    CHECK(role IN ('owner','manager','team_lead','worker','admin','director','lander')),
                password_hash TEXT NOT NULL,
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                manager_id INTEGER NULL
            );
            
            -- Zkopírování dat z původní tabulky
            INSERT INTO users_new (id, email, name, role, password_hash, active, created_at, manager_id)
            SELECT id, email, name, role, password_hash, active, created_at, manager_id
            FROM users;
            
            -- Drop staré tabulky a přejmenování nové
            DROP TABLE users;
            ALTER TABLE users_new RENAME TO users;
            
            -- Obnovení foreign key constraints
            PRAGMA foreign_keys = ON;
            
            -- Obnovení indexů
            CREATE INDEX IF NOT EXISTS idx_users_email ON users(email);
            CREATE INDEX IF NOT EXISTS idx_users_role ON users(role);
            CREATE INDEX IF NOT EXISTS idx_users_active ON users(active);
            """,
        ]),

        # v29: UserSettings tabulka pro mobile_mode a další preference
        (29, [
            """
            CREATE TABLE IF NOT EXISTS user_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL UNIQUE,
                mobile_mode TEXT DEFAULT 'auto' CHECK(mobile_mode IN ('field', 'full', 'auto')),
                updated_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            
            CREATE INDEX IF NOT EXISTS idx_user_settings_user_id ON user_settings(user_id);
            """,
        ]),

        # v30: UserDashboardLayout tabulka pro custom widget layouts
        (30, [
            """
            CREATE TABLE IF NOT EXISTS user_dashboard_layout (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL UNIQUE,
                field_widgets TEXT DEFAULT '[]',  -- JSON array widget IDs
                full_widgets TEXT DEFAULT '[]',    -- JSON array widget IDs
                updated_at TEXT DEFAULT (datetime('now')),
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            
            CREATE INDEX IF NOT EXISTS idx_dashboard_layout_user_id ON user_dashboard_layout(user_id);
            """,
        ]),

        # v31: ProcessedEvents tabulka pro deduplikaci offline queue
        (31, [
            """
            CREATE TABLE IF NOT EXISTS processed_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id TEXT UNIQUE NOT NULL,
                event_type TEXT NOT NULL,
                result_id INTEGER,
                created_at TEXT DEFAULT (datetime('now'))
            );
            
            CREATE INDEX IF NOT EXISTS idx_processed_events_event_id ON processed_events(event_id);
            CREATE INDEX IF NOT EXISTS idx_processed_events_type ON processed_events(event_type);
            """,
        ]),
    ]

    for version, alters in migrations:
        if version in applied:
            continue
        for item in alters:
            # Allow either (table, col, sql) ALTERs or raw DDL scripts as strings
            if isinstance(item, str):
                try:
                    db.executescript(item)
                except Exception:
                    pass
                continue
            table, col, sql = item
            # Skip ALTERs for tables that do not exist yet (fresh DB before ensure_schema).
            if not _table_exists(db, table):
                continue
            if not _table_has_column(db, table, col):
                try:
                    db.execute(sql)
                except Exception:
                    # last-resort: ignore to avoid breaking startup
                    pass
        db.execute("INSERT OR IGNORE INTO schema_migrations(version) VALUES (?)", (version,))
        db.commit()

def audit_event(user_id, action: str, entity_type: str, entity_id=None, before=None, after=None, meta=None):
    """Write a single audit log entry. Never raises (to avoid breaking user flows)."""
    try:
        db = get_db()
        db.execute(
            "INSERT INTO audit_log(user_id, action, entity_type, entity_id, before_json, after_json, meta_json) VALUES (?,?,?,?,?,?,?)",
            (
                user_id,
                action,
                entity_type,
                entity_id,
                json.dumps(before, ensure_ascii=False) if before is not None else None,
                json.dumps(after, ensure_ascii=False) if after is not None else None,
                json.dumps(meta, ensure_ascii=False) if meta is not None else None,
            ),
        )
        db.commit()
    except Exception:
        pass


# ----------------- Notifications & Delegation -----------------
def _employee_user_id(db, employee_id: int):
    """Return linked user_id for an employee, or None."""
    try:
        r = db.execute("SELECT user_id FROM employees WHERE id=?", (int(employee_id),)).fetchone()
        return int(r[0]) if r and r[0] is not None else None
    except Exception:
        return None


def create_notification(*, user_id=None, employee_id=None, kind="info", title="", body="", entity_type=None, entity_id=None):
    """Create an in-app notification. Never raises."""
    try:
        db = get_db()
        if employee_id is not None and user_id is None:
            user_id = _employee_user_id(db, int(employee_id))
        db.execute(
            "INSERT INTO notifications(user_id, employee_id, kind, title, body, entity_type, entity_id) VALUES (?,?,?,?,?,?,?)",
            (
                int(user_id) if user_id is not None else None,
                int(employee_id) if employee_id is not None else None,
                str(kind or "info"),
                str(title or ""),
                str(body or ""),
                str(entity_type) if entity_type is not None else None,
                int(entity_id) if entity_id is not None else None,
            ),
        )
        db.commit()
    except Exception:
        pass


def _expand_assignees_with_delegate(db, employee_ids):
    """Expand assignees by adding one-hop delegates.

    Returns: (expanded_ids, delegations)
      - expanded_ids: list[int]
      - delegations: list[{'from': int, 'to': int}]

    Note: intentionally non-recursive to avoid cycles and surprises.
    """
    try:
        ids = [int(x) for x in (employee_ids or []) if str(x).strip()]
    except Exception:
        ids = []
    seen = set(ids)
    delegations = []

    for eid in list(ids):
        try:
            r = db.execute(
                "SELECT delegate_employee_id FROM employees WHERE id=?",
                (int(eid),),
            ).fetchone()
            did = int(r[0]) if r and r[0] is not None else None
        except Exception:
            did = None

        if did and did != eid and did not in seen:
            ids.append(did)
            seen.add(did)
            delegations.append({"from": int(eid), "to": int(did)})

    return ids, delegations


def _notify_assignees(entity_type: str, entity_id: int, assignee_ids, title: str, body: str, actor_user_id=None):
    """Notify assignees (employees) - skips actor if mapped to same employee user."""
    db = get_db()
    # Map actor user_id -> employee_id (best-effort)
    actor_employee_id = None
    try:
        if actor_user_id:
            r = db.execute("SELECT id FROM employees WHERE user_id=?", (int(actor_user_id),)).fetchone()
            actor_employee_id = int(r[0]) if r else None
    except Exception:
        actor_employee_id = None

    for eid in assignee_ids or []:
        try:
            eid = int(eid)
        except Exception:
            continue
        if actor_employee_id and eid == actor_employee_id:
            continue
        create_notification(
            employee_id=eid,
            kind="assignment",
            title=title,
            body=body,
            entity_type=entity_type,
            entity_id=int(entity_id),
        )


# ----------------- Utility functions -----------------
def _normalize_date(v):
    if not v:
        return v
    s = str(v).strip()
    import re as _re
    m = _re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", s)
    if m:
        y, M, d = m.groups()
        return f"{int(y):04d}-{int(M):02d}-{int(d):02d}"
    m = _re.match(r"^(\d{1,2})[\.\s-](\d{1,2})[\.\s-](\d{4})$", s)
    if m:
        d, M, y = m.groups()
        return f"{int(y):04d}-{int(M):02d}-{int(d):02d}"
    return s





@app.route("/archive")
def view_archive():
    u, err = require_auth()
    if err:
        return err
    db = get_db()
    rows = [dict(r) for r in db.execute(
        "SELECT id,title,client,city,code,status,date,completed_at "
        "FROM jobs "
        "WHERE lower(status) LIKE 'dokon%' "
        "ORDER BY COALESCE(date(completed_at), date(date)) DESC"
    ).fetchall()]
    months = {}
    for r in rows:
        src = r.get("completed_at") or r.get("date") or ""
        ym = ""
        try:
            # normalize to YYYY-MM
            if src and len(src) >= 7:
                ym = f"{src[0:4]}-{src[5:7]}"
        except Exception:
            ym = ""
        months.setdefault(ym or "unknown", []).append(r)
    # group by year for template
    by_year = {}
    for ym, items in months.items():
        y = (ym or "unknown").split("-")[0]
        by_year.setdefault(y, []).append((ym, items))
    # sort
    for y in by_year:
        by_year[y].sort(key=lambda x: x[0], reverse=True)
    years = sorted(by_year.items(), key=lambda x: x[0], reverse=True)
    return render_template("archive.html", me=u, years=years)



@app.route("/api/jobs/archive")
def api_jobs_archive():
    u, err = require_auth()
    if err: 
        return err
    db = get_db()
    rows = [dict(r) for r in db.execute("SELECT id,title,client,city,code,status,date,completed_at FROM jobs WHERE lower(status) LIKE 'dokon%' ORDER BY COALESCE(date(completed_at), date(date)) DESC").fetchall()]
    months = {}
    for r in rows:
        # prefer completed_at month, fallback to scheduled date
        src = r.get("completed_at") or r.get("date") or ""
        ym = ""
        try:
            # normalize to YYYY-MM
            if src and len(src)>=7:
                ym = f"{src[0:4]}-{src[5:7]}"
        except Exception:
            ym = ""
        months.setdefault(ym or "unknown", []).append(r)
    return jsonify({"ok": True, "months": months})

# --- Team/Employees aliases (to avoid 404) ---
from flask import render_template

@app.route("/team")
@app.route("/team/")
def team_page():
    # Crew Control System - sloučená sekce Tým
    return send_from_directory(".", "team.html")

@app.route("/employees")
@app.route("/employees/")
@app.route("/zamestnanci")
@app.route("/zamestnanci/")
def employees_redirect():
    # Přesměrování ze starého employees na novou sekci Tým
    return redirect("/team")

@app.route("/team.html")
def team_html_direct():
    return send_from_directory(".", "team.html")

@app.route("/admin/roles")
@app.route("/admin/roles/")
@app.route("/admin_roles.html")
def admin_roles():
    """Stránka pro správu rolí - pouze pro ownera"""
    u, err = require_auth()
    if err:
        return redirect('/login')
    if u and u.get('role') not in ('owner', 'admin'):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    return render_template("admin_roles.html")
# --- end aliases ---

# ----------------- Migration -----------------
def _migrate_completed_at():
    """Add missing columns to jobs table if they don't exist"""
    db = get_db()
    try:
        if not _table_exists(db, "jobs"):
            return
        cols = [r[1] for r in db.execute("PRAGMA table_info(jobs)").fetchall()]
        if "completed_at" not in cols:
            db.execute("ALTER TABLE jobs ADD COLUMN completed_at TEXT")
            db.commit()
            print("[DB] Added column: completed_at")
        if "created_date" not in cols:
            db.execute("ALTER TABLE jobs ADD COLUMN created_date TEXT")
            db.commit()
            print("[DB] Added column: created_date")
        if "start_date" not in cols:
            db.execute("ALTER TABLE jobs ADD COLUMN start_date TEXT")
            db.commit()
            print("[DB] Added column: start_date")
        if "progress" not in cols:
            db.execute("ALTER TABLE jobs ADD COLUMN progress INTEGER DEFAULT 0")
            db.commit()
            print("[DB] Added column: progress")
    except Exception as e:
        print(f"[DB] Migration warning: {e}")

def _migrate_employees_enhanced():
    """Add enhanced columns to employees table and create new tables"""
    db = get_db()
    try:
        if not _table_exists(db, "employees"):
            return
        cols = [r[1] for r in db.execute("PRAGMA table_info(employees)").fetchall()]
        
        # Add new columns to employees table
        new_cols = {
            "birth_date": "TEXT",
            "contract_type": "TEXT DEFAULT 'HPP'",
            "start_date": "TEXT",
            "hourly_rate": "REAL",
            "salary": "REAL",
            "skills": "TEXT",
            "location": "TEXT",
            "status": "TEXT DEFAULT 'active'",
            "rating": "REAL DEFAULT 0",
            "avatar_url": "TEXT"

            ,"delegate_employee_id": "INTEGER NULL"
            ,"approver_delegate_employee_id": "INTEGER NULL"
        
            ,"user_id": "INTEGER NULL"
        }
        
        for col_name, col_def in new_cols.items():
            if col_name not in cols:
                db.execute(f"ALTER TABLE employees ADD COLUMN {col_name} {col_def}")
                db.commit()
                print(f"[DB] Added column: employees.{col_name}")
        
        # Create employee_documents table
        db.execute("""
            CREATE TABLE IF NOT EXISTS employee_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                type TEXT NOT NULL,
                file_url TEXT NOT NULL,
                uploaded_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
            )
        """)
        db.commit()
        print("[DB] Created table: employee_documents")
        
        # Create employee_timeline table
        db.execute("""
            CREATE TABLE IF NOT EXISTS employee_timeline (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER NOT NULL,
                action TEXT NOT NULL,
                description TEXT,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
            )
        """)
        db.commit()
        print("[DB] Created table: employee_timeline")
        
    except Exception as e:
        print(f"[DB] Migration warning: {e}")

def _migrate_roles_and_hierarchy():
    """Migrace: přidání sloupců role a manager_id do users tabulky"""
    db = get_db()
    try:
        if not _table_exists(db, "users"):
            return
        # Zkontrolovat existující sloupce
        cols = [r[1] for r in db.execute("PRAGMA table_info(users)").fetchall()]
        
        # Přidat sloupec role pokud neexistuje nebo aktualizovat CHECK constraint
        if 'role' not in cols:
            db.execute("ALTER TABLE users ADD COLUMN role VARCHAR(20) DEFAULT 'worker'")
            db.commit()
            print("[DB] Added column: users.role")
        else:
            # Aktualizovat CHECK constraint pro nové role
            # SQLite nepodporuje ALTER COLUMN, takže musíme vytvořit novou tabulku
            # Poznámka: Pokud už existuje manager_id, přeskočíme tuto část
            if 'manager_id' not in cols:
                try:
                    # Vypnout foreign key kontroly během migrace
                    db.execute("PRAGMA foreign_keys=OFF")
                    db.execute("""
                        CREATE TABLE users_new (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            email TEXT UNIQUE NOT NULL,
                            name TEXT NOT NULL,
                            role TEXT NOT NULL DEFAULT 'worker' CHECK(role IN ('owner','admin','manager','lander','worker','team_lead')),
                            password_hash TEXT NOT NULL,
                            active INTEGER NOT NULL DEFAULT 1,
                            created_at TEXT NOT NULL DEFAULT (datetime('now')),
                            manager_id INTEGER NULL
                        )
                    """)
                    db.execute("""
                        INSERT INTO users_new (id, email, name, role, password_hash, active, created_at)
                        SELECT id, email, name, 
                               CASE 
                                   WHEN role = 'admin' THEN 'owner'
                                   ELSE role
                               END,
                               password_hash, active, created_at
                        FROM users
                    """)
                    db.execute("DROP TABLE users")
                    db.execute("ALTER TABLE users_new RENAME TO users")
                    db.execute("PRAGMA foreign_keys=ON")
                    db.commit()
                    print("[DB] Updated users table with new role system")
                except Exception as e:
                    db.execute("PRAGMA foreign_keys=ON")
                    print(f"[DB] Role migration warning: {e}")
        
        # Přidat manager_id pokud neexistuje
        if 'manager_id' not in cols:
            db.execute("ALTER TABLE users ADD COLUMN manager_id INTEGER NULL")
            db.commit()
            print("[DB] Added column: users.manager_id")
            
            # Přidat foreign key constraint (SQLite to podporuje přes CREATE INDEX)
            try:
                db.execute("CREATE INDEX IF NOT EXISTS idx_users_manager_id ON users(manager_id)")
                db.commit()
            except Exception as e:
                print(f"[DB] Index creation warning: {e}")
        
        # Aktualizovat existujícího uživatele david@greendavid.cz na owner
        try:
            db.execute("UPDATE users SET role = 'owner' WHERE email = 'david@greendavid.cz'")
            db.commit()
            print("[DB] Updated david@greendavid.cz to owner role")
        except Exception as e:
            print(f"[DB] Owner update warning: {e}")
            
    except Exception as e:
        print(f"[DB] Roles migration warning: {e}")

# NOTE: DB initialization is handled by init_db_once() (defined later),
# which ensures base schema exists before applying any migrations.

# ----------------- Authentication functions -----------------
def current_user():
    uid = session.get("uid")
    if not uid:
        return None
    db = get_db()
    # Zkontrolovat, zda existuje sloupec manager_id
    cols = [r[1] for r in db.execute("PRAGMA table_info(users)").fetchall()]
    if 'manager_id' in cols:
        row = db.execute("SELECT id,email,name,role,active,manager_id FROM users WHERE id=?", (uid,)).fetchone()
    else:
        row = db.execute("SELECT id,email,name,role,active FROM users WHERE id=?", (uid,)).fetchone()
    return dict(row) if row else None

def require_auth():
    u = current_user()
    if not u or not u.get("active"):
        return None, (jsonify({"ok": False, "error": "unauthorized"}), 401)
    return u, None

def require_role(write=False):
    u, err = require_auth()
    if err:
        return None, err
    if write and normalize_role(u["role"]) not in WRITE_ROLES:
        return None, (jsonify({"ok": False, "error": "forbidden"}), 403)
    return u, None

# ----------------- Permissions system -----------------
from functools import wraps

def requires_role(*allowed_roles):
    """Decorator pro kontrolu oprávnění podle role"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Kontrola přihlášení
            if 'uid' not in session:
                return jsonify({'ok': False, 'error': 'unauthorized'}), 401
            
            # Získat roli uživatele z DB
            db = get_db()
            user = db.execute(
                "SELECT role FROM users WHERE id = ?", 
                (session['uid'],)
            ).fetchone()
            
            if not user:
                return jsonify({'ok': False, 'error': 'unauthorized'}), 401
            
            # Fallback: pokud nemá roli nebo je NULL, považujeme za owner (pro zpětnou kompatibilitu)
            user_role = user['role'] or 'owner'
            
            if user_role not in allowed_roles:
                return jsonify({'ok': False, 'error': 'forbidden'}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_current_user():
    """Pomocná funkce pro získání aktuálního uživatele s plnými informacemi"""
    if 'uid' not in session:
        return None
    
    db = get_db()
    user = db.execute(
        "SELECT id, email, name, role, manager_id FROM users WHERE id = ?",
        (session['uid'],)
    ).fetchone()
    
    if not user:
        return None
    
    user_dict = dict(user)
    
    # Fallback: pokud nemá roli nebo je NULL, považujeme za owner (pro zpětnou kompatibilitu)
    if not user_dict.get('role') or user_dict['role'] == 'admin':
        user_dict['role'] = 'owner'
    
    return user_dict

def can_manage_employee(manager_id, employee_id):
    """Kontrola, zda má manager oprávnění spravovat zaměstnance"""
    db = get_db()
    
    # Owner může všechno
    manager = db.execute("SELECT role FROM users WHERE id = ?", (manager_id,)).fetchone()
    if manager and manager['role'] in ('owner', 'admin'):
        return True
    
    # Manager/Team lead jen své lidi
    employee = db.execute(
        "SELECT manager_id FROM users WHERE id = ?", 
        (employee_id,)
    ).fetchone()
    
    return employee and employee['manager_id'] == manager_id

# ----------------- bootstrap (subset) -----------------
def ensure_schema():
    db = get_db()
    db.executescript("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT UNIQUE NOT NULL,
        name TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'worker' CHECK(role IN ('owner','manager','team_lead','worker','admin')),
        password_hash TEXT NOT NULL,
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL DEFAULT (datetime('now')),
        manager_id INTEGER NULL
    );
    

    CREATE TABLE IF NOT EXISTS schema_migrations (
        version INTEGER PRIMARY KEY,
        applied_at TEXT NOT NULL DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        action TEXT NOT NULL,
        entity_type TEXT NOT NULL,
        entity_id INTEGER,
        before_json TEXT,
        after_json TEXT,
        meta_json TEXT,
        created_at TEXT NOT NULL DEFAULT (datetime('now'))
    );

CREATE TABLE IF NOT EXISTS employees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'worker',
        -- delegace: kam přesměrovat úkoly/problémy (např. při nepřítomnosti)
        delegate_employee_id INTEGER NULL,
        -- delegace: kam přesměrovat schvalování (výkazy apod.)
        approver_delegate_employee_id INTEGER NULL,
        phone TEXT DEFAULT '',
        email TEXT DEFAULT '',
        address TEXT DEFAULT '',
        user_id INTEGER NULL
    );
    -- prefer new schema; legacy DBs may still have jobs.name (possibly NOT NULL)
    CREATE TABLE IF NOT EXISTS jobs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT,
        name TEXT,
        client TEXT NOT NULL DEFAULT '',
        status TEXT NOT NULL DEFAULT 'Plán',
        city TEXT NOT NULL DEFAULT '',
        code TEXT NOT NULL DEFAULT '',
        date TEXT,
        note TEXT,
        created_at TEXT NOT NULL DEFAULT (datetime('now'))
    );
    
    CREATE TABLE IF NOT EXISTS job_assignments (
        job_id INTEGER NOT NULL,
        employee_id INTEGER NOT NULL,
        PRIMARY KEY (job_id, employee_id)
    );
    CREATE TABLE IF NOT EXISTS job_employees (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        job_id INTEGER NOT NULL,
        employee_id INTEGER NOT NULL,
        role TEXT DEFAULT 'worker',
        assigned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (job_id) REFERENCES jobs(id),
        FOREIGN KEY (employee_id) REFERENCES employees(id)
    );
    CREATE TABLE IF NOT EXISTS job_materials (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        job_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        qty REAL NOT NULL DEFAULT 0,
        unit TEXT NOT NULL DEFAULT 'ks'
    );
    CREATE TABLE IF NOT EXISTS job_tools (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        job_id INTEGER NOT NULL,
        name TEXT NOT NULL,
        qty REAL NOT NULL DEFAULT 0,
        unit TEXT NOT NULL DEFAULT 'ks'
    );
    CREATE TABLE IF NOT EXISTS tasks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        job_id INTEGER,
        employee_id INTEGER,
        title TEXT NOT NULL,
        description TEXT DEFAULT '',
        status TEXT NOT NULL DEFAULT 'open',
        due_date TEXT,
        created_at TEXT NOT NULL DEFAULT (datetime('now'))
    );

    CREATE TABLE IF NOT EXISTS task_assignments (
        task_id INTEGER NOT NULL,
        employee_id INTEGER NOT NULL,
        is_primary INTEGER NOT NULL DEFAULT 0,
        PRIMARY KEY (task_id, employee_id)
    );
    CREATE TABLE IF NOT EXISTS issues (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        job_id INTEGER,
        title TEXT NOT NULL,
        description TEXT DEFAULT '',
        type TEXT DEFAULT 'issue',
        status TEXT NOT NULL DEFAULT 'open',
        severity TEXT DEFAULT '',
        assigned_to INTEGER,
        created_by INTEGER,
        created_at TEXT NOT NULL DEFAULT (datetime('now'))
    );
    CREATE TABLE IF NOT EXISTS issue_assignments (
        issue_id INTEGER NOT NULL,
        employee_id INTEGER NOT NULL,
        is_primary INTEGER NOT NULL DEFAULT 0,
        PRIMARY KEY (issue_id, employee_id)
    );

    CREATE TABLE IF NOT EXISTS timesheets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_id INTEGER NOT NULL,
        job_id INTEGER NOT NULL,
        date TEXT NOT NULL,
        hours REAL NOT NULL DEFAULT 0,
        place TEXT DEFAULT '',
        activity TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS calendar_events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL,
        title TEXT NOT NULL,
        kind TEXT NOT NULL DEFAULT 'note',
        job_id INTEGER,
        start_time TEXT,
        end_time TEXT,
        note TEXT DEFAULT '',
        color TEXT DEFAULT '#2e7d32'
    );

    CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        employee_id INTEGER,
        kind TEXT NOT NULL DEFAULT 'info',
        title TEXT NOT NULL DEFAULT '',
        body TEXT NOT NULL DEFAULT '',
        entity_type TEXT,
        entity_id INTEGER,
        is_read INTEGER NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL DEFAULT (datetime('now'))
    );
    CREATE INDEX IF NOT EXISTS idx_notifications_user_read ON notifications(user_id, is_read, created_at);
    CREATE INDEX IF NOT EXISTS idx_notifications_emp_read ON notifications(employee_id, is_read, created_at);

    -- Katalog rostlin pro autocomplete
    CREATE TABLE IF NOT EXISTS plant_catalog (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        latin_name TEXT NOT NULL,
        variety TEXT,
        container_size TEXT,
        flower_color TEXT,
        flowering_time TEXT,
        leaf_color TEXT,
        height TEXT,
        light_requirements TEXT,
        site_type TEXT,
        plants_per_m2 TEXT,
        hardiness_zone TEXT,
        notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(latin_name, variety)
    );
    CREATE INDEX IF NOT EXISTS idx_plant_catalog_latin ON plant_catalog(latin_name);
    CREATE INDEX IF NOT EXISTS idx_plant_catalog_variety ON plant_catalog(variety);

    -- Rostliny ve školce (inventář)
    CREATE TABLE IF NOT EXISTS nursery_plants (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        species TEXT NOT NULL,
        variety TEXT,
        quantity INTEGER DEFAULT 0,
        unit TEXT DEFAULT 'ks',
        stage TEXT DEFAULT 'sazenice',
        location TEXT,
        planted_date TEXT,
        ready_date TEXT,
        purchase_price REAL DEFAULT 0,
        selling_price REAL DEFAULT 0,
        flower_color TEXT,
        flowering_time TEXT,
        height TEXT,
        light_requirements TEXT,
        leaf_color TEXT,
        hardiness_zone TEXT,
        site_type TEXT,
        plants_per_m2 TEXT,
        botanical_notes TEXT,
        notes TEXT,
        status TEXT DEFAULT 'active',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    );
    CREATE INDEX IF NOT EXISTS idx_nursery_plants_species ON nursery_plants(species);
    CREATE INDEX IF NOT EXISTS idx_nursery_plants_status ON nursery_plants(status);
    """)
    
    # Vytvoření FTS (Full-Text Search) tabulky pro katalog rostlin
    try:
        db.executescript("""
        CREATE VIRTUAL TABLE IF NOT EXISTS plant_catalog_fts USING fts5(
            latin_name,
            variety,
            flower_color,
            notes,
            content=plant_catalog,
            content_rowid=id
        );
        
        -- Trigger pro automatickou aktualizaci FTS při INSERT
        CREATE TRIGGER IF NOT EXISTS plant_catalog_ai AFTER INSERT ON plant_catalog BEGIN
            INSERT INTO plant_catalog_fts(rowid, latin_name, variety, flower_color, notes)
            VALUES (new.id, new.latin_name, new.variety, new.flower_color, new.notes);
        END;
        
        -- Trigger pro automatickou aktualizaci FTS při DELETE
        CREATE TRIGGER IF NOT EXISTS plant_catalog_ad AFTER DELETE ON plant_catalog BEGIN
            DELETE FROM plant_catalog_fts WHERE rowid = old.id;
        END;
        
        -- Trigger pro automatickou aktualizaci FTS při UPDATE
        CREATE TRIGGER IF NOT EXISTS plant_catalog_au AFTER UPDATE ON plant_catalog BEGIN
            UPDATE plant_catalog_fts SET 
                latin_name = new.latin_name,
                variety = new.variety,
                flower_color = new.flower_color,
                notes = new.notes
            WHERE rowid = new.id;
        END;
        """)
        db.commit()
    except Exception as fts_err:
        print(f"[DB] FTS setup note: {fts_err}")
    
    # Migrace: přidání sloupců phone, email, address do employees (pokud neexistují)
    try:
        db.execute("ALTER TABLE employees ADD COLUMN phone TEXT DEFAULT ''")
    except Exception:
        pass  # Sloupec už existuje
    try:
        db.execute("ALTER TABLE employees ADD COLUMN email TEXT DEFAULT ''")
    except Exception:
        pass  # Sloupec už existuje
    try:
        db.execute("ALTER TABLE employees ADD COLUMN address TEXT DEFAULT ''")
    except Exception:
        pass  # Sloupec už existuje
    db.commit()

def seed_admin():
    db = get_db()
    cur = db.execute("SELECT COUNT(*) c FROM users")
    if cur.fetchone()["c"] == 0:
        db.execute(
            "INSERT INTO users(email,name,role,password_hash,active,created_at) VALUES (?,?,?,?,1,?)",
            (
                os.environ.get("ADMIN_EMAIL","admin@greendavid.local"),
                os.environ.get("ADMIN_NAME","Admin"),
                "owner",  # Vytvořit jako owner místo admin
                generate_password_hash(os.environ.get("ADMIN_PASSWORD","admin123"), method='pbkdf2:sha256'),
                datetime.utcnow().isoformat()
            )
        )
        db.commit()

def _auto_upgrade_admins_to_owner():
    """Automaticky upgrade admin účtů na owner při startu"""
    db = get_db()
    try:
        # Upgrade známých admin emailů
        admin_emails = ['admin@greendavid.local', 'david@greendavid.cz', 'admin@greendavid.cz', 'admin@example.com']
        for email in admin_emails:
            result = db.execute(
                "UPDATE users SET role = 'owner' WHERE email = ? AND (role IS NULL OR role != 'owner' OR role = 'admin')",
                (email,)
            )
            if result.rowcount > 0:
                db.commit()
                print(f"[DB] Auto-upgraded {email} to owner role")
        
        # Pokud žádný owner neexistuje, upgrade prvního uživatele
        owner_exists = db.execute("SELECT COUNT(*) as cnt FROM users WHERE role = 'owner'").fetchone()
        if owner_exists['cnt'] == 0:
            first_user = db.execute("SELECT id, email FROM users ORDER BY id LIMIT 1").fetchone()
            if first_user:
                db.execute("UPDATE users SET role = 'owner' WHERE id = ?", (first_user['id'],))
                db.commit()
                print(f"[DB] Auto-upgraded first user {first_user['email']} to owner role")
        
        # Upgrade všech admin účtů na owner
        result = db.execute(
            "UPDATE users SET role = 'owner' WHERE role = 'admin' OR role IS NULL"
        )
        if result.rowcount > 0:
            db.commit()
            print(f"[DB] Auto-upgraded {result.rowcount} admin/NULL accounts to owner role")
    except Exception as e:
        print(f"[DB] Warning: Could not auto-upgrade admin: {e}")

def seed_employees():
    db = get_db()
    cur = db.execute("SELECT COUNT(*) c FROM employees")
    if cur.fetchone()["c"] == 0:
        # Původní zaměstnanci
        employees = [
            ("david", "zahradník"),
            ("vendi", "zahradník"),
            ("jason", "zahradník"),
        ]
        for name, role in employees:
            db.execute("INSERT INTO employees(name,role) VALUES (?,?)", (name, role))
        db.commit()

def seed_plant_catalog():
    """Automaticky naplní katalog rostlin z JSON souboru pokud je prázdný"""
    db = get_db()
    cur = db.execute("SELECT COUNT(*) c FROM plant_catalog")
    if cur.fetchone()["c"] == 0:
        # Hledej JSON soubor s daty
        import json
        json_paths = [
            'plant_catalog_data.json',
            os.path.join(os.path.dirname(__file__), 'plant_catalog_data.json'),
            '/app/plant_catalog_data.json',
        ]
        
        for json_path in json_paths:
            if os.path.exists(json_path):
                try:
                    with open(json_path, 'r', encoding='utf-8') as f:
                        plants = json.load(f)
                    
                    imported = 0
                    for plant in plants:
                        try:
                            db.execute('''
                                INSERT OR IGNORE INTO plant_catalog 
                                (latin_name, variety, container_size, flower_color, flowering_time,
                                 leaf_color, height, light_requirements, site_type, plants_per_m2,
                                 hardiness_zone, notes)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                plant.get('latin_name'),
                                plant.get('variety'),
                                plant.get('container_size'),
                                plant.get('flower_color'),
                                plant.get('flowering_time'),
                                plant.get('leaf_color'),
                                plant.get('height'),
                                plant.get('light_requirements'),
                                plant.get('site_type'),
                                plant.get('plants_per_m2'),
                                plant.get('hardiness_zone'),
                                plant.get('notes')
                            ))
                            imported += 1
                        except Exception as e:
                            pass  # Skip duplicates or errors
                    
                    db.commit()
                    
                    # Rebuild FTS index
                    try:
                        db.execute("INSERT INTO plant_catalog_fts(plant_catalog_fts) VALUES('rebuild')")
                        db.commit()
                    except Exception:
                        pass
                    
                    print(f"[DB] Imported {imported} plants from {json_path}")
                    break
                except Exception as e:
                    print(f"[DB] Error importing plant catalog: {e}")

@app.before_request
def _ensure():
    # Run schema checks / migrations once per process to avoid unnecessary locking
    if not getattr(_ensure, "_schema_ready", False):
        ensure_schema()
        try:
            apply_migrations()
        except Exception as e:
            # Never break the app startup/runtime on a migration helper issue
            print(f"[DB] Migration failed: {e}")
        # Legacy migrations (kept for backward compatibility with older app.db variants)
        try:
            _migrate_completed_at()
            _migrate_employees_enhanced()
            _migrate_roles_and_hierarchy()
            _migrate_crew_control_tables()  # New: Crew Control System tables
        except Exception as e:
            print(f"[DB] Migration warning: {e}")
        _ensure._schema_ready = True
    seed_admin()
    _auto_upgrade_admins_to_owner()
    seed_employees()
    seed_plant_catalog()

def _migrate_crew_control_tables():
    """Create Crew Control System tables if they don't exist"""
    db = get_db()
    try:
        # Check if tables exist
        tables_to_create = [
            ("employee_skills", '''
                CREATE TABLE IF NOT EXISTS employee_skills (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL,
                    skill_type TEXT NOT NULL,
                    skill_name TEXT NOT NULL,
                    level INTEGER DEFAULT 1,
                    certified INTEGER DEFAULT 0,
                    certified_date TEXT,
                    certified_by TEXT,
                    notes TEXT,
                    created_at TEXT DEFAULT (datetime('now')),
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
                )
            '''),
            ("employee_certifications", '''
                CREATE TABLE IF NOT EXISTS employee_certifications (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL,
                    cert_name TEXT NOT NULL,
                    cert_type TEXT,
                    issued_date TEXT,
                    expiry_date TEXT,
                    issuer TEXT,
                    document_url TEXT,
                    status TEXT DEFAULT 'active',
                    created_at TEXT DEFAULT (datetime('now')),
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
                )
            '''),
            ("employee_preferences", '''
                CREATE TABLE IF NOT EXISTS employee_preferences (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL UNIQUE,
                    preferred_work_types TEXT,
                    avoided_work_types TEXT,
                    max_weekly_hours REAL DEFAULT 40,
                    preferred_locations TEXT,
                    travel_radius_km INTEGER DEFAULT 50,
                    prefers_team_work INTEGER DEFAULT 1,
                    prefers_solo_work INTEGER DEFAULT 0,
                    notes TEXT,
                    updated_at TEXT DEFAULT (datetime('now')),
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
                )
            '''),
            ("employee_capacity", '''
                CREATE TABLE IF NOT EXISTS employee_capacity (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL,
                    week_start TEXT NOT NULL,
                    planned_hours REAL DEFAULT 40,
                    assigned_hours REAL DEFAULT 0,
                    actual_hours REAL DEFAULT 0,
                    utilization_pct REAL DEFAULT 0,
                    status TEXT DEFAULT 'available',
                    notes TEXT,
                    created_at TEXT DEFAULT (datetime('now')),
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE,
                    UNIQUE(employee_id, week_start)
                )
            '''),
            ("employee_performance", '''
                CREATE TABLE IF NOT EXISTS employee_performance (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL,
                    period_start TEXT NOT NULL,
                    period_end TEXT NOT NULL,
                    total_hours REAL DEFAULT 0,
                    completed_tasks INTEGER DEFAULT 0,
                    on_time_rate REAL DEFAULT 100,
                    quality_score REAL DEFAULT 0,
                    efficiency_score REAL DEFAULT 0,
                    reliability_score REAL DEFAULT 0,
                    notes TEXT,
                    calculated_at TEXT DEFAULT (datetime('now')),
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
                )
            '''),
            ("employee_ai_scores", '''
                CREATE TABLE IF NOT EXISTS employee_ai_scores (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL,
                    score_date TEXT NOT NULL,
                    balance_score REAL DEFAULT 50,
                    workload_score REAL DEFAULT 50,
                    burnout_risk REAL DEFAULT 0,
                    recommendation TEXT,
                    factors TEXT,
                    created_at TEXT DEFAULT (datetime('now')),
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
                )
            '''),
            ("employee_availability", '''
                CREATE TABLE IF NOT EXISTS employee_availability (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    employee_id INTEGER NOT NULL,
                    date TEXT NOT NULL,
                    availability_type TEXT NOT NULL,
                    start_time TEXT,
                    end_time TEXT,
                    all_day INTEGER DEFAULT 1,
                    notes TEXT,
                    approved INTEGER DEFAULT 0,
                    approved_by INTEGER,
                    created_at TEXT DEFAULT (datetime('now')),
                    FOREIGN KEY (employee_id) REFERENCES employees(id) ON DELETE CASCADE
                )
            ''')
        ]
        
        for table_name, create_sql in tables_to_create:
            if not _table_exists(db, table_name):
                db.execute(create_sql)
                db.commit()
                print(f"[DB] Created table: {table_name}")
        
        # Add new columns to employees table
        if _table_exists(db, "employees"):
            cols = [r[1] for r in db.execute("PRAGMA table_info(employees)").fetchall()]
            new_cols = {
                "weekly_capacity": "REAL DEFAULT 40",
                "current_workload": "REAL DEFAULT 0",
                "ai_balance_score": "REAL DEFAULT 50",
                "burnout_risk": "REAL DEFAULT 0",
                "reliability_score": "REAL DEFAULT 0",
                "specializations": "TEXT",
                "availability_status": "TEXT DEFAULT 'available'",
            }
            for col_name, col_def in new_cols.items():
                if col_name not in cols:
                    try:
                        db.execute(f"ALTER TABLE employees ADD COLUMN {col_name} {col_def}")
                        db.commit()
                        print(f"[DB] Added column: employees.{col_name}")
                    except Exception as e:
                        pass  # Column might already exist
                        
    except Exception as e:
        print(f"[DB] Crew Control migration warning: {e}")

# ----------------- helpers for jobs schema compat -----------------
def _jobs_info():
    rows = get_db().execute("PRAGMA table_info(jobs)").fetchall()
    # rows: cid, name, type, notnull, dflt_value, pk
    return {r[1]: {"notnull": int(r[3])} for r in rows}

def _job_title_col():
    info = _jobs_info()
    if "title" in info:
        return "title"
    return "name" if "name" in info else "title"

def _job_select_all():
    info = _jobs_info()
    base_cols = "id, client, status, city, code, date, note"
    date_cols = ", created_date, start_date" if "created_date" in info else ""
    deadline_col = ", deadline" if "deadline" in info else ""
    address_col = ", address" if "address" in info else ""
    progress_col = ", progress" if "progress" in info else ""
    time_col = ", time_spent_minutes" if "time_spent_minutes" in info else ""
    budget_col = ", budget" if "budget" in info else ""
    cost_col = ", cost_spent" if "cost_spent" in info else ", actual_cost" if "actual_cost" in info else ""
    if "title" in info:
        return f"SELECT title, {base_cols}{date_cols}{deadline_col}{address_col}{progress_col}{time_col}{budget_col}{cost_col} FROM jobs"
    if "name" in info:
        return f"SELECT name AS title, {base_cols}{date_cols}{deadline_col}{address_col}{progress_col}{time_col}{budget_col}{cost_col} FROM jobs"
    return f"SELECT '' AS title, {base_cols}{date_cols}{deadline_col}{address_col}{progress_col}{time_col}{budget_col}{cost_col} FROM jobs"

def _job_insert_cols_and_vals(title, client, status, city, code, dt, note, owner_id=None):
    info = _jobs_info()
    cols = []
    vals = []
    # Keep legacy 'name' in sync if present
    if "title" in info:
        cols.append("title"); vals.append(title)
    if "name" in info:
        cols.append("name"); vals.append(title)
    cols += ["client","status","city","code","date","note"]
    vals += [client, status, city, code, dt, note]
    # legacy NOT NULL columns without defaults
    now = datetime.utcnow().isoformat()
    if "created_at" in info:
        cols.append("created_at"); vals.append(now)
    if "updated_at" in info:
        cols.append("updated_at"); vals.append(now)
    # legacy owner_id
    if "owner_id" in info:
        if owner_id is None:
            cu = current_user()
            owner_id = cu["id"] if cu else None
        cols.append("owner_id"); vals.append(int(owner_id) if owner_id is not None else None)
    return cols, vals

def _job_title_update_set(params_list, title_value):
    info = _jobs_info()
    sets = []
    if "title" in info:
        sets.append("title=?"); params_list.append(title_value)
    if "name" in info:
        sets.append("name=?"); params_list.append(title_value)
    return sets

# ----------------- static -----------------
@app.route("/")
def index():
    return send_from_directory(".", "index.html")

# Work inbox (safe standalone page)
@app.route("/inbox")
@app.route("/inbox/")
@app.route("/inbox.html")
def work_inbox_page():
    return send_from_directory(".", "inbox.html")

@app.route("/static/<path:filename>")
def static_files(filename):
    """Serve static files from static/ directory"""
    return send_from_directory("static", filename)

@app.route("/uploads/<path:name>")
def uploaded_file(name):
    safe = re.sub(r"[^a-zA-Z0-9._-]", "_", name)
    path = os.path.join(UPLOAD_DIR, safe)
    if not os.path.isfile(path):
        abort(404)
    return send_from_directory(UPLOAD_DIR, safe)

@app.route("/health")
def health():
    return {"status": "ok"}

# ----------------- APIs -----------------
@app.route("/api/me")
def api_me():
    u = current_user()
    unread = 0
    emp = None
    if u:
        try:
            db = get_db()
            # Current employee mapping (optional)
            try:
                erow = db.execute(
                    "SELECT id, name, role FROM employees WHERE user_id=? LIMIT 1",
                    (int(u["id"]),),
                ).fetchone()
                if erow:
                    emp = {"id": int(erow["id"]), "name": erow["name"], "role": erow["role"]}
                    # Backward-compatible field used by some frontends
                    u["employee_id"] = emp["id"]
            except Exception:
                emp = None
            # Prefer user_id binding; fallback also checks employee mapping.
            # NOTE: notifications table is created in ensure_schema; if legacy DB lacks it, this is safe.
            unread = db.execute(
                "SELECT COUNT(1) FROM notifications WHERE (user_id=? OR employee_id IN (SELECT id FROM employees WHERE user_id=?)) AND is_read=0",
                (int(u["id"]), int(u["id"])),
            ).fetchone()[0]
            unread = int(unread or 0)
        except Exception:
            unread = 0
    return jsonify({"ok": True, "authenticated": bool(u), "user": u, "employee": emp, "tasks_count": 0, "unread_notifications": unread})


def create_job_notification(user_id, job_id, kind, title, body, entity_type='job', entity_id=None):
    """Helper function to create a notification for a job"""
    try:
        db = get_db()
        # Check if notification already exists
        existing = db.execute("""
            SELECT id FROM notifications
            WHERE user_id = ? AND entity_type = ? AND entity_id = ? AND kind = ? AND is_read = 0
            AND created_at >= datetime('now', '-1 day')
        """, (user_id, entity_type, entity_id or job_id, kind)).fetchone()
        
        if not existing:
            db.execute("""
                INSERT INTO notifications (user_id, kind, title, body, entity_type, entity_id, is_read, created_at)
                VALUES (?, ?, ?, ?, ?, ?, 0, datetime('now'))
            """, (user_id, kind, title, body, entity_type, entity_id or job_id))
            db.commit()
            return True
    except Exception as e:
        print(f"[NOTIF] Error creating notification: {e}")
    return False

def generate_auto_notifications(user_id):
    """Automaticky generuje notifikace pro blížící se termíny a důležité události."""
    db = get_db()
    from datetime import datetime, timedelta
    now = datetime.now()
    today = now.strftime('%Y-%m-%d')
    tomorrow = (now + timedelta(days=1)).strftime('%Y-%m-%d')
    week_later = (now + timedelta(days=7)).strftime('%Y-%m-%d')
    
    # Získej existující notifikace, abychom nevytvářeli duplikáty
    existing = db.execute(
        "SELECT entity_type, entity_id, kind FROM notifications WHERE user_id=? AND created_at > datetime('now', '-24 hours')",
        (user_id,)
    ).fetchall()
    existing_keys = set((r['entity_type'], r['entity_id'], r['kind']) for r in existing)
    
    notifications_to_add = []
    
    # 1. Úkoly s blížícím se termínem (dnes nebo zítra)
    tasks = db.execute("""
        SELECT t.*, j.client as job_name FROM tasks t
        LEFT JOIN jobs j ON t.job_id = j.id
        WHERE t.status != 'done' AND t.deadline IS NOT NULL
        AND date(t.deadline) BETWEEN ? AND ?
    """, (today, tomorrow)).fetchall()
    
    for task in tasks:
        if ('task', task['id'], 'deadline') not in existing_keys:
            deadline_date = task['deadline'][:10] if task['deadline'] else ''
            is_today = deadline_date == today
            notifications_to_add.append({
                'user_id': user_id,
                'title': '⏰ Termín ' + ('DNES!' if is_today else 'zítra'),
                'body': f"{task['title']}" + (f" ({task['job_name']})" if task['job_name'] else ""),
                'kind': 'deadline',
                'entity_type': 'task',
                'entity_id': task['id']
            })
    
    # 2. Zakázky bez přiřazených zaměstnanců
    jobs_no_employees = db.execute("""
        SELECT j.* FROM jobs j
        LEFT JOIN job_employees je ON j.id = je.job_id
        WHERE j.status NOT IN ('completed', 'archived', 'cancelled')
        AND je.id IS NULL
        LIMIT 5
    """).fetchall()
    
    for job in jobs_no_employees:
        if ('job', job['id'], 'warning') not in existing_keys:
            job_name = job['client'] or job['name'] or f"#{job['id']}"
            notifications_to_add.append({
                'user_id': user_id,
                'title': '⚠️ Chybí přiřazení',
                'body': f"Zakázka '{job_name}' nemá přiřazené zaměstnance",
                'kind': 'warning',
                'entity_type': 'job',
                'entity_id': job['id']
            })
    
    # 3. Zakázky s blížícím se termínem (do 7 dnů)
    jobs_deadline = db.execute("""
        SELECT * FROM jobs 
        WHERE status NOT IN ('completed', 'archived', 'cancelled')
        AND deadline IS NOT NULL
        AND date(deadline) BETWEEN ? AND ?
    """, (today, week_later)).fetchall()
    
    for job in jobs_deadline:
        if ('job', job['id'], 'deadline') not in existing_keys:
            deadline_date = job['deadline'][:10] if job['deadline'] else ''
            days_left = (datetime.strptime(deadline_date, '%Y-%m-%d') - now).days
            job_name = job['client'] or job['name'] or f"Zakázka #{job['id']}"
            notifications_to_add.append({
                'user_id': user_id,
                'title': '📅 Termín zakázky' + (' DNES!' if days_left == 0 else f' za {days_left} dní'),
                'body': job_name,
                'kind': 'deadline',
                'entity_type': 'job',
                'entity_id': job['id']
            })
    
    # 4. Nízké zásoby materiálu
    try:
        low_stock = db.execute("""
            SELECT * FROM warehouse_items 
            WHERE quantity <= min_quantity AND min_quantity > 0
            LIMIT 5
        """).fetchall()
        
        for item in low_stock:
            if ('stock', item['id'], 'stock') not in existing_keys:
                notifications_to_add.append({
                    'user_id': user_id,
                    'title': '📦 Nízké zásoby',
                    'body': f"{item['name']}: zbývá {item['quantity']} {item.get('unit', 'ks')}",
                    'kind': 'stock',
                    'entity_type': 'stock',
                    'entity_id': item['id']
                })
    except:
        pass  # Tabulka warehouse_items nemusí existovat
    
    # Job-specific critical notifications
    try:
        critical_jobs = db.execute("""
            SELECT j.id, j.title, j.deadline, j.budget, j.actual_value, j.status
            FROM jobs j
            WHERE j.status NOT IN ('completed', 'archived', 'cancelled')
        """, ()).fetchall()
        
        for job in critical_jobs:
            job = dict(job)
            deadline = job.get('deadline')
            budget = job.get('budget') or 0
            actual = job.get('actual_value') or 0
            
            # Deadline critical
            if deadline:
                try:
                    deadline_dt = datetime.strptime(deadline[:10], '%Y-%m-%d').date()
                    days_until = (deadline_dt - today).days
                    
                    if days_until < 0:
                        create_job_notification(
                            user_id, job['id'], 'deadline',
                            f'🚨 Zakázka "{job.get("title") or job.get("name") or "Bez názvu"}" je po termínu',
                            f'Zakázka je {abs(days_until)} dní po termínu. Okamžitá akce vyžadována.',
                            'job', job['id']
                        )
                    elif days_until <= 3:
                        create_job_notification(
                            user_id, job['id'], 'deadline',
                            f'⏰ Kritický deadline: "{job.get("title") or job.get("name") or "Bez názvu"}"',
                            f'Zbývá pouze {days_until} dní do deadline. Zkontrolujte průběh.',
                            'job', job['id']
                        )
                except:
                    pass
            
            # Budget critical
            if budget > 0:
                spent_pct = (actual / budget) * 100
                if spent_pct > 90:
                    create_job_notification(
                        user_id, job['id'], 'budget',
                        f'💰 Rozpočet téměř vyčerpán: "{job.get("title") or job.get("name") or "Bez názvu"}"',
                        f'Rozpočet je vyčerpán na {spent_pct:.1f}%. Zkontrolujte další výdaje.',
                        'job', job['id']
                    )
    except Exception as e:
        print(f"[NOTIF] Job notifications error: {e}")
    
    # Vlož nové notifikace
    for n in notifications_to_add[:10]:  # Max 10 najednou
        try:
            db.execute("""
                INSERT INTO notifications (user_id, title, body, kind, entity_type, entity_id, is_read, created_at)
                VALUES (?, ?, ?, ?, ?, ?, 0, datetime('now'))
            """, (n['user_id'], n['title'], n['body'], n['kind'], n['entity_type'], n['entity_id']))
        except:
            pass
    
    if notifications_to_add:
        db.commit()


@app.route("/api/notifications", methods=["GET", "PATCH", "DELETE"])
def api_notifications():
    """In-app notifications for the current signed-in user.

    GET: list notifications
      - unread_only=1
      - limit=50 (max 200)
    PATCH: mark read
      - id: single notification id
      - all=1: mark all as read
    DELETE: delete
      - id: single id
      - all=1: delete all (dangerous - owner only)
    """
    u, err = require_auth()
    if err:
        return err
    db = get_db()

    if request.method == "GET":
        # Automaticky generuj notifikace
        try:
            generate_auto_notifications(int(u["id"]))
        except Exception as e:
            print(f"[NOTIF] Auto-generate error: {e}")
        
        unread_only = str(request.args.get("unread_only") or "").strip() in ("1", "true", "yes")
        limit = request.args.get("limit", type=int) or 50
        limit = max(1, min(int(limit), 200))

        conds = ["(n.user_id=? OR n.employee_id IN (SELECT id FROM employees WHERE user_id=?))"]
        params = [int(u["id"]), int(u["id"])]
        if unread_only:
            conds.append("n.is_read=0")

        q = "SELECT n.* FROM notifications n WHERE " + " AND ".join(conds) + " ORDER BY datetime(n.created_at) DESC, n.id DESC LIMIT ?"
        params.append(limit)
        rows = [dict(r) for r in db.execute(q, params).fetchall()]
        return jsonify({"ok": True, "rows": rows})

    if request.method == "PATCH":
        data = request.get_json(force=True, silent=True) or {}
        nid = data.get("id") or request.args.get("id", type=int)
        mark_all = str(data.get("all") or request.args.get("all") or "").strip() in ("1", "true", "yes")

        try:
            if mark_all:
                db.execute(
                    "UPDATE notifications SET is_read=1 WHERE (user_id=? OR employee_id IN (SELECT id FROM employees WHERE user_id=?))",
                    (int(u["id"]), int(u["id"])),
                )
                db.commit()
                return jsonify({"ok": True})
            if not nid:
                return jsonify({"ok": False, "error": "missing_id"}), 400
            db.execute(
                "UPDATE notifications SET is_read=1 WHERE id=? AND (user_id=? OR employee_id IN (SELECT id FROM employees WHERE user_id=?))",
                (int(nid), int(u["id"]), int(u["id"])),
            )
            db.commit()
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            return jsonify({"ok": False, "error": str(e)}), 500

    # DELETE
    data = request.get_json(force=True, silent=True) or {}
    nid = data.get("id") or request.args.get("id", type=int)
    delete_all = str(data.get("all") or request.args.get("all") or "").strip() in ("1", "true", "yes")

    # For safety: allow deleting all only for owner.
    if delete_all and normalize_role(u.get("role")) not in ("owner", "admin"):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    try:
        if delete_all:
            db.execute(
                "DELETE FROM notifications WHERE (user_id=? OR employee_id IN (SELECT id FROM employees WHERE user_id=?))",
                (int(u["id"]), int(u["id"])),
            )
            db.commit()
            return jsonify({"ok": True})
        if not nid:
            return jsonify({"ok": False, "error": "missing_id"}), 400
        db.execute(
            "DELETE FROM notifications WHERE id=? AND (user_id=? OR employee_id IN (SELECT id FROM employees WHERE user_id=?))",
            (int(nid), int(u["id"]), int(u["id"])),
        )
        db.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

# auth
@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    db = get_db()
    row = db.execute("SELECT id,email,name,role,password_hash,active FROM users WHERE email=?", (email,)).fetchone()
    
    if not row or not row["active"]:
        return jsonify({"ok": False, "error": "invalid_credentials"}), 401
    
    # Safe password verification - handle scrypt incompatibility on Python 3.9
    try:
        password_valid = check_password_hash(row["password_hash"], password)
    except (AttributeError, ValueError) as e:
        # Scrypt not available in Python 3.9 - regenerate hash with pbkdf2
        print(f"[AUTH] Password hash error for {email}: {e}")
        # For security, reject and force password reset
        return jsonify({"ok": False, "error": "password_hash_incompatible"}), 401
    
    if not password_valid:
        return jsonify({"ok": False, "error": "invalid_credentials"}), 401
    
    # Normalizovat roli (admin -> owner, NULL -> owner)
    user_role = row["role"] or "owner"
    if user_role == "admin":
        user_role = "owner"
    
    # Auto-upgrade admin účtů na owner při přihlášení
    if row["role"] in (None, "admin"):
        db.execute("UPDATE users SET role = 'owner' WHERE id = ?", (row["id"],))
        db.commit()
        user_role = "owner"
        print(f"[DB] Auto-upgraded {email} to owner role on login")
    
    # store user id and role in session
    session["uid"] = row["id"]
    session["user_name"] = row["name"]
    session["user_email"] = row["email"]
    session["user_role"] = user_role
    return jsonify({"ok": True})

@app.route("/api/logout", methods=["POST"])
def api_logout():
    # remove user id from session
    session.pop("uid", None)
    return jsonify({"ok": True})

# employees
@app.route("/api/employees", methods=["GET","POST","PATCH","DELETE"])
@requires_role('owner', 'admin', 'manager', 'lander', 'worker')
def api_employees():
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    
    db = get_db()
    if request.method == "GET":
        # Všichni přihlášení uživatelé mohou číst seznam zaměstnanců; u každého zaměstnance doplníme informaci o účtu (pokud existuje).
        rows = db.execute("""
            SELECT e.*,
                   u.id   AS account_user_id,
                   u.email AS account_email,
                   u.role AS account_role,
                   u.active AS account_active
            FROM employees e
            LEFT JOIN users u ON u.id = e.user_id
            ORDER BY e.id DESC
        """).fetchall()

        employees = []
        for r in rows:
            emp = dict(r)
            emp_id = emp["id"]

            # Normalizace role zaměstnance (aby UI bylo konzistentní)
            emp["role"] = normalize_employee_role(emp.get("role"))

            # Účet zaměstnance (pokud je navázaný přes employees.user_id)
            emp["has_account"] = emp.get("account_user_id") is not None
            if emp.get("account_role") is not None:
                emp["account_role"] = normalize_role(emp.get("account_role"))
            
            # Vypočítej hodiny tento týden
            from datetime import datetime, timedelta
            today = datetime.now()
            monday = today - timedelta(days=today.weekday())
            week_start = monday.strftime("%Y-%m-%d")
            week_end = (monday + timedelta(days=6)).strftime("%Y-%m-%d")
            
            timesheet_rows = db.execute(
                "SELECT SUM(hours) as total FROM timesheets WHERE employee_id=? AND date >= ? AND date <= ?",
                (emp_id, week_start, week_end)
            ).fetchone()
            emp["hours_week"] = round(timesheet_rows["total"] or 0, 1)
            
            # Spočítej aktivní zakázky (kde je zaměstnanec přiřazen)
            job_rows = db.execute(
                "SELECT COUNT(DISTINCT job_id) as count FROM job_assignments WHERE employee_id=?",
                (emp_id,)
            ).fetchone()
            emp["active_projects"] = job_rows["count"] or 0
            
            # Spočítej dokončené úkoly
            task_rows = db.execute(
                "SELECT COUNT(*) as count FROM tasks WHERE employee_id=? AND status='completed'",
                (emp_id,)
            ).fetchone()
            emp["completed_tasks"] = task_rows["count"] or 0
            
            # Status (online pokud má výkazy za posledních 24h)
            recent_timesheet = db.execute(
                "SELECT COUNT(*) as count FROM timesheets WHERE employee_id=? AND date >= date('now', '-1 day')",
                (emp_id,)
            ).fetchone()
            emp["status"] = "online" if (recent_timesheet["count"] or 0) > 0 else "offline"
            
            # NOVÉ: Rozšířený profil z Crew Control System (pokud existuje)
            try:
                profile = db.execute(
                    "SELECT * FROM team_member_profile WHERE employee_id = ?",
                    (emp_id,)
                ).fetchone()
                
                if profile:
                    profile_dict = dict(profile)
                    # Parse JSON fields
                    import json as json_lib
                    try:
                        profile_dict['skills'] = json_lib.loads(profile_dict.get('skills') or '[]')
                        profile_dict['certifications'] = json_lib.loads(profile_dict.get('certifications') or '[]')
                        profile_dict['preferred_work_types'] = json_lib.loads(profile_dict.get('preferred_work_types') or '[]')
                    except:
                        profile_dict['skills'] = []
                        profile_dict['certifications'] = []
                        profile_dict['preferred_work_types'] = []
                    
                    # Vypočítej kapacitu
                    current_week_hours = emp.get("hours_week", 0)
                    weekly_capacity = profile_dict.get('weekly_capacity_hours', 40.0)
                    capacity_percent = (current_week_hours / weekly_capacity * 100) if weekly_capacity > 0 else 0
                    
                    # Urči status kapacity
                    if capacity_percent > 90:
                        capacity_status = 'overloaded'
                    elif capacity_percent < 50:
                        capacity_status = 'underutilized'
                    else:
                        capacity_status = 'normal'
                    
                    emp["profile"] = {
                        'skills': profile_dict['skills'],
                        'certifications': profile_dict['certifications'],
                        'weekly_capacity_hours': weekly_capacity,
                        'preferred_work_types': profile_dict['preferred_work_types'],
                        'performance_stability_score': profile_dict.get('performance_stability_score', 0.5),
                        'ai_balance_score': profile_dict.get('ai_balance_score', 0.5),
                        'burnout_risk_level': profile_dict.get('burnout_risk_level', 'normal'),
                        'total_jobs_completed': profile_dict.get('total_jobs_completed', 0),
                        'current_active_jobs': profile_dict.get('current_active_jobs', 0),
                        'capacity_percent': round(capacity_percent, 1),
                        'capacity_status': capacity_status
                    }
            except Exception as e:
                # Pokud tabulka ještě neexistuje nebo je chyba, prostě přeskočíme
                pass
            
            employees.append(emp)
        
        return jsonify({"ok": True, "employees": employees})
    if request.method == "POST":
        # Pouze owner a manager mohou přidávat zaměstnance
        user_role = normalize_role(user.get('role')) or 'owner'
        if user_role not in ('owner','admin','manager'):
            return jsonify({"ok": False, "error": "forbidden"}), 403
        try:
            data = request.get_json(force=True, silent=True) or {}
            name = data.get("name")
            if not name: return jsonify({"ok": False, "error":"invalid_input"}), 400
            
            # Get all fields with defaults
            role = normalize_employee_role(data.get("role") or "worker")
            phone = data.get("phone") or ""
            email = data.get("email") or ""
            address = data.get("address") or ""
            birth_date = data.get("birth_date") or None
            contract_type = data.get("contract_type") or "HPP"
            start_date = data.get("start_date") or None
            hourly_rate = data.get("hourly_rate") or None
            salary = data.get("salary") or None
            skills = data.get("skills") or ""
            location = data.get("location") or ""
            status = data.get("status") or "active"
            rating = data.get("rating") or 0
            avatar_url = data.get("avatar_url") or ""

            # Delegace (volitelné)
            delegate_employee_id = data.get("delegate_employee_id")
            approver_delegate_employee_id = data.get("approver_delegate_employee_id")
            
            # Build INSERT query with all available columns
            cols = [r[1] for r in db.execute("PRAGMA table_info(employees)").fetchall()]
            insert_cols = ["name", "role"]
            insert_vals = [name, role]
            
            for col in ["phone", "email", "address", "birth_date", "contract_type", "start_date", 
                       "hourly_rate", "salary", "skills", "location", "status", "rating", "avatar_url",
                       "delegate_employee_id", "approver_delegate_employee_id"]:
                if col in cols:
                    insert_cols.append(col)
                    if col == "hourly_rate" or col == "salary" or col == "rating":
                        insert_vals.append(hourly_rate if col == "hourly_rate" else (salary if col == "salary" else rating))
                    else:
                        val = locals().get(col, "")
                        insert_vals.append(val if val else None)
            
            placeholders = ",".join(["?"] * len(insert_vals))
            db.execute(f"INSERT INTO employees({','.join(insert_cols)}) VALUES ({placeholders})", insert_vals)
            db.commit()
            print(f"✓ Employee '{name}' created successfully")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error creating employee: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500
    if request.method == "PATCH":
        # Pouze owner a manager mohou upravovat zaměstnance
        user_role = normalize_role(user.get('role')) or 'owner'
        if user_role not in ('owner','admin','manager'):
            return jsonify({"ok": False, "error": "forbidden"}), 403
        try:
            data = request.get_json(force=True, silent=True) or {}
            eid = data.get("id") or request.args.get("id", type=int)
            if not eid: return jsonify({"ok": False, "error":"missing_id"}), 400
            updates = []
            params = []
            # Get all available columns from schema
            cols = [r[1] for r in db.execute("PRAGMA table_info(employees)").fetchall()]
            allowed = ["name", "role", "phone", "email", "address", "birth_date", "contract_type", 
                      "start_date", "hourly_rate", "salary", "skills", "location", "status", "rating", "avatar_url",
                      "delegate_employee_id", "approver_delegate_employee_id"]
            for k in allowed:
                if k in data and k in cols:
                    updates.append(f"{k}=?")
                    # Handle numeric fields
                    if k in ["hourly_rate", "salary", "rating"]:
                        params.append(float(data[k]) if data[k] is not None else None)
                    elif k == "role":
                        params.append(normalize_employee_role(data[k]))
                    else:
                        params.append(data[k] if data[k] is not None else "")
            if not updates: return jsonify({"ok": False, "error":"no_updates"}), 400
            params.append(eid)
            db.execute(f"UPDATE employees SET {', '.join(updates)} WHERE id=?", tuple(params))
            db.commit()
            print(f"✓ Employee {eid} updated successfully")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error updating employee: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500
    if request.method == "DELETE":
        # Pouze owner a manager mohou mazat zaměstnance
        user_role = normalize_role(user.get('role')) or 'owner'
        if user_role not in ('owner','admin','manager'):
            return jsonify({"ok": False, "error": "forbidden"}), 403
        try:
            eid = request.args.get("id", type=int)
            if not eid: return jsonify({"ok": False, "error":"missing_id"}), 400
            db.execute("DELETE FROM employees WHERE id=?", (eid,))
            db.commit()
            print(f"✓ Employee {eid} deleted successfully")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error deleting employee: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

# ----------------- Users management API (system users) -----------------
@app.route("/api/users", methods=["GET"])
@requires_role('owner', 'admin', 'manager', 'lander')
def api_get_users():
    """Seznam uživatelů systému - filtrovaný podle role"""
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    
    db = get_db()
    
    if user['role'] in ('owner', 'admin', 'manager'):
        # Owner a Manager vidí všechny uživatele
        rows = db.execute("""
            SELECT u.id, u.email, u.name, u.role, u.manager_id, u.active,
                   m.name AS manager_name
            FROM users u
            LEFT JOIN users m ON m.id = u.manager_id
            ORDER BY u.name
        """).fetchall()
    elif normalize_role(user.get('role')) == 'lander':
        # Team lead vidí jen své podřízené
        rows = db.execute("""
            SELECT u.id, u.email, u.name, u.role, u.manager_id, u.active,
                   m.name AS manager_name
            FROM users u
            LEFT JOIN users m ON m.id = u.manager_id
            WHERE u.manager_id = ?
            ORDER BY u.name
        """, (user['id'],)).fetchall()
    else:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    
    users = [dict(r) for r in rows]
    return jsonify({"ok": True, "users": users})

@app.route("/api/users", methods=["POST"])
@requires_role('owner','admin', 'manager')
def api_add_user():
    """Přidání nového uživatele - jen owner a manager"""
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    
    data = request.get_json(force=True, silent=True) or {}
    name = data.get("name", "").strip()
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")
    role = normalize_role(data.get("role", "worker"))
    manager_id = data.get("manager_id")
    employee_id = data.get("employee_id")
    
    if not name or not email or not password:
        return jsonify({"ok": False, "error": "missing_fields"}), 400
    
    if normalize_role(role) not in ROLES:
        return jsonify({"ok": False, "error": "invalid_role"}), 400
    
    # Manager nemůže vytvořit owner/admin účty
    if normalize_role(user.get('role')) not in ('owner','admin') and role in ('owner','admin'):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    
    # Pokud není zadán manager_id, použije se aktuální uživatel jako manager
    if manager_id is None:
        manager_id = user['id']
    
    db = get_db()
    try:
        db.execute(
            "INSERT INTO users (name, email, role, password_hash, manager_id, active) VALUES (?, ?, ?, ?, ?, 1)",
            (name, email, role, generate_password_hash(password, method='pbkdf2:sha256'), manager_id)
        )
        db.commit()
        new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Volitelné navázání účtu na zaměstnance (employees.user_id)
        if employee_id is not None:
            try:
                db.execute("UPDATE employees SET user_id = ? WHERE id = ?", (new_id, int(employee_id)))
                db.commit()
            except Exception:
                db.rollback()
        return jsonify({"ok": True, "id": new_id})
    except sqlite3.IntegrityError:
        db.rollback()
        return jsonify({"ok": False, "error": "email_exists"}), 400
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/user/settings", methods=["GET", "PATCH"])
def api_user_settings():
    """API pro získání a aktualizaci uživatelských nastavení."""
    u, err = require_auth()
    if err:
        return err
    
    db = get_db()
    user_id = session.get('uid')
    
    if request.method == 'GET':
        # Získat nastavení
        settings = db.execute(
            "SELECT mobile_mode FROM user_settings WHERE user_id = ?",
            (user_id,)
        ).fetchone()
        
        if settings:
            return jsonify({
                'ok': True,
                'mobile_mode': settings[0]
            })
        else:
            # Vrátit default podle role
            from app.utils.mobile_mode import get_mobile_mode
            return jsonify({
                'ok': True,
                'mobile_mode': get_mobile_mode()
            })
    
    elif request.method == 'PATCH':
        # Aktualizovat nastavení
        data = request.get_json() or {}
        mobile_mode = data.get('mobile_mode')
        
        if mobile_mode not in ['field', 'full', 'auto']:
            return jsonify({'ok': False, 'error': 'Invalid mobile_mode'}), 400
        
        try:
            # Zkontroluj, zda už existuje záznam
            existing = db.execute(
                "SELECT id FROM user_settings WHERE user_id = ?",
                (user_id,)
            ).fetchone()
            
            if existing:
                # Update existujícího záznamu
                db.execute(
                    "UPDATE user_settings SET mobile_mode = ?, updated_at = datetime('now') WHERE user_id = ?",
                    (mobile_mode, user_id)
                )
            else:
                # Vytvoř nový záznam
                db.execute(
                    "INSERT INTO user_settings (user_id, mobile_mode, updated_at) VALUES (?, ?, datetime('now'))",
                    (user_id, mobile_mode)
                )
            
            db.commit()
            
            # Nastav cookie pro rychlý přístup (backup pokud DB není dostupná)
            response = jsonify({'ok': True, 'mobile_mode': mobile_mode})
            response.set_cookie('mobile_mode', mobile_mode, max_age=31536000, httponly=True, samesite='Lax')
            
            return response
        except Exception as e:
            print(f"[ERROR] Failed to update user settings: {e}")
            return jsonify({'ok': False, 'error': str(e)}), 500

@app.route("/api/users/<int:user_id>/role", methods=["PUT"])
@requires_role('owner','admin')
def api_change_user_role(user_id):
    """Změna role uživatele - jen owner"""
    data = request.get_json(force=True, silent=True) or {}
    new_role = normalize_role(data.get("role"))

    if new_role not in ROLES:
        return jsonify({"ok": False, "error": "invalid_role"}), 400
    
    db = get_db()
    try:
        db.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, user_id))
        db.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

# ----------------- Timesheets approval -----------------
@app.route("/api/timesheets/<int:timesheet_id>/approve", methods=["POST"])
@requires_role('owner', 'admin', 'manager', 'lander')
def api_approve_timesheet(timesheet_id):
    """Schválení výkazu - jen pro nadřízené"""
    user = get_current_user()
    if not user:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    
    db = get_db()
    
    # Získat timesheet
    timesheet = db.execute(
        "SELECT employee_id FROM timesheets WHERE id = ?",
        (timesheet_id,)
    ).fetchone()
    
    if not timesheet:
        return jsonify({"ok": False, "error": "not_found"}), 404
    
    # Kontrola oprávnění - owner může všechno
    if user['role'] not in ('owner', 'admin'):
        # Pro ostatní role potřebujeme zkontrolovat hierarchii
        # Pokud timesheets nemá přímý link na users, použijeme employees
        # Prozatím povolíme všem nadřízeným
        pass
    
    # Zkontrolovat, zda timesheets tabulka má sloupce pro schválení
    cols = [r[1] for r in db.execute("PRAGMA table_info(timesheets)").fetchall()]
    
    if 'approved' not in cols:
        db.execute("ALTER TABLE timesheets ADD COLUMN approved INTEGER DEFAULT 0")
        db.execute("ALTER TABLE timesheets ADD COLUMN approved_by INTEGER NULL")
        db.execute("ALTER TABLE timesheets ADD COLUMN approved_at TEXT NULL")
        db.commit()
    
    # Schválit
    try:
        db.execute(
            "UPDATE timesheets SET approved = 1, approved_by = ?, approved_at = datetime('now') WHERE id = ?",
            (user['id'], timesheet_id)
        )
        db.commit()
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500

# ✅ jobs with legacy schema compatibility
@app.route("/api/jobs", methods=["GET","POST","PATCH","DELETE"])
def api_jobs():
    db = get_db()
    if request.method == "GET":
        rows = [dict(r) for r in db.execute(_job_select_all() + " ORDER BY date(date) DESC, id DESC").fetchall()]
        for r in rows:
            if "date" in r and r["date"]:
                r["date"] = _normalize_date(r["date"])
        # hide completed jobs from main list (they are visible only in archive)
        visible = []
        for r in rows:
            status = (r.get("status") or "").strip().lower()
            if status.startswith("dokon"):  # "Dokončeno"
                continue
            visible.append(r)
        return jsonify({"ok": True, "jobs": visible})

    # write operations require manager/admin
    u, err = require_role(write=True)
    if err: return err

    data = request.get_json(force=True, silent=True) or {}
    title = (data.get("title") or "").strip()
    client = (data.get("client") or "").strip()
    status = (data.get("status") or "Plán").strip()
    city   = (data.get("city")   or "").strip()
    code   = (data.get("code")   or "").strip()
    note   = data.get("note") or ""
    dt     = _normalize_date(data.get("date"))

    if request.method == "POST":
        try:
            req = [title, city, code, dt]
            if not all((v is not None and str(v).strip()!='') for v in req):
                return jsonify({"ok": False, "error":"missing_fields"}), 400
            cols, vals = _job_insert_cols_and_vals(title, client, status, city, code, dt, note, owner_id=u["id"])
            sql = "INSERT INTO jobs(" + ",".join(cols) + ") VALUES (" + ",".join(["?"]*len(vals)) + ")"
            cur = db.execute(sql, vals)
            job_id = cur.lastrowid
            db.commit()
            audit_event(u.get("id"), "create", "job", job_id, after={"title": title, "client": client, "status": status, "city": city, "code": code, "date": dt})
            print(f"✓ Job '{title}' created successfully")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error creating job: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    if request.method == "PATCH":
        try:
            jid = data.get("id")
            if not jid: return jsonify({"ok": False, "error":"missing_id"}), 400
            updates = []; params = []
            if "title" in data and data["title"] is not None:
                updates += _job_title_update_set(params, data["title"])
            for f in ("client","status","city","code","date","note","created_date","start_date","progress","deadline","address","location","estimated_value","budget"):
                if f in data:
                    if f in ("date", "created_date", "start_date", "deadline"):
                        # Normalizuj datum - pokud je prázdné/null, uloží se jako NULL do DB
                        if data[f]:
                            v = _normalize_date(data[f])
                        else:
                            v = None  # Explicitně None pro NULL v DB
                    elif f=="progress":
                        v = int(data[f]) if data[f] is not None else 0
                    else:
                        v = data[f] if data[f] is not None else ""
                    # Přidej do updates i když je hodnota None (pro NULL v DB)
                    updates.append(f"{f}=?"); params.append(v)
                    print(f"[PATCH] Updating {f} = {v} (type: {type(v).__name__})")
            # Touch legacy updated_at if present
            info = _jobs_info()
            if "updated_at" in info:
                updates.append("updated_at=?"); params.append(datetime.utcnow().isoformat())
            # Optional owner change if present
            if "owner_id" in info and data.get("owner_id") is not None:
                updates.append("owner_id=?"); params.append(int(data.get("owner_id")))
            if not updates: return jsonify({"ok": False, "error":"nothing_to_update"}), 400
            params.append(int(jid))
            db.execute("UPDATE jobs SET " + ", ".join(updates) + " WHERE id=?", params)
            db.commit()
            # if status is Dokončeno, ensure completed_at is set
            try:
                job = db.execute("SELECT id,status,completed_at FROM jobs WHERE id=?", (jid,)).fetchone()
                if job:
                    s = (job["status"] or "").lower()
                    if s.startswith("dokon"):
                        if not job["completed_at"]:
                            today = datetime.utcnow().strftime("%Y-%m-%d")
                            db.execute("UPDATE jobs SET completed_at=? WHERE id=?", (today, jid))
                            db.commit()
            except Exception:
                pass
            print(f"✓ Job {jid} updated successfully")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error updating job: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    # DELETE
    try:
        jid = request.args.get("id", type=int)
        if not jid: return jsonify({"ok": False, "error":"missing_id"}), 400
        # audit snapshot
        before = db.execute("SELECT id, COALESCE(title,name,'') as title, client, status, city, code, date, note FROM jobs WHERE id=?", (jid,)).fetchone()
        before = dict(before) if before else None
        db.execute("DELETE FROM jobs WHERE id=?", (jid,))
        db.commit()
        audit_event(u.get("id"), "delete", "job", jid, before=before)
        print(f"✓ Job {jid} deleted successfully")
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error deleting job: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500


# -------- Job detail & materials/tools/assignments --------
@app.route("/api/jobs/<int:job_id>", methods=["GET"])
def api_job_detail(job_id):
    u, err = require_role(write=False)
    if err: return err
    db = get_db()
    title_col = _job_title_col()
    
    # Get job with all available columns
    cols = [r[1] for r in db.execute("PRAGMA table_info(jobs)").fetchall()]
    select_cols = ["id", title_col + " AS title", "client", "status", "city", "code", "date", "note"]
    optional_cols = ["budget", "cost_spent", "labor_cost_total", "time_spent_minutes", "hourly_rate", 
                     "budget_remaining", "margin", "time_planned_minutes"]
    for col in optional_cols:
        if col in cols:
            select_cols.append(col)
    
    job = db.execute(f"SELECT {', '.join(select_cols)} FROM jobs WHERE id=?", (job_id,)).fetchone()
    if not job: return jsonify({"ok": False, "error": "not_found"}), 404
    job = dict(job)
    if job.get("date"): job["date"] = _normalize_date(job["date"])
    
    # Calculate finance summary
    budget = float(job.get("budget") or 0)
    cost_spent = float(job.get("cost_spent") or job.get("labor_cost_total") or 0)
    budget_remaining = budget - cost_spent if budget > 0 else None
    margin_pct = ((budget - cost_spent) / budget * 100) if budget > 0 else None
    
    # Get worklogs summary for cost burn rate
    worklogs_summary = db.execute("""
        SELECT 
            COUNT(*) as count,
            SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) as total_minutes,
            SUM(COALESCE(labor_cost, 0)) as total_cost,
            MIN(date) as first_date,
            MAX(date) as last_date
        FROM timesheets
        WHERE job_id = ?
    """, (job_id,)).fetchone()
    
    finance_summary = {
        "budget": budget,
        "cost_spent": cost_spent,
        "budget_remaining": budget_remaining,
        "margin_pct": margin_pct,
        "worklogs_count": worklogs_summary["count"] or 0,
        "total_hours": round((worklogs_summary["total_minutes"] or 0) / 60, 1),
        "total_cost": float(worklogs_summary["total_cost"] or 0),
        "first_date": worklogs_summary["first_date"],
        "last_date": worklogs_summary["last_date"]
    }
    
    # Get material consumption from worklogs
    material_consumption = []
    missing_materials = []
    try:
        worklog_rows = db.execute("""
            SELECT material_used
            FROM timesheets
            WHERE job_id = ? AND material_used IS NOT NULL AND material_used != ''
        """, (job_id,)).fetchall()
        
        # Aggregate material consumption
        consumption_map = {}
        for row in worklog_rows:
            try:
                import json
                materials = json.loads(row['material_used']) if isinstance(row['material_used'], str) else row['material_used']
                if isinstance(materials, list):
                    for mat in materials:
                        item_id = mat.get('item_id')
                        qty = float(mat.get('qty', 0))
                        unit = mat.get('unit', 'ks')
                        name = mat.get('name', '')
                        
                        if item_id:
                            key = str(item_id)
                            if key not in consumption_map:
                                consumption_map[key] = {'item_id': item_id, 'qty': 0, 'unit': unit, 'name': name}
                            consumption_map[key]['qty'] += qty
                        elif name:
                            key = name.lower()
                            if key not in consumption_map:
                                consumption_map[key] = {'item_id': None, 'qty': 0, 'unit': unit, 'name': name}
                            consumption_map[key]['qty'] += qty
            except Exception as e:
                print(f"[MATERIAL] Error parsing material_used: {e}")
                pass
        
        # Get warehouse items info for consumed materials
        for key, cons in consumption_map.items():
            item_id = cons.get('item_id')
            if item_id:
                try:
                    wi = db.execute("SELECT name, quantity, min_quantity, unit FROM warehouse_items WHERE id=?", (item_id,)).fetchone()
                    if wi:
                        cons['name'] = wi['name']
                        cons['stock_qty'] = wi['quantity'] or 0
                        cons['min_qty'] = wi['min_quantity'] or 0
                        cons['unit'] = wi['unit'] or cons['unit']
                        
                        # Check if material is missing
                        if cons['stock_qty'] < cons['qty']:
                            missing_materials.append({
                                'name': cons['name'],
                                'consumed': cons['qty'],
                                'available': cons['stock_qty'],
                                'unit': cons['unit']
                            })
                except:
                    pass
            
            material_consumption.append(cons)
    except Exception as e:
        print(f"[MATERIAL] Error getting material consumption: {e}")
        pass
    
    mats = [dict(r) for r in db.execute("SELECT id, name, qty, unit FROM job_materials WHERE job_id=? ORDER BY id ASC", (job_id,)).fetchall()]
    tools = [dict(r) for r in db.execute("SELECT id, name, qty, unit FROM job_tools WHERE job_id=? ORDER BY id ASC", (job_id,)).fetchall()]
    assigns = [r["employee_id"] for r in db.execute("SELECT employee_id FROM job_assignments WHERE job_id=? ORDER BY employee_id ASC", (job_id,)).fetchall()]
    
    return jsonify({
        "ok": True, 
        "job": job, 
        "materials": mats, 
        "tools": tools, 
        "assignments": assigns, 
        "finance": finance_summary,
        "material_consumption": material_consumption,
        "missing_materials": missing_materials
    })


# ============================================================================
# JOBS 2.0 - Enhanced Jobs API with Metrics (Zakázky jako mikro-vesmír)
# ============================================================================

@app.route("/api/jobs/overview")
def api_jobs_metrics_overview():
    """Get all jobs with calculated metrics for Kanban/List/Timeline views"""
    u, err = require_role(write=False)
    if err: return err
    
    db = get_db()
    info = _jobs_info()
    
    # Dynamicky sestavit SELECT podle dostupných sloupců
    base_cols = ["id", "client", "status", "city", "code", "date", "note"]
    optional_cols = ["created_date", "start_date", "deadline", "address", "progress", 
                     "budget", "estimated_value", "actual_value", "budget_labor", 
                     "budget_materials", "budget_equipment", "budget_other",
                     "actual_labor_cost", "actual_material_cost", "profit_margin",
                     "budget_spent_percent", "completion_percent", "weather_dependent",
                     "priority", "hourly_rate", "estimated_hours", "actual_hours"]
    
    select_cols = base_cols.copy()
    for col in optional_cols:
        if col in info:
            select_cols.append(col)
    
    # Title/name handling
    title_col = _job_title_col()
    if title_col == "name":
        select_cols.insert(0, "name AS title")
    else:
        select_cols.insert(0, "title")
    
    sql = f"SELECT {', '.join(select_cols)} FROM jobs ORDER BY date(date) DESC, id DESC"
    rows = db.execute(sql).fetchall()
    
    jobs_with_metrics = []
    today = datetime.now().date()
    
    # Pre-fetch všechny potřebné data najednou pro efektivitu
    # 1. Materiály pro všechny zakázky
    all_materials = {}
    try:
        mat_rows = db.execute("""
            SELECT jm.job_id, jm.name, jm.qty, jm.unit, jm.status,
                   wi.quantity as stock_qty, wi.min_quantity
            FROM job_materials jm
            LEFT JOIN warehouse_items wi ON wi.name = jm.name
        """).fetchall()
        for m in mat_rows:
            jid = m['job_id']
            if jid not in all_materials:
                all_materials[jid] = []
            all_materials[jid].append(dict(m))
    except:
        pass
    
    # 2. Assignments pro všechny zakázky
    all_assignments = {}
    try:
        assign_rows = db.execute("""
            SELECT ja.job_id, ja.employee_id, e.name, e.role
            FROM job_assignments ja
            JOIN employees e ON e.id = ja.employee_id
        """).fetchall()
        for a in assign_rows:
            jid = a['job_id']
            if jid not in all_assignments:
                all_assignments[jid] = []
            all_assignments[jid].append({
                'employee_id': a['employee_id'],
                'name': a['name'],
                'role': a['role']
            })
    except:
        pass
    
    # 3. Timesheets - hodiny za posledních 7 dní per zaměstnanec
    week_ago = (today - timedelta(days=7)).isoformat()
    employee_hours_week = {}
    try:
        ts_rows = db.execute("""
            SELECT employee_id, SUM(hours) as total_hours
            FROM timesheets
            WHERE date >= ?
            GROUP BY employee_id
        """, (week_ago,)).fetchall()
        for ts in ts_rows:
            employee_hours_week[ts['employee_id']] = ts['total_hours'] or 0
    except:
        pass
    
    # 4. Timesheets per job (celkové hodiny)
    job_hours = {}
    try:
        jh_rows = db.execute("""
            SELECT job_id, SUM(hours) as total_hours, COUNT(DISTINCT employee_id) as workers
            FROM timesheets
            GROUP BY job_id
        """).fetchall()
        for jh in jh_rows:
            job_hours[jh['job_id']] = {
                'total_hours': jh['total_hours'] or 0,
                'workers': jh['workers'] or 0
            }
    except:
        pass
    
    # 5. Tasks per job
    job_tasks = {}
    try:
        task_rows = db.execute("""
            SELECT job_id, 
                   COUNT(*) as total,
                   SUM(CASE WHEN status IN ('done','completed','Dokončeno') THEN 1 ELSE 0 END) as completed,
                   SUM(CASE WHEN due_date < date('now') AND status NOT IN ('done','completed','Dokončeno') THEN 1 ELSE 0 END) as overdue
            FROM tasks
            GROUP BY job_id
        """).fetchall()
        for t in task_rows:
            job_tasks[t['job_id']] = {
                'total': t['total'] or 0,
                'completed': t['completed'] or 0,
                'overdue': t['overdue'] or 0
            }
    except:
        pass
    
    for row in rows:
        job = dict(row)
        job_id = job['id']
        
        # Normalize date
        if job.get('date'):
            job['date'] = _normalize_date(job['date'])
        
        # Skip completed jobs (they go to archive)
        status = (job.get('status') or '').strip().lower()
        if status.startswith('dokon'):
            continue
        
        # === Calculate Metrics ===
        
        # Time progress
        time_progress = 0
        start_date = job.get('start_date') or job.get('created_date')
        deadline = job.get('deadline') or job.get('date')
        days_until = None
        
        if deadline:
            try:
                deadline_dt = datetime.strptime(deadline[:10], '%Y-%m-%d').date()
                days_until = (deadline_dt - today).days
                
                if start_date:
                    start_dt = datetime.strptime(start_date[:10], '%Y-%m-%d').date()
                    total_days = max(1, (deadline_dt - start_dt).days)
                    elapsed_days = (today - start_dt).days
                    time_progress = min(100, max(0, (elapsed_days / total_days) * 100))
            except:
                pass
        
        # Budget calculations
        budget = job.get('budget') or job.get('estimated_value') or 0
        if not budget:
            # Spočítej z breakdown
            budget = (job.get('budget_labor') or 0) + (job.get('budget_materials') or 0) + \
                     (job.get('budget_equipment') or 0) + (job.get('budget_other') or 0)
        
        actual_cost = (job.get('actual_labor_cost') or 0) + (job.get('actual_material_cost') or 0)
        if not actual_cost and job.get('actual_value'):
            actual_cost = job.get('actual_value')
        
        budget_progress = (actual_cost / budget * 100) if budget > 0 else 0
        if job.get('budget_spent_percent'):
            budget_progress = job.get('budget_spent_percent')
        
        # Margin
        margin_pct = job.get('profit_margin')
        if margin_pct is None and budget > 0:
            margin_pct = ((budget - actual_cost) / budget) * 100
        margin_pct = margin_pct or 50  # default
        
        # Material status
        materials = all_materials.get(job_id, [])
        material_status = 'ok'
        missing_items = []
        for mat in materials:
            mat_status = mat.get('status', '').lower()
            stock_qty = mat.get('stock_qty') or 0
            needed_qty = mat.get('qty') or 0
            
            if mat_status in ('missing', 'chybí') or stock_qty < needed_qty:
                material_status = 'missing'
                missing_items.append({'name': mat['name'], 'needed': needed_qty, 'available': stock_qty})
            elif mat_status not in ('ok', 'reserved', 'rezervováno') and material_status != 'missing':
                material_status = 'partial'
        
        # Team & overload
        assigned = all_assignments.get(job_id, [])
        overloaded = []
        for emp in assigned:
            emp_hours = employee_hours_week.get(emp['employee_id'], 0)
            if emp_hours > 45:
                overloaded.append({
                    'employee_id': emp['employee_id'],
                    'name': emp['name'],
                    'hours_week': round(emp_hours, 1)
                })
        
        team_utilization = 0
        if assigned:
            total_team_hours = sum(employee_hours_week.get(e['employee_id'], 0) for e in assigned)
            max_hours = len(assigned) * 40  # 40h týdně standard
            team_utilization = min(100, (total_team_hours / max_hours * 100)) if max_hours > 0 else 0
        
        # Job hours from timesheets
        jh = job_hours.get(job_id, {'total_hours': 0, 'workers': 0})
        
        # Tasks
        jt = job_tasks.get(job_id, {'total': 0, 'completed': 0, 'overdue': 0})
        
        # Weather risk (simplified - based on weather_dependent flag)
        weather_dependent = job.get('weather_dependent', False)
        weather_risk = 'none'
        if weather_dependent:
            if days_until is not None and days_until <= 7:
                weather_risk = 'high'
            elif days_until is not None and days_until <= 14:
                weather_risk = 'medium'
            else:
                weather_risk = 'low'
        
        # === Risk Score (0-100) ===
        risk_score = 0
        alerts = []
        
        # Deadline risk
        if days_until is not None:
            completion = job.get('completion_percent') or job.get('progress') or 0
            if days_until < 0:
                risk_score += 40
                alerts.append({'type': 'overdue', 'severity': 'high', 'text': f'Po termínu {abs(days_until)} dní', 'icon': '⏰'})
            elif days_until <= 7 and completion < 80:
                risk_score += 30
                alerts.append({'type': 'deadline', 'severity': 'medium', 'text': f'Deadline za {days_until} dní', 'icon': '⚠️'})
            elif days_until <= 3:
                risk_score += 20
        
        # Budget risk
        if budget_progress > 100:
            risk_score += 25
            alerts.append({'type': 'budget', 'severity': 'high', 'text': f'Rozpočet překročen o {round(budget_progress-100)}%', 'icon': '💸'})
        elif budget_progress > 80 and (job.get('completion_percent') or 0) < 80:
            risk_score += 15
            alerts.append({'type': 'budget', 'severity': 'medium', 'text': 'Rozpočet téměř vyčerpán', 'icon': '💰'})
        
        # Material risk
        if material_status == 'missing':
            risk_score += 20
            alerts.append({'type': 'material', 'severity': 'medium', 'text': f'{len(missing_items)} položek chybí', 'icon': '📦'})
        
        # Team overload risk
        if overloaded:
            risk_score += 15
            alerts.append({'type': 'overload', 'severity': 'medium', 'text': f'{len(overloaded)} přetížených', 'icon': '👥'})
        
        # Weather risk
        if weather_risk == 'high':
            risk_score += 15
            alerts.append({'type': 'weather', 'severity': 'medium', 'text': 'Počasí ohrožuje termín', 'icon': '🌧️'})
        
        # Tasks overdue
        if jt['overdue'] > 0:
            risk_score += 10
            alerts.append({'type': 'tasks', 'severity': 'low', 'text': f'{jt["overdue"]} úkolů po termínu', 'icon': '📋'})
        
        risk_score = min(100, risk_score)
        
        # AI Score (inverse of risk, with bonus for good metrics)
        ai_score = max(0, 100 - risk_score)
        if margin_pct > 30:
            ai_score = min(100, ai_score + 5)
        if material_status == 'ok':
            ai_score = min(100, ai_score + 5)
        
        # Ghost plan suggestion (AI recommendation for timeline)
        ghost_plan = None
        if risk_score > 50 and days_until is not None and days_until > 0:
            suggested_delay = 2 if risk_score < 70 else 5
            if overloaded:
                ghost_plan = {
                    'type': 'delay',
                    'days': suggested_delay,
                    'reason': 'Přetížení týmu',
                    'suggested_date': (datetime.strptime(deadline[:10], '%Y-%m-%d') + timedelta(days=suggested_delay)).strftime('%Y-%m-%d')
                }
            elif material_status == 'missing':
                ghost_plan = {
                    'type': 'delay',
                    'days': suggested_delay,
                    'reason': 'Chybějící materiál',
                    'suggested_date': (datetime.strptime(deadline[:10], '%Y-%m-%d') + timedelta(days=suggested_delay)).strftime('%Y-%m-%d')
                }
        
        # Build metrics object
        job['metrics'] = {
            'margin_pct': round(margin_pct, 1),
            'risk_pct': risk_score,
            'risk_level': 'high' if risk_score >= 50 else 'medium' if risk_score >= 25 else 'low',
            'ai_score': round(ai_score),
            'alerts': alerts,
            'time_progress_pct': round(time_progress, 1),
            'budget_progress_pct': round(budget_progress, 1),
            'days_until_deadline': days_until,
            'material': {
                'status': material_status,
                'missing_count': len(missing_items),
                'missing_items': missing_items[:5]  # Top 5
            },
            'team': {
                'assigned': assigned,
                'assigned_count': len(assigned),
                'overload': overloaded,
                'utilization_pct': round(team_utilization, 1)
            },
            'hours': {
                'total': round(jh['total_hours'], 1),
                'workers': jh['workers'],
                'estimated': job.get('estimated_hours') or 0
            },
            'tasks': jt,
            'weather': {
                'dependent': weather_dependent,
                'risk': weather_risk
            },
            'ghost_plan': ghost_plan
        }
        
        jobs_with_metrics.append(job)
    
    return jsonify({'ok': True, 'jobs': jobs_with_metrics})


@app.route("/api/jobs/<int:job_id>/hub")
def api_job_hub(job_id):
    """Get detailed job hub with AI panel - mikro-vesmír zakázky"""
    u, err = require_role(write=False)
    if err: return err
    
    db = get_db()
    info = _jobs_info()
    
    # Načti základní data zakázky
    title_col = _job_title_col()
    job = db.execute(f"SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    if not job:
        return jsonify({'ok': False, 'error': 'not_found'}), 404
    
    job = dict(job)
    today = datetime.now().date()
    
    # === Materials ===
    materials = []
    try:
        mat_rows = db.execute("""
            SELECT jm.*, wi.quantity as stock_qty, wi.min_quantity, wi.unit_price
            FROM job_materials jm
            LEFT JOIN warehouse_items wi ON wi.name = jm.name
            WHERE jm.job_id = ?
        """, (job_id,)).fetchall()
        materials = [dict(m) for m in mat_rows]
    except:
        pass
    
    # === Team Assignments ===
    team = []
    try:
        team_rows = db.execute("""
            SELECT e.id, e.name, e.role, e.hourly_rate,
                   (SELECT SUM(hours) FROM timesheets WHERE employee_id = e.id AND date >= date('now', '-7 days')) as hours_week,
                   (SELECT SUM(hours) FROM timesheets WHERE employee_id = e.id AND job_id = ?) as hours_on_job
            FROM job_assignments ja
            JOIN employees e ON e.id = ja.employee_id
            WHERE ja.job_id = ?
        """, (job_id, job_id)).fetchall()
        team = [dict(t) for t in team_rows]
    except:
        pass
    
    # === Timesheets Summary ===
    timesheets_summary = {'total_hours': 0, 'total_cost': 0, 'entries': 0, 'by_employee': []}
    try:
        ts_rows = db.execute("""
            SELECT e.name, SUM(t.hours) as hours, COUNT(*) as entries
            FROM timesheets t
            JOIN employees e ON e.id = t.employee_id
            WHERE t.job_id = ?
            GROUP BY t.employee_id
            ORDER BY hours DESC
        """, (job_id,)).fetchall()
        
        total_hours = 0
        for ts in ts_rows:
            total_hours += ts['hours'] or 0
            timesheets_summary['by_employee'].append({
                'name': ts['name'],
                'hours': round(ts['hours'] or 0, 1),
                'entries': ts['entries']
            })
        timesheets_summary['total_hours'] = round(total_hours, 1)
        timesheets_summary['entries'] = sum(ts['entries'] for ts in ts_rows)
        
        # Estimate cost
        avg_rate = 200  # Default hourly rate
        timesheets_summary['total_cost'] = round(total_hours * avg_rate, 0)
    except:
        pass
    
    # === Finance Summary ===
    budget = job.get('budget') or job.get('estimated_value') or 0
    if not budget:
        budget = (job.get('budget_labor') or 0) + (job.get('budget_materials') or 0) + \
                 (job.get('budget_equipment') or 0) + (job.get('budget_other') or 0)
    
    actual_labor = timesheets_summary['total_cost']
    actual_materials = sum((m.get('qty') or 0) * (m.get('unit_price') or 0) for m in materials)
    actual_total = actual_labor + actual_materials
    
    finance_summary = {
        'budget': round(budget, 0),
        'budget_breakdown': {
            'labor': job.get('budget_labor') or 0,
            'materials': job.get('budget_materials') or 0,
            'equipment': job.get('budget_equipment') or 0,
            'other': job.get('budget_other') or 0
        },
        'actual': {
            'total': round(actual_total, 0),
            'labor': round(actual_labor, 0),
            'materials': round(actual_materials, 0)
        },
        'remaining': round(budget - actual_total, 0),
        'spent_pct': round((actual_total / budget * 100) if budget > 0 else 0, 1),
        'margin_pct': round(((budget - actual_total) / budget * 100) if budget > 0 else 0, 1)
    }
    
    # === Tasks Summary ===
    tasks_summary = {'total': 0, 'completed': 0, 'in_progress': 0, 'overdue': 0, 'items': []}
    try:
        task_rows = db.execute("""
            SELECT t.id, t.title, t.status, t.due_date, t.priority, e.name as assignee
            FROM tasks t
            LEFT JOIN employees e ON e.id = t.employee_id
            WHERE t.job_id = ?
            ORDER BY t.due_date ASC
        """, (job_id,)).fetchall()
        
        for task in task_rows:
            tasks_summary['total'] += 1
            status = (task['status'] or '').lower()
            if status in ('done', 'completed', 'dokončeno'):
                tasks_summary['completed'] += 1
            elif status in ('in_progress', 'v práci'):
                tasks_summary['in_progress'] += 1
            
            if task['due_date']:
                try:
                    due = datetime.strptime(task['due_date'][:10], '%Y-%m-%d').date()
                    if due < today and status not in ('done', 'completed', 'dokončeno'):
                        tasks_summary['overdue'] += 1
                except:
                    pass
            
            tasks_summary['items'].append(dict(task))
    except:
        pass
    
    # === AI Panel (Doporučení) ===
    ai_panel = {
        'risk_deadline': 0,
        'risk_deadline_text': '',
        'material_missing': [],
        'team_overload': [],
        'recommendations': []
    }
    
    # Deadline risk
    deadline = job.get('deadline') or job.get('date')
    completion = job.get('completion_percent') or job.get('progress') or 0
    if deadline:
        try:
            deadline_dt = datetime.strptime(deadline[:10], '%Y-%m-%d').date()
            days_until = (deadline_dt - today).days
            
            if days_until < 0:
                ai_panel['risk_deadline'] = 100
                ai_panel['risk_deadline_text'] = f'Zakázka je {abs(days_until)} dní po termínu!'
                ai_panel['recommendations'].append({
                    'type': 'urgent',
                    'icon': '🚨',
                    'text': 'Okamžitě kontaktovat klienta a domluvit nový termín'
                })
            elif days_until <= 7 and completion < 80:
                ai_panel['risk_deadline'] = 70
                ai_panel['risk_deadline_text'] = f'Zbývá {days_until} dní, dokončeno pouze {completion}%'
                ai_panel['recommendations'].append({
                    'type': 'warning',
                    'icon': '⚠️',
                    'text': 'Zvážit posílení týmu nebo přesčasy'
                })
            elif days_until <= 14:
                ai_panel['risk_deadline'] = 30
                ai_panel['risk_deadline_text'] = f'Zbývá {days_until} dní'
            else:
                ai_panel['risk_deadline'] = 0
                ai_panel['risk_deadline_text'] = f'V pořádku, zbývá {days_until} dní'
        except:
            pass
    
    # Material missing
    for mat in materials:
        stock = mat.get('stock_qty') or 0
        needed = mat.get('qty') or 0
        if stock < needed:
            ai_panel['material_missing'].append({
                'name': mat['name'],
                'needed': needed,
                'available': stock,
                'shortage': needed - stock
            })
    
    if ai_panel['material_missing']:
        ai_panel['recommendations'].append({
            'type': 'material',
            'icon': '📦',
            'text': f'Doobjednat {len(ai_panel["material_missing"])} položek materiálu'
        })
    
    # Team overload
    for member in team:
        hours_week = member.get('hours_week') or 0
        if hours_week > 45:
            ai_panel['team_overload'].append({
                'name': member['name'],
                'hours_week': round(hours_week, 1),
                'over_by': round(hours_week - 40, 1)
            })
    
    if ai_panel['team_overload']:
        ai_panel['recommendations'].append({
            'type': 'team',
            'icon': '👥',
            'text': f'Přerozdělit práci - {len(ai_panel["team_overload"])} lidí přetíženo'
        })
    
    # Budget recommendation
    if finance_summary['spent_pct'] > 80 and completion < 80:
        ai_panel['recommendations'].append({
            'type': 'budget',
            'icon': '💰',
            'text': 'Rozpočet téměř vyčerpán, zkontrolovat další výdaje'
        })
    
    # Positive recommendation if all is well
    if not ai_panel['recommendations']:
        ai_panel['recommendations'].append({
            'type': 'success',
            'icon': '✅',
            'text': 'Zakázka probíhá podle plánu'
        })
    
    return jsonify({
        'ok': True,
        'job': job,
        'materials': materials,
        'team': team,
        'timesheets': timesheets_summary,
        'finance': finance_summary,
        'tasks': tasks_summary,
        'ai_panel': ai_panel
    })

print("✅ Jobs 2.0 Overview API loaded")

@app.route("/api/jobs/<int:job_id>/ai-insights")
def api_job_ai_insights(job_id):
    """Get AI insights for job detail - health score, risks, recommendations"""
    u, err = require_role(write=False)
    if err: return err
    
    db = get_db()
    
    # Načti základní data zakázky
    job = db.execute("SELECT * FROM jobs WHERE id=?", (job_id,)).fetchone()
    if not job:
        return jsonify({'ok': False, 'error': 'not_found'}), 404
    
    job = dict(job)
    today = datetime.now().date()
    
    # === Materials ===
    materials = []
    try:
        mat_rows = db.execute("""
            SELECT jm.*, wi.quantity as stock_qty, wi.min_quantity, wi.unit_price
            FROM job_materials jm
            LEFT JOIN warehouse_items wi ON wi.name = jm.name
            WHERE jm.job_id = ?
        """, (job_id,)).fetchall()
        materials = [dict(m) for m in mat_rows]
    except:
        pass
    
    # === Team Assignments ===
    team = []
    try:
        team_rows = db.execute("""
            SELECT e.id, e.name, e.role, e.hourly_rate,
                   (SELECT SUM(hours) FROM timesheets WHERE employee_id = e.id AND date >= date('now', '-7 days')) as hours_week,
                   (SELECT SUM(hours) FROM timesheets WHERE employee_id = e.id AND job_id = ?) as hours_on_job
            FROM job_assignments ja
            JOIN employees e ON e.id = ja.employee_id
            WHERE ja.job_id = ?
        """, (job_id, job_id)).fetchall()
        team = [dict(t) for t in team_rows]
    except:
        pass
    
    # === Timesheets Summary ===
    timesheets_summary = {'total_hours': 0, 'total_cost': 0}
    try:
        ts_result = db.execute("""
            SELECT SUM(t.hours) as total_hours
            FROM timesheets t
            WHERE t.job_id = ?
        """, (job_id,)).fetchone()
        timesheets_summary['total_hours'] = round(ts_result['total_hours'] or 0, 1)
        timesheets_summary['total_cost'] = round((ts_result['total_hours'] or 0) * 200, 0)  # Default rate
    except:
        pass
    
    # === Finance Summary ===
    budget = job.get('budget') or job.get('estimated_value') or 0
    if not budget:
        budget = (job.get('budget_labor') or 0) + (job.get('budget_materials') or 0) + \
                 (job.get('budget_equipment') or 0) + (job.get('budget_other') or 0)
    
    actual_labor = timesheets_summary['total_cost']
    actual_materials = sum((m.get('qty') or 0) * (m.get('unit_price') or 0) for m in materials)
    actual_total = actual_labor + actual_materials
    
    spent_pct = round((actual_total / budget * 100) if budget > 0 else 0, 1)
    margin_pct = round(((budget - actual_total) / budget * 100) if budget > 0 else 0, 1)
    
    # === Tasks Summary ===
    tasks_overdue = 0
    tasks_total = 0
    try:
        task_rows = db.execute("""
            SELECT t.status, t.due_date
            FROM tasks t
            WHERE t.job_id = ?
        """, (job_id,)).fetchall()
        
        for task in task_rows:
            tasks_total += 1
            status = (task['status'] or '').lower()
            if task['due_date'] and status not in ('done', 'completed', 'dokončeno'):
                try:
                    due = datetime.strptime(task['due_date'][:10], '%Y-%m-%d').date()
                    if due < today:
                        tasks_overdue += 1
                except:
                    pass
    except:
        pass
    
    # === Calculate Health Score (0-100) ===
    health_score = 100
    risk_score = 0
    
    # Deadline risk
    deadline = job.get('deadline') or job.get('date')
    completion = job.get('completion_percent') or job.get('progress') or 0
    deadline_risk = 0
    deadline_text = ''
    
    if deadline:
        try:
            deadline_dt = datetime.strptime(deadline[:10], '%Y-%m-%d').date()
            days_until = (deadline_dt - today).days
            
            if days_until < 0:
                deadline_risk = 50
                deadline_text = f'Zakázka je {abs(days_until)} dní po termínu!'
            elif days_until <= 3:
                deadline_risk = 30
                deadline_text = f'Zbývá pouze {days_until} dní'
            elif days_until <= 7 and completion < 80:
                deadline_risk = 20
                deadline_text = f'Zbývá {days_until} dní, dokončeno {completion}%'
            elif days_until <= 14:
                deadline_risk = 10
                deadline_text = f'Zbývá {days_until} dní'
        except:
            pass
    
    risk_score += deadline_risk
    
    # Budget risk
    budget_risk = 0
    if spent_pct > 90:
        budget_risk = 40
    elif spent_pct > 80:
        budget_risk = 20
    elif spent_pct > 70:
        budget_risk = 10
    
    risk_score += budget_risk
    
    # Material risk
    material_risk = 0
    missing_materials = []
    for mat in materials:
        stock = mat.get('stock_qty') or 0
        needed = mat.get('qty') or 0
        if stock < needed:
            missing_materials.append({
                'name': mat['name'],
                'needed': needed,
                'available': stock,
                'shortage': needed - stock
            })
    
    if missing_materials:
        material_risk = min(30, len(missing_materials) * 10)
        risk_score += material_risk
    
    # Capacity risk
    capacity_risk = 0
    overloaded_team = []
    for member in team:
        hours_week = member.get('hours_week') or 0
        if hours_week > 45:
            overloaded_team.append({
                'name': member['name'],
                'hours_week': round(hours_week, 1),
                'over_by': round(hours_week - 40, 1)
            })
    
    if overloaded_team:
        capacity_risk = min(20, len(overloaded_team) * 5)
        risk_score += capacity_risk
    
    # Tasks overdue risk
    if tasks_overdue > 0:
        risk_score += min(20, tasks_overdue * 5)
    
    # Health score = 100 - risk_score (with bonuses)
    health_score = max(0, 100 - risk_score)
    if margin_pct > 30:
        health_score = min(100, health_score + 5)
    if not missing_materials:
        health_score = min(100, health_score + 5)
    if not overloaded_team:
        health_score = min(100, health_score + 5)
    
    # === Build Insights List ===
    insights = []
    
    if deadline_risk > 0:
        insights.append({
            'type': 'deadline',
            'score': deadline_risk,
            'text': deadline_text,
            'priority': 'high' if deadline_risk >= 30 else 'medium'
        })
    
    if budget_risk > 0:
        insights.append({
            'type': 'budget',
            'score': budget_risk,
            'text': f'Rozpočet vyčerpán na {spent_pct}%',
            'priority': 'high' if budget_risk >= 30 else 'medium'
        })
    
    if material_risk > 0:
        insights.append({
            'type': 'material',
            'score': material_risk,
            'text': f'Chybí {len(missing_materials)} materiálů',
            'priority': 'medium',
            'details': missing_materials
        })
    
    if capacity_risk > 0:
        insights.append({
            'type': 'capacity',
            'score': capacity_risk,
            'text': f'{len(overloaded_team)} členů týmu přetíženo',
            'priority': 'medium',
            'details': overloaded_team
        })
    
    if tasks_overdue > 0:
        insights.append({
            'type': 'tasks',
            'score': min(20, tasks_overdue * 5),
            'text': f'{tasks_overdue} úkolů po termínu',
            'priority': 'medium'
        })
    
    # === Build Recommendations ===
    recommendations = []
    
    if deadline_risk >= 30:
        recommendations.append({
            'type': 'urgent',
            'icon': '🚨',
            'title': 'Kritický deadline',
            'text': 'Okamžitě kontaktovat klienta a domluvit nový termín',
            'action': 'contact_client'
        })
    elif deadline_risk >= 20:
        recommendations.append({
            'type': 'warning',
            'icon': '⚠️',
            'title': 'Blížící se deadline',
            'text': 'Zvážit posílení týmu nebo přesčasy',
            'action': 'add_resources'
        })
    
    if material_risk > 0:
        recommendations.append({
            'type': 'material',
            'icon': '📦',
            'title': 'Chybí materiál',
            'text': f'Doobjednat {len(missing_materials)} položek materiálu',
            'action': 'order_materials'
        })
    
    if capacity_risk > 0:
        recommendations.append({
            'type': 'team',
            'icon': '👥',
            'title': 'Přetížení týmu',
            'text': f'Přerozdělit práci - {len(overloaded_team)} lidí přetíženo',
            'action': 'redistribute_work'
        })
    
    if budget_risk >= 20 and completion < 80:
        recommendations.append({
            'type': 'budget',
            'icon': '💰',
            'title': 'Rozpočet téměř vyčerpán',
            'text': 'Zkontrolovat další výdaje a optimalizovat náklady',
            'action': 'review_budget'
        })
    
    # Positive recommendation if all is well
    if not recommendations and health_score >= 80:
        recommendations.append({
            'type': 'success',
            'icon': '✅',
            'title': 'Vše v pořádku',
            'text': 'Zakázka probíhá podle plánu',
            'action': None
        })
    
    return jsonify({
        'ok': True,
        'health_score': round(health_score, 0),
        'risk_score': round(risk_score, 0),
        'insights': insights,
        'recommendations': recommendations
    })

@app.route("/api/jobs/<int:job_id>/materials", methods=["POST","DELETE"])
def api_job_materials(job_id):
    u, err = require_role(write=True)
    if err: return err
    db = get_db()
    if request.method == "POST":
        try:
            data = request.get_json(force=True, silent=True) or {}
            name = (data.get("name") or "").strip()
            qty  = float(data.get("qty") or 0)
            unit = (data.get("unit") or "ks").strip()
            if not name: return jsonify({"ok": False, "error":"invalid_input"}), 400
            db.execute("INSERT INTO job_materials(job_id,name,qty,unit) VALUES (?,?,?,?)", (job_id, name, qty, unit))
            db.commit()
            print(f"✓ Material '{name}' added to job {job_id}")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error adding material: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500
    try:
        mid = request.args.get("id", type=int)
        if not mid: return jsonify({"ok": False, "error":"missing_id"}), 400
        db.execute("DELETE FROM job_materials WHERE id=? AND job_id=?", (mid, job_id))
        db.commit()
        print(f"✓ Material {mid} deleted from job {job_id}")
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error deleting material: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/jobs/<int:job_id>/tools", methods=["POST","DELETE"])
def api_job_tools(job_id):
    u, err = require_role(write=True)
    if err: return err
    db = get_db()
    if request.method == "POST":
        try:
            data = request.get_json(force=True, silent=True) or {}
            name = (data.get("name") or "").strip()
            qty  = float(data.get("qty") or 0)
            unit = (data.get("unit") or "ks").strip()
            if not name: return jsonify({"ok": False, "error":"invalid_input"}), 400
            db.execute("INSERT INTO job_tools(job_id,name,qty,unit) VALUES (?,?,?,?)", (job_id, name, qty, unit))
            db.commit()
            print(f"✓ Tool '{name}' added to job {job_id}")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error adding tool: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500
    try:
        tid = request.args.get("id", type=int)
        if not tid: return jsonify({"ok": False, "error":"missing_id"}), 400
        db.execute("DELETE FROM job_tools WHERE id=? AND job_id=?", (tid, job_id))
        db.commit()
        print(f"✓ Tool {tid} deleted from job {job_id}")
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error deleting tool: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/jobs/<int:job_id>/assignments", methods=["POST"])
def api_job_assignments(job_id):
    u, err = require_role(write=True)
    if err: return err
    db = get_db()
    try:
        data = request.get_json(force=True, silent=True) or {}
        ids = data.get("employee_ids") or data.get("assignments") or []
        db.execute("DELETE FROM job_assignments WHERE job_id=?", (job_id,))
        for eid in ids:
            db.execute("INSERT OR IGNORE INTO job_assignments(job_id, employee_id) VALUES (?,?)", (job_id, int(eid)))
        db.commit()
        print(f"✓ Assignments updated for job {job_id}")
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error updating assignments: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

# ----------------- Tasks CRUD -----------------
@app.route("/api/tasks", methods=["GET","POST","PATCH","PUT","DELETE"])
def api_tasks():
    u, err = require_role(write=(request.method!="GET"))
    if err: return err
    db = get_db()

    if request.method == "GET":
        task_id = request.args.get("id", type=int)
        if task_id:
            # Return single task by ID
            row = db.execute("""SELECT t.id, t.job_id, t.employee_id, t.title, t.description, t.status, t.due_date,
                                      e.name AS employee_name
                               FROM tasks t
                               LEFT JOIN employees e ON e.id=t.employee_id
                               WHERE t.id=?""", (task_id,)).fetchone()
            if not row:
                return jsonify({"ok": False, "error": "not_found"}), 404
            
            task = dict(row)
            task["assignees"] = get_task_assignees(db, task_id)
            return jsonify({"ok": True, "task": task})
        
        jid = request.args.get("job_id", type=int)
        employee_id = request.args.get("employee_id", type=int)
        
        q = """SELECT t.id, t.job_id, t.employee_id, t.title, t.description, t.status, t.due_date,
                      e.name AS employee_name
               FROM tasks t
               LEFT JOIN employees e ON e.id=t.employee_id"""
        conds=[]; params=[]
        if jid: conds.append("t.job_id=?"); params.append(jid)
        if employee_id:
            # Pokud je zadán employee_id, hledej přes assignments
            q = """SELECT DISTINCT t.id, t.job_id, t.employee_id, t.title, t.description, t.status, t.due_date,
                          e.name AS employee_name
                   FROM tasks t
                   LEFT JOIN employees e ON e.id=t.employee_id
                   LEFT JOIN task_assignments ta ON ta.task_id = t.id"""
            conds.append("(t.employee_id=? OR ta.employee_id=?)")
            params.extend([employee_id, employee_id])
        if conds: q += " WHERE " + " AND ".join(conds)
        q += " ORDER BY COALESCE(t.due_date,''), t.id ASC"
        rows = [dict(r) for r in db.execute(q, params).fetchall()]
        
        # Přidej assignees ke každému tasku
        for task in rows:
            task["assignees"] = get_task_assignees(db, task["id"])
        
        return jsonify({"ok": True, "tasks": rows})

    if request.method == "POST":
        try:
            data = request.get_json(force=True, silent=True) or {}
            title = (data.get("title") or "").strip()
            if not title: return jsonify({"ok": False, "error":"invalid_input"}), 400
            
            # Vytvoř úkol
            db.execute("""
                INSERT INTO tasks(job_id, employee_id, title, description, status, due_date, created_by, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,datetime('now'),datetime('now'))
            """, (
                int(data.get("job_id")) if data.get("job_id") else None,
                int(data.get("employee_id")) if data.get("employee_id") else None,
                title,
                (data.get("description") or "").strip(),
                (data.get("status") or "open"),
                _normalize_date(data.get("due_date")) if data.get("due_date") else None,
                u["id"]
            ))
            db.commit()
            
            # Získej ID nového úkolu
            task_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            
            # Přiřaď zaměstnance (pokud jsou zadáni)
            assigned_ids = data.get("assigned_employees", [])
            primary_id = data.get("primary_employee")
            
            if assigned_ids:
                # Expand assignees with one-hop delegates (employee card settings)
                expanded_ids, delegations = _expand_assignees_with_delegate(db, assigned_ids)
                assign_employees_to_task(db, task_id, expanded_ids, primary_id)
                db.commit()

                # Notify assignees
                _notify_assignees(
                    "task",
                    task_id,
                    expanded_ids,
                    title="Nový úkol přiřazen",
                    body=f"Úkol: {title}",
                    actor_user_id=u.get("id"),
                )

                # Notify delegate explicitly (so it's clear it's by delegation)
                for d in delegations:
                    create_notification(
                        employee_id=d.get("to"),
                        kind="delegation",
                        title="Úkol delegován",
                        body=f"Úkol '{title}' byl delegován od zaměstnance ID {d.get('from')}",
                        entity_type="task",
                        entity_id=task_id,
                    )
            
            audit_event(u.get("id"), "create", "task", task_id, after={"title": title, "job_id": data.get("job_id"), "employee_id": data.get("employee_id"), "status": data.get("status"), "due_date": data.get("due_date")})
            print(f"✓ Task '{title}' created successfully (ID: {task_id})")
            return jsonify({"ok": True, "id": task_id})
        except Exception as e:
            db.rollback()
            print(f"✗ Error creating task: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    if request.method in ("PATCH", "PUT"):
        try:
            data = request.get_json(force=True, silent=True) or {}
            tid = data.get("id") or request.args.get("id", type=int)
            if not tid: return jsonify({"ok": False, "error":"missing_id"}), 400
            allowed = ["title","description","status","due_date","employee_id","job_id"]
            sets=[]; vals=[]
            for k in allowed:
                if k in data:
                    v = _normalize_date(data[k]) if k=="due_date" else data[k]
                    if k in ("employee_id","job_id") and v is not None:
                        v = int(v)
                    sets.append(f"{k}=?"); vals.append(v)
            if sets:
                vals.append(int(tid))
                db.execute("UPDATE tasks SET " + ", ".join(sets) + " WHERE id=?", vals)
                audit_event(u.get("id"), "update", "task", int(tid), meta={"fields": [s.split("=")[0] for s in sets]})
            
            # Update assignments if provided
            if "assigned_employees" in data:
                # Clear existing assignments
                db.execute("DELETE FROM task_assignments WHERE task_id=?", (tid,))
                
                # Add new assignments
                assigned_ids = data.get("assigned_employees", [])
                primary_id = data.get("primary_employee")
                if assigned_ids:
                    expanded_ids, delegations = _expand_assignees_with_delegate(db, assigned_ids)
                    assign_employees_to_task(db, tid, expanded_ids, primary_id)

                    _notify_assignees(
                        "task",
                        int(tid),
                        expanded_ids,
                        title="Úkol aktualizován",
                        body=f"Úkol byl upraven: {data.get('title') or ''}".strip() or "Úkol byl upraven",
                        actor_user_id=u.get("id"),
                    )
                    for d in delegations:
                        create_notification(
                            employee_id=d.get("to"),
                            kind="delegation",
                            title="Úkol delegován",
                            body=f"Úkol ID {tid} byl delegován od zaměstnance ID {d.get('from')}",
                            entity_type="task",
                            entity_id=int(tid),
                        )
            
            db.commit()
            print(f"✓ Task {tid} updated successfully")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error updating task: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    try:
        tid = request.args.get("id", type=int)
        if not tid: return jsonify({"ok": False, "error":"missing_id"}), 400
        # audit snapshot
        before = db.execute("SELECT id, job_id, employee_id, title, description, status, due_date FROM tasks WHERE id=?", (tid,)).fetchone()
        before = dict(before) if before else None
        db.execute("DELETE FROM tasks WHERE id=?", (tid,))
        db.commit()
        audit_event(u.get("id"), "delete", "task", tid, before=before)
        print(f"✓ Task {tid} deleted successfully")
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error deleting task: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

# issues CRUD
@app.route("/api/issues", methods=["GET","POST","PATCH","DELETE"])
def api_issues():
    u, err = require_role(write=(request.method!="GET"))
    if err: return err
    db = get_db()

    if request.method == "GET":
        issue_id = request.args.get("id", type=int)
        if issue_id:
            # Return single issue by ID
            row = db.execute("""
                SELECT i.*, e.name AS assigned_name, u.name AS creator_name
                FROM issues i
                LEFT JOIN employees e ON e.id = i.assigned_to
                LEFT JOIN users u ON u.id = i.created_by
                WHERE i.id = ?
            """, (issue_id,)).fetchone()
            if not row:
                return jsonify({"ok": False, "error": "not_found"}), 404
            
            issue = dict(row)
            issue["assignees"] = get_issue_assignees(db, issue_id)
            return jsonify({"ok": True, "issue": issue})
        
        # List issues with filters
        jid = request.args.get("job_id", type=int)
        assigned_to = request.args.get("assigned_to", type=int)
        status = request.args.get("status")
        
        if assigned_to:
            # Pokud filtrujeme podle přiřazení, použij JOIN přes assignments
            q = """
                SELECT DISTINCT i.*, e.name AS assigned_name, u.name AS creator_name, j.name AS job_name
                FROM issues i
                LEFT JOIN employees e ON e.id = i.assigned_to
                LEFT JOIN users u ON u.id = i.created_by
                LEFT JOIN jobs j ON j.id = i.job_id
                LEFT JOIN issue_assignments ia ON ia.issue_id = i.id
            """
        else:
            q = """
                SELECT i.*, e.name AS assigned_name, u.name AS creator_name, j.name AS job_name
                FROM issues i
                LEFT JOIN employees e ON e.id = i.assigned_to
                LEFT JOIN users u ON u.id = i.created_by
                LEFT JOIN jobs j ON j.id = i.job_id
            """
        
        conds = []
        params = []
        
        if jid:
            conds.append("i.job_id = ?")
            params.append(jid)
        if assigned_to:
            conds.append("(i.assigned_to = ? OR ia.employee_id = ?)")
            params.extend([assigned_to, assigned_to])
        if status:
            conds.append("i.status = ?")
            params.append(status)
        
        if conds:
            q += " WHERE " + " AND ".join(conds)
        
        q += " ORDER BY CASE i.status WHEN 'open' THEN 0 WHEN 'in_progress' THEN 1 ELSE 2 END, i.created_at DESC"
        
        rows = [dict(r) for r in db.execute(q, params).fetchall()]
        
        # Přidej assignees ke každému issue
        for issue in rows:
            issue["assignees"] = get_issue_assignees(db, issue["id"])
        
        return jsonify({"ok": True, "issues": rows})

    if request.method == "POST":
        try:
            data = request.get_json(force=True, silent=True) or {}
            title = (data.get("title") or "").strip()
            job_id = data.get("job_id")
            
            if not title or not job_id:
                return jsonify({"ok": False, "error": "missing_required_fields"}), 400
            
            db.execute("""
                INSERT INTO issues (job_id, title, description, type, status, severity, assigned_to, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                int(job_id),
                title,
                (data.get("description") or "").strip(),
                data.get("type") or "blocker",
                data.get("status") or "open",
                data.get("severity"),
                int(data["assigned_to"]) if data.get("assigned_to") else None,
                u["id"]
            ))
            db.commit()
            new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            
            # Přiřaď zaměstnance (pokud jsou zadáni)
            assigned_ids = data.get("assigned_employees", [])
            primary_id = data.get("primary_employee")
            
            if assigned_ids:
                expanded_ids, delegations = _expand_assignees_with_delegate(db, assigned_ids)
                assign_employees_to_issue(db, new_id, expanded_ids, primary_id)
                db.commit()

                _notify_assignees(
                    "issue",
                    new_id,
                    expanded_ids,
                    title="Nový problém přiřazen",
                    body=f"Problém: {title}",
                    actor_user_id=u.get("id"),
                )
                for d in delegations:
                    create_notification(
                        employee_id=d.get("to"),
                        kind="delegation",
                        title="Problém delegován",
                        body=f"Problém '{title}' byl delegován od zaměstnance ID {d.get('from')}",
                        entity_type="issue",
                        entity_id=new_id,
                    )
            
            audit_event(u.get("id"), "create", "issue", new_id, after={"title": title, "job_id": int(job_id), "status": data.get("status") or "open", "type": data.get("type") or "blocker"})
            print(f"✓ Issue '{title}' created (ID: {new_id})")
            return jsonify({"ok": True, "id": new_id})
        except Exception as e:
            db.rollback()
            print(f"✗ Error creating issue: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    if request.method == "PATCH":
        try:
            data = request.get_json(force=True, silent=True) or {}
            issue_id = data.get("id") or request.args.get("id", type=int)
            
            if not issue_id:
                return jsonify({"ok": False, "error": "missing_id"}), 400
            
            allowed = ["title", "description", "type", "status", "severity", "assigned_to"]
            sets = []
            vals = []
            
            for k in allowed:
                if k in data:
                    v = data[k]
                    if k == "assigned_to" and v is not None:
                        v = int(v)
                    sets.append(f"{k} = ?")
                    vals.append(v)
            
            # Auto-set resolved_at when status changes to resolved
            if "status" in data and data["status"] == "resolved":
                sets.append("resolved_at = datetime('now')")
            elif "status" in data and data["status"] != "resolved":
                sets.append("resolved_at = NULL")
            
            sets.append("updated_at = datetime('now')")
            
            if sets:
                vals.append(int(issue_id))
                db.execute("UPDATE issues SET " + ", ".join(sets) + " WHERE id = ?", vals)
            
            # Update assignments if provided
            if "assigned_employees" in data:
                # Clear existing assignments
                db.execute("DELETE FROM issue_assignments WHERE issue_id=?", (issue_id,))
                
                # Add new assignments
                assigned_ids = data.get("assigned_employees", [])
                primary_id = data.get("primary_employee")
                if assigned_ids:
                    expanded_ids, delegations = _expand_assignees_with_delegate(db, assigned_ids)
                    assign_employees_to_issue(db, issue_id, expanded_ids, primary_id)

                    _notify_assignees(
                        "issue",
                        int(issue_id),
                        expanded_ids,
                        title="Problém aktualizován",
                        body=f"Problém byl upraven: {data.get('title') or ''}".strip() or "Problém byl upraven",
                        actor_user_id=u.get("id"),
                    )
                    for d in delegations:
                        create_notification(
                            employee_id=d.get("to"),
                            kind="delegation",
                            title="Problém delegován",
                            body=f"Problém ID {issue_id} byl delegován od zaměstnance ID {d.get('from')}",
                            entity_type="issue",
                            entity_id=int(issue_id),
                        )
            
            db.commit()
            print(f"✓ Issue {issue_id} updated")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error updating issue: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

    if request.method == "DELETE":
        try:
            issue_id = request.args.get("id", type=int)
            if not issue_id:
                return jsonify({"ok": False, "error": "missing_id"}), 400
            
            # audit snapshot
            before = db.execute("SELECT id, job_id, title, description, type, status, severity, assigned_to FROM issues WHERE id=?", (issue_id,)).fetchone()
            before = dict(before) if before else None
            db.execute("DELETE FROM issues WHERE id = ?", (issue_id,))
            db.commit()
            audit_event(u.get("id"), "delete", "issue", issue_id, before=before)
            print(f"✓ Issue {issue_id} deleted")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error deleting issue: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500

# job_employees API - přiřazení zaměstnanců k zakázkám
@app.route("/api/job_employees", methods=["GET"])
def api_job_employees():
    u, err = require_role(write=False)
    if err: return err
    db = get_db()
    
    job_id = request.args.get("job_id", type=int)
    employee_id = request.args.get("employee_id", type=int)
    
    q = """SELECT je.*, e.name as employee_name, j.client as job_name
           FROM job_employees je
           LEFT JOIN employees e ON e.id = je.employee_id
           LEFT JOIN jobs j ON j.id = je.job_id"""
    conds = []
    params = []
    
    if job_id:
        conds.append("je.job_id = ?")
        params.append(job_id)
    if employee_id:
        conds.append("je.employee_id = ?")
        params.append(employee_id)
    
    if conds:
        q += " WHERE " + " AND ".join(conds)
    
    rows = db.execute(q, params).fetchall()
    return jsonify([dict(r) for r in rows])

# timesheets CRUD + export
@app.route("/api/timesheets", methods=["GET","POST","PATCH","DELETE"])
def api_timesheets():
    u, err = require_role(write=(request.method!="GET"))
    if err: return err
    db = get_db()

    if request.method == "GET":
        emp = request.args.get("employee_id", type=int)
        jid = request.args.get("job_id", type=int)
        task_id = request.args.get("task_id", type=int)
        d_from = _normalize_date(request.args.get("from"))
        d_to   = _normalize_date(request.args.get("to"))
        title_col = _job_title_col()
        
        # Zkontroluj existující sloupce
        timesheet_cols = [r[1] for r in db.execute("PRAGMA table_info(timesheets)").fetchall()]
        
        # Sestav SELECT s novými sloupci
        base_cols = "t.id,t.employee_id,t.job_id,t.date,t.hours,t.place,t.activity"
        new_cols = []
        
        if 'duration_minutes' in timesheet_cols:
            new_cols.append("COALESCE(t.duration_minutes, CAST(t.hours * 60 AS INTEGER)) AS duration_minutes")
        if 'labor_cost' in timesheet_cols:
            new_cols.append("COALESCE(t.labor_cost, 0) AS labor_cost")
        if 'work_type' in timesheet_cols:
            new_cols.append("t.work_type")
        if 'start_time' in timesheet_cols:
            new_cols.append("t.start_time")
        if 'end_time' in timesheet_cols:
            new_cols.append("t.end_time")
        if 'location' in timesheet_cols:
            new_cols.append("COALESCE(t.location, t.place) AS location")
        if 'task_id' in timesheet_cols:
            new_cols.append("t.task_id")
        if 'material_used' in timesheet_cols:
            new_cols.append("t.material_used")
        if 'weather_snapshot' in timesheet_cols:
            new_cols.append("t.weather_snapshot")
        if 'performance_signal' in timesheet_cols:
            new_cols.append("t.performance_signal")
        if 'delay_reason' in timesheet_cols:
            new_cols.append("t.delay_reason")
        if 'delay_note' in timesheet_cols:
            new_cols.append("t.delay_note")
        if 'photo_url' in timesheet_cols:
            new_cols.append("t.photo_url")
        if 'note' in timesheet_cols:
            new_cols.append("COALESCE(t.note, t.activity) AS note")
        if 'ai_flags' in timesheet_cols:
            new_cols.append("t.ai_flags")
        if 'created_at' in timesheet_cols:
            new_cols.append("t.created_at")
        
        all_cols = base_cols
        if new_cols:
            all_cols += "," + ",".join(new_cols)
        
        q = f"""SELECT {all_cols},
                      e.name AS employee_name, j.{title_col} AS job_title, j.code AS job_code
               FROM timesheets t
               LEFT JOIN employees e ON e.id=t.employee_id
               LEFT JOIN jobs j ON j.id=t.job_id"""
        conds=[]; params=[]
        if emp: conds.append("t.employee_id=?"); params.append(emp)
        if jid: conds.append("t.job_id=?"); params.append(jid)
        if task_id: conds.append("t.task_id=?"); params.append(task_id)
        if d_from and d_to:
            conds.append("date(t.date) BETWEEN date(?) AND date(?)"); params.extend([d_from, d_to])
        elif d_from:
            conds.append("date(t.date) >= date(?)"); params.append(d_from)
        elif d_to:
            conds.append("date(t.date) <= date(?)"); params.append(d_to)
        if conds: q += " WHERE " + " AND ".join(conds)
        q += " ORDER BY t.date ASC, t.id ASC"
        rows = db.execute(q, params).fetchall()
        
        # Parsuj JSON sloupce
        import json as json_lib
        result_rows = []
        for r in rows:
            row_dict = dict(r)
            # Parsuj JSON sloupce
            for json_col in ['material_used', 'weather_snapshot', 'ai_flags']:
                if json_col in row_dict and row_dict[json_col]:
                    try:
                        row_dict[json_col] = json_lib.loads(row_dict[json_col])
                    except:
                        row_dict[json_col] = None
            result_rows.append(row_dict)
        
        return jsonify({"ok": True, "rows": result_rows})

    if request.method == "POST":
        try:
            import json as json_lib
            data = request.get_json(force=True, silent=True) or {}
            emp = data.get("employee_id")
            job = data.get("job_id")  # Může být None nebo 0
            dt = data.get("date")
            hours = data.get("hours")
            duration_minutes = data.get("duration_minutes")
            
            # Validace povinných polí
            if not emp or not dt:
                return jsonify({"ok": False, "error": "missing_required_fields"}), 400
            
            # job_id může být None/0 pro výkazy bez zakázky
            if job is None:
                job = 0
            
            # Vypočti duration_minutes z hours pokud není zadáno
            if duration_minutes is None:
                if hours is not None:
                    duration_minutes = int(float(hours) * 60)
                else:
                    duration_minutes = 480  # default 8h
            
            # Vypočti hours z duration_minutes pokud není zadáno
            if hours is None:
                hours = duration_minutes / 60.0
            
            # Stará pole (zpětná kompatibilita)
            place = data.get("place") or data.get("location") or ""
            activity = data.get("activity") or data.get("note") or ""
            
            # Nová pole
            user_id = data.get("user_id")
            work_type = data.get("work_type") or "manual"
            start_time = data.get("start_time")
            end_time = data.get("end_time")
            location = data.get("location") or place
            task_id = data.get("task_id")
            material_used = json_lib.dumps(data.get("material_used")) if data.get("material_used") else None
            weather_snapshot = json_lib.dumps(data.get("weather_snapshot")) if data.get("weather_snapshot") else None
            performance_signal = data.get("performance_signal") or "normal"
            delay_reason = data.get("delay_reason")
            delay_note = data.get("delay_note")
            photo_url = data.get("photo_url")
            note = data.get("note") or activity
            
            # Vypočti labor_cost
            labor_cost = calculate_labor_cost(int(emp), int(job) if job else None, duration_minutes, db)
            
            # Detekuj anomálie
            ai_flags_data = detect_anomalies({
                'duration_minutes': duration_minutes,
                'performance_signal': performance_signal,
                'delay_reason': delay_reason
            }, db)
            ai_flags = json_lib.dumps(ai_flags_data)
            
            # Vložení do DB
            cols = ["employee_id", "job_id", "date", "hours", "duration_minutes", "place", "activity"]
            vals = [int(emp), int(job), _normalize_date(dt), float(hours), duration_minutes, place, activity]
            
            # Přidej nová pole pokud existují sloupce
            timesheet_cols = [r[1] for r in db.execute("PRAGMA table_info(timesheets)").fetchall()]
            
            if 'user_id' in timesheet_cols:
                cols.append("user_id"); vals.append(int(user_id) if user_id else None)
            if 'work_type' in timesheet_cols:
                cols.append("work_type"); vals.append(work_type)
            if 'start_time' in timesheet_cols:
                cols.append("start_time"); vals.append(start_time)
            if 'end_time' in timesheet_cols:
                cols.append("end_time"); vals.append(end_time)
            if 'location' in timesheet_cols:
                cols.append("location"); vals.append(location)
            if 'task_id' in timesheet_cols:
                cols.append("task_id"); vals.append(int(task_id) if task_id else None)
            if 'material_used' in timesheet_cols:
                cols.append("material_used"); vals.append(material_used)
            if 'weather_snapshot' in timesheet_cols:
                cols.append("weather_snapshot"); vals.append(weather_snapshot)
            if 'performance_signal' in timesheet_cols:
                cols.append("performance_signal"); vals.append(performance_signal)
            if 'delay_reason' in timesheet_cols:
                cols.append("delay_reason"); vals.append(delay_reason)
            if 'delay_note' in timesheet_cols:
                cols.append("delay_note"); vals.append(delay_note)
            if 'photo_url' in timesheet_cols:
                cols.append("photo_url"); vals.append(photo_url)
            if 'note' in timesheet_cols:
                cols.append("note"); vals.append(note)
            if 'ai_flags' in timesheet_cols:
                cols.append("ai_flags"); vals.append(ai_flags)
            if 'labor_cost' in timesheet_cols:
                cols.append("labor_cost"); vals.append(labor_cost)
            
            placeholders = ",".join("?" * len(vals))
            db.execute(f"INSERT INTO timesheets({','.join(cols)}) VALUES ({placeholders})", vals)
            db.commit()
            
            # Získat ID nového výkazu
            new_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            
            # Zpracuj materiál
            if material_used:
                process_material_usage(new_id, data.get("material_used"), None, db)
            
            # Přepočti statistiky zakázky
            if job:
                recalculate_job_stats(int(job), db)
            
            print(f"✓ Timesheet {new_id} created successfully (emp:{emp}, job:{job}, date:{dt})")
            return jsonify({"ok": True, "id": new_id})
        except Exception as e:
            db.rollback()
            print(f"✗ Error creating timesheet: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"ok": False, "error": str(e)}), 500

    if request.method == "PATCH":
        try:
            import json as json_lib
            data = request.get_json(force=True, silent=True) or {}
            tid = data.get("id")
            if not tid:
                return jsonify({"ok": False, "error": "missing_id"}), 400
            
            # Získej starý výkaz pro delta materiálu
            old_row = db.execute("SELECT job_id, material_used FROM timesheets WHERE id=?", (tid,)).fetchone()
            old_job_id = old_row[0] if old_row else None
            old_material = old_row[1] if old_row and old_row[1] else None
            
            # Získej existující sloupce
            timesheet_cols = [r[1] for r in db.execute("PRAGMA table_info(timesheets)").fetchall()]
            
            # Povolená pole (stará + nová)
            allowed = ["employee_id", "job_id", "date", "hours", "duration_minutes", "place", "activity",
                      "location", "note", "work_type", "start_time", "end_time", "task_id",
                      "material_used", "weather_snapshot", "performance_signal", "delay_reason",
                      "delay_note", "photo_url"]
            
            sets, vals = [], []
            
            # Zpracuj každé pole
            for k in allowed:
                if k not in data:
                    continue
                
                # Zkontroluj existenci sloupce
                if k not in timesheet_cols and k not in ["hours", "place", "activity"]:
                    continue
                
                v = data[k]
                
                # Speciální zpracování podle typu
                if k == "date":
                    v = _normalize_date(v)
                elif k in ("employee_id", "job_id", "task_id"):
                    v = int(v) if v else None
                elif k == "hours":
                    v = float(v) if v is not None else None
                elif k == "duration_minutes":
                    v = int(v) if v is not None else None
                elif k in ("material_used", "weather_snapshot", "ai_flags"):
                    v = json_lib.dumps(v) if v else None
                
                sets.append(f"{k}=?"); vals.append(v)
            
            # Přepočti duration_minutes z hours nebo naopak
            if "hours" in data and "duration_minutes" not in data:
                hours_val = float(data["hours"]) if data["hours"] is not None else None
                if hours_val is not None and "duration_minutes" in timesheet_cols:
                    sets.append("duration_minutes=?"); vals.append(int(hours_val * 60))
            elif "duration_minutes" in data and "hours" not in data:
                mins_val = int(data["duration_minutes"]) if data["duration_minutes"] is not None else None
                if mins_val is not None:
                    sets.append("hours=?"); vals.append(mins_val / 60.0)
            
            # Přepočti labor_cost pokud se změnily relevantní hodnoty
            if any(k in data for k in ["employee_id", "job_id", "duration_minutes", "hours"]):
                emp = data.get("employee_id")
                job = data.get("job_id")
                duration = data.get("duration_minutes")
                
                if not emp:
                    old_emp = db.execute("SELECT employee_id FROM timesheets WHERE id=?", (tid,)).fetchone()
                    emp = old_emp[0] if old_emp else None
                
                if not job and old_job_id:
                    job = old_job_id
                
                if not duration:
                    old_dur = db.execute("SELECT duration_minutes, hours FROM timesheets WHERE id=?", (tid,)).fetchone()
                    if old_dur:
                        duration = old_dur[0] if old_dur[0] else int(old_dur[1] * 60) if old_dur[1] else 480
                
                if emp and duration and "labor_cost" in timesheet_cols:
                    labor_cost = calculate_labor_cost(int(emp), int(job) if job else None, duration, db)
                    sets.append("labor_cost=?"); vals.append(labor_cost)
            
            # Aktualizuj AI flags
            if any(k in data for k in ["duration_minutes", "hours", "performance_signal", "delay_reason"]):
                old_row = db.execute(
                    "SELECT duration_minutes, hours, performance_signal, delay_reason FROM timesheets WHERE id=?",
                    (tid,)
                ).fetchone()
                
                duration_for_flags = data.get("duration_minutes") or (old_row[0] if old_row else None) or (int((data.get("hours") or (old_row[1] if old_row else 0)) * 60))
                perf_signal = data.get("performance_signal") or (old_row[2] if old_row else "normal")
                delay_reason = data.get("delay_reason") or (old_row[3] if old_row else None)
                
                if "ai_flags" in timesheet_cols:
                    ai_flags_data = detect_anomalies({
                        'duration_minutes': duration_for_flags,
                        'performance_signal': perf_signal,
                        'delay_reason': delay_reason
                    }, db)
                    sets.append("ai_flags=?"); vals.append(json_lib.dumps(ai_flags_data))
            
            if not sets:
                return jsonify({"ok": False, "error": "no_fields"}), 400
            
            vals.append(int(tid))
            db.execute("UPDATE timesheets SET " + ",".join(sets) + " WHERE id=?", vals)
            db.commit()
            
            # Zpracuj změny materiálu
            if "material_used" in data:
                new_material = data.get("material_used")
                process_material_usage(tid, new_material, old_material, db)
            
            # Přepočti statistiky zakázky (stará i nová)
            new_job_id = data.get("job_id", old_job_id)
            if new_job_id:
                recalculate_job_stats(int(new_job_id), db)
            if old_job_id and old_job_id != new_job_id:
                recalculate_job_stats(int(old_job_id), db)
            
            print(f"✓ Timesheet {tid} updated successfully")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error updating timesheet: {e}")
            import traceback
            traceback.print_exc()
            return jsonify({"ok": False, "error": str(e)}), 500

    # DELETE
    try:
        tid = request.args.get("id", type=int)
        if not tid:
            return jsonify({"ok": False, "error": "missing_id"}), 400
        
        # Získej data před smazáním pro reverz materiálu
        old_row = db.execute(
            "SELECT job_id, material_used FROM timesheets WHERE id=?",
            (tid,)
        ).fetchone()
        
        old_job_id = old_row[0] if old_row else None
        old_material = old_row[1] if old_row else None
        
        # Vrať materiál do skladu
        if old_material:
            try:
                import json as json_lib
                material_list = json_lib.loads(old_material) if isinstance(old_material, str) else old_material
                if material_list:
                    process_material_usage(tid, [], material_list, db)  # Reverz: odečti zápornou deltu
            except Exception as e:
                print(f"Warning: Could not reverse material usage: {e}")
        
        # Smaž výkaz
        db.execute("DELETE FROM timesheets WHERE id=?", (tid,))
        db.commit()
        
        # Přepočti statistiky zakázky
        if old_job_id:
            recalculate_job_stats(int(old_job_id), db)
        
        print(f"✓ Timesheet {tid} deleted successfully")
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error deleting timesheet: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

# ----------------- Worklog Service Layer -----------------
def calculate_labor_cost(employee_id, job_id, duration_minutes, db=None):
    """Spočítej náklad práce - priorita: zakázka > zaměstnanec > default"""
    if db is None:
        db = get_db()
    
    DEFAULT_HOURLY_RATE = 350.0  # Kč/hodinu
    
    rate = DEFAULT_HOURLY_RATE
    
    # Zkus získat sazbu z zakázky
    if job_id:
        job_cols = [r[1] for r in db.execute("PRAGMA table_info(jobs)").fetchall()]
        if 'hourly_rate' in job_cols:
            job_rate = db.execute("SELECT hourly_rate FROM jobs WHERE id=?", (job_id,)).fetchone()
            if job_rate and job_rate[0]:
                rate = float(job_rate[0])
    
    # Zkus získat sazbu ze zaměstnance
    if employee_id:
        emp_cols = [r[1] for r in db.execute("PRAGMA table_info(employees)").fetchall()]
        if 'hourly_rate' in emp_cols:
            emp_rate = db.execute("SELECT hourly_rate FROM employees WHERE id=?", (employee_id,)).fetchone()
            if emp_rate and emp_rate[0]:
                rate = float(emp_rate[0])
    
    return (duration_minutes / 60.0) * rate

def detect_anomalies(worklog_data, db=None):
    """Detekuj anomálie ve výkazu"""
    if db is None:
        db = get_db()
    
    flags = {
        'anomaly': False,
        'overtime': False,
        'low_performance': False
    }
    
    duration_minutes = worklog_data.get('duration_minutes') or (worklog_data.get('hours', 0) * 60)
    
    # Přesčas (>8h)
    if duration_minutes > 480:
        flags['overtime'] = True
    
    # Problémový výkon
    perf_signal = worklog_data.get('performance_signal', 'normal')
    if perf_signal in ['slow', 'problem']:
        flags['low_performance'] = True
    
    # Kombinace = anomálie
    if flags['overtime'] and flags['low_performance']:
        flags['anomaly'] = True
    
    if worklog_data.get('delay_reason'):
        flags['anomaly'] = True
    
    return flags

def recalculate_job_stats(job_id, db=None):
    """Přepočti statistiky zakázky z výkazů"""
    if db is None:
        db = get_db()
    
    if not job_id:
        return
    
    # Zkontroluj existující sloupce
    job_cols = [r[1] for r in db.execute("PRAGMA table_info(jobs)").fetchall()]
    
    # Součet hodin (duration_minutes)
    total_minutes = db.execute(
        "SELECT COALESCE(SUM(COALESCE(duration_minutes, hours * 60)), 0) FROM timesheets WHERE job_id=?",
        (job_id,)
    ).fetchone()[0] or 0
    
    if 'time_spent_minutes' in job_cols:
        db.execute("UPDATE jobs SET time_spent_minutes = ? WHERE id = ?", (int(total_minutes), job_id))
    
    # Náklady práce
    labor_cost_total = db.execute(
        "SELECT COALESCE(SUM(COALESCE(labor_cost, 0)), 0) FROM timesheets WHERE job_id=?",
        (job_id,)
    ).fetchone()[0] or 0
    
    if 'labor_cost_total' in job_cols:
        db.execute("UPDATE jobs SET labor_cost_total = ? WHERE id = ?", (float(labor_cost_total), job_id))
    elif 'cost_spent' in job_cols:
        db.execute("UPDATE jobs SET cost_spent = ? WHERE id = ?", (float(labor_cost_total), job_id))
    
    # Progress (pokud má plánované hodiny)
    if 'planned_hours' in job_cols and 'progress' in job_cols:
        planned = db.execute("SELECT planned_hours FROM jobs WHERE id=?", (job_id,)).fetchone()
        if planned and planned[0] and planned[0] > 0:
            progress = min(100, int((total_minutes / 60.0 / planned[0]) * 100))
            db.execute("UPDATE jobs SET progress = ? WHERE id = ?", (progress, job_id))
    
    db.commit()

def process_material_usage(worklog_id, new_material, old_material=None, db=None):
    """Zpracuj změny spotřeby materiálu"""
    if db is None:
        db = get_db()
    
    import json
    
    new_material = new_material or []
    old_material = old_material or []
    
    if isinstance(new_material, str):
        try:
            new_material = json.loads(new_material) if new_material else []
        except:
            new_material = []
    if isinstance(old_material, str):
        try:
            old_material = json.loads(old_material) if old_material else []
        except:
            old_material = []
    
    # Vytvoř mapu změn
    old_map = {m.get('item_id'): m.get('qty', 0) for m in old_material if m.get('item_id')}
    new_map = {m.get('item_id'): m.get('qty', 0) for m in new_material if m.get('item_id')}
    
    all_items = set(old_map.keys()) | set(new_map.keys())
    
    # Zkontroluj existenci warehouse_items tabulky
    if not _table_exists(db, 'warehouse_items'):
        return
    
    for item_id in all_items:
        old_qty = old_map.get(item_id, 0)
        new_qty = new_map.get(item_id, 0)
        delta = new_qty - old_qty
        
        if delta != 0:
            # Aktualizuj sklad (záporná delta = odečet)
            try:
                db.execute(
                    "UPDATE warehouse_items SET quantity = quantity + ? WHERE id = ?",
                    (-delta, item_id)
                )
                db.commit()
            except Exception as e:
                print(f"Warning: Could not update stock for item {item_id}: {e}")
                db.rollback()

@app.route("/api/timesheets/summary", methods=["GET"])
def api_timesheets_summary():
    """Agregační endpoint pro souhrnné statistiky"""
    u, err = require_role(write=False)
    if err:
        return err
    
    db = get_db()
    emp = request.args.get("employee_id", type=int)
    jid = request.args.get("job_id", type=int)
    d_from = _normalize_date(request.args.get("from"))
    d_to = _normalize_date(request.args.get("to"))
    
    # Build WHERE conditions
    conds = []
    params = []
    if emp:
        conds.append("t.employee_id=?")
        params.append(emp)
    if jid:
        conds.append("t.job_id=?")
        params.append(jid)
    if d_from and d_to:
        conds.append("date(t.date) BETWEEN date(?) AND date(?)")
        params.extend([d_from, d_to])
    elif d_from:
        conds.append("date(t.date) >= date(?)")
        params.append(d_from)
    elif d_to:
        conds.append("date(t.date) <= date(?)")
        params.append(d_to)
    
    where_clause = " WHERE " + " AND ".join(conds) if conds else ""
    
    # Zkontroluj existující sloupce
    timesheet_cols = [r[1] for r in db.execute("PRAGMA table_info(timesheets)").fetchall()]
    duration_col = "COALESCE(t.duration_minutes, CAST(t.hours * 60 AS INTEGER))" if 'duration_minutes' in timesheet_cols else "CAST(t.hours * 60 AS INTEGER)"
    
    # Celkem minut a hodin
    total_result = db.execute(
        f"SELECT SUM({duration_col}) as total_minutes, COUNT(DISTINCT t.date) as work_days FROM timesheets t {where_clause}",
        params
    ).fetchone()
    
    total_minutes = int(total_result[0] or 0)
    total_hours = total_minutes / 60.0
    work_days = int(total_result[1] or 0)
    avg_per_day = (total_hours / work_days) if work_days > 0 else 0
    
    # Overtime (>8h denně)
    overtime_result = db.execute(
        f"""SELECT SUM(CASE WHEN {duration_col} > 480 THEN {duration_col} - 480 ELSE 0 END) as overtime_minutes
            FROM timesheets t {where_clause}""",
        params
    ).fetchone()
    overtime_minutes = int(overtime_result[0] or 0)
    
    # Top jobs
    title_col = _job_title_col()
    top_jobs = db.execute(
        f"""SELECT t.job_id, j.{title_col} as job_name, SUM({duration_col}) as minutes
            FROM timesheets t
            LEFT JOIN jobs j ON j.id = t.job_id
            {where_clause}
            GROUP BY t.job_id, j.{title_col}
            ORDER BY minutes DESC
            LIMIT 10""",
        params
    ).fetchall()
    
    # Top work types
    top_work_types = []
    if 'work_type' in timesheet_cols:
        top_work_types = db.execute(
            f"""SELECT work_type, SUM({duration_col}) as minutes
                FROM timesheets t
                {where_clause}
                GROUP BY work_type
                ORDER BY minutes DESC""",
            params
        ).fetchall()
    
    # Anomálie count
    anomalies_count = 0
    if 'ai_flags' in timesheet_cols:
        import json as json_lib
        all_flags = db.execute(
            f"SELECT ai_flags FROM timesheets t {where_clause}",
            params
        ).fetchall()
        for flag_row in all_flags:
            if flag_row[0]:
                try:
                    flags = json_lib.loads(flag_row[0])
                    if flags.get('anomaly'):
                        anomalies_count += 1
                except:
                    pass
    
    return jsonify({
        "ok": True,
        "total_minutes": total_minutes,
        "total_hours": round(total_hours, 2),
        "avg_per_day": round(avg_per_day, 2),
        "overtime_minutes": overtime_minutes,
        "work_days": work_days,
        "top_jobs": [{"job_id": r[0], "job_name": r[1] or f"Projekt {r[0]}", "minutes": int(r[2] or 0)} for r in top_jobs],
        "top_work_types": [{"work_type": r[0] or "manual", "minutes": int(r[1] or 0)} for r in top_work_types],
        "anomalies_count": anomalies_count
    })

@app.route("/api/timesheets/heatmap", methods=["GET"])
def api_timesheets_heatmap():
    """Heatmapa dat pro vizualizaci týdenního zatížení"""
    u, err = require_role(write=False)
    if err:
        return err
    
    db = get_db()
    emp = request.args.get("employee_id", type=int)
    d_from = _normalize_date(request.args.get("from"))
    d_to = _normalize_date(request.args.get("to"))
    
    if not d_from or not d_to:
        return jsonify({"ok": False, "error": "from and to dates required"}), 400
    
    conds = ["date(t.date) BETWEEN date(?) AND date(?)"]
    params = [d_from, d_to]
    
    if emp:
        conds.append("t.employee_id=?")
        params.append(emp)
    
    where_clause = " WHERE " + " AND ".join(conds)
    
    # Zkontroluj existující sloupce
    timesheet_cols = [r[1] for r in db.execute("PRAGMA table_info(timesheets)").fetchall()]
    duration_col = "COALESCE(t.duration_minutes, CAST(t.hours * 60 AS INTEGER))" if 'duration_minutes' in timesheet_cols else "CAST(t.hours * 60 AS INTEGER)"
    
    # Agregace po dnech
    days_data = db.execute(
        f"""SELECT t.date, SUM({duration_col}) as total_minutes,
                   COUNT(*) as entries_count
            FROM timesheets t
            {where_clause}
            GROUP BY t.date
            ORDER BY t.date""",
        params
    ).fetchall()
    
    import json as json_lib
    
    days = []
    for row in days_data:
        date_str = row[0]
        total_mins = int(row[1] or 0)
        
        # Load level: 0=<4h, 1=4-6h, 2=6-8h, 3=>8h
        load_level = 0
        if total_mins >= 480:  # >=8h
            load_level = 3
        elif total_mins >= 360:  # >=6h
            load_level = 2
        elif total_mins >= 240:  # >=4h
            load_level = 1
        
        # Zkontroluj AI flags pro tento den
        day_flags = {"overtime": total_mins > 480, "anomaly": False}
        if 'ai_flags' in timesheet_cols:
            day_entries = db.execute(
                "SELECT ai_flags FROM timesheets WHERE date = ?" + (" AND employee_id = ?" if emp else ""),
                [date_str] + ([emp] if emp else [])
            ).fetchall()
            for flag_row in day_entries:
                if flag_row[0]:
                    try:
                        flags = json_lib.loads(flag_row[0])
                        if flags.get('anomaly'):
                            day_flags["anomaly"] = True
                            break
                    except:
                        pass
        
        days.append({
            "date": date_str,
            "total_minutes": total_mins,
            "load_level": load_level,
            "flags": day_flags
        })
    
    return jsonify({"ok": True, "days": days})

@app.route("/api/timesheets/ai-insights", methods=["GET"])
def api_timesheets_ai_insights():
    """AI signály a doporučení"""
    u, err = require_role(write=False)
    if err:
        return err
    
    db = get_db()
    d_from = _normalize_date(request.args.get("from"))
    d_to = _normalize_date(request.args.get("to"))
    
    if not d_from or not d_to:
        # Default: poslední týden
        from datetime import datetime, timedelta
        d_to = datetime.now().strftime('%Y-%m-%d')
        d_from = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    
    where_clause = " WHERE date(t.date) BETWEEN date(?) AND date(?)"
    params = [d_from, d_to]
    
    timesheet_cols = [r[1] for r in db.execute("PRAGMA table_info(timesheets)").fetchall()]
    duration_col = "COALESCE(t.duration_minutes, CAST(t.hours * 60 AS INTEGER))" if 'duration_minutes' in timesheet_cols else "CAST(t.hours * 60 AS INTEGER)"
    
    import json as json_lib
    
    # Workload risk (průměrné denní zatížení)
    workload_data = db.execute(
        f"""SELECT AVG({duration_col}) as avg_daily, COUNT(DISTINCT t.date) as days
            FROM timesheets t {where_clause}""",
        params
    ).fetchone()
    
    avg_daily = float(workload_data[0] or 0) / 60.0  # v hodinách
    workload_risk = min(100, int((avg_daily / 8.0) * 100))  # 0-100 skóre
    
    # Slowdown jobs (zakázky s rostoucím časem bez progressu)
    title_col = _job_title_col()
    slowdown_jobs = []
    if 'work_type' in timesheet_cols and 'performance_signal' in timesheet_cols:
        problem_jobs = db.execute(
            f"""SELECT t.job_id, j.{title_col} as job_name, COUNT(*) as problem_count
                FROM timesheets t
                LEFT JOIN jobs j ON j.id = t.job_id
                {where_clause}
                AND (t.performance_signal IN ('slow', 'problem') OR t.delay_reason IS NOT NULL)
                GROUP BY t.job_id, j.{title_col}
                HAVING problem_count >= 2
                ORDER BY problem_count DESC
                LIMIT 5""",
            params
        ).fetchall()
        
        slowdown_jobs = [
            {
                "job_id": r[0],
                "job_name": r[1] or f"Projekt {r[0]}",
                "reason": "Rostoucí čas bez progressu"
            }
            for r in problem_jobs
        ]
    
    # Anomaly days
    anomaly_days = []
    if 'ai_flags' in timesheet_cols:
        anomaly_rows = db.execute(
            f"""SELECT DISTINCT t.date, COUNT(*) as count
                FROM timesheets t
                {where_clause}
                AND t.ai_flags IS NOT NULL
                GROUP BY t.date""",
            params
        ).fetchall()
        
        for row in anomaly_rows:
            date_str = row[0]
            flags_data = db.execute(
                "SELECT ai_flags FROM timesheets WHERE date = ? AND ai_flags IS NOT NULL LIMIT 1",
                (date_str,)
            ).fetchone()
            
            if flags_data and flags_data[0]:
                try:
                    flags = json_lib.loads(flags_data[0])
                    if flags.get('anomaly'):
                        reasons = []
                        if flags.get('overtime'):
                            reasons.append("Přesčas")
                        if flags.get('low_performance'):
                            reasons.append("Problémový výkon")
                        anomaly_days.append({
                            "date": date_str,
                            "reason": " + ".join(reasons) if reasons else "Detekována anomálie"
                        })
                except:
                    pass
    
    # Recommendations
    recommendations = []
    if workload_risk > 70:
        recommendations.append("Vysoké vytížení týmu - zvažte redistribuci práce")
    if slowdown_jobs:
        recommendations.append(f"Zkontroluj zakázku '{slowdown_jobs[0]['job_name']}' - riziko skluzu")
    if anomaly_days:
        recommendations.append(f"{len(anomaly_days)} dní s anomáliemi - zkontroluj detaily")
    
    return jsonify({
        "ok": True,
        "workload_risk": workload_risk,
        "slowdown_jobs": slowdown_jobs,
        "anomaly_days": anomaly_days,
        "recommendations": recommendations
    })

@app.route("/api/timesheets/export")
def api_timesheets_export():
    u, err = require_role(write=False)
    if err: return err
    db = get_db()
    emp = request.args.get("employee_id", type=int)
    jid = request.args.get("job_id", type=int)
    d_from = _normalize_date(request.args.get("from"))
    d_to   = _normalize_date(request.args.get("to"))
    title_col = _job_title_col()
    q = f"""SELECT t.id,t.date,t.hours,t.place,t.activity,
                  e.name AS employee_name, e.id AS employee_id,
                  j.{title_col} AS job_title, j.code AS job_code, j.id AS job_id
           FROM timesheets t
           LEFT JOIN employees e ON e.id=t.employee_id
           LEFT JOIN jobs j ON j.id=t.job_id"""
    conds=[]; params=[]
    if emp: conds.append("t.employee_id=?"); params.append(emp)
    if jid: conds.append("t.job_id=?"); params.append(jid)
    if d_from and d_to:
        conds.append("date(t.date) BETWEEN date(?) AND date(?)"); params.extend([d_from, d_to])
    elif d_from:
        conds.append("date(t.date) >= date(?)"); params.append(d_from)
    elif d_to:
        conds.append("date(t.date) <= date(?)"); params.append(d_to)
    if conds: q += " WHERE " + " AND ".join(conds)
    q += " ORDER BY t.date ASC, t.id ASC"
    rows = get_db().execute(q, params).fetchall()

    output = io.StringIO()
    import csv as _csv
    writer = _csv.writer(output)
    writer.writerow(["id","date","employee_id","employee_name","job_id","job_title","job_code","hours","place","activity"])
    for r in rows:
        writer.writerow([r["id"], r["date"], r["employee_id"], r["employee_name"] or "", r["job_id"], r["job_title"] or "", r["job_code"] or "", r["hours"], r["place"] or "", r["activity"] or ""])
    mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    mem.seek(0)
    fname = "timesheets.csv"
    return send_file(mem, mimetype="text/csv", as_attachment=True, download_name=fname)


@app.route("/api/timesheets/export-advanced", methods=["GET"])
def api_timesheets_export_advanced():
    """Pokročilý export výkazů s filtry - podporuje PDF, XLSX, CSV"""
    u, err = require_role(write=False)
    if err: return err
    
    # Parametry z query stringu
    export_format = request.args.get("format", "csv").lower()  # csv, xlsx, pdf
    emp_ids = request.args.get("employees", "")  # comma-separated IDs nebo "all"
    jid = request.args.get("job_id", type=int)  # konkrétní zakázka nebo None pro všechny
    d_from = _normalize_date(request.args.get("from"))
    d_to = _normalize_date(request.args.get("to"))
    
    db = get_db()
    title_col = _job_title_col()
    
    # Sestavení SQL dotazu
    q = f"""SELECT t.id, t.date, t.hours, t.place, t.activity,
                  e.name AS employee_name, e.id AS employee_id,
                  j.{title_col} AS job_title, j.code AS job_code, j.id AS job_id
           FROM timesheets t
           LEFT JOIN employees e ON e.id = t.employee_id
           LEFT JOIN jobs j ON j.id = t.job_id"""
    
    conds = []
    params = []
    
    # Filtr zaměstnanců
    if emp_ids and emp_ids != "all":
        emp_list = [int(x.strip()) for x in emp_ids.split(",") if x.strip().isdigit()]
        if emp_list:
            placeholders = ",".join("?" * len(emp_list))
            conds.append(f"t.employee_id IN ({placeholders})")
            params.extend(emp_list)
    
    # Filtr zakázky
    if jid:
        conds.append("t.job_id = ?")
        params.append(jid)
    
    # Filtr data
    if d_from and d_to:
        conds.append("date(t.date) BETWEEN date(?) AND date(?)")
        params.extend([d_from, d_to])
    elif d_from:
        conds.append("date(t.date) >= date(?)")
        params.append(d_from)
    elif d_to:
        conds.append("date(t.date) <= date(?)")
        params.append(d_to)
    
    if conds:
        q += " WHERE " + " AND ".join(conds)
    q += " ORDER BY t.date ASC, t.id ASC"
    
    rows = db.execute(q, params).fetchall()
    
    # Generování exportu podle formátu
    if export_format == "pdf":
        return _generate_pdf_export(rows, d_from, d_to)
    elif export_format == "xlsx":
        return _generate_xlsx_export(rows, d_from, d_to)
    else:  # csv
        return _generate_csv_export(rows, d_from, d_to)


def _generate_csv_export(rows, d_from, d_to):
    """Generování CSV exportu"""
    output = io.StringIO()
    import csv as _csv
    writer = _csv.writer(output, delimiter=';')
    
    # Header
    writer.writerow(["Datum", "Zaměstnanec", "Zakázka", "Hodin", "Místo", "Činnost"])
    
    total_hours = 0
    for r in rows:
        hours = float(r["hours"] or 0)
        total_hours += hours
        writer.writerow([
            r["date"] or "",
            r["employee_name"] or "",
            (r["job_title"] or "") + (f" ({r['job_code']})" if r["job_code"] else ""),
            f"{hours:.2f}",
            r["place"] or "",
            r["activity"] or ""
        ])
    
    # Celkem
    writer.writerow([])
    writer.writerow(["CELKEM HODIN:", "", "", f"{total_hours:.2f}", "", ""])
    
    mem = io.BytesIO(output.getvalue().encode("utf-8-sig"))
    mem.seek(0)
    
    period = ""
    if d_from and d_to:
        period = f"_{d_from}__{d_to}"
    elif d_from:
        period = f"_od_{d_from}"
    elif d_to:
        period = f"_do_{d_to}"
    
    fname = f"vykazy{period}.csv"
    return send_file(mem, mimetype="text/csv", as_attachment=True, download_name=fname)


def _generate_xlsx_export(rows, d_from, d_to):
    """Generování XLSX exportu"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return jsonify({"error": "openpyxl not installed"}), 500
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Výkazy hodin"
    
    # Nadpis
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "VÝKAZY HODIN"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Období
    if d_from or d_to:
        ws.merge_cells('A2:F2')
        period_cell = ws['A2']
        if d_from and d_to:
            period_cell.value = f"Období: {d_from} – {d_to}"
        elif d_from:
            period_cell.value = f"Od: {d_from}"
        elif d_to:
            period_cell.value = f"Do: {d_to}"
        period_cell.font = Font(italic=True)
        period_cell.alignment = Alignment(horizontal='center')
    
    # Hlavička tabulky
    start_row = 4
    headers = ["Datum", "Zaměstnanec", "Zakázka", "Hodin", "Místo", "Činnost"]
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    total_hours = 0
    row_num = start_row + 1
    
    for r in rows:
        hours = float(r["hours"] or 0)
        total_hours += hours
        
        ws.cell(row=row_num, column=1, value=r["date"] or "")
        ws.cell(row=row_num, column=2, value=r["employee_name"] or "")
        job_str = (r["job_title"] or "") + (f" ({r['job_code']})" if r["job_code"] else "")
        ws.cell(row=row_num, column=3, value=job_str)
        ws.cell(row=row_num, column=4, value=hours)
        ws.cell(row=row_num, column=5, value=r["place"] or "")
        ws.cell(row=row_num, column=6, value=r["activity"] or "")
        
        row_num += 1
    
    # Celkem
    row_num += 1
    total_cell = ws.cell(row=row_num, column=1)
    total_cell.value = "CELKEM HODIN:"
    total_cell.font = Font(bold=True, size=12)
    
    total_value_cell = ws.cell(row=row_num, column=4)
    total_value_cell.value = total_hours
    total_value_cell.font = Font(bold=True, size=12, color="FF0000")
    
    # Formátování sloupců
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 30
    
    # Uložení do paměti
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    
    period = ""
    if d_from and d_to:
        period = f"_{d_from}__{d_to}"
    elif d_from:
        period = f"_od_{d_from}"
    elif d_to:
        period = f"_do_{d_to}"
    
    fname = f"vykazy{period}.xlsx"
    return send_file(mem, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                     as_attachment=True, download_name=fname)


def _generate_pdf_export(rows, d_from, d_to):
    """Generování PDF exportu"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return jsonify({"error": "reportlab not installed"}), 500
    
    mem = io.BytesIO()
    doc = SimpleDocTemplate(mem, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm,
                           topMargin=1.5*cm, bottomMargin=1*cm)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Try to register DejaVu fonts for Czech support
    try:
        # Try to find DejaVu fonts (common on most systems)
        import os
        font_paths = [
            '/System/Library/Fonts/Supplemental/Arial Unicode.ttf',  # macOS
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
            'C:\\Windows\\Fonts\\arial.ttf',  # Windows
        ]
        
        font_registered = False
        for font_path in font_paths:
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('CustomFont', font_path))
                font_registered = True
                break
        
        if font_registered:
            base_font = 'CustomFont'
        else:
            base_font = 'Helvetica'
    except:
        base_font = 'Helvetica'
    
    # Nadpis
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#2e7d32'),
        spaceAfter=12,
        alignment=1,  # center
        fontName=base_font
    )
    elements.append(Paragraph("VYKAZY HODIN", title_style))
    
    # Období
    if d_from or d_to:
        period_style = ParagraphStyle('Period', parent=styles['Normal'], fontSize=11, 
                                     textColor=colors.grey, alignment=1, fontName=base_font)
        if d_from and d_to:
            period_text = f"Obdobi: {d_from} - {d_to}"
        elif d_from:
            period_text = f"Od: {d_from}"
        elif d_to:
            period_text = f"Do: {d_to}"
        elements.append(Paragraph(period_text, period_style))
    
    elements.append(Spacer(1, 0.5*cm))
    
    # Helper function to clean Czech text
    def clean_text(text):
        if not text:
            return ""
        # Replace Czech characters with ASCII equivalents for Helvetica
        replacements = {
            'á': 'a', 'č': 'c', 'ď': 'd', 'é': 'e', 'ě': 'e', 'í': 'i', 
            'ň': 'n', 'ó': 'o', 'ř': 'r', 'š': 's', 'ť': 't', 'ú': 'u', 
            'ů': 'u', 'ý': 'y', 'ž': 'z',
            'Á': 'A', 'Č': 'C', 'Ď': 'D', 'É': 'E', 'Ě': 'E', 'Í': 'I',
            'Ň': 'N', 'Ó': 'O', 'Ř': 'R', 'Š': 'S', 'Ť': 'T', 'Ú': 'U',
            'Ů': 'U', 'Ý': 'Y', 'Ž': 'Z'
        }
        result = str(text)
        if base_font == 'Helvetica':
            for czech, ascii_char in replacements.items():
                result = result.replace(czech, ascii_char)
        return result
    
    # Tabulka
    data = [[clean_text("Datum"), clean_text("Zamestnanec"), clean_text("Zakazka"), 
             clean_text("Hodin"), clean_text("Misto"), clean_text("Cinnost")]]
    
    total_hours = 0
    for r in rows:
        hours = float(r["hours"] or 0)
        total_hours += hours
        job_str = (r["job_title"] or "") + (f" ({r['job_code']})" if r["job_code"] else "")
        data.append([
            clean_text(r["date"] or ""),
            clean_text(r["employee_name"] or ""),
            clean_text(job_str[:40]),  # zkrácení pro PDF
            f"{hours:.2f}",
            clean_text((r["place"] or "")[:20]),
            clean_text((r["activity"] or "")[:30])
        ])
    
    # Celkem
    data.append(["", "", "CELKEM:", f"{total_hours:.2f}", "", ""])
    
    table = Table(data, colWidths=[3*cm, 5*cm, 6*cm, 2.5*cm, 4*cm, 5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (3, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), base_font),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('TEXTCOLOR', (3, -1), (3, -1), colors.red),
    ]))
    
    elements.append(table)
    
    # Generování PDF
    doc.build(elements)
    mem.seek(0)
    
    period = ""
    if d_from and d_to:
        period = f"_{d_from}__{d_to}"
    elif d_from:
        period = f"_od_{d_from}"
    elif d_to:
        period = f"_do_{d_to}"
    
    fname = f"vykazy{period}.pdf"
    return send_file(mem, mimetype="application/pdf", as_attachment=True, download_name=fname)

# ----------------- Template route -----------------
@app.route("/trainings.html")
def page_trainings():
    return render_template("trainings.html")

@app.route("/timesheets.html")
def page_timesheets():
    # Render template version (keeps original Timesheets design) while using unified JS header via templates/layout.html
    return render_template("timesheets.html")
# ----------------- Standalone HTML routes -----------------
@app.route("/jobs.html")
def page_jobs():
    return send_from_directory(".", "jobs.html")

@app.route("/tasks.html")
def page_tasks():
    return send_from_directory(".", "tasks.html")

@app.route("/issues")
@app.route("/issues.html")
def page_issues():
    return send_from_directory(".", "issues.html")

@app.route("/employees.html")
def page_employees():
    return send_from_directory(".", "employees.html")

@app.route("/calendar.html")
def page_calendar():
    return send_from_directory(".", "calendar.html")

@app.route("/settings.html")
def page_settings():
    return send_from_directory(".", "settings.html")

@app.route("/warehouse.html")
def page_warehouse():
    return send_from_directory(".", "warehouse.html")

@app.route("/finance.html")
def page_finance():
    return send_from_directory(".", "finance.html")

@app.route("/documents.html")
def page_documents():
    return send_from_directory(".", "documents.html")

@app.route("/reports.html")
def page_reports():
    return send_from_directory(".", "reports.html")

# ----------------- Job detail UI routes -----------------
@app.route("/jobs/<int:job_id>")
def page_job_detail(job_id):
    # Classic server-rendered detail page with timesheet table
    return render_template("job_detail.html", job_id=job_id)

@app.route("/job.html")
def page_job_detail_query():
    jid = request.args.get("id", type=int)
    if not jid:
        # fallback to jobs tab if missing id
        return render_template("jobs_list.html")
    return render_template("job_detail.html", job_id=jid)

# ----------------- admin export / download -----------------

@app.route("/api/admin/download-db", methods=["GET"])
@requires_role('owner', 'admin')
def api_admin_download_db():
    """Stáhne aktuální SQLite databázi jako soubor (pouze owner/admin)."""
    db_path = os.environ.get("DB_PATH") or "/var/data/app.db"
    if not os.path.exists(db_path):
        return jsonify({"ok": False, "error": "db_not_found", "path": db_path}), 404
    return send_file(db_path, as_attachment=True, download_name="app.db")

@app.route("/api/admin/export-all", methods=["GET"])
@requires_role('owner', 'admin')
def api_admin_export_all():
    """Export všech tabulek do JSON (pouze owner/admin)."""
    db = get_db()
    tables = [r["name"] for r in db.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
    ).fetchall()]
    export = {
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "db_path": os.environ.get("DB_PATH") or "/var/data/app.db",
        "tables": {},
    }
    for t in tables:
        rows = db.execute(f"SELECT * FROM {t}").fetchall()
        export["tables"][t] = [dict(r) for r in rows]
    payload = json.dumps(export, ensure_ascii=False, indent=2)
    buf = io.BytesIO(payload.encode("utf-8"))
    filename = f"green-david-export-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.json"
    return send_file(buf, mimetype="application/json", as_attachment=True, download_name=filename)





@app.route("/search", methods=["GET"])
def search_page():
    q = (request.args.get("q") or "").strip()
    results = []
    if q:
        like = f"%{q}%"
        db = get_db()
        try:
            # Use title column if available, fallback to name
            title_col = _job_title_col()
            cur = db.execute(f"SELECT id, {title_col} AS title, city, code, date FROM jobs WHERE ({title_col} LIKE ? COLLATE NOCASE OR city LIKE ? COLLATE NOCASE OR code LIKE ? COLLATE NOCASE) ORDER BY id DESC LIMIT 50", (like, like, like))
            for r in cur.fetchall():
                results.append({"type":"Zakázka","id":r["id"],"title":r["title"] or "", "sub":" • ".join([x for x in [r["city"], r["code"]] if x]),"date":r["date"],"url": f"/?tab=jobs&jobId={r['id']}"})
        except Exception: pass
        try:
            cur = db.execute("SELECT id, name, role FROM employees WHERE (name LIKE ? COLLATE NOCASE OR role LIKE ? COLLATE NOCASE) ORDER BY id DESC LIMIT 50", (like, like))
            for r in cur.fetchall():
                results.append({"type":"Zaměstnanec","id":r["id"],"title":r["name"],"sub":r["role"] or "","date":"","url": "/?tab=employees"})
        except Exception: pass
    return render_template("search.html", title="Hledání", q=q, results=results)

# ----------------- Calendar API -----------------
@app.route("/api/calendar", methods=["GET", "POST", "PATCH"])
def api_calendar():
    db = get_db()
    if request.method == "GET":
        month_str = request.args.get("month")
        date_str = request.args.get("date")
        from_str = request.args.get("from")
        to_str = request.args.get("to")
        
        q = "SELECT id, date, title, kind, job_id, start_time, end_time, note, color FROM calendar_events WHERE 1=1"
        params = []
        
        if month_str:
            try:
                year, month = [int(x) for x in month_str.split("-")]
                q += " AND strftime('%Y-%m', date) = ?"
                params.append(month_str)
            except Exception:
                pass
        elif date_str:
            q += " AND date = ?"
            params.append(_normalize_date(date_str))
        elif from_str or to_str:
            if from_str:
                q += " AND date >= ?"
                params.append(_normalize_date(from_str))
            if to_str:
                q += " AND date <= ?"
                params.append(_normalize_date(to_str))
        
        q += " ORDER BY date ASC, id ASC"
        rows = [dict(r) for r in db.execute(q, params).fetchall()]
        return jsonify({"ok": True, "events": rows, "items": rows})
    
    u, err = require_auth()
    if err: return err
    
    data = request.get_json(force=True, silent=True) or {}
    
    if request.method == "POST":
        try:
            title = (data.get("title") or "").strip()
            date_str = _normalize_date(data.get("date"))
            kind = (data.get("kind") or data.get("type") or "note").strip()
            color = (data.get("color") or "#2e7d32").strip()
            note = (data.get("note") or data.get("details") or "").strip()
            job_id = data.get("job_id")
            start_time = data.get("start_time") or ""
            end_time = data.get("end_time") or ""
            
            if not title or not date_str:
                return jsonify({"ok": False, "error": "missing_fields"}), 400
            
            db.execute(
                "INSERT INTO calendar_events(date, title, kind, job_id, start_time, end_time, note, color) VALUES (?,?,?,?,?,?,?,?)",
                (date_str, title, kind, job_id, start_time, end_time, note, color)
            )
            db.commit()
            print(f"✓ Calendar event '{title}' created successfully")
            return jsonify({"ok": True})
        except Exception as e:
            db.rollback()
            print(f"✗ Error creating calendar event: {e}")
            return jsonify({"ok": False, "error": str(e)}), 500
    
    # PATCH
    try:
        ev_id = data.get("id")
        if not ev_id:
            return jsonify({"ok": False, "error": "missing_id"}), 400
        
        updates = []
        params = []
        allowed = ["title", "date", "kind", "type", "color", "note", "details", "job_id", "start_time", "end_time"]
        for k in allowed:
            if k in data:
                if k == "date":
                    params.append(_normalize_date(data[k]))
                elif k == "type":
                    params.append(data[k])
                    updates.append("kind=?")
                    continue
                elif k == "details":
                    params.append(data[k])
                    updates.append("note=?")
                    continue
                else:
                    params.append(data[k])
                updates.append(f"{k}=?")
        
        if not updates:
            return jsonify({"ok": False, "error": "nothing_to_update"}), 400
        
        params.append(int(ev_id))
        db.execute("UPDATE calendar_events SET " + ", ".join(updates) + " WHERE id=?", params)
        db.commit()
        print(f"✓ Calendar event {ev_id} updated successfully")
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error updating calendar event: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/calendar/<int:event_id>", methods=["DELETE"])
def api_calendar_delete(event_id):
    u, err = require_auth()
    if err: return err
    db = get_db()
    try:
        db.execute("DELETE FROM calendar_events WHERE id=?", (event_id,))
        db.commit()
        print(f"✓ Calendar event {event_id} deleted successfully")
        return jsonify({"ok": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error deleting calendar event: {e}")
        return jsonify({"ok": False, "error": str(e)}), 500

# ----------------- /gd/api/ aliases -----------------
@app.route("/gd/api/jobs", methods=["GET", "POST", "PATCH", "DELETE"])
def gd_api_jobs():
    return api_jobs()

@app.route("/gd/api/jobs/<int:job_id>", methods=["GET"])
def gd_api_job_detail(job_id):
    return api_job_detail(job_id)

# === ATTACHMENTS API ===
import os
import uuid
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx', 'xls', 'xlsx', 'zip', 'mp4', 'mov'}

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route("/api/tasks/<int:task_id>/attachments", methods=["GET", "POST", "DELETE"])
def api_task_attachments(task_id):
    u, err = require_role(write=(request.method != "GET"))
    if err: return err
    db = get_db()
    
    if request.method == "GET":
        rows = db.execute("""
            SELECT a.*, e.name as uploader_name
            FROM task_attachments a
            LEFT JOIN employees e ON e.id = a.uploaded_by
            WHERE a.task_id = ?
            ORDER BY a.uploaded_at DESC
        """, (task_id,)).fetchall()
        return jsonify({"ok": True, "attachments": [dict(r) for r in rows]})
    
    if request.method == "POST":
        if 'file' not in request.files:
            return jsonify({"ok": False, "error": "no_file"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"ok": False, "error": "empty_filename"}), 400
        
        if file and allowed_file(file.filename):
            original_filename = secure_filename(file.filename)
            ext = original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else ''
            filename = f"{uuid.uuid4().hex}.{ext}"
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            
            file.save(filepath)
            file_size = os.path.getsize(filepath)
            
            db.execute("""
                INSERT INTO task_attachments (task_id, filename, original_filename, file_path, file_size, mime_type, uploaded_by)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (task_id, filename, original_filename, filepath, file_size, file.content_type, u["id"]))
            db.commit()
            
            return jsonify({"ok": True, "filename": original_filename})
        
        return jsonify({"ok": False, "error": "invalid_file_type"}), 400
    
    if request.method == "DELETE":
        att_id = request.args.get("id", type=int)
        if not att_id:
            return jsonify({"ok": False, "error": "missing_id"}), 400
        
        row = db.execute("SELECT file_path FROM task_attachments WHERE id = ?", (att_id,)).fetchone()
        if row and os.path.exists(row[0]):
            os.remove(row[0])
        
        db.execute("DELETE FROM task_attachments WHERE id = ?", (att_id,))
        db.commit()
        return jsonify({"ok": True})

@app.route("/api/issues/<int:issue_id>/attachments", methods=["GET", "POST", "DELETE"])
def api_issue_attachments(issue_id):
    u, err = require_role(write=(request.method != "GET"))
    if err: return err
    db = get_db()
    
    if request.method == "GET":
        rows = db.execute("""
            SELECT a.*, e.name as uploader_name
            FROM issue_attachments a
            LEFT JOIN employees e ON e.id = a.uploaded_by
            WHERE a.issue_id = ?
            ORDER BY a.uploaded_at DESC
        """, (issue_id,)).fetchall()
        return jsonify({"ok": True, "attachments": [dict(r) for r in rows]})
    
    if request.method == "POST":
        if 'file' not in request.files:
            return jsonify({"ok": False, "error": "no_file"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"ok": False, "error": "empty_filename"}), 400
        
        if file and allowed_file(file.filename):
            original_filename = secure_filename(file.filename)
            ext = original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else ''
            filename = f"{uuid.uuid4().hex}.{ext}"
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            
            file.save(filepath)
            file_size = os.path.getsize(filepath)
            
            db.execute("""
                INSERT INTO issue_attachments (issue_id, filename, original_filename, file_path, file_size, mime_type, uploaded_by)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (issue_id, filename, original_filename, filepath, file_size, file.content_type, u["id"]))
            db.commit()
            
            return jsonify({"ok": True, "filename": original_filename})
        
        return jsonify({"ok": False, "error": "invalid_file_type"}), 400
    
    if request.method == "DELETE":
        att_id = request.args.get("id", type=int)
        if not att_id:
            return jsonify({"ok": False, "error": "missing_id"}), 400
        
        row = db.execute("SELECT file_path FROM issue_attachments WHERE id = ?", (att_id,)).fetchone()
        if row and os.path.exists(row[0]):
            os.remove(row[0])
        
        db.execute("DELETE FROM issue_attachments WHERE id = ?", (att_id,))
        db.commit()
        return jsonify({"ok": True})

# Download attachment
@app.route("/api/attachments/<path:filename>")
def download_attachment(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)

# === COMMENTS API ===
@app.route("/api/tasks/<int:task_id>/comments", methods=["GET", "POST"])
def api_task_comments(task_id):
    u, err = require_role(write=(request.method != "GET"))
    if err: return err
    db = get_db()
    
    if request.method == "GET":
        rows = db.execute("""
            SELECT c.*, e.name as author_name
            FROM task_comments c
            LEFT JOIN employees e ON e.id = c.user_id
            WHERE c.task_id = ?
            ORDER BY c.created_at ASC
        """, (task_id,)).fetchall()
        return jsonify({"ok": True, "comments": [dict(r) for r in rows]})
    
    data = request.get_json(force=True, silent=True) or {}
    comment = (data.get("comment") or "").strip()
    if not comment:
        return jsonify({"ok": False, "error": "empty_comment"}), 400
    
    db.execute("""
        INSERT INTO task_comments (task_id, user_id, comment)
        VALUES (?, ?, ?)
    """, (task_id, u["id"], comment))
    db.commit()
    return jsonify({"ok": True})

@app.route("/api/issues/<int:issue_id>/comments", methods=["GET", "POST"])
def api_issue_comments(issue_id):
    u, err = require_role(write=(request.method != "GET"))
    if err: return err
    db = get_db()
    
    if request.method == "GET":
        rows = db.execute("""
            SELECT c.*, e.name as author_name
            FROM issue_comments c
            LEFT JOIN employees e ON e.id = c.user_id
            WHERE c.issue_id = ?
            ORDER BY c.created_at ASC
        """, (issue_id,)).fetchall()
        return jsonify({"ok": True, "comments": [dict(r) for r in rows]})
    
    data = request.get_json(force=True, silent=True) or {}
    comment = (data.get("comment") or "").strip()
    if not comment:
        return jsonify({"ok": False, "error": "empty_comment"}), 400
    
    db.execute("""
        INSERT INTO issue_comments (issue_id, user_id, comment)
        VALUES (?, ?, ?)
    """, (issue_id, u["id"], comment))
    db.commit()
    return jsonify({"ok": True})

# === LOCATIONS API ===
@app.route("/api/tasks/<int:task_id>/location", methods=["GET", "POST", "DELETE"])
def api_task_location(task_id):
    u, err = require_role(write=(request.method != "GET"))
    if err: return err
    db = get_db()
    
    if request.method == "GET":
        row = db.execute("SELECT * FROM task_locations WHERE task_id = ? ORDER BY created_at DESC LIMIT 1", (task_id,)).fetchone()
        return jsonify({"ok": True, "location": dict(row) if row else None})
    
    if request.method == "POST":
        data = request.get_json(force=True, silent=True) or {}
        lat = data.get("latitude")
        lng = data.get("longitude")
        address = data.get("address", "")
        
        if lat is None or lng is None:
            return jsonify({"ok": False, "error": "missing_coordinates"}), 400
        
        # Delete old location
        db.execute("DELETE FROM task_locations WHERE task_id = ?", (task_id,))
        
        # Insert new
        db.execute("""
            INSERT INTO task_locations (task_id, latitude, longitude, address)
            VALUES (?, ?, ?, ?)
        """, (task_id, float(lat), float(lng), address))
        db.commit()
        return jsonify({"ok": True})
    
    if request.method == "DELETE":
        db.execute("DELETE FROM task_locations WHERE task_id = ?", (task_id,))
        db.commit()
        return jsonify({"ok": True})

@app.route("/api/issues/<int:issue_id>/location", methods=["GET", "POST", "DELETE"])
def api_issue_location(issue_id):
    u, err = require_role(write=(request.method != "GET"))
    if err: return err
    db = get_db()
    
    if request.method == "GET":
        row = db.execute("SELECT * FROM issue_locations WHERE issue_id = ? ORDER BY created_at DESC LIMIT 1", (issue_id,)).fetchone()
        return jsonify({"ok": True, "location": dict(row) if row else None})
    
    if request.method == "POST":
        data = request.get_json(force=True, silent=True) or {}
        lat = data.get("latitude")
        lng = data.get("longitude")
        address = data.get("address", "")
        
        if lat is None or lng is None:
            return jsonify({"ok": False, "error": "missing_coordinates"}), 400
        
        # Delete old location
        db.execute("DELETE FROM issue_locations WHERE issue_id = ?", (issue_id,))
        
        # Insert new
        db.execute("""
            INSERT INTO issue_locations (issue_id, latitude, longitude, address)
            VALUES (?, ?, ?, ?)
        """, (issue_id, float(lat), float(lng), address))
        db.commit()
        return jsonify({"ok": True})
    
    if request.method == "DELETE":
        db.execute("DELETE FROM issue_locations WHERE issue_id = ?", (issue_id,))
        db.commit()
        return jsonify({"ok": True})

@app.route("/gd/api/tasks", methods=["GET", "POST", "PATCH", "PUT", "DELETE"])
def gd_api_tasks():
    return api_tasks()

@app.route("/gd/api/employees", methods=["GET", "POST", "PATCH", "DELETE"])
def gd_api_employees():
    return api_employees()

@app.route("/gd/api/timesheets", methods=["GET", "POST", "PATCH", "DELETE"])
def gd_api_timesheets():
    return api_timesheets()

# Helper function to get hourly rate (hierarchy: job override > employee rate > global default)
def _get_hourly_rate(db, job_id=None, employee_id=None, user_id=None):
    """Get hourly rate with hierarchy: job.hourly_rate > employee.hourly_rate > global default (200)"""
    DEFAULT_HOURLY_RATE = 200.0  # Global default
    
    # 1. Try job override
    if job_id:
        try:
            cols = [r[1] for r in db.execute("PRAGMA table_info(jobs)").fetchall()]
            if 'hourly_rate' in cols:
                job_rate = db.execute("SELECT hourly_rate FROM jobs WHERE id = ?", (job_id,)).fetchone()
                if job_rate and job_rate.get('hourly_rate') is not None:
                    rate = float(job_rate['hourly_rate'])
                    if rate > 0:
                        return rate
        except:
            pass
    
    # 2. Try employee hourly_rate
    if employee_id:
        try:
            emp_rate = db.execute("SELECT hourly_rate FROM employees WHERE id = ?", (employee_id,)).fetchone()
            if emp_rate and emp_rate.get('hourly_rate') is not None:
                rate = float(emp_rate['hourly_rate'])
                if rate > 0:
                    return rate
        except:
            pass
    
    # 3. Try employee via user_id
    if user_id:
        try:
            emp = db.execute("SELECT hourly_rate FROM employees WHERE user_id = ?", (user_id,)).fetchone()
            if emp and emp.get('hourly_rate') is not None:
                rate = float(emp['hourly_rate'])
                if rate > 0:
                    return rate
        except:
            pass
    
    # 4. Return global default
    return DEFAULT_HOURLY_RATE

# Helper function to process material usage (deduct/return)
def _process_material_usage(db, materials, job_id, action="deduct"):
    """
    Process material usage: deduct from or return to warehouse
    materials: list of {item_id, qty, unit, name}
    action: "deduct" or "return"
    """
    if not materials or not isinstance(materials, list):
        return
    
    try:
        for material in materials:
            item_id = material.get("item_id")
            qty = float(material.get("qty", 0))
            unit = material.get("unit", "ks")
            
            if not item_id or qty <= 0:
                continue
            
            # Check if warehouse_items table exists
            try:
                cols = [r[1] for r in db.execute("PRAGMA table_info(warehouse_items)").fetchall()]
                if not cols:
                    print("[MATERIAL] warehouse_items table does not exist")
                    return
            except:
                print("[MATERIAL] Error checking warehouse_items table")
                return
            
            # Get current quantity
            item = db.execute("SELECT quantity FROM warehouse_items WHERE id = ?", (item_id,)).fetchone()
            if not item:
                print(f"[MATERIAL] Item {item_id} not found in warehouse")
                continue
            
            current_qty = float(item["quantity"] or 0)
            
            if action == "deduct":
                new_qty = max(0, current_qty - qty)
            elif action == "return":
                new_qty = current_qty + qty
            else:
                continue
            
            # Update warehouse quantity
            db.execute("UPDATE warehouse_items SET quantity = ?, updated_at = datetime('now') WHERE id = ?", 
                      (new_qty, item_id))
            
            # Create audit log (if warehouse_movements table exists)
            try:
                db.execute("""
                    INSERT INTO warehouse_movements (item_id, job_id, movement_type, quantity, unit, created_at)
                    VALUES (?, ?, ?, ?, ?, datetime('now'))
                """, (item_id, job_id, action, qty, unit))
            except:
                pass  # Table might not exist
            
            print(f"[MATERIAL] {action.capitalize()}ed {qty} {unit} of item {item_id} (was {current_qty}, now {new_qty})")
    except Exception as e:
        print(f"[MATERIAL] Error processing materials: {e}")

# Helper function to check and create worklog-related notifications
def _check_worklog_notifications(db, job_id, user_id=None, employee_id=None):
    """Check worklog conditions and create notifications if needed (with dedup)"""
    try:
        from datetime import datetime, timedelta
        
        # Get job info
        job = db.execute("SELECT * FROM jobs WHERE id = ?", (job_id,)).fetchone()
        if not job:
            return
        job = dict(job)
        
        # Get user_id from employee_id if needed
        if not user_id and employee_id:
            user_id = _employee_user_id(db, employee_id)
        
        # Get job assignees (for notifications)
        assignees = db.execute("SELECT employee_id FROM job_assignments WHERE job_id = ?", (job_id,)).fetchall()
        assignee_ids = [a['employee_id'] for a in assignees]
        
        # Check for existing notifications in last 24h (dedup)
        existing = db.execute(
            "SELECT entity_type, entity_id, kind FROM notifications WHERE entity_type='job' AND entity_id=? AND created_at > datetime('now', '-24 hours')",
            (job_id,)
        ).fetchall()
        existing_keys = set((r['entity_type'], r['entity_id'], r['kind']) for r in existing)
        
        today = datetime.now().date()
        
        # 1. Check for missing material (if material_used in worklogs)
        if ('job', job_id, 'missing_material') not in existing_keys:
            try:
                worklogs_with_material = db.execute("""
                    SELECT material_used FROM timesheets
                    WHERE job_id = ? AND material_used IS NOT NULL AND material_used != ''
                    ORDER BY date DESC LIMIT 10
                """, (job_id,)).fetchall()
                
                missing_items = []
                for wl in worklogs_with_material:
                    try:
                        import json
                        materials = json.loads(wl['material_used']) if isinstance(wl['material_used'], str) else wl['material_used']
                        if isinstance(materials, list):
                            for mat in materials:
                                item_id = mat.get('item_id')
                                if item_id:
                                    wi = db.execute("SELECT name, quantity FROM warehouse_items WHERE id=?", (item_id,)).fetchone()
                                    if wi and (wi['quantity'] or 0) < (mat.get('qty', 0)):
                                        missing_items.append(wi['name'])
                    except:
                        pass
                
                if missing_items:
                    for eid in assignee_ids:
                        create_notification(
                            employee_id=eid,
                            kind="warning",
                            title="Chybí materiál",
                            body=f"Pro zakázku {job.get('title') or job.get('name') or '#' + str(job_id)} chybí: {', '.join(set(missing_items[:3]))}",
                            entity_type="job",
                            entity_id=job_id
                        )
            except Exception as e:
                print(f"[NOTIF] Error checking missing material: {e}")
        
        # 2. Check for repeated overtime (3+ days in last 7 days)
        if ('job', job_id, 'repeated_overtime') not in existing_keys:
            try:
                seven_days_ago = (today - timedelta(days=7)).strftime('%Y-%m-%d')
                overtime_days = db.execute("""
                    SELECT date, SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) as total_minutes
                    FROM timesheets
                    WHERE job_id = ? AND date >= ?
                    GROUP BY date
                    HAVING total_minutes > 480
                """, (job_id, seven_days_ago)).fetchall()
                
                if len(overtime_days) >= 3:
                    for eid in assignee_ids:
                        create_notification(
                            employee_id=eid,
                            kind="warning",
                            title="Opakovaný overtime",
                            body=f"Zakázka {job.get('title') or job.get('name') or '#' + str(job_id)} má {len(overtime_days)} dní s overtime za posledních 7 dní",
                            entity_type="job",
                            entity_id=job_id
                        )
            except Exception as e:
                print(f"[NOTIF] Error checking overtime: {e}")
        
        # 3. Check for increasing risk (budget > 80% or deadline approaching)
        if ('job', job_id, 'increasing_risk') not in existing_keys:
            try:
                budget = job.get('budget') or job.get('estimated_value') or 0
                cost_spent = job.get('cost_spent') or job.get('labor_cost_total') or 0
                deadline = job.get('deadline') or job.get('date')
                
                risk_flags = []
                
                if budget > 0:
                    spent_pct = (cost_spent / budget * 100) if budget > 0 else 0
                    if spent_pct > 80:
                        risk_flags.append(f"Rozpočet vyčerpán na {spent_pct:.0f}%")
                
                if deadline:
                    try:
                        deadline_dt = datetime.strptime(deadline[:10], '%Y-%m-%d').date()
                        days_until = (deadline_dt - today).days
                        if days_until <= 3 and days_until >= 0:
                            risk_flags.append(f"Deadline za {days_until} dní")
                        elif days_until < 0:
                            risk_flags.append(f"Deadline překročen o {abs(days_until)} dní")
                    except:
                        pass
                
                if risk_flags:
                    for eid in assignee_ids:
                        create_notification(
                            employee_id=eid,
                            kind="warning",
                            title="Rostoucí riziko zakázky",
                            body=f"{job.get('title') or job.get('name') or '#' + str(job_id)}: {', '.join(risk_flags)}",
                            entity_type="job",
                            entity_id=job_id
                        )
            except Exception as e:
                print(f"[NOTIF] Error checking risk: {e}")
        
        # 4. Check for approaching deadline (if not already notified)
        if ('job', job_id, 'deadline_approaching') not in existing_keys:
            try:
                deadline = job.get('deadline') or job.get('date')
                if deadline:
                    deadline_dt = datetime.strptime(deadline[:10], '%Y-%m-%d').date()
                    days_until = (deadline_dt - today).days
                    
                    if 1 <= days_until <= 3:
                        for eid in assignee_ids:
                            create_notification(
                                employee_id=eid,
                                kind="deadline",
                                title="Blížící se deadline",
                                body=f"Zakázka {job.get('title') or job.get('name') or '#' + str(job_id)} má deadline za {days_until} dní",
                                entity_type="job",
                                entity_id=job_id
                            )
            except Exception as e:
                print(f"[NOTIF] Error checking deadline: {e}")
                
    except Exception as e:
        print(f"[NOTIF] Error in _check_worklog_notifications: {e}")

# Helper function to recalculate job aggregations after worklog changes
def _recalculate_job_aggregations(db, job_id):
    """Recalculate job time_spent_minutes and labor_cost after worklog changes"""
    try:
        # Calculate total duration_minutes (prefer duration_minutes, fallback to hours*60)
        result = db.execute("""
            SELECT 
                COALESCE(SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))), 0) as total_minutes,
                COALESCE(SUM(COALESCE(labor_cost, 0)), 0) as total_cost
            FROM timesheets
            WHERE job_id = ?
        """, (job_id,)).fetchone()
        
        total_minutes = result['total_minutes'] or 0
        total_cost = result['total_cost'] or 0
        
        # Update job if columns exist
        cols = [r[1] for r in db.execute("PRAGMA table_info(jobs)").fetchall()]
        updates = []
        params = []
        
        if 'time_spent_minutes' in cols:
            updates.append("time_spent_minutes = ?")
            params.append(total_minutes)
        
        if 'labor_cost_total' in cols:
            updates.append("labor_cost_total = ?")
            params.append(total_cost)
        elif 'cost_spent' in cols:
            updates.append("cost_spent = ?")
            params.append(total_cost)
        
        # Calculate budget_remaining and margin if budget exists
        if 'budget' in cols:
            budget_result = db.execute("SELECT budget FROM jobs WHERE id = ?", (job_id,)).fetchone()
            if budget_result and budget_result.get('budget'):
                budget = float(budget_result['budget'])
                budget_remaining = budget - total_cost
                margin = ((budget - total_cost) / budget * 100) if budget > 0 else 0
                
                if 'budget_remaining' in cols:
                    updates.append("budget_remaining = ?")
                    params.append(budget_remaining)
                if 'margin' in cols:
                    updates.append("margin = ?")
                    params.append(margin)
        
        if updates:
            params.append(job_id)
            db.execute(f"UPDATE jobs SET {', '.join(updates)} WHERE id = ?", params)
            db.commit()
    except Exception as e:
        print(f"[WORKLOG] Error recalculating job aggregations: {e}")

# Work logs API (new extended API)
@app.route("/gd/api/worklogs", methods=["GET", "POST", "PUT", "DELETE"])
def gd_api_worklogs():
    """Extended work logs API with new fields"""
    u, err = require_role(write=(request.method != "GET"))
    if err:
        return err
    db = get_db()
    
    if request.method == "GET":
        # GET: List work logs with filters
        user_id = request.args.get("user_id", type=int)
        job_id = request.args.get("job_id", type=int)
        task_id = request.args.get("task_id", type=int)
        d_from = _normalize_date(request.args.get("from"))
        d_to = _normalize_date(request.args.get("to"))
        
        title_col = _job_title_col()
        q = f"""SELECT 
                    t.id, t.user_id, t.employee_id, t.job_id, t.task_id, t.date,
                    COALESCE(t.duration_minutes, CAST(t.hours * 60 AS INTEGER)) as duration_minutes,
                    t.hours, t.work_type, t.start_time, t.end_time,
                    COALESCE(t.location, t.place) as location, t.place,
                    COALESCE(t.note, t.activity) as note, t.activity,
                    t.material_used, t.weather_snapshot, t.performance_signal,
                    t.delay_reason, t.photo_url, t.ai_flags, t.labor_cost, t.training_id, t.created_at,
                    e.name AS employee_name,
                    j.{title_col} AS job_title, j.code AS job_code
               FROM timesheets t
               LEFT JOIN employees e ON e.id = t.employee_id
               LEFT JOIN jobs j ON j.id = t.job_id"""
        
        conds = []
        params = []
        
        if user_id:
            conds.append("(t.user_id = ? OR (t.user_id IS NULL AND t.employee_id IN (SELECT id FROM employees WHERE user_id = ?)))")
            params.extend([user_id, user_id])
        if job_id:
            conds.append("t.job_id = ?")
            params.append(job_id)
        if task_id:
            conds.append("t.task_id = ?")
            params.append(task_id)
        if d_from and d_to:
            conds.append("date(t.date) BETWEEN date(?) AND date(?)")
            params.extend([d_from, d_to])
        elif d_from:
            conds.append("date(t.date) >= date(?)")
            params.append(d_from)
        elif d_to:
            conds.append("date(t.date) <= date(?)")
            params.append(d_to)
        
        if conds:
            q += " WHERE " + " AND ".join(conds)
        q += " ORDER BY t.date DESC, t.id DESC"
        
        rows = db.execute(q, params).fetchall()
        return jsonify({"ok": True, "worklogs": [dict(r) for r in rows]})
    
    if request.method == "POST":
        # POST: Create new work log
        try:
            data = request.get_json(force=True, silent=True) or {}
            
            # Required fields
            job_id = data.get("job_id")
            date = data.get("date")
            duration_minutes = data.get("duration_minutes")
            work_type = data.get("work_type", "manual")
            
            # Validation
            if not job_id:
                return jsonify({"ok": False, "error": "missing_job_id", "message": "job_id is required"}), 400
            if not date:
                return jsonify({"ok": False, "error": "missing_date", "message": "date is required"}), 400
            if duration_minutes is None:
                return jsonify({"ok": False, "error": "missing_duration", "message": "duration_minutes is required"}), 400
            if duration_minutes <= 0:
                return jsonify({"ok": False, "error": "invalid_duration", "message": "duration_minutes must be greater than 0"}), 400
            
            # Validate job exists
            job = db.execute("SELECT id FROM jobs WHERE id = ?", (job_id,)).fetchone()
            if not job:
                return jsonify({"ok": False, "error": "job_not_found", "message": f"Job {job_id} does not exist"}), 400
            
            # Validate work_type
            valid_work_types = ["manual", "machine", "planning", "supervision", "transport", "training", "other"]
            if work_type not in valid_work_types:
                return jsonify({"ok": False, "error": "invalid_work_type", "message": f"work_type must be one of: {', '.join(valid_work_types)}"}), 400
            
            # Training ID - pouze pokud je work_type = "training"
            training_id = None
            if work_type == "training":
                training_id = data.get("training_id")
                if training_id:
                    # Validate training exists
                    training = db.execute("SELECT id FROM trainings WHERE id = ?", (training_id,)).fetchone()
                    if not training:
                        return jsonify({"ok": False, "error": "training_not_found", "message": f"Training {training_id} does not exist"}), 400
            
            # Get user_id from current user or employee_id
            user_id = u.get("id") if u else None
            employee_id = data.get("employee_id")
            if not employee_id and user_id:
                # Try to find employee_id from user_id
                emp = db.execute("SELECT id FROM employees WHERE user_id = ?", (user_id,)).fetchone()
                if emp:
                    employee_id = emp["id"]
            
            if not employee_id:
                return jsonify({"ok": False, "error": "missing_employee_id", "message": "employee_id is required"}), 400
            
            # Optional fields
            start_time = data.get("start_time")
            end_time = data.get("end_time")
            location = data.get("location")
            task_id = data.get("task_id")
            material_used = json.dumps(data.get("material_used")) if data.get("material_used") else None
            weather_snapshot = json.dumps(data.get("weather_snapshot")) if data.get("weather_snapshot") else None
            performance_signal = data.get("performance_signal", "normal")
            delay_reason = data.get("delay_reason")
            photo_url = data.get("photo_url")
            note = data.get("note")
            ai_flags = json.dumps(data.get("ai_flags")) if data.get("ai_flags") else None
            
            # Validate performance_signal
            valid_signals = ["fast", "normal", "slow", "blocked"]
            if performance_signal not in valid_signals:
                performance_signal = "normal"
            
            # Calculate hours for backwards compatibility
            hours = duration_minutes / 60.0
            
            # Calculate labor_cost using hourly rate hierarchy
            labor_cost = None
            try:
                hourly_rate = _get_hourly_rate(db, job_id=job_id, employee_id=employee_id, user_id=u.get("id"))
                labor_cost = (duration_minutes / 60.0) * float(hourly_rate)
            except Exception as e:
                print(f"[WORKLOG] Error calculating labor_cost: {e}")
                pass
            
            # Insert
            worklog_id = db.execute("""
                INSERT INTO timesheets(
                    user_id, employee_id, job_id, task_id, date,
                    duration_minutes, hours, work_type,
                    start_time, end_time, location, place,
                    note, activity, material_used, weather_snapshot,
                    performance_signal, delay_reason, photo_url, ai_flags,
                    labor_cost, training_id, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
            """, (
                user_id, employee_id, job_id, task_id, _normalize_date(date),
                duration_minutes, hours, work_type,
                start_time, end_time, location, location,
                note, note, material_used, weather_snapshot,
                performance_signal, delay_reason, photo_url, ai_flags,
                labor_cost, training_id
            )).lastrowid
            db.commit()
            
            # Pokud je to školení, automaticky přidej zaměstnance jako účastníka školení
            if training_id and employee_id:
                try:
                    # Zkontroluj, jestli už není účastníkem
                    existing_attendee = db.execute(
                        "SELECT id FROM training_attendees WHERE training_id = ? AND employee_id = ?",
                        (training_id, employee_id)
                    ).fetchone()
                    
                    if not existing_attendee:
                        # Přidej jako účastníka se statusem "registered"
                        db.execute("""
                            INSERT INTO training_attendees (training_id, employee_id, status, attendance_confirmed)
                            VALUES (?, ?, 'registered', 0)
                        """, (training_id, employee_id))
                        db.commit()
                        print(f"[WORKLOG] Automaticky přidán zaměstnanec {employee_id} jako účastník školení {training_id}")
                except Exception as e:
                    print(f"[WORKLOG] Error adding training attendee: {e}")
                    # Necháme pokračovat - není kritická chyba
            
            # Process material_used: deduct from warehouse
            if material_used:
                try:
                    materials = json.loads(material_used) if isinstance(material_used, str) else material_used
                    if isinstance(materials, list):
                        _process_material_usage(db, materials, job_id, "deduct")
                except Exception as e:
                    print(f"[WORKLOG] Error processing materials: {e}")
            
            # Recalculate job aggregations
            _recalculate_job_aggregations(db, job_id)
            
            # Check for worklog-related notifications
            _check_worklog_notifications(db, job_id, user_id=user_id, employee_id=employee_id)
            
            return jsonify({"ok": True, "message": "Work log created successfully"})
        except Exception as e:
            db.rollback()
            print(f"✗ Error creating work log: {e}")
            return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500
    
    if request.method == "PUT":
        # PUT: Update work log
        try:
            data = request.get_json(force=True, silent=True) or {}
            worklog_id = data.get("id")
            
            if not worklog_id:
                return jsonify({"ok": False, "error": "missing_id", "message": "id is required"}), 400
            
            # Check if work log exists
            existing = db.execute("SELECT job_id FROM timesheets WHERE id = ?", (worklog_id,)).fetchone()
            if not existing:
                return jsonify({"ok": False, "error": "not_found", "message": f"Work log {worklog_id} not found"}), 404
            
            old_job_id = existing["job_id"]
            
            # Build update query
            allowed_fields = {
                "job_id": int,
                "task_id": lambda x: int(x) if x else None,
                "date": _normalize_date,
                "duration_minutes": int,
                "work_type": str,
                "start_time": str,
                "end_time": str,
                "location": str,
                "task_id": lambda x: int(x) if x else None,
                "performance_signal": str,
                "delay_reason": str,
                "photo_url": str,
                "note": str,
                "training_id": lambda x: int(x) if x else None,
            }
            
            sets = []
            vals = []
            
            for field, converter in allowed_fields.items():
                if field in data:
                    value = converter(data[field])
                    if field == "duration_minutes" and value is not None:
                        if value <= 0:
                            return jsonify({"ok": False, "error": "invalid_duration", "message": "duration_minutes must be greater than 0"}), 400
                        # Update hours for backwards compatibility
                        sets.append("hours = ?")
                        vals.append(value / 60.0)
                    sets.append(f"{field} = ?")
                    vals.append(value)
            
            # Handle JSON fields
            if "material_used" in data:
                # Get old material_used to return it
                old_material = db.execute("SELECT material_used FROM timesheets WHERE id = ?", (worklog_id,)).fetchone()
                if old_material and old_material.get("material_used"):
                    try:
                        old_materials = json.loads(old_material["material_used"])
                        if isinstance(old_materials, list):
                            _process_material_usage(db, old_materials, old_job_id, "return")
                    except:
                        pass
                
                # Process new material_used
                new_material_used = json.dumps(data["material_used"]) if data["material_used"] else None
                sets.append("material_used = ?")
                vals.append(new_material_used)
                
                # Deduct new materials
                if data.get("material_used"):
                    try:
                        materials = data["material_used"] if isinstance(data["material_used"], list) else json.loads(data["material_used"])
                        new_job_id = data.get("job_id", old_job_id)
                        if isinstance(materials, list):
                            _process_material_usage(db, materials, new_job_id, "deduct")
                    except:
                        pass
            if "weather_snapshot" in data:
                sets.append("weather_snapshot = ?")
                vals.append(json.dumps(data["weather_snapshot"]) if data["weather_snapshot"] else None)
            if "ai_flags" in data:
                sets.append("ai_flags = ?")
                vals.append(json.dumps(data["ai_flags"]) if data["ai_flags"] else None)
            
            # Validate work_type if provided
            if "work_type" in data:
                valid_work_types = ["manual", "machine", "planning", "supervision", "transport", "training", "other"]
                if data["work_type"] not in valid_work_types:
                    return jsonify({"ok": False, "error": "invalid_work_type", "message": f"work_type must be one of: {', '.join(valid_work_types)}"}), 400
            
            # Training ID - pouze pokud je work_type = "training"
            if "training_id" in data or "work_type" in data:
                work_type = data.get("work_type")
                if work_type == "training":
                    training_id = data.get("training_id")
                    if training_id:
                        # Validate training exists
                        training = db.execute("SELECT id FROM trainings WHERE id = ?", (training_id,)).fetchone()
                        if not training:
                            return jsonify({"ok": False, "error": "training_not_found", "message": f"Training {training_id} does not exist"}), 400
                        sets.append("training_id = ?")
                        vals.append(training_id)
                    else:
                        sets.append("training_id = ?")
                        vals.append(None)
                elif work_type and work_type != "training":
                    # Pokud se změnil work_type na něco jiného než training, vymaž training_id
                    sets.append("training_id = ?")
                    vals.append(None)
            
            # Validate performance_signal if provided
            if "performance_signal" in data:
                valid_signals = ["fast", "normal", "slow", "blocked"]
                if data["performance_signal"] not in valid_signals:
                    return jsonify({"ok": False, "error": "invalid_performance_signal", "message": f"performance_signal must be one of: {', '.join(valid_signals)}"}), 400
            
            if not sets:
                return jsonify({"ok": False, "error": "no_fields", "message": "No fields to update"}), 400
            
            # Recalculate labor_cost if duration_minutes changed
            if "duration_minutes" in data:
                try:
                    # Get hourly_rate
                    emp_id_result = db.execute("SELECT employee_id FROM timesheets WHERE id = ?", (worklog_id,)).fetchone()
                    if emp_id_result:
                        emp_rate = db.execute("SELECT hourly_rate FROM employees WHERE id = ?", (emp_id_result["employee_id"],)).fetchone()
                        hourly_rate = emp_rate["hourly_rate"] if emp_rate and emp_rate.get("hourly_rate") else 200
                        labor_cost = (data["duration_minutes"] / 60.0) * float(hourly_rate)
                        sets.append("labor_cost = ?")
                        vals.append(labor_cost)
                except:
                    pass
            
            vals.append(worklog_id)
            db.execute(f"UPDATE timesheets SET {', '.join(sets)} WHERE id = ?", vals)
            db.commit()
            
            # Pokud se změnil training_id, aktualizuj účastníky školení
            if "training_id" in data or "work_type" in data:
                try:
                    # Získej aktuální data worklogu po update
                    updated_worklog = db.execute(
                        "SELECT training_id, employee_id, work_type FROM timesheets WHERE id = ?",
                        (worklog_id,)
                    ).fetchone()
                    
                    if updated_worklog:
                        new_training_id = updated_worklog.get("training_id")
                        employee_id = updated_worklog.get("employee_id")
                        work_type = updated_worklog.get("work_type")
                        
                        # Pokud je work_type = "training" a má training_id, přidej účastníka
                        if work_type == "training" and new_training_id and employee_id:
                            existing_attendee = db.execute(
                                "SELECT id FROM training_attendees WHERE training_id = ? AND employee_id = ?",
                                (new_training_id, employee_id)
                            ).fetchone()
                            
                            if not existing_attendee:
                                db.execute("""
                                    INSERT INTO training_attendees (training_id, employee_id, status, attendance_confirmed)
                                    VALUES (?, ?, 'registered', 0)
                                """, (new_training_id, employee_id))
                                db.commit()
                                print(f"[WORKLOG] Automaticky přidán zaměstnanec {employee_id} jako účastník školení {new_training_id}")
                except Exception as e:
                    print(f"[WORKLOG] Error updating training attendee: {e}")
                    # Necháme pokračovat
            
            # Recalculate job aggregations for old and new job
            new_job_id = data.get("job_id", old_job_id)
            _recalculate_job_aggregations(db, old_job_id)
            _check_worklog_notifications(db, old_job_id, user_id=user_id, employee_id=employee_id)
            if new_job_id != old_job_id:
                _recalculate_job_aggregations(db, new_job_id)
                _check_worklog_notifications(db, new_job_id, user_id=user_id, employee_id=employee_id)
            
            return jsonify({"ok": True, "message": "Work log updated successfully"})
        except Exception as e:
            db.rollback()
            print(f"✗ Error updating work log: {e}")
            return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500
    
    # DELETE
    if request.method == "DELETE":
        try:
            worklog_id = request.args.get("id", type=int)
            if not worklog_id:
                return jsonify({"ok": False, "error": "missing_id", "message": "id parameter is required"}), 400
            
            # Get job_id and material_used before deletion for recalculation
            job_result = db.execute("SELECT job_id, material_used FROM timesheets WHERE id = ?", (worklog_id,)).fetchone()
            if not job_result:
                return jsonify({"ok": False, "error": "not_found", "message": f"Work log {worklog_id} not found"}), 404
            
            job_id = job_result["job_id"]
            material_used = job_result.get("material_used")
            
            # Return materials to warehouse
            if material_used:
                try:
                    materials = json.loads(material_used) if isinstance(material_used, str) else material_used
                    if isinstance(materials, list):
                        _process_material_usage(db, materials, job_id, "return")
                except Exception as e:
                    print(f"[WORKLOG] Error returning materials: {e}")
            
            db.execute("DELETE FROM timesheets WHERE id = ?", (worklog_id,))
            db.commit()
            
            # Recalculate job aggregations
            _recalculate_job_aggregations(db, job_id)
            
            # Check for worklog-related notifications
            _check_worklog_notifications(db, job_id, user_id=user_id, employee_id=employee_id)
            
            return jsonify({"ok": True, "message": "Work log deleted successfully"})
        except Exception as e:
            db.rollback()
            print(f"✗ Error deleting work log: {e}")
            return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500

# Warehouse items API
@app.route("/gd/api/warehouse/items", methods=["GET"])
def gd_api_warehouse_items():
    """Get warehouse items for autocomplete"""
    u, err = require_role(write=False)
    if err:
        return err
    db = get_db()
    
    try:
        search = request.args.get("search", "").strip()
        limit = request.args.get("limit", type=int) or 50
        
        q = "SELECT id, name, sku, category, quantity, unit, unit_price, location FROM warehouse_items WHERE quantity > 0"
        params = []
        
        if search:
            q += " AND (name LIKE ? OR sku LIKE ?)"
            search_term = f"%{search}%"
            params.extend([search_term, search_term])
        
        q += " ORDER BY name ASC LIMIT ?"
        params.append(limit)
        
        rows = db.execute(q, params).fetchall()
        items = []
        for row in rows:
            items.append({
                "id": row["id"],
                "name": row["name"],
                "sku": row.get("sku"),
                "category": row.get("category"),
                "quantity": float(row["quantity"] or 0),
                "unit": row.get("unit", "ks"),
                "unit_price": float(row.get("unit_price") or 0),
                "location": row.get("location")
            })
        
        return jsonify({"ok": True, "items": items})
    except Exception as e:
        print(f"✗ Error getting warehouse items: {e}")
        return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500

# ----------------- Training Service Functions -----------------
def calculate_training_cost(training_data, db=None):
    """Spočítej celkové náklady školení včetně mezd"""
    if db is None:
        db = get_db()
    
    # Přímé náklady (vždy)
    direct_costs = (
        (training_data.get('cost_training') or 0) +
        (training_data.get('cost_travel') or 0) +
        (training_data.get('cost_accommodation') or 0) +
        (training_data.get('cost_meals') or 0) +
        (training_data.get('cost_other') or 0)
    )
    
    # Typ kompenzace
    compensation_type = training_data.get('compensation_type', 'paid_workday')
    duration_hours = training_data.get('duration_hours')
    attendee_count = training_data.get('attendee_count', 0)
    
    # Mzdové náklady (pouze pokud paid_workday)
    wage_cost = 0
    if compensation_type == 'paid_workday' and duration_hours and attendee_count:
        # Průměrná hodinová mzda (nebo z konfigurace)
        hourly_wage = training_data.get('wage_cost_per_person') or 200  # Kč/hod default
        wage_cost = duration_hours * hourly_wage * attendee_count
    
    # Opportunity cost (ušlý zisk) - počítá se vždy když je pracovní doba
    opportunity_cost = 0
    if compensation_type in ['paid_workday', 'costs_only']:
        if duration_hours and attendee_count:
            # Ušlý zisk = co by ti lidi vydělali na zakázkách
            avg_billable_rate = 500  # Kč/hod - fakturační sazba
            opportunity_cost = duration_hours * avg_billable_rate * attendee_count
    
    return direct_costs, wage_cost, opportunity_cost

def update_employee_skills_from_training(training_id, db=None):
    """Aktualizuj skills zaměstnanců po dokončení školení"""
    if db is None:
        db = get_db()
    
    import json as json_lib
    
    # Získej školení
    training = db.execute(
        "SELECT skills_gained, skill_level_increase, duration_hours, date_start FROM trainings WHERE id = ?",
        (training_id,)
    ).fetchone()
    
    if not training or not training[0]:  # skills_gained
        return
    
    try:
        skills_gained = json_lib.loads(training[0]) if isinstance(training[0], str) else training[0]
    except:
        skills_gained = []
    
    if not skills_gained:
        return
    
    skill_level_increase = training[1] or 1
    duration_hours = training[2] or 0
    date_start = training[3]
    
    # Získej všechny dokončené účastníky
    attendees = db.execute(
        "SELECT employee_id FROM training_attendees WHERE training_id = ? AND status = 'completed'",
        (training_id,)
    ).fetchall()
    
    for attendee_row in attendees:
        employee_id = attendee_row[0]
        
        # Získej aktuální skills zaměstnance
        emp = db.execute(
            "SELECT skills, training_hours_total FROM employees WHERE id = ?",
            (employee_id,)
        ).fetchone()
        
        if not emp:
            continue
        
        try:
            skills = json_lib.loads(emp[0]) if emp[0] and isinstance(emp[0], str) else (emp[0] or {})
        except:
            skills = {}
        
        # Přidej/zvyš skills
        for skill in skills_gained:
            current_level = skills.get(skill, 0)
            new_level = min(current_level + skill_level_increase, 5)  # Max level 5
            skills[skill] = new_level
        
        # Aktualizuj training_hours_total
        training_hours_total = (emp[1] or 0) + duration_hours
        
        # Přepočítej celkové skill score
        skill_score = calculate_employee_skill_score(skills, training_hours_total)
        
        # Ulož zpět
        db.execute(
            "UPDATE employees SET skills = ?, training_hours_total = ?, skill_score = ?, last_training_date = ? WHERE id = ?",
            (json_lib.dumps(skills), training_hours_total, skill_score, date_start, employee_id)
        )
    
    db.commit()

def calculate_employee_skill_score(skills, training_hours_total=0):
    """Spočítej celkové skóre kompetencí zaměstnance (0-100)"""
    if not skills or len(skills) == 0:
        return 50.0  # Baseline
    
    # Průměr všech skills (každý 0-5) → převedeno na 0-100
    avg_skill = sum(skills.values()) / len(skills)
    
    # Bonus za počet hodin školení
    training_bonus = min((training_hours_total / 100) * 10, 10)  # Max +10 bodů
    
    # Výsledné skóre
    score = (avg_skill / 5 * 80) + 10 + training_bonus  # 10-100 range
    
    return round(min(score, 100), 1)

@app.route("/gd/api/worklogs/summary", methods=["GET"])
def gd_api_worklogs_summary():
    """Get summary statistics for worklogs"""
    u, err = require_role(write=False)
    if err:
        return err
    db = get_db()
    
    try:
        user_id = request.args.get("user_id", type=int)
        d_from = _normalize_date(request.args.get("from"))
        d_to = _normalize_date(request.args.get("to"))
        
        # Build WHERE clause
        conds = []
        params = []
        
        if user_id:
            conds.append("(user_id = ? OR (user_id IS NULL AND employee_id IN (SELECT id FROM employees WHERE user_id = ?)))")
            params.extend([user_id, user_id])
        if d_from and d_to:
            conds.append("date(date) BETWEEN date(?) AND date(?)")
            params.extend([d_from, d_to])
        elif d_from:
            conds.append("date(date) >= date(?)")
            params.append(d_from)
        elif d_to:
            conds.append("date(date) <= date(?)")
            params.append(d_to)
        
        where_clause = " WHERE " + " AND ".join(conds) if conds else ""
        
        # Calculate total_minutes
        total_result = db.execute(f"""
            SELECT COALESCE(SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))), 0) as total_minutes
            FROM timesheets
            {where_clause}
        """, params).fetchone()
        total_minutes = total_result['total_minutes'] or 0
        
        # Calculate avg_per_day
        days_result = db.execute(f"""
            SELECT COUNT(DISTINCT date) as days
            FROM timesheets
            {where_clause}
        """, params).fetchone()
        days_count = days_result['days'] or 1
        avg_per_day = total_minutes / days_count if days_count > 0 else 0
        
        # Calculate overtime_minutes (assuming 8h/day = 480 minutes)
        overtime_result = db.execute(f"""
            SELECT 
                date,
                SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) as day_minutes
            FROM timesheets
            {where_clause}
            GROUP BY date
            HAVING day_minutes > 480
        """, params).fetchall()
        overtime_minutes = sum(max(0, row['day_minutes'] - 480) for row in overtime_result)
        
        # Top jobs
        top_jobs_result = db.execute(f"""
            SELECT 
                job_id,
                SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) as total_minutes
            FROM timesheets
            {where_clause}
            GROUP BY job_id
            ORDER BY total_minutes DESC
            LIMIT 5
        """, params).fetchall()
        top_jobs = [{"job_id": row['job_id'], "total_minutes": row['total_minutes']} for row in top_jobs_result]
        
        # Efficiency (based on performance_signal)
        efficiency_result = db.execute(f"""
            SELECT 
                performance_signal,
                COUNT(*) as count
            FROM timesheets
            {where_clause} AND performance_signal IS NOT NULL
            GROUP BY performance_signal
        """, params).fetchall()
        
        total_signals = sum(row['count'] for row in efficiency_result)
        fast_count = next((row['count'] for row in efficiency_result if row['performance_signal'] == 'fast'), 0)
        slow_count = next((row['count'] for row in efficiency_result if row['performance_signal'] == 'slow'), 0)
        blocked_count = next((row['count'] for row in efficiency_result if row['performance_signal'] == 'blocked'), 0)
        
        efficiency = (fast_count / total_signals * 100) if total_signals > 0 else 0
        
        # Anomalies count (overtime + blocked + delay)
        anomalies_result = db.execute(f"""
            SELECT COUNT(*) as count
            FROM (
                SELECT date, SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) as day_minutes
                FROM timesheets
                {where_clause}
                GROUP BY date
                HAVING day_minutes > 480
            ) overtime_days
        """, params).fetchone()
        overtime_days = anomalies_result['count'] or 0
        
        blocked_anomalies = db.execute(f"""
            SELECT COUNT(*) as count
            FROM timesheets
            {where_clause} AND performance_signal = 'blocked'
        """, params).fetchone()
        blocked_count = blocked_anomalies['count'] or 0
        
        delay_anomalies = db.execute(f"""
            SELECT COUNT(*) as count
            FROM timesheets
            {where_clause} AND delay_reason IS NOT NULL AND delay_reason != ''
        """, params).fetchone()
        delay_count = delay_anomalies['count'] or 0
        
        anomalies_count = overtime_days + blocked_count + delay_count
        
        return jsonify({
            "ok": True,
            "summary": {
                "total_minutes": int(total_minutes),
                "avg_per_day": round(avg_per_day, 1),
                "overtime_minutes": int(overtime_minutes),
                "top_jobs": top_jobs,
                "efficiency": round(efficiency, 1),
                "anomalies_count": anomalies_count
            }
        })
    except Exception as e:
        print(f"✗ Error getting worklogs summary: {e}")
        return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500

@app.route("/gd/api/worklogs/heatmap", methods=["GET"])
def gd_api_worklogs_heatmap():
    """Get heatmap data for worklogs"""
    u, err = require_role(write=False)
    if err:
        return err
    db = get_db()
    
    try:
        user_id = request.args.get("user_id", type=int)
        d_from = _normalize_date(request.args.get("from"))
        d_to = _normalize_date(request.args.get("to"))
        
        # Default to last 30 days if no range specified
        if not d_from or not d_to:
            from datetime import datetime, timedelta
            today = datetime.now().date()
            d_to = today.strftime('%Y-%m-%d')
            d_from = (today - timedelta(days=30)).strftime('%Y-%m-%d')
        
        # Build WHERE clause with table prefixes
        conds = ["date(t.date) BETWEEN date(?) AND date(?)"]
        params = [d_from, d_to]
        
        if user_id:
            conds.append("(t.user_id = ? OR (t.user_id IS NULL AND t.employee_id IN (SELECT id FROM employees WHERE user_id = ?)))")
            params.extend([user_id, user_id])
        
        where_clause = " WHERE " + " AND ".join(conds)
        
        # Get daily aggregates
        daily_result = db.execute(f"""
            SELECT 
                t.date,
                SUM(COALESCE(t.duration_minutes, CAST(t.hours * 60 AS INTEGER))) as total_minutes,
                COUNT(*) as entries,
                SUM(CASE WHEN t.performance_signal = 'blocked' THEN 1 ELSE 0 END) as blocked_count,
                SUM(CASE WHEN t.delay_reason IS NOT NULL AND t.delay_reason != '' THEN 1 ELSE 0 END) as delay_count
            FROM timesheets t
            {where_clause}
            GROUP BY t.date
            ORDER BY t.date ASC
        """, params).fetchall()
        
        heatmap_data = []
        for row in daily_result:
            total_minutes = row['total_minutes'] or 0
            # Calculate load_level: 0: <4h, 1: 4-6h, 2: 6-8h, 3: >8h
            if total_minutes < 240:  # < 4h
                load_level = 0
            elif total_minutes < 360:  # 4-6h
                load_level = 1
            elif total_minutes <= 480:  # 6-8h
                load_level = 2
            else:  # > 8h
                load_level = 3
            
            # Build flags
            flags = []
            if total_minutes > 480:
                flags.append("overtime")
            if row['blocked_count'] > 0:
                flags.append("blocked")
            if row['delay_count'] > 0:
                flags.append("delay")
            
            heatmap_data.append({
                "date": row['date'],
                "total_minutes": int(total_minutes),
                "load_level": load_level,
                "flags": flags
            })
        
        # Get per-user daily aggregates (for timeline view)
        per_user_result = db.execute(f"""
            SELECT 
                t.date,
                COALESCE(t.user_id, (SELECT user_id FROM employees WHERE id = t.employee_id)) as user_id,
                e.name as user_name,
                SUM(COALESCE(t.duration_minutes, CAST(t.hours * 60 AS INTEGER))) as total_minutes,
                COUNT(*) as entries
            FROM timesheets t
            LEFT JOIN employees e ON e.id = t.employee_id
            {where_clause}
            GROUP BY t.date, user_id, e.name
            ORDER BY t.date ASC, user_id ASC
        """, params).fetchall()
        
        per_user_data = {}
        for row in per_user_result:
            date = row['date']
            user_id = row['user_id']
            if not date or not user_id:
                continue
            
            if date not in per_user_data:
                per_user_data[date] = []
            
            total_minutes = row['total_minutes'] or 0
            # Calculate load_level: 0: <4h, 1: 4-6h, 2: 6-8h, 3: >8h
            if total_minutes < 240:  # < 4h
                load_level = 0
            elif total_minutes < 360:  # 4-6h
                load_level = 1
            elif total_minutes <= 480:  # 6-8h
                load_level = 2
            else:  # > 8h
                load_level = 3
            
            per_user_data[date].append({
                "user_id": user_id,
                "user_name": row['user_name'] or f"User #{user_id}",
                "total_minutes": int(total_minutes),
                "load_level": load_level,
                "entries": row['entries'] or 0
            })
        
        return jsonify({
            "ok": True,
            "heatmap": heatmap_data,
            "per_user": per_user_data  # New: per-user data for timeline view
        })
    except Exception as e:
        print(f"✗ Error getting worklogs heatmap: {e}")
        return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500


@app.route("/gd/api/worklogs/ai-insights", methods=["GET"])
def gd_api_worklogs_ai_insights():
    """Get AI insights for worklogs - heuristics without LLM"""
    u, err = require_role(write=False)
    if err:
        return err
    db = get_db()
    
    try:
        from datetime import datetime, timedelta
        
        user_id = request.args.get("user_id", type=int)
        job_id = request.args.get("job_id", type=int)
        d_from = _normalize_date(request.args.get("from"))
        d_to = _normalize_date(request.args.get("to"))
        
        # Default to last 30 days if no range specified
        if not d_from or not d_to:
            today = datetime.now().date()
            d_to = today.strftime('%Y-%m-%d')
            d_from = (today - timedelta(days=30)).strftime('%Y-%m-%d')
        
        # Build WHERE clause
        conds = ["date(date) BETWEEN date(?) AND date(?)"]
        params = [d_from, d_to]
        
        if user_id:
            conds.append("(user_id = ? OR (user_id IS NULL AND employee_id IN (SELECT id FROM employees WHERE user_id = ?)))")
            params.extend([user_id, user_id])
        
        if job_id:
            conds.append("job_id = ?")
            params.append(job_id)
        
        where_clause = " WHERE " + " AND ".join(conds)
        
        # === 1. ANOMALY DAYS ===
        # anomaly_day: overtime + blocked signál + delay
        anomaly_days = []
        try:
            anomaly_rows = db.execute(f"""
                SELECT 
                    date,
                    SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) as total_minutes,
                    SUM(CASE WHEN performance_signal = 'blocked' THEN 1 ELSE 0 END) as blocked_count,
                    SUM(CASE WHEN delay_reason IS NOT NULL AND delay_reason != '' THEN 1 ELSE 0 END) as delay_count
                FROM timesheets
                {where_clause}
                GROUP BY date
                HAVING total_minutes > 480 OR blocked_count > 0 OR delay_count > 0
                ORDER BY date DESC
            """, params).fetchall()
            
            for row in anomaly_rows:
                total_minutes = row['total_minutes'] or 0
                blocked_count = row['blocked_count'] or 0
                delay_count = row['delay_count'] or 0
                
                flags = []
                if total_minutes > 480:
                    flags.append('overtime')
                if blocked_count > 0:
                    flags.append('blocked')
                if delay_count > 0:
                    flags.append('delay')
                
                anomaly_days.append({
                    "date": row['date'],
                    "total_hours": round(total_minutes / 60, 1),
                    "flags": flags,
                    "severity": "high" if total_minutes > 600 or (blocked_count > 0 and delay_count > 0) else "medium"
                })
        except Exception as e:
            print(f"[AI-INSIGHTS] Error calculating anomaly_days: {e}")
        
        # === 2. SLOWDOWN JOBS ===
        # slowdown_job: rostoucí čas bez progressu
        slowdown_jobs = []
        try:
            if job_id:
                # For specific job, check if time is increasing without progress
                job_rows = db.execute(f"""
                    SELECT 
                        date,
                        SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) as day_minutes
                    FROM timesheets
                    {where_clause}
                    GROUP BY date
                    ORDER BY date ASC
                """, params).fetchall()
                
                if len(job_rows) >= 3:
                    # Check if last 3 days show increasing time
                    last_3_days = [row['day_minutes'] or 0 for row in job_rows[-3:]]
                    if last_3_days[0] < last_3_days[1] < last_3_days[2]:
                        # Time is increasing - check if job progress is not increasing
                        job_info = db.execute("SELECT progress, completion_percent FROM jobs WHERE id = ?", (job_id,)).fetchone()
                        progress = job_info.get('progress') or job_info.get('completion_percent') or 0
                        
                        slowdown_jobs.append({
                            "job_id": job_id,
                            "trend": "increasing_time",
                            "last_3_days_hours": [round(m / 60, 1) for m in last_3_days],
                            "progress": progress,
                            "severity": "medium"
                        })
            else:
                # For all jobs, find jobs with increasing time
                job_trends = db.execute(f"""
                    SELECT 
                        job_id,
                        date,
                        SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) as day_minutes
                    FROM timesheets
                    {where_clause}
                    GROUP BY job_id, date
                    ORDER BY job_id, date ASC
                """, params).fetchall()
                
                # Group by job_id
                jobs_by_id = {}
                for row in job_trends:
                    jid = row['job_id']
                    if jid not in jobs_by_id:
                        jobs_by_id[jid] = []
                    jobs_by_id[jid].append({
                        "date": row['date'],
                        "minutes": row['day_minutes'] or 0
                    })
                
                # Check each job for slowdown pattern
                for jid, days in jobs_by_id.items():
                    if len(days) >= 3:
                        last_3 = days[-3:]
                        minutes_3 = [d['minutes'] for d in last_3]
                        if minutes_3[0] < minutes_3[1] < minutes_3[2]:
                            # Time increasing - check progress
                            job_info = db.execute("SELECT progress, completion_percent FROM jobs WHERE id = ?", (jid,)).fetchone()
                            if job_info:
                                progress = job_info.get('progress') or job_info.get('completion_percent') or 0
                                slowdown_jobs.append({
                                    "job_id": jid,
                                    "trend": "increasing_time",
                                    "last_3_days_hours": [round(m / 60, 1) for m in minutes_3],
                                    "progress": progress,
                                    "severity": "medium"
                                })
        except Exception as e:
            print(f"[AI-INSIGHTS] Error calculating slowdown_jobs: {e}")
        
        # === 3. WORKLOAD RISK ===
        # workload_risk: 7denní průměr > threshold
        workload_risk = {
            "level": "low",
            "avg_hours_7d": 0,
            "threshold": 8,
            "users_at_risk": []
        }
        try:
            seven_days_ago = (datetime.now().date() - timedelta(days=7)).strftime('%Y-%m-%d')
            today_str = datetime.now().date().strftime('%Y-%m-%d')
            
            workload_conds = ["date(date) BETWEEN date(?) AND date(?)"]
            workload_params = [seven_days_ago, today_str]
            
            if user_id:
                workload_conds.append("(user_id = ? OR (user_id IS NULL AND employee_id IN (SELECT id FROM employees WHERE user_id = ?)))")
                workload_params.extend([user_id, user_id])
            
            workload_where = " WHERE " + " AND ".join(workload_conds)
            
            user_workloads = db.execute(f"""
                SELECT 
                    COALESCE(user_id, (SELECT user_id FROM employees WHERE id = employee_id)) as user_id,
                    e.name as user_name,
                    SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) as total_minutes,
                    COUNT(DISTINCT date) as days_worked
                FROM timesheets t
                LEFT JOIN employees e ON e.id = t.employee_id
                {workload_where}
                GROUP BY user_id, e.name
            """, workload_params).fetchall()
            
            users_at_risk = []
            max_avg = 0
            
            for row in user_workloads:
                user_id_val = row['user_id']
                user_name = row['user_name'] or f"User #{user_id_val}"
                total_minutes = row['total_minutes'] or 0
                days_worked = row['days_worked'] or 1
                avg_hours = (total_minutes / 60) / days_worked
                
                if avg_hours > workload_risk['threshold']:
                    users_at_risk.append({
                        "user_id": user_id_val,
                        "user_name": user_name,
                        "avg_hours": round(avg_hours, 1),
                        "total_hours": round(total_minutes / 60, 1)
                    })
                
                max_avg = max(max_avg, avg_hours)
            
            workload_risk['avg_hours_7d'] = round(max_avg, 1)
            workload_risk['users_at_risk'] = users_at_risk
            
            if max_avg > 10:
                workload_risk['level'] = "critical"
            elif max_avg > 8:
                workload_risk['level'] = "high"
            elif max_avg > 6:
                workload_risk['level'] = "medium"
        except Exception as e:
            print(f"[AI-INSIGHTS] Error calculating workload_risk: {e}")
        
        # === 4. COST BURN RATE ===
        cost_burn_rate = {
            "daily_avg": 0,
            "weekly_total": 0,
            "trend": "stable"
        }
        try:
            cost_rows = db.execute(f"""
                SELECT 
                    date,
                    SUM(COALESCE(labor_cost, 0)) as daily_cost
                FROM timesheets
                {where_clause}
                GROUP BY date
                ORDER BY date DESC
                LIMIT 14
            """, params).fetchall()
            
            if cost_rows:
                daily_costs = [row['daily_cost'] or 0 for row in cost_rows]
                cost_burn_rate['daily_avg'] = round(sum(daily_costs) / len(daily_costs), 0)
                cost_burn_rate['weekly_total'] = round(sum(daily_costs[:7]), 0)
                
                if len(daily_costs) >= 7:
                    first_half = sum(daily_costs[7:]) / len(daily_costs[7:]) if len(daily_costs) > 7 else daily_costs[0]
                    second_half = sum(daily_costs[:7]) / 7
                    if second_half > first_half * 1.2:
                        cost_burn_rate['trend'] = "increasing"
                    elif second_half < first_half * 0.8:
                        cost_burn_rate['trend'] = "decreasing"
        except Exception as e:
            print(f"[AI-INSIGHTS] Error calculating cost_burn_rate: {e}")
        
        # === 5. PREDICTION HINT ===
        prediction_hint = None
        try:
            if anomaly_days:
                prediction_hint = f"Detekováno {len(anomaly_days)} anomálních dní s overtime nebo blokacemi"
            elif slowdown_jobs:
                prediction_hint = f"{len(slowdown_jobs)} zakázek vykazuje zpomalení"
            elif workload_risk['level'] != "low":
                prediction_hint = f"Vysoké zatížení týmu ({workload_risk['avg_hours_7d']}h/den průměr)"
            elif cost_burn_rate['trend'] == "increasing":
                prediction_hint = "Rostoucí náklady práce - zkontroluj efektivitu"
        except:
            pass
        
        return jsonify({
            "ok": True,
            "insights": {
                "workload_risk": workload_risk,
                "slowdown_jobs": slowdown_jobs,
                "anomaly_days": anomaly_days,
                "cost_burn_rate": cost_burn_rate,
                "prediction_hint": prediction_hint
            }
        })
    except Exception as e:
        print(f"✗ Error getting worklogs AI insights: {e}")
        return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500

# Helper function to recalculate job aggregations after worklog changes
def _recalculate_job_aggregations(job_id):
    """Recalculate job time_spent_minutes and cost_spent after worklog changes"""
    db = get_db()
    try:
        # Calculate total duration_minutes (prefer duration_minutes, fallback to hours*60)
        result = db.execute("""
            SELECT 
                COALESCE(SUM(CASE WHEN duration_minutes IS NOT NULL THEN duration_minutes ELSE CAST(hours * 60 AS INTEGER) END), 0) as total_minutes,
                COALESCE(SUM(labor_cost), 0) as total_cost
            FROM timesheets
            WHERE job_id = ?
        """, (job_id,)).fetchone()
        
        total_minutes = result['total_minutes'] or 0
        total_cost = result['total_cost'] or 0
        
        # Update job table if time_spent_minutes column exists
        cols = [r[1] for r in db.execute("PRAGMA table_info(jobs)").fetchall()]
        if 'time_spent_minutes' in cols:
            db.execute("UPDATE jobs SET time_spent_minutes = ? WHERE id = ?", (total_minutes, job_id))
        if 'cost_spent' in cols:
            db.execute("UPDATE jobs SET cost_spent = ? WHERE id = ?", (total_cost, job_id))
        
        db.commit()
    except Exception as e:
        print(f"[WORKLOG] Error recalculating job aggregations: {e}")
        db.rollback()

# ----------------- Trainings API -----------------
@app.route("/gd/api/trainings", methods=["GET"])
def gd_api_trainings_get():
    """Seznam školení s filtry"""
    u, err = require_role(write=False)
    if err:
        return err
    db = get_db()
    
    try:
        import json as json_lib
        
        # Filtry
        d_from = _normalize_date(request.args.get("from"))
        d_to = _normalize_date(request.args.get("to"))
        category = request.args.get("category")
        is_paid = request.args.get("is_paid")
        employee_id = request.args.get("employee_id", type=int)
        
        # Build query
        conds = []
        params = []
        
        if d_from:
            conds.append("date(t.date_start) >= date(?)")
            params.append(d_from)
        if d_to:
            conds.append("date(t.date_start) <= date(?)")
            params.append(d_to)
        if category:
            conds.append("t.category = ?")
            params.append(category)
        if is_paid is not None:
            conds.append("t.is_paid = ?")
            params.append(1 if is_paid == 'true' else 0)
        if employee_id:
            conds.append("EXISTS (SELECT 1 FROM training_attendees ta WHERE ta.training_id = t.id AND ta.employee_id = ?)")
            params.append(employee_id)
        
        where_clause = " WHERE " + " AND ".join(conds) if conds else ""
        
        # Get trainings
        trainings = db.execute(f"""
            SELECT t.*, 
                   COUNT(DISTINCT ta.id) as attendees_count,
                   COUNT(DISTINCT CASE WHEN ta.status = 'completed' THEN ta.id END) as completed_count
            FROM trainings t
            LEFT JOIN training_attendees ta ON ta.training_id = t.id
            {where_clause}
            GROUP BY t.id
            ORDER BY t.date_start DESC, t.id DESC
        """, params).fetchall()
        
        # Parse JSON fields
        result = []
        for t in trainings:
            training_dict = dict(t)
            # Parse JSON fields
            if training_dict.get('skills_gained'):
                try:
                    training_dict['skills_gained'] = json_lib.loads(training_dict['skills_gained'])
                except:
                    training_dict['skills_gained'] = []
            else:
                training_dict['skills_gained'] = []
            
            result.append(training_dict)
        
        return jsonify({"ok": True, "trainings": result, "total": len(result)})
    except Exception as e:
        print(f"✗ Error getting trainings: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500

@app.route("/gd/api/trainings", methods=["POST"])
def gd_api_trainings_post():
    """Vytvoření nového školení"""
    u, err = require_role(write=True)
    if err:
        return err
    db = get_db()
    
    try:
        import json as json_lib
        data = request.get_json(force=True, silent=True) or {}
        
        # Validate required fields - podporuj title i name
        training_name = data.get('title') or data.get('name')
        if not training_name or not data.get('date_start'):
            return jsonify({"ok": False, "error": "missing_required_fields", "message": "Chybí název nebo datum školení"}), 400
        
        # Calculate costs
        # Podporuj různé formáty účastníků
        participants_raw = data.get('participants') or data.get('attendees') or data.get('attendee_ids') or '[]'
        if isinstance(participants_raw, str):
            try:
                attendee_ids = json_lib.loads(participants_raw)
            except:
                attendee_ids = []
        elif isinstance(participants_raw, list):
            attendee_ids = participants_raw
        else:
            attendee_ids = []
        attendee_count = len(attendee_ids)
        compensation_type = data.get('compensation_type', 'paid_workday')
        
        cost_data = {
            'cost_training': data.get('cost_training', 0),
            'cost_travel': data.get('cost_travel', 0),
            'cost_accommodation': data.get('cost_accommodation', 0),
            'cost_meals': data.get('cost_meals', 0),
            'cost_other': data.get('cost_other', 0),
            'duration_hours': data.get('duration_hours'),
            'attendee_count': attendee_count,
            'compensation_type': compensation_type,
            'wage_cost_per_person': data.get('wage_cost_per_person')
        }
        
        direct_costs, wage_cost, opportunity_cost = calculate_training_cost(cost_data, db)
        cost_total = direct_costs + wage_cost
        
        # Prepare skills_gained - podporuj skills_improved i skills_gained
        skills_raw = data.get('skills_improved') or data.get('skills_gained') or '[]'
        if isinstance(skills_raw, str):
            try:
                skills_list = json_lib.loads(skills_raw)
            except:
                skills_list = []
        elif isinstance(skills_raw, list):
            skills_list = skills_raw
        else:
            skills_list = []
        skills_gained_json = json_lib.dumps(skills_list)
        
        # Prepare participants JSON - uložit do participants pole
        participants_json = json_lib.dumps(attendee_ids)
        
        # Insert training - podporuj title i name, date i date_start
        date_value = _normalize_date(data.get('date_start') or data.get('date'))
        training_id = db.execute("""
            INSERT INTO trainings (
                name, title, description, training_type, category, provider, provider_type,
                date_start, date, date_end, duration_hours,
                is_paid, cost_training, cost_travel, cost_accommodation, cost_meals, cost_other,
                cost_total, cost_opportunity,
                compensation_type, wage_cost, wage_cost_per_person,
                location, is_remote,
                has_certificate, certificate_name, certificate_valid_until,
                skills_gained, skills_improved, skill_level_increase, skill_increase,
                participants,
                created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            training_name,
            data.get('title') or training_name,  # title pro novou strukturu
            data.get('description'),
            data.get('training_type', 'external'),
            data.get('category'),
            data.get('provider'),
            data.get('provider_type'),
            date_value,
            date_value,  # date pro kompatibilitu
            _normalize_date(data.get('date_end')) if data.get('date_end') else None,
            data.get('duration_hours'),
            1 if data.get('is_paid', True) else 0,
            data.get('cost_training', 0),
            data.get('cost_travel', 0),
            data.get('cost_accommodation', 0),
            data.get('cost_meals', 0),
            data.get('cost_other', 0),
            cost_total,
            opportunity_cost,
            compensation_type,
            wage_cost,
            data.get('wage_cost_per_person'),
            data.get('location'),
            1 if data.get('is_remote', False) else 0,
            1 if data.get('has_certificate', False) else 0,
            data.get('certificate_name'),
            _normalize_date(data.get('certificate_valid_until')) if data.get('certificate_valid_until') else None,
            skills_gained_json,
            skills_gained_json,  # skills_improved pro kompatibilitu
            data.get('skill_level_increase') or data.get('skill_increase') or 1,
            data.get('skill_level_increase') or data.get('skill_increase') or 1,  # skill_increase pro kompatibilitu
            participants_json,  # Uložit participants jako JSON
            u['id'] if u else None
        )).lastrowid
        
        # Add attendees
        for emp_id in attendee_ids:
            db.execute("""
                INSERT INTO training_attendees (training_id, employee_id, status)
                VALUES (?, ?, 'registered')
            """, (training_id, emp_id))
        
        db.commit()
        
        return jsonify({"ok": True, "id": training_id, "message": "Školení vytvořeno", "success": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error creating training: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500

@app.route("/gd/api/trainings/<int:training_id>", methods=["PUT"])
def gd_api_trainings_put(training_id):
    """Úprava školení"""
    u, err = require_role(write=True)
    if err:
        return err
    db = get_db()
    
    try:
        import json as json_lib
        data = request.get_json(force=True, silent=True) or {}
        
        # Check if training exists
        training = db.execute("SELECT id FROM trainings WHERE id = ?", (training_id,)).fetchone()
        if not training:
            return jsonify({"ok": False, "error": "not_found"}), 404
        
        # Build UPDATE query
        updates = []
        params = []
        
        allowed_fields = {
            'name': str, 'description': str, 'training_type': str, 'category': str,
            'provider': str, 'provider_type': str, 'duration_hours': float,
            'is_paid': bool, 'cost_training': float, 'cost_travel': float,
            'cost_accommodation': float, 'cost_meals': float, 'cost_other': float,
            'location': str, 'is_remote': bool, 'has_certificate': bool,
            'certificate_name': str, 'rating': int, 'notes': str,
            'skill_level_increase': int, 'compensation_type': str,
            'wage_cost_per_person': float
        }
        
        for field, field_type in allowed_fields.items():
            if field in data:
                if field_type == bool:
                    updates.append(f"{field} = ?")
                    params.append(1 if data[field] else 0)
                elif field in ['date_start', 'date_end', 'certificate_valid_until']:
                    updates.append(f"{field} = ?")
                    params.append(_normalize_date(data[field]) if data[field] else None)
                else:
                    updates.append(f"{field} = ?")
                    params.append(data[field])
        
        # Handle skills_gained / skills_improved
        if 'skills_gained' in data or 'skills_improved' in data:
            skills_raw = data.get('skills_gained') or data.get('skills_improved')
            if isinstance(skills_raw, list):
                skills_json = json_lib.dumps(skills_raw)
                updates.append("skills_gained = ?")
                params.append(skills_json)
                updates.append("skills_improved = ?")
                params.append(skills_json)
            elif isinstance(skills_raw, str):
                updates.append("skills_gained = ?")
                params.append(skills_raw)
                updates.append("skills_improved = ?")
                params.append(skills_raw)
        
        # Handle participants - aktualizovat participants pole i training_attendees
        if 'participants' in data:
            participants_raw = data.get('participants')
            if isinstance(participants_raw, str):
                try:
                    attendee_ids = json_lib.loads(participants_raw)
                except:
                    attendee_ids = []
            elif isinstance(participants_raw, list):
                attendee_ids = participants_raw
            else:
                attendee_ids = []
            
            # Uložit do participants pole
            updates.append("participants = ?")
            params.append(json_lib.dumps(attendee_ids))
            
            # Aktualizovat training_attendees - smazat staré a přidat nové
            db.execute("DELETE FROM training_attendees WHERE training_id = ?", (training_id,))
            for emp_id in attendee_ids:
                db.execute("""
                    INSERT INTO training_attendees (training_id, employee_id, status)
                    VALUES (?, ?, 'registered')
                """, (training_id, emp_id))
        
        # Recalculate costs
        if any(field in data for field in ['cost_training', 'cost_travel', 'cost_accommodation', 
                                          'cost_meals', 'cost_other', 'duration_hours', 'compensation_type', 'wage_cost_per_person']):
            # Get current attendee count
            attendee_count = db.execute(
                "SELECT COUNT(*) FROM training_attendees WHERE training_id = ?",
                (training_id,)
            ).fetchone()[0] or 0
            
            # Get current compensation_type if not in update
            current_training = db.execute(
                "SELECT compensation_type, wage_cost_per_person FROM trainings WHERE id = ?",
                (training_id,)
            ).fetchone()
            
            compensation_type = data.get('compensation_type') or (current_training['compensation_type'] if current_training else 'paid_workday')
            wage_cost_per_person = data.get('wage_cost_per_person') or (current_training['wage_cost_per_person'] if current_training else None)
            
            cost_data = {
                'cost_training': data.get('cost_training') or 0,
                'cost_travel': data.get('cost_travel') or 0,
                'cost_accommodation': data.get('cost_accommodation') or 0,
                'cost_meals': data.get('cost_meals') or 0,
                'cost_other': data.get('cost_other') or 0,
                'duration_hours': data.get('duration_hours'),
                'attendee_count': attendee_count,
                'compensation_type': compensation_type,
                'wage_cost_per_person': wage_cost_per_person
            }
            
            direct_costs, wage_cost, opportunity_cost = calculate_training_cost(cost_data, db)
            updates.append("cost_total = ?")
            params.append(direct_costs + wage_cost)
            updates.append("cost_opportunity = ?")
            params.append(opportunity_cost)
            updates.append("wage_cost = ?")
            params.append(wage_cost)
        
        if not updates:
            return jsonify({"ok": False, "error": "no_fields"}), 400
        
        params.append(training_id)
        db.execute(f"UPDATE trainings SET {', '.join(updates)} WHERE id = ?", params)
        db.commit()
        
        return jsonify({"ok": True, "success": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error updating training: {e}")
        return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500

@app.route("/gd/api/trainings/<int:training_id>", methods=["DELETE"])
def gd_api_trainings_delete(training_id):
    """Smazání školení"""
    u, err = require_role(write=True)
    if err:
        return err
    db = get_db()
    
    try:
        # Check if training exists
        training = db.execute("SELECT id FROM trainings WHERE id = ?", (training_id,)).fetchone()
        if not training:
            return jsonify({"ok": False, "error": "not_found"}), 404
        
        # Delete (CASCADE will delete attendees)
        db.execute("DELETE FROM trainings WHERE id = ?", (training_id,))
        db.commit()
        
        return jsonify({"ok": True, "success": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error deleting training: {e}")
        return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500

@app.route("/gd/api/trainings/<int:training_id>/complete", methods=["POST"])
def gd_api_trainings_complete(training_id):
    """Označení školení jako dokončeného + aktualizace skills"""
    u, err = require_role(write=True)
    if err:
        return err
    db = get_db()
    
    try:
        import json as json_lib
        data = request.get_json(force=True, silent=True) or {}
        
        # Check if training exists
        training = db.execute("SELECT id FROM trainings WHERE id = ?", (training_id,)).fetchone()
        if not training:
            return jsonify({"ok": False, "error": "not_found"}), 404
        
        # Update attendees
        attendees_data = data.get('attendees', [])
        for attendee_data in attendees_data:
            employee_id = attendee_data.get('employee_id')
            if not employee_id:
                continue
            
            db.execute("""
                UPDATE training_attendees
                SET status = 'completed',
                    attendance_confirmed = 1,
                    test_score = ?,
                    certificate_issued = ?,
                    personal_rating = ?
                WHERE training_id = ? AND employee_id = ?
            """, (
                attendee_data.get('test_score'),
                1 if attendee_data.get('certificate_issued', False) else 0,
                attendee_data.get('personal_rating'),
                training_id,
                employee_id
            ))
        
        db.commit()
        
        # Update employee skills
        update_employee_skills_from_training(training_id, db)
        
        return jsonify({"ok": True, "success": True})
    except Exception as e:
        db.rollback()
        print(f"✗ Error completing training: {e}")
        return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500

# ----------------- Trainings Statistics API -----------------
@app.route("/gd/api/trainings/stats", methods=["GET"])
def gd_api_trainings_stats():
    """Statistiky školení"""
    u, err = require_role(write=False)
    if err:
        return err
    db = get_db()
    
    try:
        from datetime import datetime, timedelta
        
        # Filtry
        date_from = _normalize_date(request.args.get('from'))
        date_to = _normalize_date(request.args.get('to'))
        employee_id = request.args.get('employee_id', type=int)
        
        # Default: poslední rok
        if not date_from or not date_to:
            today = datetime.now().date()
            date_to = today.strftime('%Y-%m-%d')
            date_from = (today - timedelta(days=365)).strftime('%Y-%m-%d')
        
        # Build query
        conds = ["date(t.date_start) >= date(?)", "date(t.date_start) <= date(?)"]
        params = [date_from, date_to]
        
        if employee_id:
            conds.append("EXISTS (SELECT 1 FROM training_attendees ta WHERE ta.training_id = t.id AND ta.employee_id = ?)")
            params.append(employee_id)
        
        where_clause = " WHERE " + " AND ".join(conds)
        
        # Get trainings
        trainings = db.execute(f"""
            SELECT t.*, COUNT(DISTINCT ta.id) as attendees_count
            FROM trainings t
            LEFT JOIN training_attendees ta ON ta.training_id = t.id
            {where_clause}
            GROUP BY t.id
        """, params).fetchall()
        
        # Agregace - přístup pomocí indexů nebo dict() konverze
        total_trainings = len(trainings)
        total_cost = 0
        total_hours = 0
        total_attendees = 0
        paid_trainings = 0
        
        # Konverze Row na dict pro snadnější přístup
        trainings_dicts = []
        for t in trainings:
            t_dict = dict(t) if hasattr(t, 'keys') else t
            trainings_dicts.append(t_dict)
            total_cost += float(t_dict.get('cost_total') or t_dict.get('cost_total', 0) or 0)
            total_hours += float(t_dict.get('duration_hours') or t_dict.get('duration_hours', 0) or 0)
            total_attendees += int(t_dict.get('attendees_count') or t_dict.get('attendees_count', 0) or 0)
            if t_dict.get('is_paid') or (isinstance(t_dict.get('is_paid'), int) and t_dict.get('is_paid') == 1):
                paid_trainings += 1
        
        free_trainings = total_trainings - paid_trainings
        
        # By category
        by_category = {}
        for t_dict in trainings_dicts:
            cat = t_dict.get('category') or 'other'
            if cat not in by_category:
                by_category[cat] = {'count': 0, 'cost': 0, 'hours': 0}
            by_category[cat]['count'] += 1
            by_category[cat]['cost'] += float(t_dict.get('cost_total') or 0)
            by_category[cat]['hours'] += float(t_dict.get('duration_hours') or 0)
        
        # By compensation type
        by_compensation = {
            'paid_workday': {
                'count': 0,
                'total_cost': 0,
                'wage_cost': 0
            },
            'costs_only': {
                'count': 0,
                'total_cost': 0
            },
            'unpaid': {
                'count': 0
            }
        }
        
        for t_dict in trainings_dicts:
            # Bezpečný přístup k compensation_type
            comp_type = dict(t_dict).get('compensation_type') if hasattr(t_dict, 'keys') else (t_dict.get('compensation_type') if isinstance(t_dict, dict) else 'paid_workday')
            comp_type = comp_type or 'paid_workday'
            if comp_type not in by_compensation:
                comp_type = 'paid_workday'
            
            by_compensation[comp_type]['count'] += 1
            if comp_type != 'unpaid':
                cost_total_val = dict(t_dict).get('cost_total') if hasattr(t_dict, 'keys') else (t_dict.get('cost_total') if isinstance(t_dict, dict) else 0)
                by_compensation[comp_type]['total_cost'] += float(cost_total_val or 0)
            if comp_type == 'paid_workday':
                wage_cost_val = dict(t_dict).get('wage_cost') if hasattr(t_dict, 'keys') else (t_dict.get('wage_cost') if isinstance(t_dict, dict) else 0)
                by_compensation[comp_type]['wage_cost'] += float(wage_cost_val or 0)
        
        # ROI odhad (zjednodušený)
        # Předpoklad: 1 hodina školení = 0.5% zvýšení produktivity na 6 měsíců
        # Průměrná měsíční produktivita zaměstnance = 150 000 Kč
        estimated_productivity_gain = total_hours * 0.005 * 150000 * 6 * (total_attendees / max(total_trainings, 1))
        roi_percentage = ((estimated_productivity_gain - total_cost) / max(total_cost, 1)) * 100 if total_cost > 0 else 0
        
        return jsonify({
            "ok": True,
            "stats": {
                "count": total_trainings,
                "total_cost": round(total_cost, 2),
                "total_hours": round(total_hours, 1),
                "attendees": total_attendees,
                "roi_percentage": round(roi_percentage, 1),
                "roi_gain": round(estimated_productivity_gain, 2),
                "payback_months": round(total_cost / max(estimated_productivity_gain / 6, 1), 1) if estimated_productivity_gain > 0 else None
            },
            "period": {"from": date_from, "to": date_to},
            "summary": {
                "total_trainings": total_trainings,
                "paid_trainings": paid_trainings,
                "free_trainings": free_trainings,
                "total_cost": round(total_cost, 2),
                "total_hours": round(total_hours, 1),
                "total_attendees": total_attendees,
                "avg_cost_per_training": round(total_cost / max(total_trainings, 1), 2),
                "avg_cost_per_hour": round(total_cost / max(total_hours, 1), 2),
                "avg_cost_per_attendee": round(total_cost / max(total_attendees, 1), 2)
            },
            "by_category": by_category,
            "by_compensation": by_compensation,
            "roi": {
                "estimated_productivity_gain": round(estimated_productivity_gain, 2),
                "roi_percentage": round(roi_percentage, 1),
                "payback_months": round(total_cost / max(estimated_productivity_gain / 6, 1), 1) if estimated_productivity_gain > 0 else None
            }
        })
    except Exception as e:
        print(f"✗ Error getting training stats: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500

@app.route("/gd/api/trainings/team-skills", methods=["GET"])
def gd_api_trainings_team_skills():
    """Přehled skills celého týmu"""
    u, err = require_role(write=False)
    if err:
        return err
    db = get_db()
    
    try:
        import json as json_lib
        
        # Get active employees (použij správný sloupec nebo odstraň filtr)
        employees = db.execute("SELECT id, name, skills, skill_score, training_hours_total, last_training_date FROM employees").fetchall()
        
        # Agregace skills
        all_skills = {}
        for emp in employees:
            try:
                skills = json_lib.loads(emp['skills']) if emp['skills'] and isinstance(emp['skills'], str) else (emp['skills'] or {})
            except:
                skills = {}
            
            if skills:
                for skill, level in skills.items():
                    if skill not in all_skills:
                        all_skills[skill] = {'total_level': 0, 'count': 0, 'employees': []}
                    all_skills[skill]['total_level'] += level
                    all_skills[skill]['count'] += 1
                    all_skills[skill]['employees'].append({
                        'id': emp['id'],
                        'name': emp['name'],
                        'level': level
                    })
        
        # Průměry
        for skill in all_skills:
            all_skills[skill]['avg_level'] = round(
                all_skills[skill]['total_level'] / all_skills[skill]['count'], 1
            )
        
        # Celkové skóre týmu
        team_skill_scores = [float(emp['skill_score'] or 50) for emp in employees]
        team_avg_score = sum(team_skill_scores) / len(team_skill_scores) if team_skill_scores else 50
        
        return jsonify({
            "ok": True,
            "team_score": round(team_avg_score, 1),
            "employees_count": len(employees),
            "skills": all_skills,
            "employees": [{
                "id": emp['id'],
                "name": emp['name'],
                "skill_score": float(emp['skill_score'] or 50),
                "skills": json_lib.loads(emp['skills']) if emp['skills'] and isinstance(emp['skills'], str) else (emp['skills'] or {}),
                "training_hours": float(emp['training_hours_total'] or 0),
                "last_training": emp['last_training_date'] if emp['last_training_date'] else None
            } for emp in employees]
        })
    except Exception as e:
        print(f"✗ Error getting team skills: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": "server_error", "message": str(e)}), 500

@app.route("/gd/api/calendar", methods=["GET", "POST", "PATCH"])
def gd_api_calendar():
    return api_calendar()

@app.route("/gd/api/calendar/<int:event_id>", methods=["DELETE"])
def gd_api_calendar_delete(event_id):
    return api_calendar_delete(event_id)


# Přidej tyto endpointy do main.py

# ----------------- GLOBAL SEARCH -----------------
@app.route("/api/search", methods=["GET"])
def api_global_search():
    """Globální vyhledávání napříč zakázkami, úkoly, issues a zaměstnanci"""
    u, err = require_role()
    if err: return err
    
    query = request.args.get("q", "").strip()
    if not query or len(query) < 2:
        return jsonify({"ok": True, "results": {"jobs": [], "tasks": [], "issues": [], "employees": []}})
    
    db = get_db()
    search_term = f"%{query}%"
    
    # Choose best available timestamp for ordering jobs (schema differs across versions)
    if _table_has_column(db, "jobs", "created_at"):
        jobs_order = "datetime(created_at) DESC, id DESC"
    elif _table_has_column(db, "jobs", "created_date"):
        jobs_order = "datetime(created_date) DESC, id DESC"
    elif _table_has_column(db, "jobs", "date"):
        jobs_order = "date(date) DESC, id DESC"
    else:
        jobs_order = "id DESC"

    # Search Jobs
    # NOTE: Schéma "jobs" se v různých verzích liší. V této aplikaci má tabulka jobs
    # typicky sloupce: name/title, client, city, note, code, created_at. Původní varianta
    # používala description/customer/address, které v DB nejsou.
    jobs = db.execute(f"""
        SELECT
            id,
            COALESCE(title, name, '') AS name,
            note AS description,
            client AS customer,
            city AS address,
            status
        FROM jobs
        WHERE COALESCE(title, name, '') LIKE ?
           OR client LIKE ?
           OR city LIKE ?
           OR note LIKE ?
           OR code LIKE ?
        ORDER BY {jobs_order}
        LIMIT 10
    """, (search_term, search_term, search_term, search_term, search_term)).fetchall()
    
    # Search Tasks (including assigned employees)
    tasks = db.execute("""
        SELECT DISTINCT t.id, t.job_id, t.title, t.description, t.status, t.due_date,
               COALESCE(j.title, j.name, '') as job_name
        FROM tasks t
        LEFT JOIN jobs j ON j.id = t.job_id
        LEFT JOIN task_assignments ta ON ta.task_id = t.id
        LEFT JOIN employees e ON e.id = ta.employee_id
        WHERE t.title LIKE ? 
           OR t.description LIKE ?
           OR e.name LIKE ?
        ORDER BY t.created_at DESC
        LIMIT 10
    """, (search_term, search_term, search_term)).fetchall()
    
    # Search Issues (including assigned employees)
    issues = db.execute("""
        SELECT DISTINCT i.id, i.job_id, i.title, i.description, i.type, i.status, i.severity,
               COALESCE(j.title, j.name, '') as job_name
        FROM issues i
        LEFT JOIN jobs j ON j.id = i.job_id
        LEFT JOIN issue_assignments ia ON ia.issue_id = i.id
        LEFT JOIN employees e ON e.id = ia.employee_id
        WHERE i.title LIKE ? 
           OR i.description LIKE ?
           OR e.name LIKE ?
        ORDER BY i.created_at DESC
        LIMIT 10
    """, (search_term, search_term, search_term)).fetchall()
    
    # Search Employees
    employees = db.execute("""
        SELECT id, name, email, phone, role
        FROM employees
        WHERE name LIKE ? OR email LIKE ? OR phone LIKE ?
        ORDER BY name
        LIMIT 10
    """, (search_term, search_term, search_term)).fetchall()
    
    return jsonify({
        "ok": True,
        "query": query,
        "results": {
            "jobs": [dict(r) for r in jobs],
            "tasks": [dict(r) for r in tasks],
            "issues": [dict(r) for r in issues],
            "employees": [dict(r) for r in employees]
        },
        "total": len(jobs) + len(tasks) + len(issues) + len(employees)
    })


# ----------------- SMART FILTERS -----------------
@app.route("/api/filters/tasks", methods=["GET"])
def api_task_filters():
    """Chytré filtry pro úkoly"""
    u, err = require_role()
    if err: return err
    
    filter_type = request.args.get("filter", "all")
    db = get_db()
    
    # Base query
    base_query = """
        SELECT t.*, j.name as job_name
        FROM tasks t
        LEFT JOIN jobs j ON j.id = t.job_id
    """
    
    # Apply filters
    if filter_type == "my_today":
        # Moje úkoly s deadlinem dnes
        rows = db.execute(base_query + """
            INNER JOIN task_assignments ta ON ta.task_id = t.id
            WHERE ta.employee_id = ?
            AND DATE(t.deadline) = DATE('now')
            AND t.status != 'completed'
            ORDER BY t.priority DESC, t.deadline ASC
        """, (u["id"],)).fetchall()
    
    elif filter_type == "my_overdue":
        # Moje přetažené úkoly
        rows = db.execute(base_query + """
            INNER JOIN task_assignments ta ON ta.task_id = t.id
            WHERE ta.employee_id = ?
            AND DATE(t.deadline) < DATE('now')
            AND t.status != 'completed'
            ORDER BY t.deadline ASC
        """, (u["id"],)).fetchall()
    
    elif filter_type == "my_week":
        # Moje úkoly tento týden
        rows = db.execute(base_query + """
            INNER JOIN task_assignments ta ON ta.task_id = t.id
            WHERE ta.employee_id = ?
            AND DATE(t.deadline) BETWEEN DATE('now') AND DATE('now', '+7 days')
            AND t.status != 'completed'
            ORDER BY t.deadline ASC
        """, (u["id"],)).fetchall()
    
    elif filter_type == "high_priority":
        # Vysoká priorita
        rows = db.execute(base_query + """
            WHERE t.priority = 'high'
            AND t.status != 'completed'
            ORDER BY t.deadline ASC
        """).fetchall()
    
    elif filter_type == "unassigned":
        # Nepřiřazené úkoly
        rows = db.execute(base_query + """
            LEFT JOIN task_assignments ta ON ta.task_id = t.id
            WHERE ta.id IS NULL
            AND t.status != 'completed'
            ORDER BY t.created_at DESC
        """).fetchall()
    
    else:
        # All tasks
        rows = db.execute(base_query + """
            WHERE t.status != 'completed'
            ORDER BY t.created_at DESC
            LIMIT 50
        """).fetchall()
    
    tasks = []
    for task in rows:
        task_dict = dict(task)
        task_dict["assignees"] = get_task_assignees(db, task["id"])
        tasks.append(task_dict)
    
    return jsonify({"ok": True, "tasks": tasks, "filter": filter_type})


@app.route("/api/filters/issues", methods=["GET"])
def api_issue_filters():
    """Chytré filtry pro issues"""
    u, err = require_role()
    if err: return err
    
    filter_type = request.args.get("filter", "all")
    db = get_db()
    
    base_query = """
        SELECT i.*, j.name as job_name
        FROM issues i
        LEFT JOIN jobs j ON j.id = i.job_id
    """
    
    if filter_type == "blockers":
        # Blokující issues
        rows = db.execute(base_query + """
            WHERE i.type = 'blocker'
            AND i.status = 'open'
            ORDER BY i.created_at DESC
        """).fetchall()
    
    elif filter_type == "my_issues":
        # Moje issues
        rows = db.execute(base_query + """
            INNER JOIN issue_assignments ia ON ia.issue_id = i.id
            WHERE ia.employee_id = ?
            AND i.status = 'open'
            ORDER BY i.created_at DESC
        """, (u["id"],)).fetchall()
    
    elif filter_type == "critical":
        # Kritické severity
        rows = db.execute(base_query + """
            WHERE i.severity = 'critical'
            AND i.status = 'open'
            ORDER BY i.created_at DESC
        """).fetchall()
    
    elif filter_type == "recent":
        # Nedávno vytvořené (48h)
        rows = db.execute(base_query + """
            WHERE datetime(i.created_at) >= datetime('now', '-2 days')
            ORDER BY i.created_at DESC
        """).fetchall()
    
    else:
        # All open issues
        rows = db.execute(base_query + """
            WHERE i.status = 'open'
            ORDER BY i.created_at DESC
            LIMIT 50
        """).fetchall()
    
    issues = []
    for issue in rows:
        issue_dict = dict(issue)
        issue_dict["assignees"] = get_issue_assignees(db, issue["id"])
        issues.append(issue_dict)
    
    return jsonify({"ok": True, "issues": issues, "filter": filter_type})


# ----------------- BULK OPERATIONS -----------------
@app.route("/api/bulk/tasks", methods=["POST"])
def api_bulk_tasks():
    """Hromadné operace na úkolech"""
    u, err = require_role(write=True)
    if err: return err
    
    data = request.get_json(force=True, silent=True) or {}
    task_ids = data.get("task_ids", [])
    action = data.get("action", "")
    
    if not task_ids or not action:
        return jsonify({"ok": False, "error": "missing_params"}), 400
    
    db = get_db()
    affected = 0
    
    try:
        if action == "complete":
            # Označit jako dokončené
            placeholders = ','.join('?' * len(task_ids))
            db.execute(f"""
                UPDATE tasks 
                SET status = 'completed', completed_at = datetime('now')
                WHERE id IN ({placeholders})
            """, task_ids)
            affected = db.total_changes
        
        elif action == "delete":
            # Smazat úkoly
            placeholders = ','.join('?' * len(task_ids))
            db.execute(f"DELETE FROM tasks WHERE id IN ({placeholders})", task_ids)
            affected = db.total_changes
        
        elif action == "assign":
            # Přiřadit zaměstnance
            employee_ids = data.get("employee_ids", [])
            if not employee_ids:
                return jsonify({"ok": False, "error": "missing_employees"}), 400
            
            for task_id in task_ids:
                # Remove old assignments
                db.execute("DELETE FROM task_assignments WHERE task_id = ?", (task_id,))
                # Add new assignments
                for idx, emp_id in enumerate(employee_ids):
                    db.execute("""
                        INSERT INTO task_assignments (task_id, employee_id, is_primary)
                        VALUES (?, ?, ?)
                    """, (task_id, emp_id, 1 if idx == 0 else 0))
            affected = len(task_ids)
        
        elif action == "change_status":
            # Změnit stav
            new_status = data.get("status", "")
            if new_status not in ["pending", "in_progress", "completed", "blocked"]:
                return jsonify({"ok": False, "error": "invalid_status"}), 400
            
            placeholders = ','.join('?' * len(task_ids))
            db.execute(f"""
                UPDATE tasks 
                SET status = ?
                WHERE id IN ({placeholders})
            """, [new_status] + task_ids)
            affected = db.total_changes
        
        elif action == "change_priority":
            # Změnit prioritu
            new_priority = data.get("priority", "")
            if new_priority not in ["low", "medium", "high"]:
                return jsonify({"ok": False, "error": "invalid_priority"}), 400
            
            placeholders = ','.join('?' * len(task_ids))
            db.execute(f"""
                UPDATE tasks 
                SET priority = ?
                WHERE id IN ({placeholders})
            """, [new_priority] + task_ids)
            affected = db.total_changes
        
        else:
            return jsonify({"ok": False, "error": "unknown_action"}), 400
        
        db.commit()
        return jsonify({"ok": True, "affected": affected})
    
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


@app.route("/api/bulk/issues", methods=["POST"])
def api_bulk_issues():
    """Hromadné operace na issues"""
    u, err = require_role(write=True)
    if err: return err
    
    data = request.get_json(force=True, silent=True) or {}
    issue_ids = data.get("issue_ids", [])
    action = data.get("action", "")
    
    if not issue_ids or not action:
        return jsonify({"ok": False, "error": "missing_params"}), 400
    
    db = get_db()
    affected = 0
    
    try:
        if action == "resolve":
            # Vyřešit issues
            placeholders = ','.join('?' * len(issue_ids))
            db.execute(f"""
                UPDATE issues 
                SET status = 'resolved', resolved_at = datetime('now')
                WHERE id IN ({placeholders})
            """, issue_ids)
            affected = db.total_changes
        
        elif action == "delete":
            # Smazat issues
            placeholders = ','.join('?' * len(issue_ids))
            db.execute(f"DELETE FROM issues WHERE id IN ({placeholders})", issue_ids)
            affected = db.total_changes
        
        elif action == "assign":
            # Přiřadit zaměstnance
            employee_ids = data.get("employee_ids", [])
            if not employee_ids:
                return jsonify({"ok": False, "error": "missing_employees"}), 400
            
            for issue_id in issue_ids:
                db.execute("DELETE FROM issue_assignments WHERE issue_id = ?", (issue_id,))
                for idx, emp_id in enumerate(employee_ids):
                    db.execute("""
                        INSERT INTO issue_assignments (issue_id, employee_id, is_primary)
                        VALUES (?, ?, ?)
                    """, (issue_id, emp_id, 1 if idx == 0 else 0))
            affected = len(issue_ids)
        
        elif action == "change_severity":
            # Změnit závažnost
            new_severity = data.get("severity", "")
            if new_severity not in ["low", "medium", "high", "critical"]:
                return jsonify({"ok": False, "error": "invalid_severity"}), 400
            
            placeholders = ','.join('?' * len(issue_ids))
            db.execute(f"""
                UPDATE issues 
                SET severity = ?
                WHERE id IN ({placeholders})
            """, [new_severity] + issue_ids)
            affected = db.total_changes
        
        else:
            return jsonify({"ok": False, "error": "unknown_action"}), 400
        
        db.commit()
        return jsonify({"ok": True, "affected": affected})
    
    except Exception as e:
        db.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500


# ============================================================================
# JOBS EXTENDED API - Rozšířené zakázky
# ============================================================================

def dict_from_db_row(row):
    """Převede sqlite3.Row na dict"""
    if row is None:
        return None
    try:
        return dict(row)
    except:
        # Fallback pro starší verze
        return {key: row[key] for key in row.keys()}

@app.route('/api/jobs/<int:job_id>/complete', methods=['GET'])
def get_job_complete(job_id):
    """Získá kompletní informace o zakázce"""
    try:
        db = get_db()
        cursor = db.cursor()
        
        cursor.execute("SELECT * FROM jobs WHERE id = ?", (job_id,))
        job = dict_from_db_row(cursor.fetchone())
        
        if not job:
            return jsonify({"error": "Job not found"}), 404
        
        cursor.execute("SELECT * FROM job_clients WHERE job_id = ?", (job_id,))
        client = dict_from_db_row(cursor.fetchone())
        
        cursor.execute("SELECT * FROM job_locations WHERE job_id = ?", (job_id,))
        location = dict_from_db_row(cursor.fetchone())
        
        cursor.execute("SELECT * FROM job_milestones WHERE job_id = ? ORDER BY order_num, planned_date", (job_id,))
        milestones = [dict_from_db_row(row) for row in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM job_materials WHERE job_id = ?", (job_id,))
        materials = [dict_from_db_row(row) for row in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM job_equipment WHERE job_id = ? ORDER BY date_from", (job_id,))
        equipment = [dict_from_db_row(row) for row in cursor.fetchall()]
        
        cursor.execute("""
            SELECT jta.*, e.name as employee_name, e.position as employee_position
            FROM job_team_assignments jta
            LEFT JOIN employees e ON jta.employee_id = e.id
            WHERE jta.job_id = ? AND jta.is_active = 1
        """, (job_id,))
        team = [dict_from_db_row(row) for row in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM job_subcontractors WHERE job_id = ?", (job_id,))
        subcontractors = [dict_from_db_row(row) for row in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM job_risks WHERE job_id = ? AND status != 'closed'", (job_id,))
        risks = [dict_from_db_row(row) for row in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM job_payments WHERE job_id = ? ORDER BY planned_date", (job_id,))
        payments = [dict_from_db_row(row) for row in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM job_photos WHERE job_id = ? LIMIT 50", (job_id,))
        photos = [dict_from_db_row(row) for row in cursor.fetchall()]
        
        result = {
            "job": job,
            "client": client,
            "location": location,
            "milestones": milestones,
            "materials": materials,
            "equipment": equipment,
            "team": team,
            "subcontractors": subcontractors,
            "risks": risks,
            "payments": payments,
            "photos": photos,
            "summary": {
                "milestones_count": len(milestones),
                "materials_count": len(materials),
                "team_size": len(team),
                "photos_count": len(photos)
            }
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/<int:job_id>', methods=['PUT'])
def update_job(job_id):
    """Aktualizace základních údajů zakázky (pouze owner/manager/leader)"""
    db = get_db()
    
    try:
        # Kontrola oprávnění
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({"error": "Unauthorized"}), 401
        
        cursor = db.cursor()
        cursor.execute("SELECT role FROM employees WHERE id = ?", (user_id,))
        user = cursor.fetchone()
        
        if not user or user[0] not in ['owner', 'manager', 'leader']:
            return jsonify({"error": "Forbidden - requires owner/manager/leader role"}), 403
        
        data = request.get_json()
        
        # Sestavit UPDATE query dynamicky podle toho co je v data
        allowed_fields = ['title', 'client', 'city', 'address', 'date', 'deadline', 
                         'start_date', 'description', 'status', 'estimated_value', 
                         'actual_value', 'priority', 'type']
        
        updates = []
        values = []
        
        for field in allowed_fields:
            if field in data:
                updates.append(f"{field} = ?")
                values.append(data[field])
        
        if not updates:
            return jsonify({"error": "No fields to update"}), 400
        
        values.append(job_id)
        query = f"UPDATE jobs SET {', '.join(updates)} WHERE id = ?"
        
        db.execute(query, values)
        db.commit()
        
        return jsonify({"success": True}), 200
        
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/<int:job_id>/client', methods=['GET', 'POST', 'PUT'])
def manage_job_client(job_id):
    """Správa klienta"""
    db = get_db()
    
    try:
        if request.method == 'GET':
            cursor = db.cursor()
            cursor.execute("SELECT * FROM job_clients WHERE job_id = ?", (job_id,))
            client = dict_from_db_row(cursor.fetchone())
            return jsonify(client if client else {}), 200
            
        elif request.method in ['POST', 'PUT']:
            data = request.get_json()
            
            if not data.get('name'):
                return jsonify({"error": "Name is required"}), 400
            
            cursor = db.cursor()
            cursor.execute("SELECT id FROM job_clients WHERE job_id = ?", (job_id,))
            exists = cursor.fetchone()
            
            if exists:
                db.execute("""
                    UPDATE job_clients SET
                        name = ?, company = ?, ico = ?, dic = ?,
                        email = ?, phone = ?, phone_secondary = ?,
                        billing_street = ?, billing_city = ?, billing_zip = ?
                    WHERE job_id = ?
                """, (
                    data.get('name'), data.get('company'), data.get('ico'), data.get('dic'),
                    data.get('email'), data.get('phone'), data.get('phone_secondary'),
                    data.get('billing_street'), data.get('billing_city'), data.get('billing_zip'),
                    job_id
                ))
            else:
                db.execute("""
                    INSERT INTO job_clients (
                        job_id, name, company, ico, dic, email, phone, 
                        phone_secondary, billing_street, billing_city, billing_zip
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    job_id, data.get('name'), data.get('company'), data.get('ico'), data.get('dic'),
                    data.get('email'), data.get('phone'), data.get('phone_secondary'),
                    data.get('billing_street'), data.get('billing_city'), data.get('billing_zip')
                ))
            
            db.commit()
            return jsonify({"success": True}), 200
            
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/<int:job_id>/location', methods=['GET', 'POST', 'PUT'])
def manage_job_location(job_id):
    """Správa lokace"""
    db = get_db()
    
    try:
        if request.method == 'GET':
            cursor = db.cursor()
            cursor.execute("SELECT * FROM job_locations WHERE job_id = ?", (job_id,))
            location = dict_from_db_row(cursor.fetchone())
            return jsonify(location if location else {}), 200
            
        elif request.method in ['POST', 'PUT']:
            data = request.get_json()
            
            cursor = db.cursor()
            cursor.execute("SELECT id FROM job_locations WHERE job_id = ?", (job_id,))
            exists = cursor.fetchone()
            
            if exists:
                db.execute("""
                    UPDATE job_locations SET
                        street = ?, city = ?, zip = ?, lat = ?, lng = ?,
                        parking = ?, access_notes = ?, gate_code = ?
                    WHERE job_id = ?
                """, (
                    data.get('street'), data.get('city'), data.get('zip'),
                    data.get('lat'), data.get('lng'), data.get('parking'),
                    data.get('access_notes'), data.get('gate_code'), job_id
                ))
            else:
                db.execute("""
                    INSERT INTO job_locations (
                        job_id, street, city, zip, lat, lng,
                        parking, access_notes, gate_code
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    job_id, data.get('street'), data.get('city'), data.get('zip'),
                    data.get('lat'), data.get('lng'), data.get('parking'),
                    data.get('access_notes'), data.get('gate_code')
                ))
            
            db.commit()
            return jsonify({"success": True}), 200
            
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/<int:job_id>/milestones', methods=['GET', 'POST'])
def manage_milestones(job_id):
    """Seznam a vytvoření milníků"""
    db = get_db()
    
    try:
        if request.method == 'GET':
            cursor = db.cursor()
            cursor.execute("SELECT * FROM job_milestones WHERE job_id = ? ORDER BY order_num", (job_id,))
            milestones = [dict_from_db_row(row) for row in cursor.fetchall()]
            return jsonify(milestones), 200
            
        elif request.method == 'POST':
            data = request.get_json()
            
            if not data.get('name'):
                return jsonify({"error": "Name is required"}), 400
            
            cursor = db.cursor()
            cursor.execute("SELECT COALESCE(MAX(order_num), 0) + 1 FROM job_milestones WHERE job_id = ?", (job_id,))
            next_order = cursor.fetchone()[0]
            
            db.execute("""
                INSERT INTO job_milestones (
                    job_id, name, description, planned_date, status, order_num
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, (
                job_id, data.get('name'), data.get('description'),
                data.get('planned_date'), data.get('status', 'pending'), next_order
            ))
            
            db.commit()
            return jsonify({"success": True, "id": db.cursor().lastrowid}), 201
            
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/<int:job_id>/milestones/<int:milestone_id>', methods=['PUT', 'DELETE'])
def manage_milestone(job_id, milestone_id):
    """Update nebo delete milníku"""
    db = get_db()
    
    try:
        if request.method == 'PUT':
            data = request.get_json()
            
            db.execute("""
                UPDATE job_milestones SET
                    name = ?, description = ?, planned_date = ?,
                    actual_date = ?, status = ?, completion_percent = ?
                WHERE id = ? AND job_id = ?
            """, (
                data.get('name'), data.get('description'), data.get('planned_date'),
                data.get('actual_date'), data.get('status'), data.get('completion_percent', 0),
                milestone_id, job_id
            ))
            
            db.commit()
            return jsonify({"success": True}), 200
            
        elif request.method == 'DELETE':
            db.execute("DELETE FROM job_milestones WHERE id = ? AND job_id = ?", (milestone_id, job_id))
            db.commit()
            return jsonify({"success": True}), 200
            
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/<int:job_id>/materials', methods=['GET', 'POST'])
def manage_materials(job_id):
    """Seznam a vytvoření materiálu"""
    db = get_db()
    
    try:
        if request.method == 'GET':
            cursor = db.cursor()
            cursor.execute("SELECT * FROM job_materials WHERE job_id = ?", (job_id,))
            materials = [dict_from_db_row(row) for row in cursor.fetchall()]
            
            total_cost = sum(m.get('total_price', 0) or 0 for m in materials)
            
            return jsonify({
                "materials": materials,
                "summary": {"total": len(materials), "total_cost": total_cost}
            }), 200
            
        elif request.method == 'POST':
            data = request.get_json()
            
            if not data.get('name'):
                return jsonify({"error": "Name is required"}), 400
            
            quantity = float(data.get('quantity', 0))
            price_per_unit = float(data.get('price_per_unit', 0))
            total_price = quantity * price_per_unit
            
            db.execute("""
                INSERT INTO job_materials (
                    job_id, name, quantity, unit, price_per_unit, 
                    total_price, supplier, ordered, delivery_status
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                job_id, data.get('name'), quantity, data.get('unit', 'ks'),
                price_per_unit, total_price, data.get('supplier'),
                data.get('ordered', False), data.get('delivery_status', 'pending')
            ))
            
            db.commit()
            return jsonify({"success": True, "id": db.cursor().lastrowid}), 201
            
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/<int:job_id>/materials/<int:material_id>', methods=['PUT', 'DELETE'])
def manage_material(job_id, material_id):
    """Update nebo delete materiálu"""
    db = get_db()
    
    try:
        if request.method == 'PUT':
            data = request.get_json()
            
            quantity = float(data.get('quantity', 0))
            price_per_unit = float(data.get('price_per_unit', 0))
            total_price = quantity * price_per_unit
            
            db.execute("""
                UPDATE job_materials SET
                    name = ?, quantity = ?, unit = ?,
                    price_per_unit = ?, total_price = ?,
                    supplier = ?, ordered = ?, delivery_status = ?
                WHERE id = ? AND job_id = ?
            """, (
                data.get('name'), quantity, data.get('unit'),
                price_per_unit, total_price, data.get('supplier'),
                data.get('ordered'), data.get('delivery_status'),
                material_id, job_id
            ))
            
            db.commit()
            return jsonify({"success": True}), 200
            
        elif request.method == 'DELETE':
            db.execute("DELETE FROM job_materials WHERE id = ? AND job_id = ?", (material_id, job_id))
            db.commit()
            return jsonify({"success": True}), 200
            
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/api/jobs/<int:job_id>/team', methods=['GET', 'POST'])
def manage_team(job_id):
    """Seznam a přiřazení týmu"""
    db = get_db()
    
    try:
        if request.method == 'GET':
            cursor = db.cursor()
            cursor.execute("""
                SELECT jta.*, e.name as employee_name, e.position as employee_position
                FROM job_team_assignments jta
                LEFT JOIN employees e ON jta.employee_id = e.id
                WHERE jta.job_id = ? AND jta.is_active = 1
            """, (job_id,))
            team = [dict_from_db_row(row) for row in cursor.fetchall()]
            return jsonify({"team": team, "summary": {"size": len(team)}}), 200
            
        elif request.method == 'POST':
            data = request.get_json()
            
            if not data.get('employee_id'):
                return jsonify({"error": "Employee ID is required"}), 400
            
            db.execute("""
                INSERT INTO job_team_assignments (
                    job_id, employee_id, role, hours_planned, hours_actual
                ) VALUES (?, ?, ?, ?, ?)
            """, (
                job_id, data.get('employee_id'), data.get('role', 'worker'),
                data.get('hours_planned', 0), data.get('hours_actual', 0)
            ))
            
            db.commit()
            return jsonify({"success": True}), 201
            
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500

print("✅ Jobs Extended API loaded")

# ----------------- Planning Module API -----------------
# Import planning functions directly
import planning_api

# Make get_db available to planning_api
planning_api.get_db = get_db

# Planning routes
@app.route("/api/planning/timeline")
def api_planning_timeline():
    return planning_api.get_planning_timeline()

@app.route("/api/planning/daily")
@app.route("/api/planning/daily/<target_date>")
def api_planning_daily(target_date=None):
    return planning_api.get_planning_daily(target_date)

@app.route("/api/planning/week")
def api_planning_week():
    return planning_api.get_planning_week()

@app.route("/api/planning/costs")
@app.route("/api/planning/costs/<int:job_id>")
def api_planning_costs(job_id=None):
    return planning_api.get_planning_costs(job_id)

@app.route("/api/action-items", methods=["POST"])
def api_create_action_item():
    return planning_api.create_action_item()

@app.route("/api/planning/actions/my")
def api_my_action_items():
    return planning_api.get_my_action_items()

@app.route("/api/material-delivery", methods=["POST"])
def api_create_material_delivery():
    return planning_api.create_material_delivery()

@app.route("/api/planning/assign", methods=["POST"])
def api_assign_employee():
    return planning_api.assign_employee_to_day()

@app.route("/api/planning/employee/<int:employee_id>")
def api_employee_dashboard(employee_id):
    return planning_api.get_employee_dashboard(employee_id)

@app.route("/api/planning/notifications")
def api_planning_notifications():
    return planning_api.get_planning_notifications()

@app.route("/api/planning/action-items/<int:id>/complete", methods=["POST"])
def api_complete_action_item(id):
    request.view_args = {'id': id}
    return planning_api.quick_complete_action_item()

@app.route("/api/planning/tasks/<int:id>/reschedule", methods=["POST"])
def api_reschedule_task(id):
    request.view_args = {'id': id}
    return planning_api.reschedule_task()

@app.route("/api/planning/suggestions")
def api_planning_suggestions():
    return planning_api.get_smart_suggestions()

# Planning HTML pages
@app.route("/planning/timeline")
def planning_timeline_page():
    return send_from_directory(".", "planning-timeline.html")

@app.route("/planning/daily")
def planning_daily_page():
    return send_from_directory(".", "planning-daily.html")

@app.route("/planning/week")
def planning_week_page():
    return send_from_directory(".", "planning-week.html")

@app.route("/planning/costs")
def planning_costs_page():
    return send_from_directory(".", "planning-costs.html")

print("✅ Planning Module loaded")

# ================================================================
# AI OPERÁTOR - Digitální mozek firmy 🧠
# ================================================================

# AI Operator API modul
import ai_operator_api
ai_operator_api.get_db = get_db
ai_operator_api.register_routes(app)

# AI Operator Rule Engine (nová verze)
try:
    import ai_operator_rule_engine
    ai_operator_rule_engine.get_db = get_db
    ai_operator_rule_engine.register_ai_operator_routes(app)
    print("✅ AI Operátor Rule Engine loaded")
except ImportError as e:
    print(f"⚠️ AI Operátor Rule Engine not available: {e}")

# AI Operator RBAC & Notifikace
try:
    import ai_operator_notifications
    ai_operator_notifications.get_db = get_db
    ai_operator_notifications.register_notification_routes(app)
    print("✅ AI Operátor Notifications loaded")
except ImportError as e:
    print(f"⚠️ AI Operátor Notifications not available: {e}")

# Migrace pro nové tabulky - spustí se při prvním requestu
@app.before_request
def _run_ai_migrations_once():
    if not getattr(app, '_ai_migrations_done', False):
        try:
            from ai_operator_migrations import apply_ai_operator_migrations
            apply_ai_operator_migrations(get_db())
            print("✅ AI Operátor migrations applied")
        except Exception as e:
            print(f"⚠️ AI Operátor migrations: {e}")
        
        # Notifikační tabulky
        try:
            from ai_operator_notifications import apply_notification_migrations
            apply_notification_migrations(get_db())
            print("✅ AI Notification migrations applied")
        except Exception as e:
            print(f"⚠️ AI Notification migrations: {e}")
        
        app._ai_migrations_done = True

# AI Operátor HTML stránka
@app.route('/ai')
@app.route('/ai-operator')
def ai_operator_page():
    return send_from_directory('.', 'ai-operator.html')

print("✅ AI Operátor Module loaded")

# ================================================================
# VELITELSKÝ PANEL - Command Reports 🛰️
# ================================================================

@app.route('/reports')
@app.route('/reports-hub')
def reports_hub_page():
    return send_from_directory('.', 'reports-hub.html')

@app.route('/reports-daily')
def reports_daily_page():
    return send_from_directory('.', 'reports-daily.html')

@app.route('/reports-week')
def reports_week_page():
    return send_from_directory('.', 'reports-week.html')

@app.route('/reports-project')
def reports_project_page():
    return send_from_directory('.', 'reports-project.html')

# ================================================================
# REPORTS GENERATOR API - KOMPLETNÍ FUNKČNÍ VERZE
# ================================================================

@app.route('/api/reports/generate', methods=['POST'])
def api_generate_report():
    """Generate comprehensive report with real data from all modules"""
    u, err = require_role()
    if err: return err
    
    try:
        data = request.get_json() or {}
        report_type = data.get('type', 'weekly')
        date_from = data.get('dateFrom')
        date_to = data.get('dateTo')
        project_id = data.get('projectId')
        content_sections = data.get('content', [])
        detail_sections = data.get('details', [])
        export_format = data.get('format', 'json')
        
        # Filtry zaměstnanců a zakázek
        employee_ids = data.get('employeeIds')  # None = všichni, [] = žádní, [1,2] = konkrétní
        job_ids = data.get('jobIds')  # None = všechny, [] = žádné, [1,2] = konkrétní
        
        db = get_db()
        report_data = {
            'type': report_type,
            'generated_at': datetime.now().isoformat(),
            'date_from': date_from,
            'date_to': date_to,
            'filters': {
                'employees': employee_ids,
                'jobs': job_ids
            },
            'sections': {},
            'summary': {}
        }
        
        # Helper pro employee filter
        def emp_filter_sql(prefix=''):
            if not employee_ids:
                return '', []
            placeholders = ','.join(['?' for _ in employee_ids])
            return f' AND {prefix}employee_id IN ({placeholders})', employee_ids
        
        # Helper pro job filter
        def job_filter_sql(prefix='', col='job_id'):
            if not job_ids:
                return '', []
            placeholders = ','.join(['?' for _ in job_ids])
            return f' AND {prefix}{col} IN ({placeholders})', job_ids
        
        # ============================================================
        # WEEKLY REPORT - Týdenní přehled
        # ============================================================
        if report_type == 'weekly':
            
            # Hodiny - souhrn
            if 'hours_summary' in content_sections:
                try:
                    emp_sql, emp_params = emp_filter_sql()
                    job_sql, job_params = job_filter_sql()
                    result = db.execute(f'''
                        SELECT 
                            COALESCE(SUM(hours), 0) as total_hours,
                            COUNT(DISTINCT employee_id) as unique_employees,
                            COUNT(DISTINCT job_id) as unique_jobs,
                            COUNT(*) as total_entries
                        FROM timesheets
                        WHERE date BETWEEN ? AND ?
                        {emp_sql} {job_sql}
                    ''', (date_from, date_to) + tuple(emp_params) + tuple(job_params)).fetchone()
                    report_data['sections']['hours_summary'] = {
                        'total_hours': round(result['total_hours'] or 0, 1),
                        'unique_employees': result['unique_employees'] or 0,
                        'unique_jobs': result['unique_jobs'] or 0,
                        'total_entries': result['total_entries'] or 0
                    }
                except Exception as e:
                    report_data['sections']['hours_summary'] = {'error': str(e)}
            
            # Hodiny dle projektů
            if 'hours_by_project' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            j.id, 
                            COALESCE(j.name, j.title, 'Bez názvu') as project_name,
                            j.client,
                            j.status,
                            COALESCE(SUM(t.hours), 0) as hours,
                            COUNT(DISTINCT t.employee_id) as workers
                        FROM jobs j
                        LEFT JOIN timesheets t ON t.job_id = j.id AND t.date BETWEEN ? AND ?
                        GROUP BY j.id
                        HAVING hours > 0
                        ORDER BY hours DESC
                    ''', (date_from, date_to)).fetchall()
                    report_data['sections']['hours_by_project'] = [
                        {'id': r['id'], 'name': r['project_name'], 'client': r['client'] or '-', 
                         'status': r['status'], 'hours': round(r['hours'], 1), 'workers': r['workers']}
                        for r in rows
                    ]
                except Exception as e:
                    report_data['sections']['hours_by_project'] = []
            
            # Hodiny dle zaměstnanců
            if 'hours_by_employee' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            e.id,
                            e.name,
                            e.role,
                            COALESCE(SUM(t.hours), 0) as hours,
                            COUNT(DISTINCT t.job_id) as projects,
                            COUNT(DISTINCT t.date) as days_worked
                        FROM employees e
                        LEFT JOIN timesheets t ON t.employee_id = e.id AND t.date BETWEEN ? AND ?
                        WHERE e.active = 1
                        GROUP BY e.id
                        HAVING hours > 0
                        ORDER BY hours DESC
                    ''', (date_from, date_to)).fetchall()
                    report_data['sections']['hours_by_employee'] = [
                        {'id': r['id'], 'name': r['name'], 'role': r['role'] or '-',
                         'hours': round(r['hours'], 1), 'projects': r['projects'], 'days': r['days_worked']}
                        for r in rows
                    ]
                except Exception as e:
                    report_data['sections']['hours_by_employee'] = []
            
            # Dokončené úkoly
            if 'tasks_completed' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            t.id, t.title, t.status,
                            COALESCE(j.name, j.title, '-') as job_name,
                            e.name as assignee
                        FROM tasks t
                        LEFT JOIN jobs j ON j.id = t.job_id
                        LEFT JOIN employees e ON e.id = t.employee_id
                        WHERE t.status IN ('done', 'completed', 'Dokončeno')
                        AND t.created_at BETWEEN ? AND ?
                        ORDER BY t.id DESC
                        LIMIT 50
                    ''', (date_from, date_to + ' 23:59:59')).fetchall()
                    report_data['sections']['tasks_completed'] = [dict(r) for r in rows]
                except Exception as e:
                    report_data['sections']['tasks_completed'] = []
            
            # Rozpracované úkoly
            if 'tasks_pending' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            t.id, t.title, t.status, t.due_date,
                            COALESCE(j.name, j.title, '-') as job_name,
                            e.name as assignee
                        FROM tasks t
                        LEFT JOIN jobs j ON j.id = t.job_id
                        LEFT JOIN employees e ON e.id = t.employee_id
                        WHERE t.status IN ('open', 'in_progress', 'Otevřený', 'V práci')
                        ORDER BY t.due_date ASC, t.id DESC
                        LIMIT 50
                    ''').fetchall()
                    report_data['sections']['tasks_pending'] = [dict(r) for r in rows]
                except Exception as e:
                    report_data['sections']['tasks_pending'] = []
            
            # Úkoly po termínu
            if 'tasks_overdue' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            t.id, t.title, t.status, t.due_date,
                            COALESCE(j.name, j.title, '-') as job_name,
                            e.name as assignee,
                            julianday('now') - julianday(t.due_date) as days_overdue
                        FROM tasks t
                        LEFT JOIN jobs j ON j.id = t.job_id
                        LEFT JOIN employees e ON e.id = t.employee_id
                        WHERE t.due_date < date('now')
                        AND t.status NOT IN ('done', 'completed', 'Dokončeno', 'cancelled')
                        ORDER BY t.due_date ASC
                    ''').fetchall()
                    report_data['sections']['tasks_overdue'] = [
                        {**dict(r), 'days_overdue': int(r['days_overdue'] or 0)}
                        for r in rows
                    ]
                except Exception as e:
                    report_data['sections']['tasks_overdue'] = []
            
            # Denní breakdown (detaily)
            if 'daily_breakdown' in detail_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            date,
                            SUM(hours) as hours,
                            COUNT(DISTINCT employee_id) as workers,
                            COUNT(DISTINCT job_id) as projects
                        FROM timesheets
                        WHERE date BETWEEN ? AND ?
                        GROUP BY date
                        ORDER BY date
                    ''', (date_from, date_to)).fetchall()
                    report_data['sections']['daily_breakdown'] = [
                        {'date': r['date'], 'hours': round(r['hours'], 1), 
                         'workers': r['workers'], 'projects': r['projects']}
                        for r in rows
                    ]
                except Exception as e:
                    report_data['sections']['daily_breakdown'] = []
            
            # Issues/problémy
            if 'issues_reported' in detail_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            i.id, i.title, i.status, i.severity, i.type,
                            COALESCE(j.name, j.title, '-') as job_name,
                            e.name as assigned_to
                        FROM issues i
                        LEFT JOIN jobs j ON j.id = i.job_id
                        LEFT JOIN employees e ON e.id = i.assigned_to
                        WHERE i.created_at BETWEEN ? AND ?
                        ORDER BY i.created_at DESC
                    ''', (date_from, date_to + ' 23:59:59')).fetchall()
                    report_data['sections']['issues'] = [dict(r) for r in rows]
                except Exception as e:
                    report_data['sections']['issues'] = []
        
        # ============================================================
        # MONTHLY REPORT - Měsíční přehled
        # ============================================================
        elif report_type == 'monthly':
            
            # Finanční přehled
            if 'financial' in content_sections:
                try:
                    # Celkové hodiny a náklady na práci
                    labor = db.execute('''
                        SELECT 
                            COALESCE(SUM(t.hours), 0) as total_hours,
                            COALESCE(SUM(t.hours * COALESCE(e.hourly_rate, 200)), 0) as labor_cost
                        FROM timesheets t
                        LEFT JOIN employees e ON e.id = t.employee_id
                        WHERE t.date BETWEEN ? AND ?
                    ''', (date_from, date_to)).fetchone()
                    
                    # Počet zakázek dle statusu
                    jobs_stats = db.execute('''
                        SELECT 
                            COUNT(*) as total_jobs,
                            SUM(CASE WHEN status IN ('completed', 'Dokončeno', 'done') THEN 1 ELSE 0 END) as completed,
                            SUM(CASE WHEN status IN ('active', 'Aktivní', 'V práci') THEN 1 ELSE 0 END) as active
                        FROM jobs
                    ''').fetchone()
                    
                    report_data['sections']['financial'] = {
                        'total_hours': round(labor['total_hours'] or 0, 1),
                        'labor_cost': round(labor['labor_cost'] or 0, 0),
                        'total_jobs': jobs_stats['total_jobs'] or 0,
                        'completed_jobs': jobs_stats['completed'] or 0,
                        'active_jobs': jobs_stats['active'] or 0
                    }
                except Exception as e:
                    report_data['sections']['financial'] = {'error': str(e)}
            
            # Top projekty
            if 'top_projects' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            j.id,
                            COALESCE(j.name, j.title, 'Bez názvu') as name,
                            j.client,
                            j.status,
                            COALESCE(SUM(t.hours), 0) as hours,
                            COUNT(DISTINCT t.employee_id) as workers,
                            COALESCE(SUM(t.hours * COALESCE(e.hourly_rate, 200)), 0) as cost
                        FROM jobs j
                        LEFT JOIN timesheets t ON t.job_id = j.id AND t.date BETWEEN ? AND ?
                        LEFT JOIN employees e ON e.id = t.employee_id
                        GROUP BY j.id
                        ORDER BY hours DESC
                        LIMIT 15
                    ''', (date_from, date_to)).fetchall()
                    report_data['sections']['top_projects'] = [
                        {'id': r['id'], 'name': r['name'], 'client': r['client'] or '-',
                         'status': r['status'], 'hours': round(r['hours'], 1), 
                         'workers': r['workers'], 'cost': round(r['cost'], 0)}
                        for r in rows
                    ]
                except Exception as e:
                    report_data['sections']['top_projects'] = []
            
            # Produktivita týmu
            if 'productivity' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            e.id, e.name, e.role,
                            COALESCE(e.hourly_rate, 200) as hourly_rate,
                            COALESCE(SUM(t.hours), 0) as hours,
                            COUNT(DISTINCT t.job_id) as projects,
                            COUNT(DISTINCT t.date) as days_worked
                        FROM employees e
                        LEFT JOIN timesheets t ON t.employee_id = e.id AND t.date BETWEEN ? AND ?
                        WHERE e.active = 1
                        GROUP BY e.id
                        ORDER BY hours DESC
                    ''', (date_from, date_to)).fetchall()
                    report_data['sections']['productivity'] = [
                        {'id': r['id'], 'name': r['name'], 'role': r['role'] or '-',
                         'hourly_rate': r['hourly_rate'], 'hours': round(r['hours'], 1),
                         'projects': r['projects'], 'days': r['days_worked'],
                         'earnings': round(r['hours'] * r['hourly_rate'], 0)}
                        for r in rows
                    ]
                except Exception as e:
                    report_data['sections']['productivity'] = []
            
            # Sklad - stav
            if 'warehouse' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            id, name, category, quantity, unit, min_quantity,
                            COALESCE(unit_price, 0) as unit_price
                        FROM warehouse_items
                        ORDER BY category, name
                    ''').fetchall()
                    total_value = sum(r['quantity'] * r['unit_price'] for r in rows)
                    low_stock = [r for r in rows if r['min_quantity'] and r['quantity'] <= r['min_quantity']]
                    
                    report_data['sections']['warehouse'] = {
                        'total_items': len(rows),
                        'total_value': round(total_value, 0),
                        'low_stock_count': len(low_stock),
                        'items': [dict(r) for r in rows[:30]]
                    }
                except Exception as e:
                    report_data['sections']['warehouse'] = {'total_items': 0, 'items': []}
            
            # Rozpad příjmů dle projektů
            if 'revenue' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            j.id,
                            COALESCE(j.name, j.title, 'Bez názvu') as name,
                            j.client,
                            COALESCE(j.budget, 0) as budget,
                            COALESCE(SUM(t.hours * COALESCE(e.hourly_rate, 200)), 0) as labor_cost,
                            COALESCE(j.budget, 0) - COALESCE(SUM(t.hours * COALESCE(e.hourly_rate, 200)), 0) as margin
                        FROM jobs j
                        LEFT JOIN timesheets t ON t.job_id = j.id AND t.date BETWEEN ? AND ?
                        LEFT JOIN employees e ON e.id = t.employee_id
                        WHERE j.budget > 0
                        GROUP BY j.id
                        ORDER BY j.budget DESC
                        LIMIT 20
                    ''', (date_from, date_to)).fetchall()
                    report_data['sections']['revenue'] = [
                        {'id': r['id'], 'name': r['name'], 'client': r['client'] or '-',
                         'budget': round(r['budget'], 0), 'labor_cost': round(r['labor_cost'], 0),
                         'margin': round(r['margin'], 0)}
                        for r in rows
                    ]
                except Exception as e:
                    report_data['sections']['revenue'] = {'info': 'Data nejsou k dispozici - vyžaduje rozpočty u zakázek'}
            
            # Rozpad nákladů
            if 'costs' in content_sections:
                try:
                    # Náklady na práci
                    labor = db.execute('''
                        SELECT COALESCE(SUM(t.hours * COALESCE(e.hourly_rate, 200)), 0) as total
                        FROM timesheets t
                        LEFT JOIN employees e ON e.id = t.employee_id
                        WHERE t.date BETWEEN ? AND ?
                    ''', (date_from, date_to)).fetchone()
                    
                    # Spotřeba materiálu (z job_materials)
                    materials = db.execute('''
                        SELECT COALESCE(SUM(jm.qty * COALESCE(wi.unit_price, 0)), 0) as total
                        FROM job_materials jm
                        LEFT JOIN warehouse_items wi ON wi.name = jm.name
                        LEFT JOIN jobs j ON j.id = jm.job_id
                    ''').fetchone()
                    
                    report_data['sections']['costs'] = {
                        'labor_cost': round(labor['total'] or 0, 0),
                        'materials_cost': round(materials['total'] or 0, 0),
                        'total_cost': round((labor['total'] or 0) + (materials['total'] or 0), 0)
                    }
                except Exception as e:
                    report_data['sections']['costs'] = {'info': 'Data nejsou k dispozici'}
            
            # Ziskovost
            if 'profitability' in content_sections:
                try:
                    # Celkové příjmy (rozpočty dokončených zakázek)
                    revenue = db.execute('''
                        SELECT COALESCE(SUM(budget), 0) as total
                        FROM jobs
                        WHERE status IN ('completed', 'Dokončeno', 'done')
                    ''').fetchone()
                    
                    # Celkové náklady
                    costs = db.execute('''
                        SELECT COALESCE(SUM(t.hours * COALESCE(e.hourly_rate, 200)), 0) as total
                        FROM timesheets t
                        LEFT JOIN employees e ON e.id = t.employee_id
                        WHERE t.date BETWEEN ? AND ?
                    ''', (date_from, date_to)).fetchone()
                    
                    total_revenue = revenue['total'] or 0
                    total_costs = costs['total'] or 0
                    profit = total_revenue - total_costs
                    margin = (profit / total_revenue * 100) if total_revenue > 0 else 0
                    
                    report_data['sections']['profitability'] = {
                        'total_revenue': round(total_revenue, 0),
                        'total_costs': round(total_costs, 0),
                        'profit': round(profit, 0),
                        'margin_percent': round(margin, 1)
                    }
                except Exception as e:
                    report_data['sections']['profitability'] = {'info': 'Data nejsou k dispozici - vyžaduje rozpočty u zakázek'}
            
            # Přehled klientů
            if 'clients' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            j.client,
                            COUNT(DISTINCT j.id) as jobs_count,
                            COALESCE(SUM(j.budget), 0) as total_budget,
                            COALESCE(SUM(t.hours), 0) as total_hours
                        FROM jobs j
                        LEFT JOIN timesheets t ON t.job_id = j.id AND t.date BETWEEN ? AND ?
                        WHERE j.client IS NOT NULL AND j.client != ''
                        GROUP BY j.client
                        ORDER BY total_budget DESC
                        LIMIT 20
                    ''', (date_from, date_to)).fetchall()
                    report_data['sections']['clients'] = [
                        {'client': r['client'], 'jobs_count': r['jobs_count'],
                         'total_budget': round(r['total_budget'], 0), 'total_hours': round(r['total_hours'], 1)}
                        for r in rows
                    ]
                except Exception as e:
                    report_data['sections']['clients'] = {'info': 'Data nejsou k dispozici'}
            
            # Srovnání s minulým obdobím
            if 'comparison' in content_sections:
                try:
                    df = datetime.strptime(date_from, '%Y-%m-%d')
                    dt = datetime.strptime(date_to, '%Y-%m-%d')
                    period_days = (dt - df).days + 1
                    prev_from = (df - timedelta(days=period_days)).strftime('%Y-%m-%d')
                    prev_to = (df - timedelta(days=1)).strftime('%Y-%m-%d')
                    
                    # Aktuální období
                    current = db.execute('''
                        SELECT COALESCE(SUM(hours), 0) as hours, COUNT(DISTINCT job_id) as jobs
                        FROM timesheets WHERE date BETWEEN ? AND ?
                    ''', (date_from, date_to)).fetchone()
                    
                    # Předchozí období
                    previous = db.execute('''
                        SELECT COALESCE(SUM(hours), 0) as hours, COUNT(DISTINCT job_id) as jobs
                        FROM timesheets WHERE date BETWEEN ? AND ?
                    ''', (prev_from, prev_to)).fetchone()
                    
                    curr_hours = current['hours'] or 0
                    prev_hours = previous['hours'] or 0
                    hours_change = ((curr_hours - prev_hours) / prev_hours * 100) if prev_hours > 0 else 0
                    
                    report_data['sections']['comparison'] = {
                        'current_hours': round(curr_hours, 1),
                        'previous_hours': round(prev_hours, 1),
                        'hours_change_percent': round(hours_change, 1),
                        'current_jobs': current['jobs'] or 0,
                        'previous_jobs': previous['jobs'] or 0
                    }
                except Exception as e:
                    report_data['sections']['comparison'] = {'info': 'Data nejsou k dispozici'}
            
            # Detail nákladů - Náklady na zaměstnance
            if 'salaries' in detail_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            e.id, e.name, e.role,
                            COALESCE(e.hourly_rate, 200) as hourly_rate,
                            COALESCE(SUM(t.hours), 0) as hours,
                            COALESCE(SUM(t.hours * COALESCE(e.hourly_rate, 200)), 0) as total_cost
                        FROM employees e
                        LEFT JOIN timesheets t ON t.employee_id = e.id AND t.date BETWEEN ? AND ?
                        WHERE e.active = 1
                        GROUP BY e.id
                        HAVING hours > 0
                        ORDER BY total_cost DESC
                    ''', (date_from, date_to)).fetchall()
                    report_data['sections']['salaries'] = [
                        {'name': r['name'], 'role': r['role'] or '-', 'hourly_rate': r['hourly_rate'],
                         'hours': round(r['hours'], 1), 'total_cost': round(r['total_cost'], 0)}
                        for r in rows
                    ]
                except Exception as e:
                    report_data['sections']['salaries'] = []
            
            # Detail nákladů - Náklady na materiál
            if 'materials' in detail_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            jm.name, jm.qty, jm.unit,
                            COALESCE(wi.unit_price, 0) as unit_price,
                            jm.qty * COALESCE(wi.unit_price, 0) as total_price,
                            COALESCE(j.name, j.title, '-') as job_name
                        FROM job_materials jm
                        LEFT JOIN warehouse_items wi ON wi.name = jm.name
                        LEFT JOIN jobs j ON j.id = jm.job_id
                        ORDER BY total_price DESC
                        LIMIT 50
                    ''').fetchall()
                    report_data['sections']['materials'] = [
                        {'name': r['name'], 'qty': r['qty'], 'unit': r['unit'] or 'ks',
                         'unit_price': round(r['unit_price'], 0), 'total_price': round(r['total_price'], 0),
                         'job_name': r['job_name']}
                        for r in rows
                    ]
                except Exception as e:
                    report_data['sections']['materials'] = []
            
            # Faktury (placeholder - závisí na struktuře DB)
            if 'invoices' in detail_sections:
                report_data['sections']['invoices'] = {'info': 'Modul faktur není aktivní - data nejsou k dispozici'}
            
            # Status plateb (placeholder)
            if 'payments' in detail_sections:
                report_data['sections']['payments'] = {'info': 'Modul plateb není aktivní - data nejsou k dispozici'}
            
            # Detail výdajů (placeholder)
            if 'expenses' in detail_sections:
                report_data['sections']['expenses'] = {'info': 'Modul výdajů není aktivní - data nejsou k dispozici'}
            
            # Subdodavatelé (placeholder)
            if 'subcontractors' in detail_sections:
                report_data['sections']['subcontractors'] = {'info': 'Modul subdodavatelů není aktivní - data nejsou k dispozici'}
        
        # ============================================================
        # PROJECT REPORT - Report konkrétního projektu
        # ============================================================
        elif report_type == 'project' and project_id:
            
            # Základní info o projektu
            if 'summary' in content_sections:
                try:
                    job = db.execute('''
                        SELECT * FROM jobs WHERE id = ?
                    ''', (project_id,)).fetchone()
                    if job:
                        report_data['sections']['project_info'] = dict(job)
                except Exception as e:
                    pass
            
            # Timeline - všechny výkazy
            if 'timeline' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            t.id, t.date, t.hours, t.activity, t.place,
                            e.name as employee
                        FROM timesheets t
                        LEFT JOIN employees e ON e.id = t.employee_id
                        WHERE t.job_id = ?
                        ORDER BY t.date DESC
                    ''', (project_id,)).fetchall()
                    report_data['sections']['timesheets'] = [dict(r) for r in rows]
                    
                    # Souhrn
                    total_hours = sum(r['hours'] for r in rows)
                    report_data['sections']['hours_total'] = round(total_hours, 1)
                except Exception as e:
                    report_data['sections']['timesheets'] = []
            
            # Náklady na projekt
            if 'budget' in content_sections:
                try:
                    result = db.execute('''
                        SELECT 
                            COALESCE(SUM(t.hours), 0) as total_hours,
                            COALESCE(SUM(t.hours * COALESCE(e.hourly_rate, 200)), 0) as labor_cost
                        FROM timesheets t
                        LEFT JOIN employees e ON e.id = t.employee_id
                        WHERE t.job_id = ?
                    ''', (project_id,)).fetchone()
                    
                    report_data['sections']['costs'] = {
                        'total_hours': round(result['total_hours'] or 0, 1),
                        'labor_cost': round(result['labor_cost'] or 0, 0)
                    }
                except Exception as e:
                    report_data['sections']['costs'] = {}
            
            # Tým na projektu
            if 'team' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            e.id, e.name, e.role,
                            SUM(t.hours) as hours,
                            COUNT(DISTINCT t.date) as days
                        FROM timesheets t
                        JOIN employees e ON e.id = t.employee_id
                        WHERE t.job_id = ?
                        GROUP BY e.id
                        ORDER BY hours DESC
                    ''', (project_id,)).fetchall()
                    report_data['sections']['team'] = [
                        {'id': r['id'], 'name': r['name'], 'role': r['role'] or '-',
                         'hours': round(r['hours'], 1), 'days': r['days']}
                        for r in rows
                    ]
                except Exception as e:
                    report_data['sections']['team'] = []
            
            # Úkoly projektu
            if 'tasks' in detail_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            t.id, t.title, t.status, t.due_date, t.description,
                            e.name as assignee
                        FROM tasks t
                        LEFT JOIN employees e ON e.id = t.employee_id
                        WHERE t.job_id = ?
                        ORDER BY 
                            CASE t.status 
                                WHEN 'open' THEN 1 
                                WHEN 'in_progress' THEN 2 
                                ELSE 3 
                            END,
                            t.id DESC
                    ''', (project_id,)).fetchall()
                    report_data['sections']['tasks'] = [dict(r) for r in rows]
                except Exception as e:
                    report_data['sections']['tasks'] = []
            
            # Materiály použité
            if 'materials' in detail_sections:
                try:
                    rows = db.execute('''
                        SELECT id, name, qty, unit
                        FROM job_materials
                        WHERE job_id = ?
                    ''', (project_id,)).fetchall()
                    report_data['sections']['materials'] = [dict(r) for r in rows]
                except Exception as e:
                    report_data['sections']['materials'] = []
            
            # Issues/problémy
            if 'risks' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT 
                            i.id, i.title, i.status, i.severity, i.type, i.description,
                            e.name as assigned_to
                        FROM issues i
                        LEFT JOIN employees e ON e.id = i.assigned_to
                        WHERE i.job_id = ?
                        ORDER BY i.created_at DESC
                    ''', (project_id,)).fetchall()
                    report_data['sections']['issues'] = [dict(r) for r in rows]
                except Exception as e:
                    report_data['sections']['issues'] = []
        
        # ============================================================
        # CUSTOM REPORT - Vlastní kombinace
        # ============================================================
        elif report_type == 'custom':
            
            # Hodiny celkem
            if 'hours' in content_sections:
                try:
                    result = db.execute('''
                        SELECT COALESCE(SUM(hours), 0) as total
                        FROM timesheets
                        WHERE date BETWEEN ? AND ?
                    ''', (date_from, date_to)).fetchone()
                    report_data['sections']['hours'] = {'total': round(result['total'] or 0, 1)}
                except Exception as e:
                    report_data['sections']['hours'] = {'total': 0}
            
            # Úkoly
            if 'tasks' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT status, COUNT(*) as count
                        FROM tasks
                        GROUP BY status
                    ''').fetchall()
                    total = sum(r['count'] for r in rows)
                    report_data['sections']['tasks'] = {
                        'total': total,
                        'by_status': [dict(r) for r in rows]
                    }
                except Exception as e:
                    report_data['sections']['tasks'] = {'total': 0, 'by_status': []}
            
            # Projekty
            if 'projects' in content_sections:
                try:
                    rows = db.execute('''
                        SELECT status, COUNT(*) as count
                        FROM jobs
                        GROUP BY status
                    ''').fetchall()
                    total = sum(r['count'] for r in rows)
                    report_data['sections']['projects'] = {
                        'total': total,
                        'by_status': [dict(r) for r in rows]
                    }
                except Exception as e:
                    report_data['sections']['projects'] = {'total': 0, 'by_status': []}
            
            # Tým
            if 'team' in content_sections:
                try:
                    result = db.execute('''
                        SELECT 
                            COUNT(*) as total,
                            SUM(CASE WHEN active = 1 THEN 1 ELSE 0 END) as active
                        FROM employees
                    ''').fetchone()
                    report_data['sections']['team'] = {
                        'total': result['total'] or 0,
                        'active': result['active'] or 0
                    }
                except Exception as e:
                    report_data['sections']['team'] = {'total': 0, 'active': 0}
            
            # Sklad
            if 'warehouse' in content_sections:
                try:
                    result = db.execute('''
                        SELECT 
                            COUNT(*) as items,
                            COALESCE(SUM(quantity * COALESCE(unit_price, 0)), 0) as value
                        FROM warehouse_items
                    ''').fetchone()
                    report_data['sections']['warehouse'] = {
                        'items': result['items'] or 0,
                        'value': round(result['value'] or 0, 0)
                    }
                except Exception as e:
                    report_data['sections']['warehouse'] = {'items': 0, 'value': 0}
            
            # Školka
            if 'nursery' in content_sections:
                try:
                    result = db.execute('''
                        SELECT 
                            COUNT(*) as total_plants,
                            COALESCE(SUM(quantity), 0) as total_quantity
                        FROM nursery_plants
                    ''').fetchone()
                    report_data['sections']['nursery'] = {
                        'total_plants': result['total_plants'] or 0,
                        'total_quantity': result['total_quantity'] or 0
                    }
                except Exception as e:
                    report_data['sections']['nursery'] = {'total_plants': 0, 'total_quantity': 0}
        
        # Přidat souhrn
        report_data['summary'] = {
            'sections_count': len(report_data['sections']),
            'has_data': any(report_data['sections'].values())
        }
        
        return jsonify(report_data)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/reports/projects')
def api_reports_projects():
    """Get list of projects for report selector"""
    u, err = require_role()
    if err: return err
    
    db = get_db()
    jobs = db.execute('''
        SELECT id, COALESCE(name, title, 'Projekt ' || id) as name, status, client, city
        FROM jobs
        ORDER BY id DESC
    ''').fetchall()
    return jsonify([dict(r) for r in jobs])

@app.route('/api/reports/employees')
def api_reports_employees():
    """Get list of employees for report selector"""
    u, err = require_role()
    if err: return err
    
    db = get_db()
    # Zjisti jestli existuje sloupec active
    cols = {r[1] for r in db.execute("PRAGMA table_info(employees)").fetchall()}
    
    if 'active' in cols:
        employees = db.execute('''
            SELECT id, name, role, active
            FROM employees
            WHERE active = 1
            ORDER BY name
        ''').fetchall()
    else:
        employees = db.execute('''
            SELECT id, name, role, 1 as active
            FROM employees
            ORDER BY name
        ''').fetchall()
    return jsonify([dict(r) for r in employees])

print("✅ Reports API loaded")

print("✅ Velitelský Panel loaded")

# ================================================================
# PLANNING EXTENDED ROUTES - All New Features
# ================================================================
import planning_extended_api as ext_api
ext_api.get_db = get_db

# Nursery - Trvalkové školka 🌸
@app.route('/nursery')
def nursery_page():
    return send_from_directory('.', 'nursery.html')

# Timeline / Gantt Chart 📈
@app.route('/timeline')
def timeline_page():
    return render_template('timeline.html')

@app.route('/api/nursery/overview')
def api_nursery_overview():
    return ext_api.get_nursery_overview()

@app.route('/api/nursery/plants')
def api_nursery_plants():
    return ext_api.get_nursery_plants()

@app.route('/api/nursery/plants', methods=['POST'])
def api_create_nursery_plant():
    return ext_api.create_nursery_plant()

@app.route('/api/nursery/plants/<int:plant_id>', methods=['PUT'])
def api_update_nursery_plant(plant_id):
    return ext_api.update_nursery_plant()

@app.route('/api/nursery/plants/<int:plant_id>/to-warehouse', methods=['POST'])
def api_move_plant_to_warehouse(plant_id):
    return ext_api.move_to_warehouse()

@app.route('/api/nursery/warehouse-transfers', methods=['GET'])
def api_get_warehouse_transfers():
    return ext_api.get_warehouse_transfer_history()

@app.route('/api/nursery/watering', methods=['POST'])
def api_log_watering():
    return ext_api.log_watering()

# Recurring tasks 🔄
@app.route('/recurring-tasks')
def recurring_tasks_page():
    return send_from_directory('.', 'recurring-tasks.html')

@app.route('/api/recurring/templates')
def api_recurring_templates():
    return ext_api.get_recurring_templates()

@app.route('/api/recurring/templates', methods=['POST'])
def api_create_recurring_template():
    return ext_api.create_recurring_template()

@app.route('/api/recurring/generate', methods=['POST'])
def api_generate_recurring():
    return ext_api.generate_recurring_tasks()

# Materials 📦  
@app.route('/materials')
@app.route('/warehouse.html')
@app.route('/warehouse')
def materials_page():
    return send_from_directory('.', 'warehouse.html')

@app.route('/api/materials')
def api_materials():
    return ext_api.get_materials()

@app.route('/api/materials', methods=['POST'])
def api_create_material():
    return ext_api.create_material()

@app.route('/api/materials/<int:material_id>', methods=['PUT'])
def api_update_material(material_id):
    request.view_args = {'material_id': material_id}
    return ext_api.update_material()

@app.route('/api/materials/movement', methods=['POST'])
def api_material_movement():
    return ext_api.add_material_movement()

@app.route('/api/materials/movements')
def api_material_movements():
    return ext_api.get_material_movements()

# Photos 📸
@app.route('/api/tasks/<int:task_id>/photos', methods=['POST'])
def api_upload_task_photo(task_id):
    request.view_args = {'task_id': task_id}
    return ext_api.upload_task_photo()

@app.route('/api/tasks/<int:task_id>/photos')
def api_get_task_photos(task_id):
    return ext_api.get_task_photos(task_id)

# Plant database 🌺
@app.route('/plant-database')
def plant_database_page():
    return send_from_directory('.', 'plant-database.html')

@app.route('/api/plant-species')
def api_plant_species():
    return ext_api.get_plant_species()

# Maintenance contracts 📋
@app.route('/api/contracts')
def api_contracts():
    return ext_api.get_maintenance_contracts()

# Seasonal planner 🌱
@app.route('/api/seasonal-tasks')
def api_seasonal():
    return ext_api.get_seasonal_tasks()

print("✅ Planning Extended Routes loaded")

# WAREHOUSE EXTENDED - Nadčasové rozšíření skladu
# ================================================================

# Import warehouse extended module
import warehouse_extended

# Set get_db reference
warehouse_extended.get_db = get_db

# Apply warehouse migrations on startup
@app.before_request
def _ensure_warehouse_schema():
    if not getattr(_ensure_warehouse_schema, "_applied", False):
        try:
            warehouse_extended.apply_warehouse_migrations()
            _ensure_warehouse_schema._applied = True
            print("✅ Warehouse Extended migrations applied")
        except Exception as e:
            print(f"[WAREHOUSE] Migration error: {e}")

# -------- WAREHOUSE ITEMS --------
@app.route("/api/warehouse/items", methods=["GET"])
def api_warehouse_items_list():
    u, err = require_auth()
    if err: return err
    try:
        db = get_db()
        low_stock_only = request.args.get('low_stock') in ('1', 'true', 'yes')
        
        if low_stock_only:
            # Return only items below minimum stock
            items = db.execute("""
                SELECT * FROM warehouse_items 
                WHERE status = 'active' AND qty <= minStock AND minStock > 0
                ORDER BY (qty * 1.0 / NULLIF(minStock, 0)) ASC, name
            """).fetchall()
        else:
            items = db.execute("""
                SELECT * FROM warehouse_items 
                WHERE status = 'active' 
                ORDER BY name
            """).fetchall()
        
        # Normalize field names for frontend
        result = []
        for i in items:
            item = dict(i)
            item['quantity'] = item.get('qty', 0)
            item['min_quantity'] = item.get('minStock', 0)
            result.append(item)
        
        return jsonify({"success": True, "ok": True, "items": result})
    except Exception as e:
        print(f"[ERROR] Warehouse items: {e}")
        return jsonify({"success": False, "ok": False, "error": str(e)}), 500

@app.route("/api/warehouse/items", methods=["POST"])
def api_warehouse_items_create():
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    try:
        data = request.json
        db = get_db()
        cursor = db.execute("""
            INSERT INTO warehouse_items (name, sku, category, location, qty, unit, price, minStock, note, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'active')
        """, (
            data.get('name'),
            data.get('sku', ''),
            data.get('category', ''),
            data.get('location', ''),
            data.get('qty', 0),
            data.get('unit', 'ks'),
            data.get('price', 0),
            data.get('minStock', 10),
            data.get('note', '')
        ))
        db.commit()
        return jsonify({"success": True, "id": cursor.lastrowid})
    except Exception as e:
        print(f"[ERROR] Create warehouse item: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/warehouse/items/<int:item_id>", methods=["PUT"])
def api_warehouse_items_update(item_id):
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    try:
        data = request.json
        db = get_db()
        db.execute("""
            UPDATE warehouse_items 
            SET name=?, sku=?, category=?, location=?, qty=?, unit=?, price=?, minStock=?, note=?, updated_at=datetime('now')
            WHERE id=?
        """, (
            data.get('name'),
            data.get('sku', ''),
            data.get('category', ''),
            data.get('location', ''),
            data.get('qty', 0),
            data.get('unit', 'ks'),
            data.get('price', 0),
            data.get('minStock', 10),
            data.get('note', ''),
            item_id
        ))
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        print(f"[ERROR] Update warehouse item: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/warehouse/items/<int:item_id>", methods=["DELETE"])
def api_warehouse_items_delete(item_id):
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    try:
        db = get_db()
        db.execute("UPDATE warehouse_items SET status='deleted' WHERE id=?", (item_id,))
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

# -------- WAREHOUSE LOCATIONS --------
@app.route("/api/warehouse/locations", methods=["GET"])
def api_warehouse_locations():
    u, err = require_auth()
    if err: return err
    return warehouse_extended.get_locations()

@app.route("/api/warehouse/locations", methods=["POST"])
def api_warehouse_locations_create():
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    return warehouse_extended.create_location()

@app.route("/api/warehouse/locations/<int:location_id>", methods=["PATCH"])
def api_warehouse_locations_update(location_id):
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    return warehouse_extended.update_location(location_id)

@app.route("/api/warehouse/locations/<int:location_id>", methods=["DELETE"])
def api_warehouse_locations_delete(location_id):
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    return warehouse_extended.delete_location(location_id)

# -------- WAREHOUSE MOVEMENTS --------
@app.route("/api/warehouse/movements", methods=["GET"])
def api_warehouse_movements():
    u, err = require_auth()
    if err: return err
    return warehouse_extended.get_movements()

@app.route("/api/warehouse/movements", methods=["POST"])
def api_warehouse_movements_create():
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    return warehouse_extended.create_movement()

@app.route("/api/warehouse/jobs/<int:job_id>/materials", methods=["GET"])
def api_warehouse_job_materials(job_id):
    u, err = require_auth()
    if err: return err
    return warehouse_extended.get_job_materials(job_id)

# -------- WAREHOUSE RESERVATIONS --------
@app.route("/api/warehouse/reservations", methods=["GET"])
def api_warehouse_reservations():
    u, err = require_auth()
    if err: return err
    return warehouse_extended.get_reservations()

@app.route("/api/warehouse/reservations", methods=["POST"])
def api_warehouse_reservations_create():
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    return warehouse_extended.create_reservation()

@app.route("/api/warehouse/reservations/<int:reservation_id>", methods=["PATCH"])
def api_warehouse_reservations_update(reservation_id):
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    return warehouse_extended.update_reservation(reservation_id)

# -------- WAREHOUSE INVENTORY --------
@app.route("/api/warehouse/inventory", methods=["GET"])
def api_warehouse_inventory():
    u, err = require_auth()
    if err: return err
    return warehouse_extended.get_inventories()

@app.route("/api/warehouse/inventory/start", methods=["POST"])
def api_warehouse_inventory_start():
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    return warehouse_extended.start_inventory()

@app.route("/api/warehouse/inventory/<int:inventory_id>/items", methods=["GET"])
def api_warehouse_inventory_items(inventory_id):
    u, err = require_auth()
    if err: return err
    return warehouse_extended.get_inventory_items(inventory_id)

@app.route("/api/warehouse/inventory/items/<int:inventory_item_id>", methods=["PATCH"])
def api_warehouse_inventory_items_update(inventory_item_id):
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    return warehouse_extended.update_inventory_item(inventory_item_id)

@app.route("/api/warehouse/inventory/<int:inventory_id>/complete", methods=["POST"])
def api_warehouse_inventory_complete(inventory_id):
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    return warehouse_extended.complete_inventory(inventory_id)

# -------- WAREHOUSE MERGE & RENAME --------
@app.route("/api/warehouse/items/merge", methods=["POST"])
def api_warehouse_items_merge():
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    return warehouse_extended.merge_items()

@app.route("/api/warehouse/items/<int:item_id>/rename", methods=["PATCH"])
def api_warehouse_items_rename(item_id):
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    return warehouse_extended.rename_item(item_id)

# -------- WAREHOUSE STATS --------
@app.route("/api/warehouse/stats", methods=["GET"])
def api_warehouse_stats():
    u, err = require_auth()
    if err: return err
    return warehouse_extended.get_warehouse_stats()


# -------- WAREHOUSE ITEMS CRUD --------
@app.route("/api/items", methods=["GET"])
def api_warehouse_get_items():
    u, err = require_auth()
    if err: return err
    try:
        db = get_db()
        site = request.args.get('site', '')
        
        if site == 'warehouse':
            try:
                items = db.execute("SELECT * FROM warehouse_items WHERE status = 'active' ORDER BY name").fetchall()
                return jsonify({"success": True, "items": [dict(i) for i in items]})
            except Exception as e:
                print(f"[ERROR] Loading warehouse_items: {e}")
                return jsonify({"success": True, "items": []})
        else:
            return jsonify({"success": True, "items": []})
    except Exception as e:
        print(f"[ERROR] api_get_items: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/items", methods=["POST"])
def api_warehouse_create_item():
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    
    try:
        data = request.json
        db = get_db()
        
        cursor = db.execute(
            "INSERT INTO warehouse_items (name, sku, category, location, qty, unit, price, minStock, batch_number, expiration_date, image, note) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (data.get('name', ''), data.get('sku', ''), data.get('category', ''), data.get('location', ''),
             float(data.get('qty', 0)), data.get('unit', 'ks'), float(data.get('price', 0)),
             float(data.get('minStock', 10)), data.get('batch_number', ''), data.get('expiration_date', ''),
             data.get('image', ''), data.get('note', ''))
        )
        
        db.commit()
        print(f"[SUCCESS] Created warehouse item: {data.get('name')} (ID: {cursor.lastrowid})")
        return jsonify({"success": True, "item_id": cursor.lastrowid})
    except Exception as e:
        print(f"[ERROR] api_create_item: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/items", methods=["PATCH"])
def api_warehouse_update_item():
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    
    try:
        data = request.json
        db = get_db()
        item_id = data.get('id')
        
        db.execute(
            "UPDATE warehouse_items SET name = ?, sku = ?, category = ?, location = ?, qty = ?, unit = ?, price = ?, minStock = ?, batch_number = ?, expiration_date = ?, image = ?, note = ?, updated_at = datetime('now') WHERE id = ?",
            (data.get('name', ''), data.get('sku', ''), data.get('category', ''), data.get('location', ''),
             float(data.get('qty', 0)), data.get('unit', 'ks'), float(data.get('price', 0)),
             float(data.get('minStock', 10)), data.get('batch_number', ''), data.get('expiration_date', ''),
             data.get('image', ''), data.get('note', ''), item_id)
        )
        
        db.commit()
        print(f"[SUCCESS] Updated warehouse item ID: {item_id}")
        return jsonify({"success": True})
    except Exception as e:
        print(f"[ERROR] api_update_item: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


print("✅ Warehouse Extended Routes loaded")

# ----------------- Main Entry Point -----------------
# -------- WAREHOUSE ITEMS CRUD --------
@app.route("/api/items", methods=["GET"])
def api_get_items():
    u, err = require_auth()
    if err: return err
    try:
        db = get_db()
        site = request.args.get('site', '')
        
        if site == 'warehouse':
            try:
                items = db.execute("SELECT * FROM warehouse_items WHERE status = 'active' ORDER BY name").fetchall()
                return jsonify({"success": True, "items": [dict(i) for i in items]})
            except Exception:
                return jsonify({"success": True, "items": []})
        else:
            return jsonify({"success": True, "items": []})
    except Exception as e:
        print(f"[ERROR] api_get_items: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/items", methods=["POST"])
def api_create_item():
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    
    try:
        data = request.json
        db = get_db()
        
        cursor = db.execute(
            "INSERT INTO warehouse_items (name, sku, category, location, qty, unit, price, minStock, batch_number, expiration_date, image, note) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (data.get('name', ''), data.get('sku', ''), data.get('category', ''), data.get('location', ''),
             float(data.get('qty', 0)), data.get('unit', 'ks'), float(data.get('price', 0)),
             float(data.get('minStock', 10)), data.get('batch_number', ''), data.get('expiration_date', ''),
             data.get('image', ''), data.get('note', ''))
        )
        
        db.commit()
        return jsonify({"success": True, "item_id": cursor.lastrowid})
    except Exception as e:
        print(f"[ERROR] api_create_item: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/items", methods=["PATCH"])
def api_update_item():
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    
    try:
        data = request.json
        db = get_db()
        item_id = data.get('id')
        
        db.execute(
            "UPDATE warehouse_items SET name = ?, sku = ?, category = ?, location = ?, qty = ?, unit = ?, price = ?, minStock = ?, batch_number = ?, expiration_date = ?, image = ?, note = ?, updated_at = datetime('now') WHERE id = ?",
            (data.get('name', ''), data.get('sku', ''), data.get('category', ''), data.get('location', ''),
             float(data.get('qty', 0)), data.get('unit', 'ks'), float(data.get('price', 0)),
             float(data.get('minStock', 10)), data.get('batch_number', ''), data.get('expiration_date', ''),
             data.get('image', ''), data.get('note', ''), item_id)
        )
        
        db.commit()
        return jsonify({"success": True})
    except Exception as e:
        print(f"[ERROR] api_update_item: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

print("✅ Warehouse Items API loaded")

# PŘIDEJ DO main.py před 
# ========================================
# ============ WAREHOUSE <-> JOBS INTEGRATION ============

@app.route("/api/warehouse/search", methods=["GET"])
def api_warehouse_search():
    """Vyhledání položek ve skladu pro autocomplete"""
    u, err = require_auth()
    if err: return err
    
    try:
        db = get_db()
        query = request.args.get('q', '').strip()
        
        if not query or len(query) < 2:
            return jsonify({"items": []})
        
        # Vyhledej položky podle názvu
        sql = """
            SELECT 
                id, name, sku, category, qty, unit, price, location,
                reserved_qty,
                (qty - COALESCE(reserved_qty, 0)) as available_qty
            FROM warehouse_items 
            WHERE status = 'active' 
            AND (name LIKE ? OR sku LIKE ? OR category LIKE ?)
            ORDER BY name
            LIMIT 20
        """
        
        pattern = f"%{query}%"
        items = db.execute(sql, (pattern, pattern, pattern)).fetchall()
        
        result = []
        for item in items:
            result.append({
                "id": item["id"],
                "name": item["name"],
                "sku": item["sku"],
                "category": item["category"],
                "qty": item["qty"],
                "unit": item["unit"],
                "price": item["price"],
                "location": item["location"],
                "reserved_qty": item["reserved_qty"] or 0,
                "available_qty": item["available_qty"] or item["qty"]
            })
        
        return jsonify({"items": result})
    except Exception as e:
        print(f"[ERROR] warehouse_search: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/jobs/<int:job_id>/materials/reserve", methods=["POST"])
def api_job_reserve_material(job_id):
    """Přidání materiálu do zakázky s rezervací ze skladu"""
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    
    try:
        db = get_db()
        data = request.get_json()
        
        warehouse_item_id = data.get("warehouse_item_id")
        qty = float(data.get("qty", 0))
        
        if not warehouse_item_id or qty <= 0:
            return jsonify({"error": "Invalid data"}), 400
        
        # Načti položku ze skladu
        item = db.execute("""
            SELECT id, name, unit, qty, reserved_qty, location, price 
            FROM warehouse_items 
            WHERE id = ? AND status = 'active'
        """, (warehouse_item_id,)).fetchone()
        
        if not item:
            return jsonify({"error": "Item not found"}), 404
        
        available = (item["qty"] or 0) - (item["reserved_qty"] or 0)
        if available < qty:
            return jsonify({"error": f"Not enough in stock. Available: {available} {item['unit']}"}), 400
        
        # Přidej do job_materials
        db.execute("""
            INSERT INTO job_materials 
            (job_id, name, qty, unit, warehouse_item_id, reserved_qty, warehouse_location)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (job_id, item["name"], qty, item["unit"], item["id"], qty, item["location"]))
        
        # Aktualizuj rezervaci ve skladu
        new_reserved = (item["reserved_qty"] or 0) + qty
        db.execute("""
            UPDATE warehouse_items 
            SET reserved_qty = ?, updated_at = datetime('now')
            WHERE id = ?
        """, (new_reserved, item["id"]))
        
        db.commit()
        
        return jsonify({"ok": True, "message": "Material reserved successfully"})
    except Exception as e:
        db.rollback()
        print(f"[ERROR] reserve_material: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/jobs/<int:job_id>/materials/<int:material_id>", methods=["PATCH"])
def api_job_material_update(job_id, material_id):
    """Úprava materiálu v zakázce (množství, cena, dodavatel, status)"""
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    
    try:
        db = get_db()
        data = request.get_json()
        
        # Načti aktuální materiál
        material = db.execute("""
            SELECT * FROM job_materials 
            WHERE id = ? AND job_id = ?
        """, (material_id, job_id)).fetchone()
        
        if not material:
            return jsonify({"error": "Material not found"}), 404
        
        # Připrav update
        updates = []
        params = []
        
        # Možné fieldy k updatu
        if "quantity" in data:
            new_qty = float(data["quantity"])
            if new_qty < 0:
                return jsonify({"error": "Quantity cannot be negative"}), 400
            
            # Pokud je materiál rezervovaný ze skladu, zkontroluj dostupnost
            if material["warehouse_item_id"]:
                warehouse_item = db.execute("""
                    SELECT qty, reserved_qty FROM warehouse_items WHERE id = ?
                """, (material["warehouse_item_id"],)).fetchone()
                
                if warehouse_item:
                    # Dostupné = celkové - (rezervované - toto rezervování)
                    available = warehouse_item["qty"] - (warehouse_item["reserved_qty"] - material["qty"])
                    if new_qty > available:
                        return jsonify({
                            "error": f"Not enough in stock. Available: {available:.2f}"
                        }), 400
            
            updates.append("qty = ?")
            params.append(new_qty)
            
            # Update také reserved_qty pokud je materiál ze skladu
            if material["warehouse_item_id"]:
                updates.append("reserved_qty = ?")
                params.append(new_qty)
        
        if "price_per_unit" in data:
            updates.append("price_per_unit = ?")
            params.append(float(data["price_per_unit"]))
        
        if "supplier" in data:
            updates.append("supplier = ?")
            params.append(data["supplier"].strip())
        
        if "status" in data:
            new_status = data["status"]
            if new_status not in ['planned', 'ordered', 'delivered', 'used']:
                return jsonify({"error": "Invalid status"}), 400
            updates.append("status = ?")
            params.append(new_status)
        
        if not updates:
            return jsonify({"error": "No fields to update"}), 400
        
        # Proved update
        updates.append("updated_at = datetime('now')")
        params.extend([material_id, job_id])
        
        sql = f"""
            UPDATE job_materials 
            SET {', '.join(updates)}
            WHERE id = ? AND job_id = ?
        """
        
        db.execute(sql, params)
        
        # Spočítej total_price
        if "quantity" in data or "price_per_unit" in data:
            db.execute("""
                UPDATE job_materials
                SET total_price = qty * price_per_unit
                WHERE id = ? AND job_id = ?
            """, (material_id, job_id))
        
        db.commit()
        
        return jsonify({"ok": True, "message": "Material updated successfully"})
    except Exception as e:
        db.rollback()
        print(f"[ERROR] update_material: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/jobs/<int:job_id>/materials/<int:material_id>", methods=["DELETE"])
def api_job_material_delete(job_id, material_id):
    """Smazání materiálu ze zakázky (automaticky uvolní rezervaci ze skladu)"""
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    
    try:
        db = get_db()
        
        # Načti materiál (pro kontrolu)
        material = db.execute("""
            SELECT warehouse_item_id, reserved_qty, status 
            FROM job_materials 
            WHERE id = ? AND job_id = ?
        """, (material_id, job_id)).fetchone()
        
        if not material:
            return jsonify({"error": "Material not found"}), 404
        
        # Smaž materiál (trigger automaticky upraví reserved_qty ve warehouse_items)
        db.execute("""
            DELETE FROM job_materials 
            WHERE id = ? AND job_id = ?
        """, (material_id, job_id))
        
        db.commit()
        
        message = "Material deleted"
        if material["warehouse_item_id"]:
            message += " and reservation released"
        
        return jsonify({"ok": True, "message": message})
    except Exception as e:
        db.rollback()
        print(f"[ERROR] delete_material: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@app.route("/api/jobs/<int:job_id>/materials/<int:material_id>/release", methods=["POST"])
def api_job_release_material(job_id, material_id):
    """Uvolnění materiálu z zakázky zpět do skladu"""
    u, err = require_auth()
    if err: return err
    if u["role"] not in WRITE_ROLES:
        return jsonify({"error": "Forbidden"}), 403
    
    try:
        db = get_db()
        
        # Načti materiál
        material = db.execute("""
            SELECT warehouse_item_id, reserved_qty 
            FROM job_materials 
            WHERE id = ? AND job_id = ?
        """, (material_id, job_id)).fetchone()
        
        if not material:
            return jsonify({"error": "Material not found"}), 404
        
        # Pokud je propojený se skladem, uvolni rezervaci
        if material["warehouse_item_id"]:
            db.execute("""
                UPDATE warehouse_items 
                SET reserved_qty = reserved_qty - ?,
                    updated_at = datetime('now')
                WHERE id = ?
            """, (material["reserved_qty"] or 0, material["warehouse_item_id"]))
        
        # Smaž z job_materials
        db.execute("DELETE FROM job_materials WHERE id = ? AND job_id = ?", (material_id, job_id))
        
        db.commit()
        return jsonify({"ok": True, "message": "Material released successfully"})
    except Exception as e:
        db.rollback()
        print(f"[ERROR] release_material: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/api/warehouse/items/<int:item_id>/reservations", methods=["GET"])
def api_warehouse_item_reservations(item_id):
    """Zobrazení všech rezervací pro položku skladu"""
    u, err = require_auth()
    if err: return err
    
    try:
        db = get_db()
        
        reservations = db.execute("""
            SELECT 
                jm.id, jm.job_id, jm.qty, jm.reserved_qty,
                j.title as job_title, j.code as job_code, j.status as job_status
            FROM job_materials jm
            JOIN jobs j ON jm.job_id = j.id
            WHERE jm.warehouse_item_id = ?
            ORDER BY j.date DESC
        """, (item_id,)).fetchall()
        
        result = []
        for r in reservations:
            result.append({
                "id": r["id"],
                "job_id": r["job_id"],
                "job_title": r["job_title"],
                "job_code": r["job_code"],
                "job_status": r["job_status"],
                "qty": r["qty"],
                "reserved_qty": r["reserved_qty"]
            })
        
        return jsonify({"reservations": result})
    except Exception as e:
        print(f"[ERROR] get_reservations: {e}")
        return jsonify({"error": str(e)}), 500



# ========== PLANT CATALOG API ==========

@app.route('/api/plant-catalog/search', methods=['GET'])
def api_plant_catalog_search():
    """Vyhledávání v katalogu rostlin (autocomplete)"""
    u, err = require_auth()
    if err: return err
    
    query = request.args.get('q', '').strip()
    limit = request.args.get('limit', 20, type=int)
    
    if not query or len(query) < 2:
        return jsonify({
            'success': False,
            'message': 'Zadej alespoň 2 znaky pro vyhledávání'
        }), 400
    
    try:
        db = get_db()
        
        # Pokus o FTS, pokud není dostupné, použij LIKE
        try:
            results = db.execute('''
                SELECT pc.id, pc.latin_name, pc.variety, pc.container_size,
                       pc.flower_color, pc.flowering_time, pc.height,
                       pc.light_requirements, pc.hardiness_zone
                FROM plant_catalog_fts fts
                JOIN plant_catalog pc ON pc.id = fts.rowid
                WHERE plant_catalog_fts MATCH ?
                ORDER BY rank
                LIMIT ?
            ''', (f'{query}*', limit)).fetchall()
        except Exception as fts_error:
            # FTS není dostupné, použij LIKE fallback
            print(f"[INFO] FTS not available, using LIKE: {fts_error}")
            search_pattern = f'%{query}%'
            results = db.execute('''
                SELECT id, latin_name, variety, container_size,
                       flower_color, flowering_time, height,
                       light_requirements, hardiness_zone
                FROM plant_catalog
                WHERE latin_name LIKE ? OR variety LIKE ? OR notes LIKE ?
                ORDER BY latin_name
                LIMIT ?
            ''', (search_pattern, search_pattern, search_pattern, limit)).fetchall()
        
        plants = []
        for row in results:
            full_name = row['latin_name']
            if row['variety']:
                full_name += f" '{row['variety']}'"
            if row['container_size']:
                full_name += f" - {row['container_size']}"
            
            plants.append({
                'id': row['id'],
                'full_name': full_name,
                'latin_name': row['latin_name'],
                'variety': row['variety'],
                'container_size': row['container_size'],
                'flower_color': row['flower_color'],
                'flowering_time': row['flowering_time'],
                'height': row['height'],
                'light_requirements': row['light_requirements'],
                'hardiness_zone': row['hardiness_zone']
            })
        
        return jsonify({
            'success': True,
            'plants': plants,
            'count': len(plants)
        })
        
    except Exception as e:
        print(f"[ERROR] plant_catalog_search: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Chyba při vyhledávání: {str(e)}'
        }), 500


@app.route('/api/plant-catalog/<int:plant_id>', methods=['GET'])
def api_plant_catalog_detail(plant_id):
    """Detail rostliny z katalogu"""
    u, err = require_auth()
    if err: return err
    
    try:
        db = get_db()
        plant = db.execute('''
            SELECT * FROM plant_catalog WHERE id = ?
        ''', (plant_id,)).fetchone()
        
        if not plant:
            return jsonify({
                'success': False,
                'message': 'Rostlina nenalezena'
            }), 404
        
        return jsonify({
            'success': True,
            'plant': dict(plant)
        })
        
    except Exception as e:
        print(f"[ERROR] plant_catalog_detail: {e}")
        return jsonify({
            'success': False,
            'message': f'Chyba při načítání: {str(e)}'
        }), 500


@app.route('/api/plant-catalog/stats', methods=['GET'])
def api_plant_catalog_stats():
    """Statistiky katalogu"""
    u, err = require_auth()
    if err: return err
    
    try:
        db = get_db()
        
        stats = db.execute('''
            SELECT 
                COUNT(*) as total_plants,
                COUNT(DISTINCT latin_name) as species_count,
                COUNT(DISTINCT CASE WHEN variety IS NOT NULL THEN latin_name END) as varieties_count
            FROM plant_catalog
        ''').fetchone()
        
        return jsonify({
            'success': True,
            'stats': dict(stats)
        })
        
    except Exception as e:
        print(f"[ERROR] plant_catalog_stats: {e}")
        return jsonify({
            'success': False,
            'message': f'Chyba: {str(e)}'
        }), 500


@app.route('/api/plant-catalog/by-name', methods=['GET'])
def api_plant_catalog_by_name():
    """Najdi rostlinu přesně podle názvu a odrůdy"""
    u, err = require_auth()
    if err: return err
    
    latin = request.args.get('latin', '').strip()
    variety = request.args.get('variety', '').strip() or None
    
    if not latin:
        return jsonify({
            'success': False,
            'message': 'Chybí latinský název'
        }), 400
    
    try:
        db = get_db()
        
        if variety:
            plant = db.execute('''
                SELECT * FROM plant_catalog 
                WHERE latin_name = ? AND variety = ?
                LIMIT 1
            ''', (latin, variety)).fetchone()
        else:
            plant = db.execute('''
                SELECT * FROM plant_catalog 
                WHERE latin_name = ? AND variety IS NULL
                LIMIT 1
            ''', (latin,)).fetchone()
        
        if plant:
            return jsonify({
                'success': True,
                'plant': dict(plant)
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Rostlina v katalogu nenalezena'
            }), 404
            
    except Exception as e:
        print(f"[ERROR] plant_catalog_by_name: {e}")
        return jsonify({
            'success': False,
            'message': f'Chyba: {str(e)}'
        }), 500


# ========== NURSERY PLANTS API ==========

@app.route('/api/nursery/plants', methods=['POST'])
def api_nursery_add_plant():
    """Přidání nové rostliny do školky"""
    u, err = require_auth()
    if err: return err
    
    if u["role"] not in ["owner", "manager", "employee"]:
        return jsonify({"error": "Forbidden"}), 403
    
    try:
        data = request.get_json()
        
        # Validace povinných polí
        if not data.get('species'):
            return jsonify({
                'success': False,
                'message': 'Chybí název rostliny (species)'
            }), 400
        
        if not data.get('quantity'):
            return jsonify({
                'success': False,
                'message': 'Chybí počet kusů (quantity)'
            }), 400
        
        db = get_db()
        
        # Vlož rostlinu do databáze
        cursor = db.execute('''
            INSERT INTO nursery_plants (
                species, variety, quantity, unit, stage, location,
                planted_date, ready_date, selling_price, purchase_price,
                notes, flower_color, flowering_time, height,
                light_requirements, leaf_color, hardiness_zone,
                site_type, plants_per_m2, botanical_notes, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data.get('species'),
            data.get('variety'),
            data.get('quantity', 0),
            data.get('unit', 'ks'),
            data.get('stage', 'sazenice'),
            data.get('location'),
            data.get('planted_date'),
            data.get('ready_date'),
            data.get('selling_price'),
            data.get('purchase_price'),
            data.get('notes'),
            data.get('flower_color'),
            data.get('flowering_time'),
            data.get('height'),
            data.get('light_requirements'),
            data.get('leaf_color'),
            data.get('hardiness_zone'),
            data.get('site_type'),
            data.get('plants_per_m2'),
            data.get('botanical_notes'),
            'active'
        ))
        
        plant_id = cursor.lastrowid
        db.commit()
        
        return jsonify({
            'success': True,
            'message': 'Rostlina byla úspěšně přidána',
            'plant_id': plant_id
        })
        
    except Exception as e:
        print(f"[ERROR] api_nursery_add_plant: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Chyba při ukládání: {str(e)}'
        }), 500


@app.route('/api/nursery/plants', methods=['GET'])
def api_nursery_list_plants():
    """Seznam všech rostlin ve školce"""
    u, err = require_auth()
    if err: return err
    
    try:
        db = get_db()
        
        # Filtry z query parametrů
        stage = request.args.get('stage')
        status = request.args.get('status', 'active')
        location = request.args.get('location')
        
        query = 'SELECT * FROM nursery_plants WHERE status = ?'
        params = [status]
        
        if stage:
            query += ' AND stage = ?'
            params.append(stage)
        
        if location:
            query += ' AND location = ?'
            params.append(location)
        
        query += ' ORDER BY created_at DESC'
        
        plants = db.execute(query, params).fetchall()
        
        result = []
        for p in plants:
            result.append({
                'id': p['id'],
                'species': p['species'],
                'variety': p['variety'],
                'quantity': p['quantity'],
                'unit': p['unit'],
                'stage': p['stage'],
                'location': p['location'],
                'planted_date': p['planted_date'],
                'selling_price': p['selling_price'],
                'flower_color': p['flower_color'],
                'flowering_time': p['flowering_time'],
                'height': p['height'],
                'created_at': p['created_at']
            })
        
        return jsonify({
            'success': True,
            'plants': result,
            'count': len(result)
        })
        
    except Exception as e:
        print(f"[ERROR] api_nursery_list_plants: {e}")
        return jsonify({
            'success': False,
            'message': f'Chyba: {str(e)}'
        }), 500


@app.route('/api/nursery/plants/<int:plant_id>', methods=['GET'])
def api_nursery_get_plant(plant_id):
    """Detail rostliny"""
    u, err = require_auth()
    if err: return err
    
    try:
        db = get_db()
        plant = db.execute('SELECT * FROM nursery_plants WHERE id = ?', (plant_id,)).fetchone()
        
        if not plant:
            return jsonify({
                'success': False,
                'message': 'Rostlina nenalezena'
            }), 404
        
        return jsonify({
            'success': True,
            'plant': dict(plant)
        })
        
    except Exception as e:
        print(f"[ERROR] api_nursery_get_plant: {e}")
        return jsonify({
            'success': False,
            'message': f'Chyba: {str(e)}'
        }), 500


@app.route('/api/nursery/plants/<int:plant_id>', methods=['PUT'])
def api_nursery_update_plant(plant_id):
    """Aktualizace rostliny"""
    u, err = require_auth()
    if err: return err
    
    if u["role"] not in ["owner", "manager", "employee"]:
        return jsonify({"error": "Forbidden"}), 403
    
    try:
        data = request.get_json()
        db = get_db()
        
        # Zkontroluj, že rostlina existuje
        plant = db.execute('SELECT id FROM nursery_plants WHERE id = ?', (plant_id,)).fetchone()
        if not plant:
            return jsonify({
                'success': False,
                'message': 'Rostlina nenalezena'
            }), 404
        
        # Aktualizuj rostlinu
        db.execute('''
            UPDATE nursery_plants SET
                species = ?, variety = ?, quantity = ?, stage = ?,
                location = ?, planted_date = ?, selling_price = ?,
                notes = ?, flower_color = ?, flowering_time = ?,
                height = ?, light_requirements = ?, leaf_color = ?,
                hardiness_zone = ?, site_type = ?, plants_per_m2 = ?,
                botanical_notes = ?
            WHERE id = ?
        ''', (
            data.get('species'),
            data.get('variety'),
            data.get('quantity'),
            data.get('stage'),
            data.get('location'),
            data.get('planted_date'),
            data.get('selling_price'),
            data.get('notes'),
            data.get('flower_color'),
            data.get('flowering_time'),
            data.get('height'),
            data.get('light_requirements'),
            data.get('leaf_color'),
            data.get('hardiness_zone'),
            data.get('site_type'),
            data.get('plants_per_m2'),
            data.get('botanical_notes'),
            plant_id
        ))
        
        db.commit()
        
        return jsonify({
            'success': True,
            'message': 'Rostlina byla aktualizována'
        })
        
    except Exception as e:
        print(f"[ERROR] api_nursery_update_plant: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Chyba: {str(e)}'
        }), 500


@app.route('/api/nursery/plants/<int:plant_id>', methods=['DELETE'])
def api_nursery_delete_plant(plant_id):
    """Smazání rostliny (soft delete)"""
    u, err = require_auth()
    if err: return err
    
    if u["role"] not in ["owner", "manager"]:
        return jsonify({"error": "Forbidden"}), 403
    
    try:
        db = get_db()
        
        # Soft delete - změň status na 'deleted'
        result = db.execute(
            'UPDATE nursery_plants SET status = ? WHERE id = ?',
            ('deleted', plant_id)
        )
        
        if result.rowcount == 0:
            return jsonify({
                'success': False,
                'message': 'Rostlina nenalezena'
            }), 404
        
        db.commit()
        
        return jsonify({
            'success': True,
            'message': 'Rostlina byla smazána'
        })
        
    except Exception as e:
        print(f"[ERROR] api_nursery_delete_plant: {e}")
        return jsonify({
            'success': False,
            'message': f'Chyba: {str(e)}'
        }), 500


# ==================== WEATHER API ====================
@app.route('/api/weather', methods=['GET'])
def api_weather():
    """Získá aktuální počasí pro zadané město nebo GPS koordináty (default: Příbram)"""
    u, err = require_auth()
    if err: return err
    
    # Podpora GPS koordinátů (priorita) nebo města
    lat = request.args.get('lat')
    lon = request.args.get('lon')
    city = request.args.get('city', 'Příbram')
    source = request.args.get('source', 'fallback')  # 'gps', 'manual', 'fallback'
    
    # Open-Meteo API - free, no API key required
    # Koordináty pro Příbram, CZ (fallback)
    locations = {
        'Příbram': {'lat': 49.6897, 'lon': 14.0101},
        'Praha': {'lat': 50.0755, 'lon': 14.4378},
        'Brno': {'lat': 49.1951, 'lon': 16.6068},
        'Ostrava': {'lat': 49.8209, 'lon': 18.2625},
        'Plzeň': {'lat': 49.7384, 'lon': 13.3736},
    }
    
    # Použít GPS koordináty pokud jsou poskytnuty, jinak město
    if lat and lon:
        try:
            coords = {'lat': float(lat), 'lon': float(lon)}
            # Pokud není city poskytnut, použít z locations nebo default
            if not city or city == 'Příbram':
                # Najít nejbližší město nebo použít GPS souřadnice
                city = city if city != 'Příbram' else 'GPS lokace'
        except ValueError:
            coords = locations.get(city, locations['Příbram'])
            source = 'fallback'
    else:
        coords = locations.get(city, locations['Příbram'])
        source = 'fallback' if city == 'Příbram' else 'manual'
    
    try:
        import urllib.request
        import urllib.error
        
        url = f"https://api.open-meteo.com/v1/forecast?latitude={coords['lat']}&longitude={coords['lon']}&current=temperature_2m,relative_humidity_2m,weather_code,wind_speed_10m&daily=weather_code,temperature_2m_max,temperature_2m_min,precipitation_probability_max&timezone=Europe%2FPrague&forecast_days=5"
        
        req = urllib.request.Request(url, headers={'User-Agent': 'GreenDavidApp/1.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            data = json.loads(response.read().decode('utf-8'))
        
        # Weather code to description and icon
        weather_codes = {
            0: {'desc': 'Jasno', 'icon': '☀️'},
            1: {'desc': 'Převážně jasno', 'icon': '🌤️'},
            2: {'desc': 'Polojasno', 'icon': '⛅'},
            3: {'desc': 'Zataženo', 'icon': '☁️'},
            45: {'desc': 'Mlha', 'icon': '🌫️'},
            48: {'desc': 'Mrznoucí mlha', 'icon': '🌫️'},
            51: {'desc': 'Mrholení', 'icon': '🌧️'},
            53: {'desc': 'Mrholení', 'icon': '🌧️'},
            55: {'desc': 'Mrholení', 'icon': '🌧️'},
            61: {'desc': 'Déšť', 'icon': '🌧️'},
            63: {'desc': 'Déšť', 'icon': '🌧️'},
            65: {'desc': 'Silný déšť', 'icon': '🌧️'},
            71: {'desc': 'Sněžení', 'icon': '🌨️'},
            73: {'desc': 'Sněžení', 'icon': '🌨️'},
            75: {'desc': 'Husté sněžení', 'icon': '🌨️'},
            80: {'desc': 'Přeháňky', 'icon': '🌦️'},
            81: {'desc': 'Přeháňky', 'icon': '🌦️'},
            82: {'desc': 'Silné přeháňky', 'icon': '🌦️'},
            95: {'desc': 'Bouřka', 'icon': '⛈️'},
            96: {'desc': 'Bouřka s krupobitím', 'icon': '⛈️'},
            99: {'desc': 'Bouřka s krupobitím', 'icon': '⛈️'},
        }
        
        current = data.get('current', {})
        daily = data.get('daily', {})
        
        weather_code = current.get('weather_code', 0)
        weather_info = weather_codes.get(weather_code, {'desc': 'Neznámo', 'icon': '❓'})
        
        # Forecast for next days
        forecast = []
        if daily:
            for i in range(min(5, len(daily.get('time', [])))):
                day_code = daily['weather_code'][i] if i < len(daily.get('weather_code', [])) else 0
                day_info = weather_codes.get(day_code, {'desc': 'Neznámo', 'icon': '❓'})
                forecast.append({
                    'date': daily['time'][i],
                    'temp_max': daily['temperature_2m_max'][i] if i < len(daily.get('temperature_2m_max', [])) else None,
                    'temp_min': daily['temperature_2m_min'][i] if i < len(daily.get('temperature_2m_min', [])) else None,
                    'precipitation_prob': daily['precipitation_probability_max'][i] if i < len(daily.get('precipitation_probability_max', [])) else 0,
                    'description': day_info['desc'],
                    'icon': day_info['icon']
                })
        
        # Gardening advice based on weather
        advice = []
        temp = current.get('temperature_2m', 15)
        humidity = current.get('relative_humidity_2m', 50)
        wind = current.get('wind_speed_10m', 0)
        
        if weather_code in [61, 63, 65, 80, 81, 82]:
            advice.append("🌧️ Dnes prší - ideální den pro práci ve skleníku")
        elif weather_code in [71, 73, 75]:
            advice.append("🌨️ Sněží - pozor na mráz, zkontrolujte ochranu rostlin")
        elif weather_code in [0, 1] and temp > 20:
            advice.append("☀️ Horký den - nezapomeňte zavlažovat")
        
        if temp < 5:
            advice.append("🥶 Nízká teplota - pozor na mráz citlivých rostlin")
        elif temp > 30:
            advice.append("🌡️ Tropické teploty - ranní nebo večerní práce doporučena")
        
        if wind > 30:
            advice.append("💨 Silný vítr - odložte postřiky a mulčování")
        
        if humidity > 80 and temp > 15:
            advice.append("💧 Vysoká vlhkost - sledujte plísňové choroby")
        
        # Vypočítat pocitovou teplotu (wind chill)
        temp = current.get('temperature_2m', 15)
        wind_speed = current.get('wind_speed_10m', 0)
        feels_like = temp  # Zjednodušeně, Open-Meteo nemá feels_like
        if wind_speed > 0 and temp < 10:
            # Wind chill aproximace
            feels_like = round(13.12 + 0.6215 * temp - 11.37 * (wind_speed ** 0.16) + 0.3965 * temp * (wind_speed ** 0.16))
        
        # Generovat alerty pro zahradníky
        alerts = []
        if temp <= 0:
            alerts.append({
                'type': 'frost',
                'icon': '❄️',
                'text': f'Mráz ({round(temp)}°C) - nevhodné pro výsadbu!'
            })
        if weather_code in [61, 63, 65, 80, 81, 82]:
            alerts.append({
                'type': 'rain',
                'icon': '🌧️',
                'text': 'Déšť - zvažte přesunutí venkovních prací'
            })
        if wind_speed * 3.6 > 40:
            alerts.append({
                'type': 'wind',
                'icon': '💨',
                'text': f'Silný vítr ({round(wind_speed * 3.6)} km/h) - pozor na práci ve výškách'
            })
        if temp >= 30:
            alerts.append({
                'type': 'heat',
                'icon': '🥵',
                'text': 'Extrémní vedro - zajistěte dostatek vody!'
            })
        
        return jsonify({
            'success': True,
            'city': city,
            'source': source,  # 'gps', 'manual', 'fallback'
            'coordinates': coords,
            'current': {
                'temperature': round(temp),
                'feels_like': round(feels_like),
                'humidity': current.get('relative_humidity_2m', 50),
                'wind_speed': round(wind_speed * 3.6),  # m/s -> km/h
                'weather_code': weather_code,
                'description': weather_info['desc'],
                'icon': weather_info['icon']
            },
            'forecast': forecast,
            'advice': advice,
            'alerts': alerts
        })
        
    except urllib.error.URLError as e:
        print(f"[WARNING] Weather API network error: {e}")
        return jsonify({
            'success': False,
            'message': 'Nelze načíst počasí - zkontrolujte připojení',
            'offline': True
        }), 503
    except Exception as e:
        print(f"[ERROR] api_weather: {e}")
        return jsonify({
            'success': False,
            'message': f'Chyba při načítání počasí: {str(e)}'
        }), 500


# ==================== DASHBOARD STATS API ====================
@app.route('/api/dashboard/stats', methods=['GET'])
def api_dashboard_stats():
    """Rychlé statistiky pro dashboard"""
    u, err = require_auth()
    if err: return err
    
    try:
        db = get_db()
        today = date.today().isoformat()
        week_start = (date.today() - timedelta(days=date.today().weekday())).isoformat()
        
        # Active jobs
        active_jobs = db.execute(
            "SELECT COUNT(*) as cnt FROM jobs WHERE status IN ('Plán', 'Probíhá')"
        ).fetchone()['cnt']
        
        # Pending tasks
        pending_tasks = db.execute(
            "SELECT COUNT(*) as cnt FROM tasks WHERE status IN ('open', 'in-progress')"
        ).fetchone()['cnt']
        
        # Urgent tasks (due today)
        urgent_tasks = db.execute(
            "SELECT COUNT(*) as cnt FROM tasks WHERE due_date = ? AND status != 'completed'",
            (today,)
        ).fetchone()['cnt']
        
        # Hours this week
        hours_week = db.execute(
            "SELECT COALESCE(SUM(hours), 0) as total FROM timesheets WHERE date >= ?",
            (week_start,)
        ).fetchone()['total']
        
        # Employees count
        emp_count = db.execute("SELECT COUNT(*) as cnt FROM employees").fetchone()['cnt']
        
        # Low stock items
        low_stock = db.execute(
            "SELECT COUNT(*) as cnt FROM warehouse_items WHERE quantity <= min_quantity AND min_quantity > 0"
        ).fetchone()['cnt']
        
        return jsonify({
            'success': True,
            'stats': {
                'active_jobs': active_jobs,
                'pending_tasks': pending_tasks,
                'urgent_tasks': urgent_tasks,
                'hours_week': round(hours_week, 1),
                'employee_count': emp_count,
                'low_stock_items': low_stock
            }
        })
        
    except Exception as e:
        print(f"[ERROR] api_dashboard_stats: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


# ==================== DEADLINE CHECK & NOTIFICATIONS ====================
@app.route('/api/check-deadlines', methods=['POST'])
def api_check_deadlines():
    """Check for upcoming deadlines and create notifications.
    Can be called periodically (e.g., once per day via cron or manual trigger).
    """
    u, err = require_auth()
    if err: return err
    
    # Only admin/owner can trigger this
    if normalize_role(u.get("role")) not in ("owner", "admin"):
        return jsonify({"ok": False, "error": "forbidden"}), 403
    
    try:
        db = get_db()
        today = date.today()
        tomorrow = today + timedelta(days=1)
        
        notifications_created = 0
        
        # Tasks due today
        tasks_today = db.execute("""
            SELECT t.id, t.title, t.employee_id, ta.employee_id as assigned_emp
            FROM tasks t
            LEFT JOIN task_assignees ta ON t.id = ta.task_id
            WHERE t.due_date = ? AND t.status != 'completed'
        """, (today.isoformat(),)).fetchall()
        
        for task in tasks_today:
            emp_id = task['assigned_emp'] or task['employee_id']
            if emp_id:
                create_notification(
                    employee_id=emp_id,
                    kind="deadline",
                    title="⏰ Deadline dnes!",
                    body=f"Úkol '{task['title']}' má deadline dnes",
                    entity_type="task",
                    entity_id=task['id']
                )
                notifications_created += 1
        
        # Tasks due tomorrow
        tasks_tomorrow = db.execute("""
            SELECT t.id, t.title, t.employee_id, ta.employee_id as assigned_emp
            FROM tasks t
            LEFT JOIN task_assignees ta ON t.id = ta.task_id
            WHERE t.due_date = ? AND t.status != 'completed'
        """, (tomorrow.isoformat(),)).fetchall()
        
        for task in tasks_tomorrow:
            emp_id = task['assigned_emp'] or task['employee_id']
            if emp_id:
                create_notification(
                    employee_id=emp_id,
                    kind="deadline",
                    title="📅 Deadline zítra",
                    body=f"Úkol '{task['title']}' má deadline zítra",
                    entity_type="task",
                    entity_id=task['id']
                )
                notifications_created += 1
        
        # Jobs ending soon (within 3 days)
        soon = today + timedelta(days=3)
        jobs_soon = db.execute("""
            SELECT id, title, date FROM jobs
            WHERE date BETWEEN ? AND ? AND status != 'Dokončeno'
        """, (today.isoformat(), soon.isoformat())).fetchall()
        
        # Notify all admins about jobs ending soon
        admins = db.execute("SELECT id FROM users WHERE role IN ('owner', 'admin')").fetchall()
        for job in jobs_soon:
            for admin in admins:
                create_notification(
                    user_id=admin['id'],
                    kind="job",
                    title="🏗️ Zakázka končí brzy",
                    body=f"Zakázka '{job['title']}' má termín {job['date']}",
                    entity_type="job",
                    entity_id=job['id']
                )
                notifications_created += 1
        
        # Low stock alerts
        low_stock = db.execute("""
            SELECT id, name, quantity, min_quantity FROM warehouse_items
            WHERE quantity <= min_quantity AND min_quantity > 0
        """).fetchall()
        
        for item in low_stock:
            for admin in admins:
                create_notification(
                    user_id=admin['id'],
                    kind="stock",
                    title="📦 Nízký stav skladu",
                    body=f"'{item['name']}' - zbývá {item['quantity']} ks (min: {item['min_quantity']})",
                    entity_type="warehouse",
                    entity_id=item['id']
                )
                notifications_created += 1
        
        return jsonify({
            'ok': True,
            'notifications_created': notifications_created
        })
        
    except Exception as e:
        print(f"[ERROR] api_check_deadlines: {e}")
        return jsonify({
            'ok': False,
            'error': str(e)
        }), 500


# ==================== QUICK ACTIONS API ====================
@app.route('/api/quick-add', methods=['POST'])
def api_quick_add():
    """Rychlé přidání položky (task, note, timesheet) z dashboardu"""
    u, err = require_auth()
    if err: return err
    
    try:
        data = request.get_json(force=True, silent=True) or {}
        item_type = data.get('type', 'task')
        
        db = get_db()
        
        if item_type == 'task':
            title = (data.get('title') or '').strip()
            if not title:
                return jsonify({'ok': False, 'error': 'Zadejte název úkolu'}), 400
            
            db.execute("""
                INSERT INTO tasks(title, status, created_by, created_at, updated_at)
                VALUES (?, 'open', ?, datetime('now'), datetime('now'))
            """, (title, u['id']))
            db.commit()
            task_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
            
            return jsonify({'ok': True, 'id': task_id, 'type': 'task'})
        
        elif item_type == 'timesheet':
            hours = data.get('hours')
            job_id = data.get('job_id')
            
            if not hours:
                return jsonify({'ok': False, 'error': 'Zadejte počet hodin'}), 400
            
            # Get employee_id for current user
            emp = db.execute("SELECT id FROM employees WHERE user_id = ?", (u['id'],)).fetchone()
            emp_id = emp['id'] if emp else None
            
            db.execute("""
                INSERT INTO timesheets(employee_id, job_id, date, hours, note, created_by)
                VALUES (?, ?, date('now'), ?, ?, ?)
            """, (emp_id, job_id, float(hours), data.get('note', ''), u['id']))
            db.commit()
            
            return jsonify({'ok': True, 'type': 'timesheet'})
        
        return jsonify({'ok': False, 'error': 'Neznámý typ'}), 400
        
    except Exception as e:
        print(f"[ERROR] api_quick_add: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


# ============================================================
# GPS CHECK-IN/OUT API
# ============================================================

@app.route('/api/gps/checkin', methods=['POST'])
def api_gps_checkin():
    """Record GPS check-in to a job."""
    user, err = require_auth()
    if err:
        return err
    
    try:
        data = request.get_json() or {}
        job_id = data.get('job_id')
        check_in_time = data.get('check_in_time')
        lat = data.get('lat')
        lng = data.get('lng')
        
        if not job_id:
            return jsonify({'ok': False, 'error': 'job_id required'}), 400
        
        db = get_db()
        
        # Create gps_logs table if not exists
        db.execute('''CREATE TABLE IF NOT EXISTS gps_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            job_id INTEGER,
            check_in_time TEXT,
            check_out_time TEXT,
            check_in_lat REAL,
            check_in_lng REAL,
            check_out_lat REAL,
            check_out_lng REAL,
            hours_worked REAL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )''')
        
        db.execute('''
            INSERT INTO gps_logs (user_id, job_id, check_in_time, check_in_lat, check_in_lng)
            VALUES (?, ?, ?, ?, ?)
        ''', [session.get('user_id'), job_id, check_in_time, lat, lng])
        db.commit()
        
        return jsonify({'ok': True, 'message': 'Check-in recorded'})
    except Exception as e:
        print(f"[ERROR] api_gps_checkin: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/api/gps/checkout', methods=['POST'])
def api_gps_checkout():
    """Record GPS check-out and create timesheet entry."""
    user, err = require_auth()
    if err:
        return err
    
    try:
        data = request.get_json() or {}
        job_id = data.get('job_id')
        check_in_time = data.get('check_in_time')
        check_out_time = data.get('check_out_time')
        hours_worked = data.get('hours_worked', 0)
        lat = data.get('lat')
        lng = data.get('lng')
        
        db = get_db()
        
        # Update GPS log
        db.execute('''
            UPDATE gps_logs 
            SET check_out_time = ?, check_out_lat = ?, check_out_lng = ?, hours_worked = ?
            WHERE job_id = ? AND user_id = ? AND check_out_time IS NULL
            ORDER BY id DESC LIMIT 1
        ''', [check_out_time, lat, lng, hours_worked, job_id, session.get('user_id')])
        
        # Auto-create timesheet entry
        if hours_worked > 0:
            today = check_out_time.split('T')[0] if check_out_time else None
            db.execute('''
                INSERT INTO timesheets (job_id, employee_id, date, hours, description)
                VALUES (?, ?, ?, ?, ?)
            ''', [job_id, session.get('user_id'), today, round(hours_worked, 2), 'GPS auto-záznam'])
        
        db.commit()
        
        return jsonify({'ok': True, 'hours_logged': round(hours_worked, 2)})
    except Exception as e:
        print(f"[ERROR] api_gps_checkout: {e}")
        return jsonify({'ok': False, 'error': str(e)}), 500


# ============================================================
# CLIENT PORTAL API
# ============================================================

@app.route('/portal')
def client_portal():
    """Serve client portal page."""
    return render_template('client-portal.html')


@app.route('/api/portal/job')
def api_portal_job():
    """Get job by client code (no auth required)."""
    code = request.args.get('code', '').strip()
    
    if not code:
        return jsonify({'error': 'Code required'}), 400
    
    db = get_db()
    
    # Search by code or client name
    job = db.execute('''
        SELECT * FROM jobs 
        WHERE code = ? OR client LIKE ? OR id = ?
        LIMIT 1
    ''', [code, f'%{code}%', code if code.isdigit() else -1]).fetchone()
    
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    
    job_dict = dict(job)
    
    # Get task stats
    tasks = db.execute('''
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'done' THEN 1 ELSE 0 END) as completed
        FROM tasks WHERE job_id = ?
    ''', [job['id']]).fetchone()
    
    job_dict['tasks_total'] = tasks['total'] or 0
    job_dict['tasks_completed'] = tasks['completed'] or 0
    
    # Calculate progress
    if job_dict['tasks_total'] > 0:
        job_dict['progress'] = int((job_dict['tasks_completed'] / job_dict['tasks_total']) * 100)
    else:
        job_dict['progress'] = job.get('progress', 0) or 0
    
    return jsonify(job_dict)


@app.route('/api/portal/job/<int:job_id>/updates')
def api_portal_updates(job_id):
    """Get job updates/timeline for portal."""
    db = get_db()
    
    # Get completed tasks as updates
    tasks = db.execute('''
        SELECT title, 'Úkol dokončen' as description, 
               COALESCE(updated_at, created_at) as date
        FROM tasks 
        WHERE job_id = ? AND status = 'done'
        ORDER BY date DESC
        LIMIT 10
    ''', [job_id]).fetchall()
    
    updates = [{'title': t['title'], 'description': t['description'], 'date': t['date']} for t in tasks]
    
    # Add job notes as updates if available
    job = db.execute('SELECT note, created_at FROM jobs WHERE id = ?', [job_id]).fetchone()
    if job and job['note']:
        updates.insert(0, {
            'title': 'Poznámka k zakázce',
            'description': job['note'][:200],
            'date': job['created_at']
        })
    
    return jsonify(updates)


@app.route('/api/portal/job/<int:job_id>/photos')
def api_portal_photos(job_id):
    """Get job photos for portal."""
    db = get_db()
    
    # Check if job_photos table exists
    try:
        photos = db.execute('''
            SELECT url, thumbnail, created_at as date
            FROM job_photos
            WHERE job_id = ?
            ORDER BY created_at DESC
        ''', [job_id]).fetchall()
        
        return jsonify([dict(p) for p in photos])
    except:
        # Table doesn't exist
        return jsonify([])


# ============================================================
# AI ESTIMATE API (for future AI integration)
# ============================================================

@app.route('/api/ai/estimate', methods=['POST'])
def api_ai_estimate():
    """Generate AI estimate for a job based on description/photo."""
    user, err = require_auth()
    if err:
        return err
    
    try:
        data = request.get_json() or {}
        description = data.get('description', '')
        photo_base64 = data.get('photo')
        job_type = data.get('type', 'general')
        
        # Simple rule-based estimates (can be replaced with actual AI)
        estimates = {
            'plot': {'days': 3, 'material': 12000, 'price': 35000},
            'zahrada': {'days': 5, 'material': 8000, 'price': 45000},
            'terasa': {'days': 4, 'material': 15000, 'price': 50000},
            'general': {'days': 2, 'material': 5000, 'price': 20000}
        }
        
        # Parse description for keywords
        desc_lower = description.lower()
        estimate_type = 'general'
        
        for keyword in ['plot', 'oplocení', 'plotov']:
            if keyword in desc_lower:
                estimate_type = 'plot'
                break
        for keyword in ['zahrad', 'trávník', 'vysázení']:
            if keyword in desc_lower:
                estimate_type = 'zahrada'
                break
        for keyword in ['teras', 'dlažb', 'dřev']:
            if keyword in desc_lower:
                estimate_type = 'terasa'
                break
        
        # Extract numbers from description (e.g., "50m plot")
        import re
        numbers = re.findall(r'(\d+)\s*m', desc_lower)
        multiplier = 1
        if numbers:
            meters = int(numbers[0])
            multiplier = max(1, meters / 20)  # Base estimate is for ~20m
        
        base = estimates[estimate_type]
        
        result = {
            'type': estimate_type,
            'estimated_days': round(base['days'] * multiplier),
            'estimated_material': round(base['material'] * multiplier),
            'estimated_price': round(base['price'] * multiplier),
            'confidence': 0.7,
            'note': f'Odhad pro typ "{estimate_type}" na základě popisu'
        }
        
        return jsonify(result)
    except Exception as e:
        print(f"[ERROR] api_ai_estimate: {e}")
        return jsonify({'error': str(e)}), 500


# ============================================================
# CREW CONTROL SYSTEM API
# ============================================================

def calculate_capacity_percent(profile_dict, current_hours):
    """Vypočítá procento využití kapacity"""
    weekly_capacity = profile_dict.get('weekly_capacity_hours', 40.0)
    if weekly_capacity <= 0:
        return 0
    return min(100, (current_hours / weekly_capacity) * 100)

def calculate_capacity_status(capacity_percent):
    """Určí status kapacity na základě procenta"""
    if capacity_percent > 90:
        return 'overloaded'
    elif capacity_percent < 50:
        return 'underutilized'
    elif capacity_percent > 75:
        return 'high'
    else:
        return 'normal'

def calculate_burnout_risk(profile_dict, capacity_percent, weeks_overloaded=0):
    """Vypočítá úroveň rizika burnout"""
    if capacity_percent > 95 and weeks_overloaded >= 3:
        return 'critical'
    elif capacity_percent > 90 and weeks_overloaded >= 2:
        return 'high'
    elif capacity_percent > 85:
        return 'normal'
    else:
        return 'low'

# ============================================================
# TEAM METRICS HELPER FUNCTIONS
# ============================================================

def get_active_members_count(db):
    """Počet aktivních členů týmu"""
    result = db.execute("SELECT COUNT(*) as count FROM employees WHERE status = 'active' OR status IS NULL").fetchone()
    return result['count'] if result else 0

def get_worked_hours_this_week(db, employee_id, week_start, week_end):
    """Odpracované hodiny tento týden pro zaměstnance"""
    result = db.execute(
        """SELECT SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) / 60.0 as total 
           FROM timesheets 
           WHERE employee_id=? AND date >= ? AND date <= ?""",
        (employee_id, week_start, week_end)
    ).fetchone()
    return round(result['total'] or 0, 1) if result else 0.0

def get_average_utilization(db, week_start, week_end):
    """Průměrné vytížení týmu v procentech"""
    employees = db.execute("SELECT id FROM employees WHERE status = 'active' OR status IS NULL").fetchall()
    if not employees:
        return 0.0
    
    total_utilization = 0
    counted_members = 0
    
    for emp in employees:
        emp_id = emp['id']
        
        # Kapacita člena (default 40h/týden)
        profile = db.execute(
            "SELECT weekly_capacity_hours FROM team_member_profile WHERE employee_id = ?",
            (emp_id,)
        ).fetchone()
        capacity = 40.0
        if profile and profile['weekly_capacity_hours']:
            capacity = profile['weekly_capacity_hours']
        
        if capacity == 0:
            continue
        
        # Odpracované hodiny tento týden
        worked_hours = get_worked_hours_this_week(db, emp_id, week_start, week_end)
        
        utilization = (worked_hours / capacity) * 100 if capacity > 0 else 0
        total_utilization += utilization
        counted_members += 1
    
    if counted_members == 0:
        return 0.0
    
    return round(total_utilization / counted_members, 1)

def get_overloaded_count(db, week_start, week_end):
    """Počet přetížených členů (vytížení > 100%)"""
    employees = db.execute("SELECT id FROM employees WHERE status = 'active' OR status IS NULL").fetchall()
    overloaded = 0
    
    for emp in employees:
        emp_id = emp['id']
        
        # Kapacita člena
        profile = db.execute(
            "SELECT weekly_capacity_hours FROM team_member_profile WHERE employee_id = ?",
            (emp_id,)
        ).fetchone()
        capacity = 40.0
        if profile and profile['weekly_capacity_hours']:
            capacity = profile['weekly_capacity_hours']
        
        if capacity == 0:
            continue
        
        # Odpracované hodiny tento týden
        worked_hours = get_worked_hours_this_week(db, emp_id, week_start, week_end)
        
        utilization = (worked_hours / capacity) * 100 if capacity > 0 else 0
        
        if utilization > 100:
            overloaded += 1
    
    return overloaded

def get_ai_balance_score(db, week_start, week_end):
    """AI Balance Skóre - rovnoměrnost rozdělení práce (0-100%)"""
    employees = db.execute("SELECT id FROM employees WHERE status = 'active' OR status IS NULL").fetchall()
    if len(employees) < 2:
        return None
    
    utilizations = []
    
    for emp in employees:
        emp_id = emp['id']
        
        # Kapacita člena
        profile = db.execute(
            "SELECT weekly_capacity_hours FROM team_member_profile WHERE employee_id = ?",
            (emp_id,)
        ).fetchone()
        capacity = 40.0
        if profile and profile['weekly_capacity_hours']:
            capacity = profile['weekly_capacity_hours']
        
        if capacity == 0:
            continue
        
        # Odpracované hodiny tento týden
        worked_hours = get_worked_hours_this_week(db, emp_id, week_start, week_end)
        
        utilization = (worked_hours / capacity) * 100 if capacity > 0 else 0
        utilizations.append(utilization)
    
    if len(utilizations) < 2:
        return None
    
    # Vypočítat směrodatnou odchylku
    import statistics
    try:
        std_dev = statistics.stdev(utilizations)
        mean = statistics.mean(utilizations)
        
        # Převést na skóre 0-100 (nižší odchylka = vyšší skóre)
        if std_dev >= 50:
            return 0
        
        balance_score = 100 - (std_dev * 2)
        return max(0, min(100, round(balance_score)))
    except:
        return None

def count_overloaded_weeks(db, employee_id, weeks=4):
    """Kolik týdnů byl člen přetížený"""
    from datetime import datetime, timedelta
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())
    overloaded_count = 0
    
    # Kapacita člena
    profile = db.execute(
        "SELECT weekly_capacity_hours FROM team_member_profile WHERE employee_id = ?",
        (employee_id,)
    ).fetchone()
    capacity = 40.0
    if profile and profile['weekly_capacity_hours']:
        capacity = profile['weekly_capacity_hours']
    
    for i in range(weeks):
        week_start = (monday - timedelta(weeks=i)).strftime("%Y-%m-%d")
        week_end = (monday - timedelta(weeks=i) + timedelta(days=6)).strftime("%Y-%m-%d")
        
        worked_hours = get_worked_hours_this_week(db, employee_id, week_start, week_end)
        
        if capacity > 0 and (worked_hours / capacity * 100) > 100:
            overloaded_count += 1
    
    return overloaded_count

@app.route('/api/team/stats', methods=['GET'])
@requires_role('owner', 'admin', 'manager', 'lander')
def api_team_stats():
    """API endpoint pro všechny team metriky"""
    u, err = require_auth()
    if err:
        return err
    
    db = get_db()
    try:
        from datetime import datetime, timedelta
        
        today = datetime.now()
        monday = today - timedelta(days=today.weekday())
        week_start = monday.strftime("%Y-%m-%d")
        week_end = (monday + timedelta(days=6)).strftime("%Y-%m-%d")
        
        stats = {
            'active_count': get_active_members_count(db),
            'average_utilization': get_average_utilization(db, week_start, week_end),
            'overloaded_count': get_overloaded_count(db, week_start, week_end),
            'ai_balance_score': get_ai_balance_score(db, week_start, week_end)
        }
        
        return jsonify({'ok': True, 'stats': stats})
    except Exception as e:
        print(f"[ERROR] api_team_stats: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

# ============================================================
# TEAM METRICS HELPER FUNCTIONS
# ============================================================

def get_active_members_count(db):
    """Počet aktivních členů týmu"""
    result = db.execute("SELECT COUNT(*) as count FROM employees WHERE status = 'active' OR status IS NULL").fetchone()
    return result['count'] if result else 0

def get_worked_hours_this_week(db, employee_id, week_start, week_end):
    """Odpracované hodiny tento týden pro zaměstnance"""
    result = db.execute(
        """SELECT SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) / 60.0 as total 
           FROM timesheets 
           WHERE employee_id=? AND date >= ? AND date <= ?""",
        (employee_id, week_start, week_end)
    ).fetchone()
    return round(result['total'] or 0, 1) if result else 0.0

def get_average_utilization(db, week_start, week_end):
    """Průměrné vytížení týmu v procentech"""
    employees = db.execute("SELECT id FROM employees WHERE status = 'active' OR status IS NULL").fetchall()
    if not employees:
        return 0.0
    
    total_utilization = 0
    counted_members = 0
    
    for emp in employees:
        emp_id = emp['id']
        
        # Kapacita člena (default 40h/týden)
        profile = db.execute(
            "SELECT weekly_capacity_hours FROM team_member_profile WHERE employee_id = ?",
            (emp_id,)
        ).fetchone()
        capacity = 40.0
        if profile and profile['weekly_capacity_hours']:
            capacity = profile['weekly_capacity_hours']
        
        if capacity == 0:
            continue
        
        # Odpracované hodiny tento týden
        worked_hours = get_worked_hours_this_week(db, emp_id, week_start, week_end)
        
        utilization = (worked_hours / capacity) * 100 if capacity > 0 else 0
        total_utilization += utilization
        counted_members += 1
    
    if counted_members == 0:
        return 0.0
    
    return round(total_utilization / counted_members, 1)

def get_overloaded_count(db, week_start, week_end):
    """Počet přetížených členů (vytížení > 100%)"""
    employees = db.execute("SELECT id FROM employees WHERE status = 'active' OR status IS NULL").fetchall()
    overloaded = 0
    
    for emp in employees:
        emp_id = emp['id']
        
        # Kapacita člena
        profile = db.execute(
            "SELECT weekly_capacity_hours FROM team_member_profile WHERE employee_id = ?",
            (emp_id,)
        ).fetchone()
        capacity = 40.0
        if profile and profile['weekly_capacity_hours']:
            capacity = profile['weekly_capacity_hours']
        
        if capacity == 0:
            continue
        
        # Odpracované hodiny tento týden
        worked_hours = get_worked_hours_this_week(db, emp_id, week_start, week_end)
        
        utilization = (worked_hours / capacity) * 100 if capacity > 0 else 0
        
        if utilization > 100:
            overloaded += 1
    
    return overloaded

def get_ai_balance_score(db, week_start, week_end):
    """AI Balance Skóre - rovnoměrnost rozdělení práce (0-100%)"""
    employees = db.execute("SELECT id FROM employees WHERE status = 'active' OR status IS NULL").fetchall()
    if len(employees) < 2:
        return None
    
    utilizations = []
    
    for emp in employees:
        emp_id = emp['id']
        
        # Kapacita člena
        profile = db.execute(
            "SELECT weekly_capacity_hours FROM team_member_profile WHERE employee_id = ?",
            (emp_id,)
        ).fetchone()
        capacity = 40.0
        if profile and profile['weekly_capacity_hours']:
            capacity = profile['weekly_capacity_hours']
        
        if capacity == 0:
            continue
        
        # Odpracované hodiny tento týden
        worked_hours = get_worked_hours_this_week(db, emp_id, week_start, week_end)
        
        utilization = (worked_hours / capacity) * 100 if capacity > 0 else 0
        utilizations.append(utilization)
    
    if len(utilizations) < 2:
        return None
    
    # Vypočítat směrodatnou odchylku
    import statistics
    try:
        std_dev = statistics.stdev(utilizations)
        mean = statistics.mean(utilizations)
        
        # Převést na skóre 0-100 (nižší odchylka = vyšší skóre)
        if std_dev >= 50:
            return 0
        
        balance_score = 100 - (std_dev * 2)
        return max(0, min(100, round(balance_score)))
    except:
        return None

def count_overloaded_weeks(db, employee_id, weeks=4):
    """Kolik týdnů byl člen přetížený"""
    from datetime import datetime, timedelta
    today = datetime.now()
    monday = today - timedelta(days=today.weekday())
    overloaded_count = 0
    
    # Kapacita člena
    profile = db.execute(
        "SELECT weekly_capacity_hours FROM team_member_profile WHERE employee_id = ?",
        (employee_id,)
    ).fetchone()
    capacity = 40.0
    if profile and profile['weekly_capacity_hours']:
        capacity = profile['weekly_capacity_hours']
    
    for i in range(weeks):
        week_start = (monday - timedelta(weeks=i)).strftime("%Y-%m-%d")
        week_end = (monday - timedelta(weeks=i) + timedelta(days=6)).strftime("%Y-%m-%d")
        
        worked_hours = get_worked_hours_this_week(db, employee_id, week_start, week_end)
        
        if capacity > 0 and (worked_hours / capacity * 100) > 100:
            overloaded_count += 1
    
    return overloaded_count

@app.route('/api/team/capacity-overview', methods=['GET'])
@requires_role('owner', 'admin', 'manager', 'lander')
def api_team_capacity_overview():
    """Přehled kapacit celého týmu"""
    u, err = require_auth()
    if err:
        return err
    
    db = get_db()
    try:
        from datetime import datetime, timedelta
        
        # Získej aktivní zaměstnance
        employees = db.execute("SELECT * FROM employees ORDER BY name").fetchall()
        
        today = datetime.now()
        monday = today - timedelta(days=today.weekday())
        week_start = monday.strftime("%Y-%m-%d")
        week_end = (monday + timedelta(days=6)).strftime("%Y-%m-%d")
        
        overview = {
            'total_capacity': 0,
            'used_capacity': 0,
            'overloaded_count': 0,
            'underutilized_count': 0,
            'members': []
        }
        
        for emp in employees:
            emp_dict = dict(emp)
            emp_id = emp_dict['id']
            
            # Získej profil
            profile = db.execute(
                "SELECT * FROM team_member_profile WHERE employee_id = ?",
                (emp_id,)
            ).fetchone()
            
            if not profile:
                # Vytvoř defaultní profil
                weekly_capacity = 40.0
            else:
                profile_dict = dict(profile)
                weekly_capacity = profile_dict.get('weekly_capacity_hours', 40.0)
            
            # Vypočítej aktuální hodiny tento týden
            timesheet_rows = db.execute(
                "SELECT SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) / 60.0 as total FROM timesheets WHERE employee_id=? AND date >= ? AND date <= ?",
                (emp_id, week_start, week_end)
            ).fetchone()
            current_hours = round(timesheet_rows["total"] or 0, 1)
            
            capacity_percent = calculate_capacity_percent({'weekly_capacity_hours': weekly_capacity}, current_hours)
            capacity_status = calculate_capacity_status(capacity_percent)
            
            overview['total_capacity'] += weekly_capacity
            overview['used_capacity'] += current_hours
            
            if capacity_status == 'overloaded':
                overview['overloaded_count'] += 1
            elif capacity_status == 'underutilized':
                overview['underutilized_count'] += 1
            
            overview['members'].append({
                'id': emp_id,
                'name': emp_dict.get('name', ''),
                'capacity_percent': round(capacity_percent, 1),
                'status': capacity_status,
                'current_hours': current_hours,
                'weekly_capacity': weekly_capacity
            })
        
        overview['utilization_percent'] = round((overview['used_capacity'] / overview['total_capacity'] * 100) if overview['total_capacity'] > 0 else 0, 1)
        
        return jsonify({'ok': True, 'overview': overview})
    except Exception as e:
        print(f"[ERROR] api_team_capacity_overview: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/team/<int:employee_id>/ai-analysis', methods=['GET'])
@requires_role('owner', 'admin', 'manager', 'lander')
def api_team_member_ai_analysis(employee_id):
    """AI analýza člena týmu"""
    u, err = require_auth()
    if err:
        return err
    
    db = get_db()
    try:
        from datetime import datetime, timedelta
        
        # Získej zaměstnance
        employee = db.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not employee:
            return jsonify({'ok': False, 'error': 'employee_not_found'}), 404
        
        employee_dict = dict(employee)
        
        # Získej nebo vytvoř profil
        profile = db.execute(
            "SELECT * FROM team_member_profile WHERE employee_id = ?",
            (employee_id,)
        ).fetchone()
        
        if not profile:
            # Vytvoř defaultní profil
            db.execute("""
                INSERT INTO team_member_profile (employee_id, weekly_capacity_hours)
                VALUES (?, 40.0)
            """, (employee_id,))
            db.commit()
            profile = db.execute(
                "SELECT * FROM team_member_profile WHERE employee_id = ?",
                (employee_id,)
            ).fetchone()
        
        profile_dict = dict(profile)
        
        # Parse JSON fields
        import json as json_lib
        try:
            profile_dict['skills'] = json_lib.loads(profile_dict.get('skills') or '[]')
            profile_dict['certifications'] = json_lib.loads(profile_dict.get('certifications') or '[]')
            profile_dict['preferred_work_types'] = json_lib.loads(profile_dict.get('preferred_work_types') or '[]')
        except:
            profile_dict['skills'] = []
            profile_dict['certifications'] = []
            profile_dict['preferred_work_types'] = []
        
        # Shromáždě data pro analýzu
        today = datetime.now()
        monday = today - timedelta(days=today.weekday())
        week_start = monday.strftime("%Y-%m-%d")
        week_end = (monday + timedelta(days=6)).strftime("%Y-%m-%d")
        
        # Aktuální hodiny tento týden
        timesheet_rows = db.execute(
            "SELECT SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) / 60.0 as total FROM timesheets WHERE employee_id=? AND date >= ? AND date <= ?",
            (employee_id, week_start, week_end)
        ).fetchone()
        current_hours = round(timesheet_rows["total"] or 0, 1)
        
        # Počet týdnů přetížení (posledních 4 týdny)
        weeks_overloaded = 0
        for i in range(4):
            week_start_check = (monday - timedelta(weeks=i)).strftime("%Y-%m-%d")
            week_end_check = (monday - timedelta(weeks=i) + timedelta(days=6)).strftime("%Y-%m-%d")
            week_hours = db.execute(
                "SELECT SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) / 60.0 as total FROM timesheets WHERE employee_id=? AND date >= ? AND date <= ?",
                (employee_id, week_start_check, week_end_check)
            ).fetchone()
            week_total = round(week_hours["total"] or 0, 1)
            weekly_capacity = profile_dict.get('weekly_capacity_hours', 40.0)
            if (week_total / weekly_capacity * 100) > 90:
                weeks_overloaded += 1
        
        capacity_percent = calculate_capacity_percent(profile_dict, current_hours)
        capacity_status = calculate_capacity_status(capacity_percent)
        burnout_risk = calculate_burnout_risk(profile_dict, capacity_percent, weeks_overloaded)
        
        # Aktuální aktivní zakázky
        active_jobs = db.execute(
            "SELECT COUNT(DISTINCT job_id) as count FROM job_assignments WHERE employee_id=?",
            (employee_id,)
        ).fetchone()
        current_active_jobs = active_jobs["count"] or 0
        
        # Generuj doporučení
        recommendations = []
        warnings = []
        
        if burnout_risk in ['high', 'critical']:
            warnings.append({
                'type': 'burnout_warning',
                'priority': 'high',
                'message': f'{employee_dict.get("name", "Zaměstnanec")} je {weeks_overloaded} týdny přetížený/á. Doporučuji snížit zátěž.'
            })
        
        if capacity_status == 'underutilized' and weeks_overloaded == 0:
            recommendations.append({
                'type': 'underutilized',
                'priority': 'low',
                'message': f'{employee_dict.get("name", "Zaměstnanec")} má volnou kapacitu. Vhodný/á pro nové zakázky.'
            })
        
        if current_active_jobs > 5:
            warnings.append({
                'type': 'too_many_jobs',
                'priority': 'medium',
                'message': f'{employee_dict.get("name", "Zaměstnanec")} má {current_active_jobs} aktivních zakázek. Zvažte redistribuci.'
            })
        
        analysis_data = {
            'employee_id': employee_id,
            'name': employee_dict.get('name', ''),
            'current_hours': current_hours,
            'weekly_capacity': profile_dict.get('weekly_capacity_hours', 40.0),
            'capacity_percent': round(capacity_percent, 1),
            'capacity_status': capacity_status,
            'weeks_overloaded': weeks_overloaded,
            'burnout_risk_level': burnout_risk,
            'current_active_jobs': current_active_jobs,
            'performance_stability_score': profile_dict.get('performance_stability_score', 0.5),
            'ai_balance_score': profile_dict.get('ai_balance_score', 0.5),
            'skills': profile_dict['skills'],
            'certifications': profile_dict['certifications']
        }
        
        return jsonify({
            'ok': True,
            'employee_id': employee_id,
            'analysis': analysis_data,
            'recommendations': recommendations,
            'warnings': warnings
        })
    except Exception as e:
        print(f"[ERROR] api_team_member_ai_analysis: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/team/ai-crew-assistant', methods=['GET'])
@requires_role('owner', 'admin', 'manager', 'lander')
def api_ai_crew_assistant():
    """AI Crew Assistant - celkový přehled a doporučení"""
    u, err = require_auth()
    if err:
        return err
    
    db = get_db()
    try:
        from datetime import datetime, timedelta
        
        employees = db.execute("SELECT * FROM employees WHERE status = 'active' OR status IS NULL ORDER BY name").fetchall()
        
        today = datetime.now()
        monday = today - timedelta(days=today.weekday())
        week_start = monday.strftime("%Y-%m-%d")
        week_end = (monday + timedelta(days=6)).strftime("%Y-%m-%d")
        
        insights = []
        warnings = []
        recommendations = []
        
        # Celková kapacita a plánované hodiny
        total_capacity = 0
        total_planned = 0
        
        for emp in employees:
            emp_dict = dict(emp)
            emp_id = emp_dict['id']
            emp_name = emp_dict.get('name', 'Zaměstnanec')
            
            # Získej profil
            profile = db.execute(
                "SELECT * FROM team_member_profile WHERE employee_id = ?",
                (emp_id,)
            ).fetchone()
            
            if not profile:
                weekly_capacity = 40.0
                skills = []
            else:
                profile_dict = dict(profile)
                weekly_capacity = profile_dict.get('weekly_capacity_hours', 40.0)
                skills_json = profile_dict.get('skills', '[]')
                try:
                    import json
                    skills = json.loads(skills_json) if isinstance(skills_json, str) else (skills_json or [])
                except:
                    skills = []
            
            total_capacity += weekly_capacity
            
            # Aktuální hodiny tento týden
            worked_hours = get_worked_hours_this_week(db, emp_id, week_start, week_end)
            total_planned += worked_hours
            
            utilization = (worked_hours / weekly_capacity * 100) if weekly_capacity > 0 else 0
            
            # === PRAVIDLO 1: Přetížení (> 100%) ===
            if utilization > 100:
                warnings.append({
                    'type': 'alert',
                    'priority': 'high',
                    'icon': 'warning',
                    'member_id': emp_id,
                    'member_name': emp_name,
                    'message': f'{emp_name} je přetížen/a ({utilization:.0f}% kapacity)',
                    'suggestion': 'Přerozdělit úkoly nebo odložit termíny',
                    'action': 'redistribute_tasks'
                })
            
            # === PRAVIDLO 2: Vysoké vytížení (85-100%) ===
            elif utilization > 85:
                warnings.append({
                    'type': 'warning',
                    'priority': 'medium',
                    'icon': 'attention',
                    'member_id': emp_id,
                    'member_name': emp_name,
                    'message': f'{emp_name} má vysoké vytížení ({utilization:.0f}%)',
                    'suggestion': 'Sledovat a nepřidávat další úkoly',
                    'action': 'monitor'
                })
            
            # === PRAVIDLO 3: Nevyužitá kapacita (< 30%) ===
            elif utilization < 30:
                recommendations.append({
                    'type': 'info',
                    'priority': 'low',
                    'icon': 'opportunity',
                    'member_id': emp_id,
                    'member_name': emp_name,
                    'message': f'{emp_name} má volnou kapacitu ({100-utilization:.0f}% volno)',
                    'suggestion': 'Vhodný/á pro nové zakázky nebo pomoc kolegům',
                    'action': 'assign_more'
                })
            
            # === PRAVIDLO 4: Dlouhodobé přetížení (burnout risk) ===
            weeks_overloaded = count_overloaded_weeks(db, emp_id, weeks=4)
            if weeks_overloaded >= 2:
                warnings.append({
                    'type': 'alert',
                    'priority': 'high',
                    'icon': 'burnout',
                    'member_id': emp_id,
                    'member_name': emp_name,
                    'message': f'{emp_name} je přetížen/a už {weeks_overloaded} týdny',
                    'suggestion': 'Riziko vyhoření - zvážit volno nebo snížení zátěže',
                    'action': 'prevent_burnout'
                })
        
        # === PRAVIDLO 5: Kapacitní mezera tento týden ===
        if total_planned > total_capacity:
            gap = total_planned - total_capacity
            warnings.append({
                'type': 'alert',
                'priority': 'high',
                'icon': 'capacity_gap',
                'message': f'Tento týden chybí {gap:.0f} hodin kapacity',
                'suggestion': 'Zvážit přesun termínů nebo externí pomoc',
                'action': 'resolve_capacity'
            })
        
        # === PRAVIDLO 6: Nerovnoměrné rozdělení ===
        balance_score = get_ai_balance_score(db, week_start, week_end)
        if balance_score is not None and balance_score < 50:
            warnings.append({
                'type': 'warning',
                'priority': 'medium',
                'icon': 'imbalance',
                'message': 'Práce je nerovnoměrně rozdělena v týmu',
                'suggestion': 'Některé členy přetěžujete, jiní mají volno',
                'action': 'rebalance_team'
            })
        
        # Seřadit podle priority
        priority_order = {'high': 0, 'medium': 1, 'low': 2}
        all_recommendations = warnings + recommendations
        all_recommendations.sort(key=lambda x: priority_order.get(x.get('priority', 'low'), 99))
        
        return jsonify({
            'ok': True,
            'insights': insights,
            'warnings': [r for r in all_recommendations if r.get('type') in ('alert', 'warning')],
            'recommendations': [r for r in all_recommendations if r.get('type') in ('info', 'suggestion')],
            'capacity_forecast': {
                'total_capacity': total_capacity,
                'total_planned': total_planned,
                'utilization_percent': round((total_planned / total_capacity * 100) if total_capacity > 0 else 0, 1)
            }
        })
    except Exception as e:
        print(f"[ERROR] api_ai_crew_assistant: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/team/<int:employee_id>/skills', methods=['POST', 'PATCH'])
@requires_role('owner', 'admin', 'manager', 'lander')
def api_team_member_skills(employee_id):
    """Aktualizace dovedností člena týmu"""
    u, err = require_auth()
    if err:
        return err
    
    db = get_db()
    try:
        import json as json_lib
        
        # Zkontroluj, že zaměstnanec existuje
        employee = db.execute("SELECT id FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not employee:
            return jsonify({'ok': False, 'error': 'employee_not_found'}), 404
        
        data = request.get_json(force=True, silent=True) or {}
        
        # Získej nebo vytvoř profil
        profile = db.execute(
            "SELECT * FROM team_member_profile WHERE employee_id = ?",
            (employee_id,)
        ).fetchone()
        
        if not profile:
            # Vytvoř nový profil
            db.execute("""
                INSERT INTO team_member_profile (employee_id, weekly_capacity_hours)
                VALUES (?, 40.0)
            """, (employee_id,))
            db.commit()
            profile = db.execute(
                "SELECT * FROM team_member_profile WHERE employee_id = ?",
                (employee_id,)
            ).fetchone()
        
        # Aktualizuj data
        updates = []
        params = []
        
        if 'skills' in data:
            updates.append("skills = ?")
            params.append(json_lib.dumps(data['skills'] if isinstance(data['skills'], list) else []))
        
        if 'certifications' in data:
            updates.append("certifications = ?")
            params.append(json_lib.dumps(data['certifications'] if isinstance(data['certifications'], list) else []))
        
        if 'preferred_work_types' in data:
            updates.append("preferred_work_types = ?")
            params.append(json_lib.dumps(data['preferred_work_types'] if isinstance(data['preferred_work_types'], list) else []))
        
        if 'weekly_capacity_hours' in data:
            updates.append("weekly_capacity_hours = ?")
            params.append(float(data['weekly_capacity_hours']))
        
        if updates:
            updates.append("updated_at = datetime('now')")
            params.append(employee_id)
            
            db.execute(
                f"UPDATE team_member_profile SET {', '.join(updates)} WHERE employee_id = ?",
                params
            )
            db.commit()
        
        # Vrať aktualizovaný profil
        updated_profile = db.execute(
            "SELECT * FROM team_member_profile WHERE employee_id = ?",
            (employee_id,)
        ).fetchone()
        
        profile_dict = dict(updated_profile)
        try:
            profile_dict['skills'] = json_lib.loads(profile_dict.get('skills') or '[]')
            profile_dict['certifications'] = json_lib.loads(profile_dict.get('certifications') or '[]')
            profile_dict['preferred_work_types'] = json_lib.loads(profile_dict.get('preferred_work_types') or '[]')
        except:
            profile_dict['skills'] = []
            profile_dict['certifications'] = []
            profile_dict['preferred_work_types'] = []
        
        return jsonify({'ok': True, 'profile': profile_dict})
    except Exception as e:
        print(f"[ERROR] api_team_member_skills: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/team/member/<int:employee_id>', methods=['GET'])
@requires_role('owner', 'admin', 'manager', 'lander', 'worker')
def api_team_member_detail(employee_id):
    """Detail člena týmu s rozšířeným profilem"""
    u, err = require_auth()
    if err:
        return err
    
    db = get_db()
    try:
        import json as json_lib
        
        # Získej zaměstnance
        employee = db.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)).fetchone()
        if not employee:
            return jsonify({'ok': False, 'error': 'employee_not_found'}), 404
        
        employee_dict = dict(employee)
        
        # Normalizace role
        employee_dict['role'] = normalize_employee_role(employee_dict.get('role'))
        
        # Účet zaměstnance
        if employee_dict.get('user_id'):
            user = db.execute("SELECT * FROM users WHERE id = ?", (employee_dict['user_id'],)).fetchone()
            if user:
                employee_dict['has_account'] = True
                employee_dict['account_role'] = normalize_role(dict(user).get('role'))
            else:
                employee_dict['has_account'] = False
        else:
            employee_dict['has_account'] = False
        
        # Vypočítej hodiny tento týden
        from datetime import datetime, timedelta
        today = datetime.now()
        monday = today - timedelta(days=today.weekday())
        week_start = monday.strftime("%Y-%m-%d")
        week_end = (monday + timedelta(days=6)).strftime("%Y-%m-%d")
        
        timesheet_rows = db.execute(
            "SELECT SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) / 60.0 as total FROM timesheets WHERE employee_id=? AND date >= ? AND date <= ?",
            (employee_id, week_start, week_end)
        ).fetchone()
        employee_dict['hours_week'] = round(timesheet_rows["total"] or 0, 1)
        
        # Aktivní zakázky
        job_rows = db.execute(
            "SELECT COUNT(DISTINCT job_id) as count FROM job_assignments WHERE employee_id=?",
            (employee_id,)
        ).fetchone()
        employee_dict['active_projects'] = job_rows["count"] or 0
        
        # Získej nebo vytvoř profil
        profile = db.execute(
            "SELECT * FROM team_member_profile WHERE employee_id = ?",
            (employee_id,)
        ).fetchone()
        
        if not profile:
            # Vytvoř defaultní profil
            db.execute("""
                INSERT INTO team_member_profile (employee_id, weekly_capacity_hours)
                VALUES (?, 40.0)
            """, (employee_id,))
            db.commit()
            profile = db.execute(
                "SELECT * FROM team_member_profile WHERE employee_id = ?",
                (employee_id,)
            ).fetchone()
        
        profile_dict = dict(profile)
        
        # Parse JSON fields
        try:
            profile_dict['skills'] = json_lib.loads(profile_dict.get('skills') or '[]')
            profile_dict['certifications'] = json_lib.loads(profile_dict.get('certifications') or '[]')
            profile_dict['preferred_work_types'] = json_lib.loads(profile_dict.get('preferred_work_types') or '[]')
        except:
            profile_dict['skills'] = []
            profile_dict['certifications'] = []
            profile_dict['preferred_work_types'] = []
        
        # Vypočítej kapacitu
        current_week_hours = employee_dict.get('hours_week', 0)
        weekly_capacity = profile_dict.get('weekly_capacity_hours', 40.0)
        capacity_percent = calculate_capacity_percent(profile_dict, current_week_hours)
        capacity_status = calculate_capacity_status(capacity_percent)
        
        # Přidej vypočítané hodnoty do profilu
        profile_dict['capacity_percent'] = round(capacity_percent, 1)
        profile_dict['capacity_status'] = capacity_status
        profile_dict['current_week_hours'] = current_week_hours
        
        employee_dict['profile'] = profile_dict
        
        return jsonify({'ok': True, 'employee': employee_dict})
    except Exception as e:
        print(f"[ERROR] api_team_member_detail: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/team/ai-recommendations', methods=['GET'])
@requires_role('owner', 'admin', 'manager', 'lander')
def api_team_ai_recommendations():
    """AI doporučení pro tým (alias pro ai-crew-assistant s jiným formátem)"""
    u, err = require_auth()
    if err:
        return err
    
    db = get_db()
    try:
        from datetime import datetime, timedelta
        
        employees = db.execute("SELECT * FROM employees ORDER BY name").fetchall()
        recommendations = []
        
        today = datetime.now()
        monday = today - timedelta(days=today.weekday())
        week_start = monday.strftime("%Y-%m-%d")
        week_end = (monday + timedelta(days=6)).strftime("%Y-%m-%d")
        
        for emp in employees:
            emp_dict = dict(emp)
            emp_id = emp_dict['id']
            
            # Získej profil
            profile = db.execute(
                "SELECT * FROM team_member_profile WHERE employee_id = ?",
                (emp_id,)
            ).fetchone()
            
            if not profile:
                continue
            
            profile_dict = dict(profile)
            
            # Aktuální hodiny
            timesheet_rows = db.execute(
                "SELECT SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) / 60.0 as total FROM timesheets WHERE employee_id=? AND date >= ? AND date <= ?",
                (emp_id, week_start, week_end)
            ).fetchone()
            current_hours = round(timesheet_rows["total"] or 0, 1)
            
            weekly_capacity = profile_dict.get('weekly_capacity_hours', 40.0)
            capacity_percent = calculate_capacity_percent(profile_dict, current_hours)
            
            # Počet týdnů přetížení
            weeks_overloaded = 0
            for i in range(4):
                week_start_check = (monday - timedelta(weeks=i)).strftime("%Y-%m-%d")
                week_end_check = (monday - timedelta(weeks=i) + timedelta(days=6)).strftime("%Y-%m-%d")
                week_hours = db.execute(
                    "SELECT SUM(COALESCE(duration_minutes, CAST(hours * 60 AS INTEGER))) / 60.0 as total FROM timesheets WHERE employee_id=? AND date >= ? AND date <= ?",
                    (emp_id, week_start_check, week_end_check)
                ).fetchone()
                week_total = round(week_hours["total"] or 0, 1)
                if (week_total / weekly_capacity * 100) > 90:
                    weeks_overloaded += 1
            
            # Pravidlo 1: Přetížení
            if capacity_percent > 90:
                recommendations.append({
                    'type': 'warning',
                    'priority': 'high',
                    'employee_id': emp_id,
                    'employee_name': emp_dict.get('name', ''),
                    'message': f'{emp_dict.get("name", "Zaměstnanec")} je přetížen/a ({int(capacity_percent)}% kapacity)',
                    'suggestion': 'Doporučuji přerozdělit úkoly'
                })
            
            # Pravidlo 2: Burnout risk
            burnout_risk = profile_dict.get('burnout_risk_level', 'normal')
            if burnout_risk in ['high', 'critical']:
                recommendations.append({
                    'type': 'alert',
                    'priority': 'high',
                    'employee_id': emp_id,
                    'employee_name': emp_dict.get('name', ''),
                    'message': f'{emp_dict.get("name", "Zaměstnanec")} má vysoké riziko vyhoření',
                    'suggestion': 'Zvažte snížení zátěže nebo volno'
                })
            
            # Pravidlo 3: Nevyužitý
            if capacity_percent < 40:
                recommendations.append({
                    'type': 'info',
                    'priority': 'low',
                    'employee_id': emp_id,
                    'employee_name': emp_dict.get('name', ''),
                    'message': f'{emp_dict.get("name", "Zaměstnanec")} má volnou kapacitu ({int(capacity_percent)}%)',
                    'suggestion': 'Vhodný/á pro nové zakázky'
                })
        
        # Seřadit podle priority
        priority_order = {'high': 0, 'medium': 1, 'low': 2}
        recommendations.sort(key=lambda x: priority_order.get(x['priority'], 99))
        
        return jsonify({'ok': True, 'recommendations': recommendations})
    except Exception as e:
        print(f"[ERROR] api_team_ai_recommendations: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

def init_team_profiles():
    """Vytvoří profily pro všechny existující zaměstnance"""
    db = get_db()
    try:
        employees = db.execute("SELECT id FROM employees").fetchall()
        created_count = 0
        
        for emp in employees:
            emp_id = emp['id']
            
            # Zkontroluj, jestli už profil existuje
            existing = db.execute(
                "SELECT id FROM team_member_profile WHERE employee_id = ?",
                (emp_id,)
            ).fetchone()
            
            if not existing:
                db.execute("""
                    INSERT INTO team_member_profile (employee_id, weekly_capacity_hours)
                    VALUES (?, 40.0)
                """, (emp_id,))
                created_count += 1
        
        db.commit()
        print(f"[INIT] Vytvořeno {created_count} profilů pro {len(employees)} zaměstnanců")
        return created_count
    except Exception as e:
        print(f"[ERROR] init_team_profiles: {e}")
        db.rollback()
        return 0

@app.route('/api/team/init-profiles', methods=['POST'])
@requires_role('owner', 'admin')
def api_init_team_profiles():
    """Inicializační endpoint pro vytvoření profilů"""
    u, err = require_auth()
    if err:
        return err
    
    count = init_team_profiles()
    return jsonify({'ok': True, 'created': count, 'message': f'Vytvořeno {count} profilů'})


# ============================================================
# MOBILE UI ROUTES
# ============================================================

@app.route('/mobile/dashboard')
def mobile_dashboard():
    """Mobile dashboard s widgety."""
    # Pro demo povolíme i bez autentizace
    u = None
    user_id = None
    user_role = 'worker'
    
    if 'uid' in session:
        u, err = require_auth()
        if err:
            return err
        user_id = session.get('uid')
        user_role = session.get('user_role', 'worker')
    
    from app.utils.mobile_mode import get_mobile_mode
    from app.utils.widgets import get_user_widgets, get_available_widgets_for_user
    from config.widgets import WIDGET_REGISTRY, ROLE_DEFAULT_WIDGETS
    from config.permissions import normalize_role
    
    mode = request.args.get('mode') or (get_mobile_mode() if user_id else 'field')
    
    # Získej widgety pro uživatele nebo použij demo widgety
    if user_id:
        widgets = get_user_widgets(user_id, user_role, mode)
    else:
        # Demo widgety podle módu
        if mode == 'field':
            widgets = ['current_job', 'quick_log', 'my_tasks', 'add_photo', 'report_blocker', 'offline_status']
        else:
            widgets = ['notifications', 'jobs_risk', 'overdue_jobs', 'team_load', 'stock_alerts', 'budget_burn']
    
    # Context data pro widgety
    db = get_db()
    context = {}
    
    if user_id:
        # Aktuální zakázka
        current_job = db.execute(
            "SELECT * FROM jobs WHERE status NOT LIKE 'dokon%' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        if current_job:
            context['current_job'] = dict(current_job)
        
        # Úkoly pro uživatele
        employee_id = db.execute(
            "SELECT id FROM employees WHERE user_id = ?", (user_id,)
        ).fetchone()
        if employee_id:
            tasks = db.execute("""
                SELECT t.*, j.title as job_name, j.client 
                FROM tasks t 
                LEFT JOIN jobs j ON t.job_id = j.id 
                WHERE t.employee_id = ? AND t.status != 'done'
                ORDER BY t.due_date ASC
                LIMIT 10
            """, (employee_id[0],)).fetchall()
            context['tasks'] = [dict(t) for t in tasks]
    else:
        # Demo data
        context['current_job'] = {
            'id': 123,
            'title': 'Zahrada u Nováků',
            'status': 'V realizaci',
            'city': 'Praha 5'
        }
        context['tasks'] = [
            {'id': 1, 'title': 'Zasadit keře', 'status': 'open', 'job_name': 'Zahrada u Nováků', 'due_date': '2025-01-15'},
            {'id': 2, 'title': 'Posekat trávu', 'status': 'open', 'job_name': 'Zahrada u Nováků', 'due_date': '2025-01-16'}
        ]
    
    # Přidej demo data pro FULL mode widgety pokud není přihlášen
    if not user_id and mode == 'full':
        context.update({
            'notifications': [
                {'id': 1, 'title': 'Nová zakázka', 'body': 'Přidána zakázka #124', 'kind': 'info', 'read': False, 'created_at': '2025-01-14'},
                {'id': 2, 'title': 'Blocker nahlášen', 'body': 'Problém na zakázce #123', 'kind': 'warning', 'read': False, 'created_at': '2025-01-13'}
            ],
            'risky_jobs': [
                {'id': 123, 'title': 'Zahrada u Nováků', 'risk_level': 'high', 'risk_score': 75, 'risk_factors': ['Zpoždění', 'Chybí materiál']},
                {'id': 124, 'title': 'Park u školy', 'risk_level': 'medium', 'risk_score': 45, 'risk_factors': ['Blíží se termín']}
            ],
            'overdue_jobs': [
                {'id': 120, 'title': 'Zahrada u Svobodů', 'days_overdue': 3, 'due_date': '2025-01-11'}
            ],
            'team_members': [
                {'name': 'Jan Novák', 'role_label': 'Lander', 'load_level': 'high', 'load_percent': 85, 'hours_today': 8},
                {'name': 'Petr Svoboda', 'role_label': 'Worker', 'load_level': 'normal', 'load_percent': 60, 'hours_today': 6},
                {'name': 'Marie Dvořáková', 'role_label': 'Worker', 'load_level': 'low', 'load_percent': 40, 'hours_today': 4}
            ],
            'alerts': [
                {'material_name': 'Hnojivo', 'message': 'Nízká zásoba', 'severity': 'warning', 'quantity': 5, 'unit': 'kg'},
                {'material_name': 'Semena', 'message': 'Kritická zásoba', 'severity': 'critical', 'quantity': 2, 'unit': 'ks'}
            ],
            'budget_data': {
                'total_budget': 500000,
                'spent': 325000,
                'remaining': 175000,
                'percent_spent': 65,
                'jobs': [
                    {'id': 123, 'name': 'Zahrada u Nováků', 'spent': 150000},
                    {'id': 124, 'name': 'Park u školy', 'spent': 100000}
                ]
            }
        })
    
    return render_template('mobile/dashboard.html',
        mobile_mode=mode,
        widgets=widgets,
        widget_registry=WIDGET_REGISTRY,
        context=context,
        tasks=context.get('tasks', []),
        notifications=context.get('notifications', []),
        risky_jobs=context.get('risky_jobs', []),
        overdue_jobs=context.get('overdue_jobs', []),
        team_members=context.get('team_members', []),
        alerts=context.get('alerts', []),
        budget_data=context.get('budget_data'),
        work_types=[{'id': 1, 'name': 'Zahradnické práce'}, {'id': 2, 'name': 'Údržba'}] if not user_id else [],
        active_tab='dashboard'
    )

@app.route('/mobile/today')
def mobile_today():
    """Mobile Today screen (FIELD mode)."""
    # Pro demo povolíme i bez autentizace
    u = None
    user_id = None
    user_role = 'worker'
    
    if 'uid' in session:
        u, err = require_auth()
        if err:
            return err
        user_id = session.get('uid')
        user_role = session.get('user_role', 'worker')
    
    from app.utils.mobile_mode import get_mobile_mode
    from app.utils.widgets import get_user_widgets
    from config.widgets import WIDGET_REGISTRY
    
    mode = 'field'  # Today screen je vždy field mode
    
    # Získej widgety pro field mode nebo použij demo widgety
    if user_id:
        widgets = get_user_widgets(user_id, user_role, mode)
    else:
        # Demo widgety pro field mode
        widgets = ['current_job', 'quick_log', 'my_tasks', 'add_photo', 'report_blocker', 'offline_status']
    
    # Context data
    db = get_db()
    context = {}
    
    # Aktuální zakázka
    if user_id:
        current_job = db.execute(
            "SELECT * FROM jobs WHERE status NOT LIKE 'dokon%' ORDER BY id DESC LIMIT 1"
        ).fetchone()
        if current_job:
            context['current_job'] = dict(current_job)
        
        # Úkoly
        employee_id = db.execute(
            "SELECT id FROM employees WHERE user_id = ?", (user_id,)
        ).fetchone()
        if employee_id:
            tasks = db.execute("""
                SELECT t.*, j.title as job_name 
                FROM tasks t 
                LEFT JOIN jobs j ON t.job_id = j.id 
                WHERE t.employee_id = ? AND t.status != 'done'
                ORDER BY t.due_date ASC
                LIMIT 5
            """, (employee_id[0],)).fetchall()
            context['tasks'] = [dict(t) for t in tasks]
    else:
        # Demo data
        context['current_job'] = {
            'id': 123,
            'title': 'Zahrada u Nováků',
            'status': 'V realizaci',
            'city': 'Praha 5'
        }
        context['tasks'] = [
            {'id': 1, 'title': 'Zasadit keře', 'status': 'open', 'job_name': 'Zahrada u Nováků', 'due_date': '2025-01-15'},
            {'id': 2, 'title': 'Posekat trávu', 'status': 'open', 'job_name': 'Zahrada u Nováků', 'due_date': '2025-01-16'}
        ]
    
    # Work types pro quick log
    try:
        work_types = db.execute("SELECT id, name FROM work_types LIMIT 10").fetchall()
    except:
        work_types = []
    
    if not work_types:
        work_types = [{'id': 1, 'name': 'Zahradnické práce'}, {'id': 2, 'name': 'Údržba'}]
    
    return render_template('mobile/dashboard.html',
        mobile_mode=mode,
        widgets=widgets,
        widget_registry=WIDGET_REGISTRY,
        context=context,
        work_types=[dict(wt) for wt in work_types],
        active_tab='today'
    )

@app.route('/mobile/edit-dashboard')
def mobile_edit_dashboard():
    """Editor widget layoutu."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.mobile_mode import get_mobile_mode
    from app.utils.widgets import get_user_widgets, get_available_widgets_for_user
    from config.widgets import WIDGET_REGISTRY
    
    mode = request.args.get('mode') or get_mobile_mode()
    user_id = session.get('uid')
    user_role = session.get('user_role', 'worker')
    
    # Aktuální widgety
    current_widgets = get_user_widgets(user_id, user_role, mode)
    
    # Všechny dostupné widgety
    available_widget_ids = get_available_widgets_for_user(user_role, mode)
    available_widgets = {wid: WIDGET_REGISTRY[wid] for wid in available_widget_ids if wid in WIDGET_REGISTRY}
    
    return render_template('mobile/edit_dashboard.html',
        mobile_mode=mode,
        current_widgets=current_widgets,
        available_widgets=available_widgets,
        widget_registry=WIDGET_REGISTRY
    )

@app.route('/mobile/demo')
def mobile_demo():
    """Demo stránka pro ukázku mobilního UI."""
    # Demo stránka nevyžaduje autentizaci
    u = None
    if 'uid' in session:
        u, _ = require_auth()
    
    from app.utils.mobile_mode import get_mobile_mode
    from config.widgets import WIDGET_REGISTRY
    
    mode = request.args.get('mode', 'field')
    
    # Demo data pro widgety
    context = {
        'current_job': {
            'id': 123,
            'title': 'Zahrada u Nováků',
            'status': 'V realizaci',
            'city': 'Praha 5'
        },
        'tasks': [
            {'id': 1, 'title': 'Zasadit keře', 'status': 'open', 'job_name': 'Zahrada u Nováků', 'due_date': '2025-01-15'},
            {'id': 2, 'title': 'Posekat trávu', 'status': 'open', 'job_name': 'Zahrada u Nováků', 'due_date': '2025-01-16'}
        ],
        'notifications': [
            {'id': 1, 'title': 'Nová zakázka', 'body': 'Přidána zakázka #124', 'kind': 'info', 'read': False, 'created_at': '2025-01-14'},
            {'id': 2, 'title': 'Blocker nahlášen', 'body': 'Problém na zakázce #123', 'kind': 'warning', 'read': False, 'created_at': '2025-01-13'}
        ],
        'risky_jobs': [
            {'id': 123, 'title': 'Zahrada u Nováků', 'risk_level': 'high', 'risk_score': 75, 'risk_factors': ['Zpoždění', 'Chybí materiál']},
            {'id': 124, 'title': 'Park u školy', 'risk_level': 'medium', 'risk_score': 45, 'risk_factors': ['Blíží se termín']}
        ],
        'overdue_jobs': [
            {'id': 120, 'title': 'Zahrada u Svobodů', 'days_overdue': 3, 'due_date': '2025-01-11'}
        ],
        'team_members': [
            {'name': 'Jan Novák', 'role_label': 'Lander', 'load_level': 'high', 'load_percent': 85, 'hours_today': 8},
            {'name': 'Petr Svoboda', 'role_label': 'Worker', 'load_level': 'normal', 'load_percent': 60, 'hours_today': 6},
            {'name': 'Marie Dvořáková', 'role_label': 'Worker', 'load_level': 'low', 'load_percent': 40, 'hours_today': 4}
        ],
        'alerts': [
            {'material_name': 'Hnojivo', 'message': 'Nízká zásoba', 'severity': 'warning', 'quantity': 5, 'unit': 'kg'},
            {'material_name': 'Semena', 'message': 'Kritická zásoba', 'severity': 'critical', 'quantity': 2, 'unit': 'ks'}
        ],
        'budget_data': {
            'total_budget': 500000,
            'spent': 325000,
            'remaining': 175000,
            'percent_spent': 65,
            'jobs': [
                {'id': 123, 'name': 'Zahrada u Nováků', 'spent': 150000},
                {'id': 124, 'name': 'Park u školy', 'spent': 100000}
            ]
        }
    }
    
    # Demo widgety podle módu
    if mode == 'field':
        demo_widgets = ['current_job', 'quick_log', 'my_tasks', 'add_photo', 'report_blocker', 'offline_status']
    else:
        demo_widgets = ['notifications', 'jobs_risk', 'overdue_jobs', 'team_load', 'stock_alerts', 'budget_burn']
    
    return render_template('mobile/dashboard.html',
        mobile_mode=mode,
        widgets=demo_widgets,
        widget_registry=WIDGET_REGISTRY,
        context=context,
        work_types=[{'id': 1, 'name': 'Zahradnické práce'}, {'id': 2, 'name': 'Údržba'}],
        notifications=context.get('notifications', []),
        risky_jobs=context.get('risky_jobs', []),
        overdue_jobs=context.get('overdue_jobs', []),
        team_members=context.get('team_members', []),
        alerts=context.get('alerts', []),
        budget_data=context.get('budget_data'),
        tasks=context.get('tasks', []),
        active_tab='demo'
    )

@app.route('/mobile-demo.html')
def mobile_demo_index():
    """Index stránka pro mobile demo."""
    return send_from_directory('.', 'mobile-demo.html')

@app.route('/mobile/tasks')
def mobile_tasks():
    """Stránka s úkoly pro mobilní UI."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.mobile_mode import get_mobile_mode
    from app.utils.permissions import get_user_role
    
    mode = get_mobile_mode()
    user_id = session.get('uid')
    user_role = get_user_role()
    
    db = get_db()
    
    # Získej employee_id
    employee = db.execute(
        "SELECT id FROM employees WHERE user_id = ?",
        (user_id,)
    ).fetchone()
    
    tasks = []
    if employee:
        employee_id = employee[0]
        tasks = db.execute("""
            SELECT t.*, j.title as job_title, j.code as job_code,
                   e.name as assignee_name
            FROM tasks t
            LEFT JOIN jobs j ON t.job_id = j.id
            LEFT JOIN employees e ON t.employee_id = e.id
            WHERE t.employee_id = ? AND t.status != 'done'
            ORDER BY 
                CASE WHEN t.due_date IS NULL THEN 1 ELSE 0 END,
                t.due_date ASC,
                t.created_at DESC
            LIMIT 50
        """, (employee_id,)).fetchall()
    
    return render_template('mobile/tasks.html',
        mobile_mode=mode,
        tasks=[dict(t) for t in tasks],
        user_role=user_role
    )

@app.route('/mobile/photos')
def mobile_photos():
    """Stránka s fotografiemi pro mobilní UI."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.mobile_mode import get_mobile_mode
    
    mode = get_mobile_mode()
    user_id = session.get('uid')
    job_id = request.args.get('job_id', type=int)
    
    db = get_db()
    
    # Získej fotky
    if job_id:
        photos = db.execute("""
            SELECT a.*, j.title as job_title, j.code as job_code
            FROM attachments a
            LEFT JOIN jobs j ON a.entity_id = j.id AND a.entity_type = 'job'
            WHERE a.entity_type = 'job' AND a.entity_id = ?
              AND a.mimetype LIKE 'image/%'
            ORDER BY a.created_at DESC
            LIMIT 100
        """, (job_id,)).fetchall()
    else:
        # Všechny fotky uživatele
        photos = db.execute("""
            SELECT a.*, j.title as job_title, j.code as job_code
            FROM attachments a
            LEFT JOIN jobs j ON a.entity_id = j.id AND a.entity_type = 'job'
            WHERE a.uploaded_by = ? AND a.mimetype LIKE 'image/%'
            ORDER BY a.created_at DESC
            LIMIT 100
        """, (user_id,)).fetchall()
    
    return render_template('mobile/photos.html',
        mobile_mode=mode,
        photos=[dict(p) for p in photos],
        job_id=job_id
    )

@app.route('/mobile/notifications')
def mobile_notifications():
    """Stránka s notifikacemi pro mobilní UI."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.mobile_mode import get_mobile_mode
    
    mode = get_mobile_mode()
    user_id = session.get('uid')
    
    db = get_db()
    
    # Získej notifikace
    notifications = db.execute("""
        SELECT * FROM notifications
        WHERE user_id = ?
        ORDER BY created_at DESC
        LIMIT 50
    """, (user_id,)).fetchall()
    
    # Označ jako přečtené
    db.execute("""
        UPDATE notifications
        SET is_read = 1
        WHERE user_id = ? AND is_read = 0
    """, (user_id,))
    db.commit()
    
    return render_template('mobile/notifications.html',
        mobile_mode=mode,
        notifications=[dict(n) for n in notifications]
    )

@app.route('/mobile/queue')
def mobile_queue():
    """Stránka pro zobrazení offline fronty."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.mobile_mode import get_mobile_mode
    from app.utils.permissions import get_user_role
    
    mode = get_mobile_mode()
    user_role = get_user_role()
    
    # Pouze director, manager, lander mohou vidět queue
    if user_role not in ['director', 'manager', 'lander']:
        return jsonify({'ok': False, 'error': 'Nedostatečná oprávnění'}), 403
    
    db = get_db()
    
    # Načti zpracované eventy z DB (pro historii)
    recent_events = db.execute("""
        SELECT * FROM processed_events
        WHERE created_at >= datetime('now', '-1 day')
        ORDER BY created_at DESC
        LIMIT 50
    """).fetchall()
    
    return render_template('mobile/queue.html',
        recent_events=[dict(e) for e in recent_events],
        mobile_mode=mode,
        user_role=user_role
    )

# ============================================================
# MOBILE API ENDPOINTS
# ============================================================

@app.route('/api/user/dashboard-layout', methods=['GET', 'PATCH', 'DELETE'])
def api_dashboard_layout():
    """API pro správu widget layoutu."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.widgets import get_user_widgets, save_user_widgets
    from app.utils.mobile_mode import get_mobile_mode
    
    user_id = session.get('uid')
    user_role = session.get('user_role', 'worker')
    mode = request.args.get('mode') or request.get_json().get('mode') or get_mobile_mode()
    
    if request.method == 'GET':
        widgets = get_user_widgets(user_id, user_role, mode)
        return jsonify({'ok': True, 'widgets': widgets, 'mode': mode})
    
    elif request.method == 'PATCH':
        data = request.get_json() or {}
        widget_ids = data.get('widgets', [])
        
        if save_user_widgets(user_id, mode, widget_ids):
            return jsonify({'ok': True, 'widgets': widget_ids})
        else:
            return jsonify({'ok': False, 'error': 'Failed to save'}), 500
    
    elif request.method == 'DELETE':
        # Reset na default - smaž user override
        db = get_db()
        db.execute(
            "DELETE FROM user_dashboard_layout WHERE user_id = ?",
            (user_id,)
        )
        db.commit()
        return jsonify({'ok': True})

@app.route('/api/widgets/current-job')
def api_widget_current_job():
    """API pro widget aktuální zakázky."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.permissions import require_permission
    # Worker může vidět svou aktuální zakázku
    
    db = get_db()
    job = db.execute(
        "SELECT * FROM jobs WHERE status NOT LIKE 'dokon%' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    
    if job:
        return jsonify({'ok': True, 'job': dict(job)})
    else:
        return jsonify({'ok': True, 'job': None})

@app.route('/api/widgets/my-tasks')
def api_widget_my_tasks():
    """API pro widget moje úkoly."""
    u, err = require_auth()
    if err:
        return err
    
    db = get_db()
    user_id = session.get('uid')
    employee_id = db.execute(
        "SELECT id FROM employees WHERE user_id = ?", (user_id,)
    ).fetchone()
    
    if employee_id:
        tasks = db.execute("""
            SELECT t.*, j.title as job_name 
            FROM tasks t 
            LEFT JOIN jobs j ON t.job_id = j.id 
            WHERE t.employee_id = ? AND t.status != 'done'
            ORDER BY t.due_date ASC
            LIMIT 10
        """, (employee_id[0],)).fetchall()
        return jsonify({'ok': True, 'tasks': [dict(t) for t in tasks]})
    else:
        return jsonify({'ok': True, 'tasks': []})

@app.route('/api/widgets/notifications')
def api_widget_notifications():
    """API pro widget notifikace."""
    u, err = require_auth()
    if err:
        return err
    
    db = get_db()
    user_id = session.get('uid')
    notifications = db.execute("""
        SELECT * FROM notifications 
        WHERE user_id = ? 
        ORDER BY created_at DESC 
        LIMIT 10
    """, (user_id,)).fetchall()
    
    return jsonify({'ok': True, 'notifications': [dict(n) for n in notifications]})

@app.route('/api/widgets/jobs-risk')
def api_widget_jobs_risk():
    """API pro widget rizikové zakázky."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.permissions import has_permission
    if not has_permission('view_reports'):
        return jsonify({'ok': False, 'error': 'Nedostatečná oprávnění'}), 403
    
    db = get_db()
    # Zjednodušená logika - zakázky s blokery nebo zpožděním
    risky_jobs = db.execute("""
        SELECT j.*, 
               COUNT(DISTINCT i.id) as blocker_count,
               CASE 
                   WHEN COUNT(DISTINCT i.id) > 2 THEN 'high'
                   WHEN COUNT(DISTINCT i.id) > 0 THEN 'medium'
                   ELSE 'low'
               END as risk_level,
               (COUNT(DISTINCT i.id) * 25) as risk_score
        FROM jobs j
        LEFT JOIN issues i ON i.job_id = j.id AND i.status != 'resolved'
        WHERE j.status NOT LIKE 'dokon%'
        GROUP BY j.id
        HAVING blocker_count > 0 OR j.date < date('now')
        ORDER BY risk_score DESC, blocker_count DESC
        LIMIT 5
    """).fetchall()
    
    result = []
    for job in risky_jobs:
        job_dict = dict(job)
        # Získej faktory rizika
        blockers = db.execute("""
            SELECT description FROM issues 
            WHERE job_id = ? AND status != 'resolved'
            LIMIT 3
        """, (job_dict['id'],)).fetchall()
        job_dict['risk_factors'] = [b[0] for b in blockers if b[0]]
        result.append(job_dict)
    
    return jsonify({'ok': True, 'risky_jobs': result})

@app.route('/api/widgets/overdue-jobs')
def api_widget_overdue_jobs():
    """API pro widget zpožděné zakázky."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.permissions import has_permission
    if not has_permission('view_reports'):
        return jsonify({'ok': False, 'error': 'Nedostatečná oprávnění'}), 403
    
    db = get_db()
    overdue_jobs = db.execute("""
        SELECT j.*,
               julianday('now') - julianday(j.date) as days_overdue
        FROM jobs j
        WHERE j.status NOT LIKE 'dokon%'
          AND j.date < date('now')
        ORDER BY j.date ASC
        LIMIT 10
    """).fetchall()
    
    return jsonify({'ok': True, 'overdue_jobs': [dict(j) for j in overdue_jobs]})

@app.route('/api/widgets/team-load')
def api_widget_team_load():
    """API pro widget vytížení týmu."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.permissions import has_permission
    if not has_permission('assign_people'):
        return jsonify({'ok': False, 'error': 'Nedostatečná oprávnění'}), 403
    
    db = get_db()
    today = datetime.now().strftime('%Y-%m-%d')
    
    team_members = db.execute("""
        SELECT e.id, e.name, e.role,
               COALESCE(SUM(t.duration_minutes), 0) / 60.0 as hours_today,
               CASE 
                   WHEN COALESCE(SUM(t.duration_minutes), 0) / 60.0 > 7 THEN 'high'
                   WHEN COALESCE(SUM(t.duration_minutes), 0) / 60.0 > 5 THEN 'normal'
                   ELSE 'low'
               END as load_level,
               CASE 
                   WHEN COALESCE(SUM(t.duration_minutes), 0) / 60.0 > 8 THEN 100
                   ELSE (COALESCE(SUM(t.duration_minutes), 0) / 60.0 / 8.0 * 100)
               END as load_percent
        FROM employees e
        LEFT JOIN timesheets t ON t.employee_id = e.id AND t.date = ?
        WHERE e.status = 'active' OR e.status IS NULL
        GROUP BY e.id
        ORDER BY hours_today DESC
        LIMIT 10
    """, (today,)).fetchall()
    
    result = []
    for member in team_members:
        m = dict(member)
        # Mapování rolí na label
        role_map = {
            'director': 'Ředitel',
            'manager': 'Manažer',
            'lander': 'Vedoucí',
            'worker': 'Pracovník'
        }
        m['role_label'] = role_map.get(m.get('role', 'worker'), 'Pracovník')
        result.append(m)
    
    return jsonify({'ok': True, 'team_members': result})

@app.route('/api/widgets/stock-alerts')
def api_widget_stock_alerts():
    """API pro widget skladové výstrahy."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.permissions import has_permission
    if not has_permission('log_material'):
        return jsonify({'ok': False, 'error': 'Nedostatečná oprávnění'}), 403
    
    db = get_db()
    
    # Zkontroluj, zda existuje warehouse_items tabulka
    try:
        alerts = db.execute("""
            SELECT COALESCE(name, 'Neznámá položka') as material_name, 
                   quantity,
                   unit,
                   CASE 
                       WHEN quantity <= 0 THEN 'critical'
                       WHEN quantity < min_stock THEN 'warning'
                       ELSE 'info'
                   END as severity,
                   'Nízká zásoba' as message
            FROM warehouse_items
            WHERE (quantity <= 0 OR quantity < COALESCE(min_stock, 5))
              AND (status = 'active' OR status IS NULL)
            ORDER BY 
                CASE WHEN quantity <= 0 THEN 0 ELSE 1 END,
                quantity ASC
            LIMIT 10
        """).fetchall()
    except:
        # Fallback pokud tabulka neexistuje
        alerts = []
    
    return jsonify({'ok': True, 'alerts': [dict(a) for a in alerts]})

@app.route('/api/widgets/budget-burn')
def api_widget_budget_burn():
    """API pro widget čerpání rozpočtu."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.permissions import has_permission
    if not has_permission('view_finance'):
        return jsonify({'ok': False, 'error': 'Nedostatečná oprávnění'}), 403
    
    db = get_db()
    
    # Zkontroluj, zda existuje budget sloupec v jobs
    try:
        cols = [r[1] for r in db.execute("PRAGMA table_info(jobs)").fetchall()]
        has_budget = 'budget' in cols
        
        if has_budget:
            budget_data = db.execute("""
                SELECT 
                    SUM(COALESCE(budget, 0)) as total_budget,
                    SUM(COALESCE(labor_cost, 0)) as spent,
                    SUM(COALESCE(budget, 0)) - SUM(COALESCE(labor_cost, 0)) as remaining
                FROM jobs
                WHERE status NOT LIKE 'dokon%'
            """).fetchone()
            
            if budget_data and budget_data[0]:
                total_budget = budget_data[0] or 0
                spent = budget_data[1] or 0
                remaining = budget_data[2] or 0
                percent_spent = (spent / total_budget * 100) if total_budget > 0 else 0
                
                # Top spenders
                top_jobs = db.execute("""
                    SELECT id, title, name, client, labor_cost as spent
                    FROM jobs
                    WHERE status NOT LIKE 'dokon%'
                      AND labor_cost > 0
                    ORDER BY labor_cost DESC
                    LIMIT 5
                """).fetchall()
                
                return jsonify({
                    'ok': True,
                    'budget_data': {
                        'total_budget': total_budget,
                        'spent': spent,
                        'remaining': remaining,
                        'percent_spent': round(percent_spent, 1),
                        'jobs': [dict(j) for j in top_jobs]
                    }
                })
    except Exception as e:
        print(f"[BUDGET] Error: {e}")
    
    return jsonify({'ok': True, 'budget_data': None})

# ============================================================
# QUICK ACTIONS API ENDPOINTS
# ============================================================

@app.route('/api/worklogs', methods=['POST'])
def api_worklogs_create():
    """API pro vytvoření worklogu z mobilu (quick log)."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.permissions import has_permission
    if not has_permission('log_work'):
        return jsonify({'ok': False, 'error': 'Nedostatečná oprávnění'}), 403
    
    # Kontrola deduplikace podle X-Event-ID
    event_id = request.headers.get('X-Event-ID')
    if event_id:
        db = get_db()
        existing = db.execute(
            "SELECT result_id FROM processed_events WHERE event_id = ?",
            (event_id,)
        ).fetchone()
        if existing:
            return jsonify({
                'ok': True,
                'status': 'duplicate',
                'id': existing[0],
                'message': 'Worklog již existuje'
            }), 409
    
    try:
        data = request.get_json() or {}
        db = get_db()
        
        job_id = data.get('job_id')
        work_type_id = data.get('work_type_id')
        duration = data.get('duration')  # v minutách
        note = data.get('note', '')
        created_at = data.get('created_at')
        
        if not job_id:
            return jsonify({'ok': False, 'error': 'missing_job_id'}), 400
        if not duration or duration <= 0:
            return jsonify({'ok': False, 'error': 'invalid_duration'}), 400
        
        # Získej employee_id z user_id
        employee = db.execute(
            "SELECT id FROM employees WHERE user_id = ?",
            (user_id,)
        ).fetchone()
        
        if not employee:
            return jsonify({'ok': False, 'error': 'employee_not_found'}), 400
        
        employee_id = employee[0]
        date_str = created_at[:10] if created_at else datetime.now().strftime('%Y-%m-%d')
        hours = duration / 60.0
        
        # Vytvoř worklog
        worklog_id = db.execute("""
            INSERT INTO timesheets (
                user_id, employee_id, job_id, date,
                duration_minutes, hours, work_type, note, activity, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            user_id, employee_id, job_id, date_str,
            duration, hours, 'manual', note, note,
            created_at or datetime.now().isoformat()
        )).lastrowid
        
        # Zaznamenej processed event pro deduplikaci
        if event_id:
            db.execute("""
                INSERT INTO processed_events (event_id, event_type, result_id, created_at)
                VALUES (?, 'worklog_create', ?, datetime('now'))
            """, (event_id, worklog_id))
        
        db.commit()
        
        return jsonify({
            'ok': True,
            'id': worklog_id,
            'status': 'created',
            'message': 'Worklog vytvořen'
        }), 201
        
    except Exception as e:
        print(f"[WORKLOG] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/photos', methods=['POST'])
def api_photos_create():
    """API pro nahrání fotky z mobilu."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.permissions import has_permission
    if not has_permission('add_photo'):
        return jsonify({'ok': False, 'error': 'Nedostatečná oprávnění'}), 403
    
    # Kontrola deduplikace
    event_id = request.headers.get('X-Event-ID')
    if event_id:
        db = get_db()
        existing = db.execute(
            "SELECT result_id FROM processed_events WHERE event_id = ?",
            (event_id,)
        ).fetchone()
        if existing:
            return jsonify({
                'ok': True,
                'status': 'duplicate',
                'id': existing[0]
            }), 409
    
    try:
        data = request.get_json() or {}
        db = get_db()
        
        job_id = data.get('job_id')
        image_data = data.get('image_data')  # base64
        tag = data.get('tag', 'progress')
        created_at = data.get('created_at')
        
        if not job_id:
            return jsonify({'ok': False, 'error': 'missing_job_id'}), 400
        if not image_data:
            return jsonify({'ok': False, 'error': 'missing_image_data'}), 400
        
        # Ulož obrázek
        import base64
        import os
        from datetime import datetime
        
        # Dekóduj base64
        if ',' in image_data:
            image_data = image_data.split(',')[1]
        
        image_bytes = base64.b64decode(image_data)
        
        # Vytvoř složku pro uploads pokud neexistuje
        upload_dir = os.path.join(UPLOAD_DIR, 'photos', str(job_id))
        os.makedirs(upload_dir, exist_ok=True)
        
        # Vytvoř název souboru
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"photo_{job_id}_{timestamp}.jpg"
        filepath = os.path.join(upload_dir, filename)
        
        # Ulož soubor
        with open(filepath, 'wb') as f:
            f.write(image_bytes)
        
        # Vytvoř attachment záznam
        attachment_id = db.execute("""
            INSERT INTO attachments (
                entity_type, entity_id, filename, filepath,
                filesize, mimetype, uploaded_by, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            'job', job_id, filename, filepath,
            len(image_bytes), 'image/jpeg', user_id,
            created_at or datetime.now().isoformat()
        )).lastrowid
        
        # Aktualizuj worklog s photo_url pokud existuje poslední worklog
        db.execute("""
            UPDATE timesheets 
            SET photo_url = ?
            WHERE job_id = ? AND employee_id = (
                SELECT id FROM employees WHERE user_id = ?
            )
            ORDER BY created_at DESC LIMIT 1
        """, (filepath, job_id, user_id))
        
        # Zaznamenej processed event
        if event_id:
            db.execute("""
                INSERT INTO processed_events (event_id, event_type, result_id, created_at)
                VALUES (?, 'photo_add', ?, datetime('now'))
            """, (event_id, attachment_id))
        
        db.commit()
        
        return jsonify({
            'ok': True,
            'id': attachment_id,
            'photo_url': filepath,
            'status': 'created'
        }), 201
        
    except Exception as e:
        print(f"[PHOTO] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/materials/use', methods=['POST'])
def api_materials_use():
    """API pro odepsání materiálu z mobilu."""
    u, err = require_auth()
    if err:
        return err
    
    from app.utils.permissions import has_permission
    if not has_permission('log_material'):
        return jsonify({'ok': False, 'error': 'Nedostatečná oprávnění'}), 403
    
    # Kontrola deduplikace
    event_id = request.headers.get('X-Event-ID')
    if event_id:
        db = get_db()
        existing = db.execute(
            "SELECT result_id FROM processed_events WHERE event_id = ?",
            (event_id,)
        ).fetchone()
        if existing:
            return jsonify({
                'ok': True,
                'status': 'duplicate',
                'id': existing[0]
            }), 409
    
    try:
        data = request.get_json() or {}
        db = get_db()
        
        material_id = data.get('material_id')
        quantity = data.get('quantity', 1)
        job_id = data.get('job_id')
        
        if not material_id:
            return jsonify({'ok': False, 'error': 'missing_material_id'}), 400
        if not job_id:
            return jsonify({'ok': False, 'error': 'missing_job_id'}), 400
        
        # Zkontroluj, zda existuje warehouse_items tabulka
        try:
            # Získej aktuální množství
            material = db.execute(
                "SELECT quantity, unit FROM warehouse_items WHERE id = ?",
                (material_id,)
            ).fetchone()
            
            if not material:
                return jsonify({'ok': False, 'error': 'material_not_found'}), 404
            
            current_qty = material[0] or 0
            new_qty = max(0, current_qty - quantity)
            
            # Aktualizuj množství
            db.execute("""
                UPDATE warehouse_items 
                SET quantity = ?, updated_at = datetime('now')
                WHERE id = ?
            """, (new_qty, material_id))
            
            # Zaznamenej do job_materials
            db.execute("""
                INSERT INTO job_materials (job_id, name, qty, unit)
                SELECT ?, name, ?, unit
                FROM warehouse_items
                WHERE id = ?
            """, (job_id, quantity, material_id))
            
            # Zaznamenej processed event
            if event_id:
                db.execute("""
                    INSERT INTO processed_events (event_id, event_type, result_id, created_at)
                    VALUES (?, 'material_use', ?, datetime('now'))
                """, (event_id, material_id))
            
            db.commit()
            
            return jsonify({
                'ok': True,
                'material_id': material_id,
                'quantity_used': quantity,
                'remaining': new_qty,
                'status': 'deducted'
            }), 200
            
        except Exception as e:
            print(f"[MATERIAL] Error: {e}")
            # Fallback - jen zaznamenej do job_materials
            db.execute("""
                INSERT INTO job_materials (job_id, name, qty, unit)
                VALUES (?, ?, ?, 'ks')
            """, (job_id, f"Material #{material_id}", quantity))
            db.commit()
            
            return jsonify({
                'ok': True,
                'status': 'recorded',
                'message': 'Materiál zaznamenán (sklad nepodporován)'
            }), 200
            
    except Exception as e:
        print(f"[MATERIAL] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/blockers', methods=['POST'])
def api_blockers_create():
    """API pro vytvoření blockeru/problému z mobilu."""
    # Povolíme i bez autentizace pro demo
    user_id = session.get('uid')
    if not user_id:
        return jsonify({'ok': False, 'error': 'unauthorized', 'queued': True}), 202
    
    u, err = require_auth()
    if err:
        return err
    
    # Kontrola deduplikace
    event_id = request.headers.get('X-Event-ID')
    if event_id:
        db = get_db()
        existing = db.execute(
            "SELECT result_id FROM processed_events WHERE event_id = ?",
            (event_id,)
        ).fetchone()
        if existing:
            return jsonify({
                'ok': True,
                'status': 'duplicate',
                'id': existing[0]
            }), 409
    
    try:
        data = request.get_json() or {}
        db = get_db()
        
        job_id = data.get('job_id')
        blocker_type = data.get('type', 'other')
        description = data.get('description', '')
        photo_data = data.get('photo_data')  # base64
        
        if not job_id:
            return jsonify({'ok': False, 'error': 'missing_job_id'}), 400
        if not description:
            return jsonify({'ok': False, 'error': 'missing_description'}), 400
        
        # Získej employee_id
        employee = db.execute(
            "SELECT id FROM employees WHERE user_id = ?",
            (user_id,)
        ).fetchone()
        
        employee_id = employee[0] if employee else None
        
        # Vytvoř issue/blocker
        blocker_id = db.execute("""
            INSERT INTO issues (
                job_id, title, description, type, status,
                severity, assigned_to, created_by, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
        """, (
            job_id,
            f"Blocker: {blocker_type}",
            description,
            'blocker',
            'open',
            'high',
            employee_id,
            employee_id
        )).lastrowid
        
        # Pokud je foto, ulož ho jako attachment
        if photo_data:
            try:
                import base64
                import os
                
                if ',' in photo_data:
                    photo_data = photo_data.split(',')[1]
                
                image_bytes = base64.b64decode(photo_data)
                upload_dir = os.path.join(UPLOAD_DIR, 'blockers', str(blocker_id))
                os.makedirs(upload_dir, exist_ok=True)
                
                filename = f"blocker_{blocker_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                filepath = os.path.join(upload_dir, filename)
                
                with open(filepath, 'wb') as f:
                    f.write(image_bytes)
                
                db.execute("""
                    INSERT INTO attachments (
                        entity_type, entity_id, filename, filepath,
                        filesize, mimetype, uploaded_by, created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
                """, (
                    'issue', blocker_id, filename, filepath,
                    len(image_bytes), 'image/jpeg', user_id
                ))
            except Exception as e:
                print(f"[BLOCKER] Error saving photo: {e}")
        
        # Zaznamenej processed event
        if event_id:
            db.execute("""
                INSERT INTO processed_events (event_id, event_type, result_id, created_at)
                VALUES (?, 'blocker_create', ?, datetime('now'))
            """, (event_id, blocker_id))
        
        db.commit()
        
        # Vytvoř notifikaci pro managera
        try:
            managers = db.execute("""
                SELECT u.id FROM users u
                WHERE u.role IN ('director', 'manager', 'admin')
                  AND u.active = 1
            """).fetchall()
            
            for manager in managers:
                create_notification(
                    user_id=manager[0],
                    kind='warning',
                    title='Nový blocker',
                    body=f'Blocker na zakázce #{job_id}: {description[:50]}',
                    entity_type='issue',
                    entity_id=blocker_id
                )
        except Exception as e:
            print(f"[BLOCKER] Error creating notifications: {e}")
        
        return jsonify({
            'ok': True,
            'id': blocker_id,
            'status': 'created',
            'message': 'Blocker vytvořen'
        }), 201
        
    except Exception as e:
        print(f"[BLOCKER] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/materials/search', methods=['GET'])
def api_materials_search():
    """API pro vyhledávání materiálu."""
    u, err = require_auth()
    if err:
        return err
    
    query = request.args.get('q', '').strip()
    if len(query) < 2:
        return jsonify({'ok': True, 'items': []})
    
    db = get_db()
    
    try:
        # Zkus warehouse_items
        materials = db.execute("""
            SELECT id, name, quantity, unit
            FROM warehouse_items
            WHERE (name LIKE ? OR name LIKE ?)
              AND (status = 'active' OR status IS NULL)
            ORDER BY name
            LIMIT 20
        """, (f'%{query}%', f'{query}%')).fetchall()
        
        return jsonify({
            'ok': True,
            'items': [dict(m) for m in materials]
        })
    except:
        # Fallback - vrať prázdný seznam
        return jsonify({'ok': True, 'items': []})

# ============================================================
# OFFLINE QUEUE API ENDPOINTS
# ============================================================

@app.route('/api/offline/queue', methods=['POST'])
def api_offline_queue_sync():
    """API pro synchronizaci offline queue - zpracování více eventů najednou."""
    u, err = require_auth()
    if err:
        return err
    
    try:
        data = request.get_json() or {}
        events = data.get('events', [])
        
        if not events:
            return jsonify({'ok': True, 'processed': [], 'failed': []})
        
        db = get_db()
        processed = []
        failed = []
        
        for event in events:
            event_id = event.get('id')
            event_type = event.get('type')
            event_data = event.get('data', {})
            
            if not event_id or not event_type:
                failed.append({
                    'id': event_id,
                    'error': 'missing_id_or_type'
                })
                continue
            
            # Zkontroluj deduplikaci
            existing = db.execute(
                "SELECT result_id FROM processed_events WHERE event_id = ?",
                (event_id,)
            ).fetchone()
            
            if existing:
                processed.append({
                    'id': event_id,
                    'status': 'duplicate',
                    'result_id': existing[0]
                })
                continue
            
            # Zpracuj event podle typu
            try:
                result_id = None
                
                if event_type == 'worklog_create':
                    # Vytvoř worklog
                    job_id = event_data.get('job_id')
                    duration = event_data.get('duration')
                    work_type_id = event_data.get('work_type_id')
                    note = event_data.get('note', '')
                    created_at = event_data.get('created_at')
                    
                    if job_id and duration:
                        user_id = session.get('uid')
                        employee = db.execute(
                            "SELECT id FROM employees WHERE user_id = ?",
                            (user_id,)
                        ).fetchone()
                        
                        if employee:
                            employee_id = employee[0]
                            date_str = created_at[:10] if created_at else datetime.now().strftime('%Y-%m-%d')
                            hours = duration / 60.0
                            
                            result_id = db.execute("""
                                INSERT INTO timesheets (
                                    user_id, employee_id, job_id, date,
                                    duration_minutes, hours, work_type, note, activity, created_at
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (
                                user_id, employee_id, job_id, date_str,
                                duration, hours, 'manual', note, note,
                                created_at or datetime.now().isoformat()
                            )).lastrowid
                            
                            db.commit()
                
                elif event_type == 'photo_add':
                    # Vytvoř photo attachment
                    job_id = event_data.get('job_id')
                    image_data = event_data.get('image_data')
                    
                    if job_id and image_data:
                        user_id = session.get('uid')
                        import base64
                        
                        if ',' in image_data:
                            image_data = image_data.split(',')[1]
                        
                        image_bytes = base64.b64decode(image_data)
                        upload_dir = os.path.join(UPLOAD_DIR, 'photos', str(job_id))
                        os.makedirs(upload_dir, exist_ok=True)
                        
                        filename = f"photo_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                        filepath = os.path.join(upload_dir, filename)
                        
                        with open(filepath, 'wb') as f:
                            f.write(image_bytes)
                        
                        result_id = db.execute("""
                            INSERT INTO attachments (
                                entity_type, entity_id, filename, filepath,
                                filesize, mimetype, uploaded_by, created_at
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            'job', job_id, filename, filepath,
                            len(image_bytes), 'image/jpeg', user_id,
                            datetime.now().isoformat()
                        )).lastrowid
                        
                        db.commit()
                
                elif event_type == 'material_use':
                    # Odepiš materiál
                    material_id = event_data.get('material_id')
                    quantity = event_data.get('quantity', 1)
                    job_id = event_data.get('job_id')
                    
                    if material_id and job_id:
                        try:
                            material = db.execute(
                                "SELECT quantity, unit FROM warehouse_items WHERE id = ?",
                                (material_id,)
                            ).fetchone()
                            
                            if material:
                                current_qty = material[0] or 0
                                new_qty = max(0, current_qty - quantity)
                                
                                db.execute("""
                                    UPDATE warehouse_items 
                                    SET quantity = ?, updated_at = datetime('now')
                                    WHERE id = ?
                                """, (new_qty, material_id))
                                
                                db.execute("""
                                    INSERT INTO job_materials (job_id, name, qty, unit)
                                    SELECT ?, name, ?, unit
                                    FROM warehouse_items
                                    WHERE id = ?
                                """, (job_id, quantity, material_id))
                                
                                result_id = material_id
                                db.commit()
                        except Exception as e:
                            print(f"[QUEUE] Material error: {e}")
                
                elif event_type == 'blocker_create':
                    # Vytvoř blocker
                    job_id = event_data.get('job_id')
                    blocker_type = event_data.get('type', 'other')
                    description = event_data.get('description', '')
                    
                    if job_id and description:
                        user_id = session.get('uid')
                        employee = db.execute(
                            "SELECT id FROM employees WHERE user_id = ?",
                            (user_id,)
                        ).fetchone()
                        
                        employee_id = employee[0] if employee else None
                        
                        result_id = db.execute("""
                            INSERT INTO issues (
                                job_id, title, description, type, status,
                                severity, assigned_to, created_by, created_at
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, datetime('now'))
                        """, (
                            job_id,
                            f"Blocker: {blocker_type}",
                            description,
                            'blocker',
                            'open',
                            'high',
                            employee_id,
                            employee_id
                        )).lastrowid
                        
                        db.commit()
                
                # Zaznamenej processed event
                if result_id:
                    db.execute("""
                        INSERT INTO processed_events (event_id, event_type, result_id, created_at)
                        VALUES (?, ?, ?, datetime('now'))
                    """, (event_id, event_type, result_id))
                    db.commit()
                    
                    processed.append({
                        'id': event_id,
                        'status': 'success',
                        'result_id': result_id
                    })
                else:
                    failed.append({
                        'id': event_id,
                        'error': 'processing_failed'
                    })
                    
            except Exception as e:
                print(f"[QUEUE] Error processing event {event_id}: {e}")
                import traceback
                traceback.print_exc()
                failed.append({
                    'id': event_id,
                    'error': str(e)
                })
        
        return jsonify({
            'ok': True,
            'processed': processed,
            'failed': failed,
            'total': len(events),
            'success_count': len(processed),
            'failed_count': len(failed)
        })
        
    except Exception as e:
        print(f"[QUEUE] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/offline/status', methods=['GET'])
def api_offline_status():
    """API pro kontrolu offline statusu a synchronizace."""
    u, err = require_auth()
    if err:
        return err
    
    # Zkontroluj, zda je server online
    db = get_db()
    
    try:
        # Test DB připojení
        db.execute("SELECT 1").fetchone()
        
        return jsonify({
            'ok': True,
            'online': True,
            'server_time': datetime.now().isoformat(),
            'message': 'Server je online'
        })
    except Exception as e:
        return jsonify({
            'ok': False,
            'online': False,
            'error': str(e),
            'message': 'Server není dostupný'
        }), 503

# ============================================================
# SMART PLANNER API
# ============================================================

@app.route('/api/planner/morning')
def api_morning_planner():
    """Get smart morning planning data."""
    user, err = require_auth()
    if err:
        return err
    
    try:
        db = get_db()
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Today's tasks
        tasks = db.execute('''
            SELECT t.*, j.client as job_name
            FROM tasks t
            LEFT JOIN jobs j ON t.job_id = j.id
            WHERE t.deadline = ? AND t.status != 'done'
            ORDER BY t.priority DESC
        ''', [today]).fetchall()
        
        # Urgent tasks
        urgent = db.execute('''
            SELECT t.*, j.client as job_name
            FROM tasks t
            LEFT JOIN jobs j ON t.job_id = j.id
            WHERE t.priority = 'urgent' AND t.status != 'done'
        ''').fetchall()
        
        # Active jobs
        jobs = db.execute('''
            SELECT * FROM jobs 
            WHERE status IN ('active', 'in_progress')
        ''').fetchall()
        
        # Overdue
        overdue = db.execute('''
            SELECT t.*, j.client as job_name
            FROM tasks t
            LEFT JOIN jobs j ON t.job_id = j.id
            WHERE t.deadline < ? AND t.status != 'done'
        ''', [today]).fetchall()
        
        return jsonify({
            'today_tasks': [dict(t) for t in tasks],
            'urgent_tasks': [dict(t) for t in urgent],
            'active_jobs': [dict(j) for j in jobs],
            'overdue_tasks': [dict(t) for t in overdue]
        })
    except Exception as e:
        print(f"[ERROR] api_morning_planner: {e}")
        return jsonify({'error': str(e)}), 500


# Template filter pro event_type_label
@app.template_filter('event_type_label')
def event_type_label_filter(event_type):
    """Jinja filter pro zobrazení labelu typu eventu."""
    labels = {
        'worklog_create': 'Zápis práce',
        'photo_add': 'Fotografie',
        'material_use': 'Výdej materiálu',
        'blocker_create': 'Nahlášení problému'
    }
    return labels.get(event_type, event_type)

if __name__ == "__main__":
    # Pro lokální vývoj použij 127.0.0.1, pro Render použij 0.0.0.0
    is_render = os.environ.get("RENDER") or os.environ.get("RENDER_EXTERNAL_HOSTNAME")
    host = "0.0.0.0" if is_render else "127.0.0.1"
    port = int(os.environ.get("PORT", 5000))
    debug = not is_render  # Debug mode jen lokálně
    
    print(f"[Server] Starting Flask app on {host}:{port} (debug={debug})")
    app.run(host=host, port=port, debug=debug)

